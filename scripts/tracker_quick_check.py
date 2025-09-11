import os
from pathlib import Path
import shutil
import zipfile
import importlib.util
import pandas as pd
from openpyxl import load_workbook


def _load_module():
    root = Path(__file__).resolve().parents[1]
    mod_path = root / "1.1.0_birha.py"
    spec = importlib.util.spec_from_file_location("birha_mod", str(mod_path))
    mod = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    spec.loader.exec_module(mod)
    return mod


def main():
    mod = _load_module()

    tmp = Path("tmp_tracker_quick_check.xlsx").resolve()
    try:
        if tmp.exists():
            tmp.unlink()
    except Exception:
        pass

    # 1) create tracker
    mod.ensure_word_tracker(str(tmp))

    # 2) append once
    mod.append_to_word_tracker(
        str(tmp),
        words_rows=[{
            "word": "foo",
            "word_key_norm": "foo",
            "listed_by_user": True,
            "listed_at": pd.Timestamp.now(),
            "sequence_index": 0,
        }],
        progress_rows=[{
            "word": "foo",
            "word_key_norm": "foo",
            "verse": "test verse",
            "selected_for_analysis": True,
            "selected_at": pd.Timestamp.now(),
        }]
    )

    # 3) add a non-spec sheet with a formula, then append again and verify it survives
    wb = load_workbook(str(tmp), keep_vba=True)
    if "Dashboard" not in wb.sheetnames:
        ws = wb.create_sheet("Dashboard")
        ws["A1"].value = "=SUM(1,2,3)"
        ws["B1"].value = "keep-me"
        wb.save(str(tmp))

    # 4) append again
    mod.append_to_word_tracker(
        str(tmp),
        words_rows=[{"word": "bar", "word_key_norm": "bar", "listed_by_user": False, "sequence_index": 1}],
        progress_rows=[{"word": "bar", "word_key_norm": "bar", "verse": "v2", "selected_for_analysis": False}]
    )

    # 5) verify
    wb = load_workbook(str(tmp), keep_vba=True, data_only=False)
    assert "Words" in wb.sheetnames and "Progress" in wb.sheetnames, "missing spec sheets"
    assert "Dashboard" in wb.sheetnames, "non-spec sheet lost"
    assert wb["Dashboard"]["A1"].value and str(wb["Dashboard"]["A1"].value).startswith("="), "formula stripped"
    assert wb["Dashboard"]["B1"].value == "keep-me", "cell value changed"
    # Check headers in spec sheets present at row 1
    words_headers = [c.value for c in wb["Words"][1]]
    progress_headers = [c.value for c in wb["Progress"][1]]
    for required in mod._WORDS_COLUMNS:
        assert required in words_headers, f"Words header missing: {required}"
    for required in mod._PROGRESS_COLUMNS:
        assert required in progress_headers, f"Progress header missing: {required}"

    print("OK: tracker quick check passed")

    # Optional: macro-enabled quick check if a template is available
    macro_tpl = os.getenv("TRACKER_VBA_TEMPLATE") or "macro_template.xlsm"
    tpl_path = Path(macro_tpl)
    if tpl_path.exists():
        xlsm = Path("tmp_tracker_macro_check.xlsm").resolve()
        try:
            if xlsm.exists():
                xlsm.unlink()
        except Exception:
            pass
        shutil.copyfile(str(tpl_path), str(xlsm))

        # ensure twice; verify vbaProject.bin preserved
        mod.ensure_word_tracker(str(xlsm))
        mod.ensure_word_tracker(str(xlsm))
        # check for vba_project presence in the zip
        has_vba = False
        try:
            with zipfile.ZipFile(str(xlsm), 'r') as zf:
                namelist = set(zf.namelist())
                has_vba = any(name.lower().endswith("vbaProject.bin".lower()) or name.lower()=="xl/vbaproject.bin" for name in namelist)
        except Exception:
            has_vba = False
        assert has_vba, "Macro project missing after ensure_word_tracker on .xlsm"
        print("OK: macro-enabled tracker VBA preserved")


if __name__ == "__main__":
    main()
