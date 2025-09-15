import csv
import os
from tkinter import messagebox, scrolledtext
import pandas as pd
import ast
import re
import math
import unicodedata
import pyperclip
import tkinter as tk
import tkinter.font as tkfont
from tkinter import ttk
import threading
from rapidfuzz import fuzz, process
import numpy as np
import textwrap
import webbrowser
import platform
try:
    import ctypes
    from ctypes import wintypes
except Exception:
    ctypes = None
from datetime import datetime, timezone
import subprocess
import json
import webbrowser
from openpyxl import load_workbook
import tempfile
import shutil
import zipfile


# ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬
# GLOBAL HELPER  ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ  build live noun-morphology lookup
# ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬
from functools import lru_cache

# ---------------------------
# Window Manager (maximize/restore/toggle)
# ---------------------------
class WindowManager:
    """Cross-platform window manager for Tk windows.

    - maximize(): resizes to usable screen area.
      On Windows, uses SystemParametersInfoW(SPI_GETWORKAREA) to exclude taskbar.
      Else, falls back to winfo_screenwidth/height.
    - restore(): returns to the previous geometry.
    - toggle(): F11-style toggle between maximize and restore.

    Automatically binds F11 to toggle on the given window.
    """
    def __init__(self, win):
        self.win = win
        self.prev_geometry = None
        self.maximized = False
        # If a client-margin maximize is used, remember the preferred bottom margin
        self._client_margin_px = None
        try:
            self.win.bind("<F11>", lambda e: self.toggle())
        except Exception:
            pass

    def _usable_area(self):
        """Return (x, y, w, h) for the usable work area of the monitor containing this window.

        Windows:
          - Prefer MonitorFromWindow + GetMonitorInfoW for per-monitor work areas (excludes taskbar).
          - Fallback to SPI_GETWORKAREA (primary monitor) if monitor lookup fails.
        Other OS: fall back to full screen dimensions from Tk.
        """
        # Windows per-monitor work area
        try:
            if os.name == 'nt' and ctypes is not None:
                user32 = ctypes.windll.user32

                # Try per-monitor APIs first
                try:
                    MONITOR_DEFAULTTONEAREST = 2

                    class RECT(ctypes.Structure):
                        _fields_ = [
                            ("left", wintypes.LONG),
                            ("top", wintypes.LONG),
                            ("right", wintypes.LONG),
                            ("bottom", wintypes.LONG),
                        ]

                    class MONITORINFO(ctypes.Structure):
                        _fields_ = [
                            ("cbSize", wintypes.DWORD),
                            ("rcMonitor", RECT),
                            ("rcWork", RECT),
                            ("dwFlags", wintypes.DWORD),
                        ]

                    hwnd = int(self.win.winfo_id())
                    hmon = user32.MonitorFromWindow(hwnd, MONITOR_DEFAULTTONEAREST)
                    if hmon:
                        mi = MONITORINFO()
                        mi.cbSize = ctypes.sizeof(MONITORINFO)
                        ok = user32.GetMonitorInfoW(hmon, ctypes.byref(mi))
                        if ok:
                            x = int(mi.rcWork.left)
                            y = int(mi.rcWork.top)
                            w = int(mi.rcWork.right - mi.rcWork.left)
                            h = int(mi.rcWork.bottom - mi.rcWork.top)
                            return x, y, w, h
                except Exception:
                    # Fall back to primary work area
                    pass

                # Primary monitor work area via SPI
                try:
                    SPI_GETWORKAREA = 0x0030
                    rect = RECT()
                    ok = user32.SystemParametersInfoW(SPI_GETWORKAREA, 0, ctypes.byref(rect), 0)
                    if ok:
                        x = int(rect.left)
                        y = int(rect.top)
                        w = int(rect.right - rect.left)
                        h = int(rect.bottom - rect.top)
                        return x, y, w, h
                except Exception:
                    pass
        except Exception:
            pass

        # Fallback: full screen size
        try:
            self.win.update_idletasks()
        except Exception:
            pass
        try:
            w = int(self.win.winfo_screenwidth())
            h = int(self.win.winfo_screenheight())
            return 0, 0, w, h
        except Exception:
            return 0, 0, 1024, 768

    def maximize(self):
        try:
            if not self.maximized:
                try:
                    self.win.update_idletasks()
                    self.prev_geometry = self.win.winfo_geometry()
                except Exception:
                    self.prev_geometry = None
            # A normal maximize clears client-margin preference
            self._client_margin_px = None
            # First try per-monitor usable area geometry
            x, y, w, h = self._usable_area()
            if w and h:
                # On Windows, prefer native zoom after moving to the target monitor to avoid
                # off-by-frame overlaps with taskbar and to respect system window metrics.
                if os.name == 'nt' and ctypes is not None:
                    try:
                        # Move to monitor's work-area origin, then let OS maximize precisely.
                        self.win.geometry(f"+{x}+{y}")
                        self.win.state('zoomed')
                        self.maximized = True
                        return
                    except Exception:
                        # Fallback: compute adjusted client area using AdjustWindowRectEx
                        try:
                            user32 = ctypes.windll.user32
                            GWL_STYLE = -16
                            GWL_EXSTYLE = -20
                            style = user32.GetWindowLongW(int(self.win.winfo_id()), GWL_STYLE)
                            exstyle = user32.GetWindowLongW(int(self.win.winfo_id()), GWL_EXSTYLE)

                            class RECT(ctypes.Structure):
                                _fields_ = [
                                    ("left", wintypes.LONG),
                                    ("top", wintypes.LONG),
                                    ("right", wintypes.LONG),
                                    ("bottom", wintypes.LONG),
                                ]
                            r = RECT()
                            r.left = 0
                            r.top = 0
                            r.right = 0
                            r.bottom = 0
                            ok = user32.AdjustWindowRectEx(ctypes.byref(r), style, False, exstyle)
                            if ok:
                                dw = int(r.right - r.left)
                                dh = int(r.bottom - r.top)
                                cw = max(1, int(w - dw))
                                ch = max(1, int(h - dh))
                                self.win.geometry(f"{cw}x{ch}+{x}+{y}")
                                self.maximized = True
                                return
                        except Exception:
                            pass
                # Non-Windows or if above fails: set geometry to work area
                self.win.geometry(f"{w}x{h}+{x}+{y}")
                self.maximized = True
                return
            # If no geometry computed, fall back to native zoom behavior (often per-monitor)
            try:
                self.win.state('zoomed')
                self.maximized = True
                return
            except Exception:
                pass

        except Exception:
            pass

    def restore(self):
        try:
            # Ensure window is back to normal state so geometry takes effect (Windows ignores
            # geometry changes while in 'zoomed' state)
            try:
                self.win.state('normal')
            except Exception:
                pass
            if self.prev_geometry:
                self.win.geometry(self.prev_geometry)
            self.maximized = False
        except Exception:
            pass

    def toggle(self):
        if self.maximized:
            self.restore()
        else:
            # Preserve taskbar-safe geometry if a client-margin maximize was used before
            try:
                if isinstance(self._client_margin_px, int) and self._client_margin_px > 0:
                    self.maximize_client(bottom_margin_px=self._client_margin_px)
                else:
                    self.maximize()
            except Exception:
                self.maximize()

    def maximize_client(self, bottom_margin_px: int = 48):
        """Like maximize(), but prefers explicit client-geometry sizing to avoid
        edge overlaps (e.g., auto-hide taskbar). Shrinks the available height by
        a small bottom margin so bottom buttons remain visible.

        - Captures prev_geometry before changing size.
        - On Windows, uses per-monitor rcWork and AdjustWindowRectEx to account for
          non-client borders/caption.
        - Elsewhere, falls back to setting geometry directly with the margin.
        """
        try:
            if not self.maximized:
                try:
                    self.win.update_idletasks()
                    self.prev_geometry = self.win.winfo_geometry()
                except Exception:
                    self.prev_geometry = None

            x, y, w, h = self._usable_area()
            if not (w and h):
                # Fall back to native maximize if geometry not available
                return self.maximize()

            # Apply safety margin at the bottom to avoid taskbar overlap
            safe_h = max(100, int(h) - int(bottom_margin_px or 0))

            if os.name == 'nt' and ctypes is not None:
                try:
                    user32 = ctypes.windll.user32
                    GWL_STYLE = -16
                    GWL_EXSTYLE = -20
                    style = user32.GetWindowLongW(int(self.win.winfo_id()), GWL_STYLE)
                    exstyle = user32.GetWindowLongW(int(self.win.winfo_id()), GWL_EXSTYLE)

                    class RECT(ctypes.Structure):
                        _fields_ = [
                            ("left", wintypes.LONG),
                            ("top", wintypes.LONG),
                            ("right", wintypes.LONG),
                            ("bottom", wintypes.LONG),
                        ]
                    r = RECT()
                    ok = user32.AdjustWindowRectEx(ctypes.byref(r), style, False, exstyle)
                    if ok:
                        dw = int(r.right - r.left)
                        dh = int(r.bottom - r.top)
                        cw = max(1, int(w) - dw)
                        ch = max(1, int(safe_h) - dh)
                        self.win.geometry(f"{cw}x{ch}+{x}+{y}")
                        self.maximized = True
                        self._client_margin_px = int(bottom_margin_px or 0)
                        return
                except Exception:
                    pass

            # Non-Windows or failed AdjustWindowRectEx: best-effort geometry
            self.win.geometry(f"{w}x{safe_h}+{x}+{y}")
            self.maximized = True
            self._client_margin_px = int(bottom_margin_px or 0)
        except Exception:
            pass

# ------------------------------------------------------------------
# CSV helper: load Predefined-only keyset for Top Matches filtering
# ------------------------------------------------------------------
def _normalize_headers(headers):
    try:
        return [str(h).replace('\ufeff', '').strip() for h in headers]
    except Exception:
        return list(headers)

@lru_cache(maxsize=8)
def load_predefined_keyset(csv_path: str = "1.1.1_birha.csv") -> set[tuple[str, str, str, str, str]]:
    """Pure helper. Read CSV and return a keyset of
    (Number, Grammar, Gender, WordRoot, Type) for rows where Evaluation == "Predefined".

    - Robust to UTF-8 BOM in headers (e.g., "\ufeffVowel Ending").
    - Tolerates NaN/empty cells; coerces to empty strings.
    - Column name lookup is case-insensitive and trims whitespace.
    - Evaluation comparison is exact match after str.strip().
    """
    try:
        # utf-8-sig handles BOM at file start; we still strip stray BOM from headers explicitly.
        df = pd.read_csv(csv_path, encoding="utf-8-sig")
    except Exception:
        # Fallback if encoding not recognized; keep behavior permissive
        df = pd.read_csv(csv_path)

    df.columns = _normalize_headers(df.columns)
    # Case-insensitive column resolver
    colmap = {str(c).strip().lower(): c for c in df.columns}
    def col(*cands):
        for c in cands:
            k = str(c).strip().lower()
            if k in colmap:
                return colmap[k]
        return None

    eval_col   = col("Evaluation")
    # Prefer Punjabi headers; accept both with/without "options"; fall back to legacy names
    num_col    = col(COL_NUMBER, f"{COL_NUMBER} options", "Number / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ options", "Number / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨", "Number")
    gram_col   = col(COL_GRAMMAR, f"{COL_GRAMMAR} options", "Grammar Case / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¯ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£ options", "Grammar Case / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¯ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£", "Grammar")
    gen_col    = col(COL_GENDER, f"{COL_GENDER} options", "Gender / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â options", "Gender / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â", "Gender")
    root_col   = col("Word Root", "Root")
    type_col   = col("Type", "Word Type")

    if eval_col is None or None in (num_col, gram_col, gen_col, root_col, type_col):
        # Missing expected columns; nothing qualifies
        return set()

    # Normalize values and filter strictly by Evaluation == "Predefined" after strip
    safe = df.copy()
    for c in (eval_col, num_col, gram_col, gen_col, root_col, type_col):
        # Important: fillna before astype(str) so NaN does not become the literal string "nan"
        safe[c] = safe[c].fillna("").astype(str).map(lambda s: s.strip())

    mask = safe[eval_col].map(lambda s: s.strip()) == "Predefined"
    pre = safe.loc[mask, [num_col, gram_col, gen_col, root_col, type_col]]

    keyset: set[tuple[str, str, str, str, str]] = set()
    for _, r in pre.iterrows():
        key = (
            r[num_col] or "",
            r[gram_col] or "",
            r[gen_col] or "",
            r[root_col] or "",
            r[type_col] or "",
        )
        # Already stripped; keep as-is (case-sensitive values preserved)
        keyset.add(tuple(key))
    return keyset

# Debug flag for translation auto-fill matching. Enable via env var BIRHA_DEBUG_AUTOFILL=1
DEBUG_AUTOFILL = False
try:
    DEBUG_AUTOFILL = str(os.getenv('BIRHA_DEBUG_AUTOFILL', '0')).strip().lower() in {'1', 'true', 'yes', 'on'}
except Exception:
    DEBUG_AUTOFILL = False

# ------------------------------------------------------------------
# Shared UI spacing constants
# ------------------------------------------------------------------
# Bottom padding (in pixels) to keep action buttons off the window edge.
# Matches the gap used by user_input_grammar's bottom button row.
BOTTOM_PAD = 46

# Canonical column labels used across Verse and ABW flows
COL_NUMBER  = "Number / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨"
COL_GRAMMAR = "Grammar / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¯ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£"
COL_GENDER  = "Gender / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â"

def _resolve_col(df: pd.DataFrame, *candidates: str) -> str | None:
    """Return the actual column name in df matching any of the candidate labels (case-insensitive)."""
    if df is None or not hasattr(df, 'columns'):
        return None
    colmap = {str(c).strip().lower(): c for c in df.columns}
    for c in candidates:
        k = str(c).strip().lower()
        if k in colmap:
            return colmap[k]
    return None

def _dbg_autofill(self, msg: str):
    try:
        if DEBUG_AUTOFILL or bool(getattr(self, '_debug_autofill', False)):
            print(msg)
    except Exception:
        # Best-effort logging only
        if DEBUG_AUTOFILL:
            try:
                print(msg)
            except Exception:
                pass

# ID used to gate the one-time "What's New" prompt.
# Prefer the latest UI tag from Git if available; otherwise, fall back.
def _compute_whats_new_id():
    try:
        # Get UI tags sorted by create date (newest first)
        result = subprocess.run(
            ["git", "tag", "--list", "ui-*", "--sort=-creatordate"],
            capture_output=True, text=True, check=False
        )
        if result.returncode == 0:
            for line in result.stdout.splitlines():
                tag = line.strip()
                if tag:
                    return tag
    except Exception:
        pass
    return "ui-2025-09-07-cards-layout"

WHATS_NEW_ID = _compute_whats_new_id()

def extract_darpan_translation(text: str) -> str:
    """Extract the Darpan Translation content from a labeled text.

    Rules:
    - Recognize exact labels at start of a line: "Verse:", "Padarth:", "Arth:", "Chhand:", "Bhav:".
    - Discard Verse and Padarth blocks entirely.
    - Concatenate remaining blocks in the order: Arth, Chhand, Bhav (if present).
    - Preserve line breaks but collapse excessive in-line whitespace (spaces/tabs).
    - If none of Arth/Chhand/Bhav are found, return the original text unchanged.

    This function is pure and has no side effects.
    """
    try:
        original = "" if text is None else str(text)
    except Exception:
        original = ""

    labels = ("Verse", "Padarth", "Arth", "Chhand", "Bhav")
    label_re = re.compile(r"^(Verse|Padarth|Arth|Chhand|Bhav):\s*$")
    blocks = {k: [] for k in labels}
    current = None

    for raw_line in original.splitlines():
        m = label_re.match(raw_line.strip())
        if m:
            current = m.group(1)
            continue
        if current is not None:
            blocks[current].append(raw_line)

    def _clean_join(lines):
        cleaned_lines = []
        for ln in lines:
            s = ln.strip()
            if s == "":
                cleaned_lines.append("")
            else:
                s = re.sub(r"[ \t]+", " ", s)
                cleaned_lines.append(s)
        # Trim leading/trailing empties
        first = 0
        last = len(cleaned_lines) - 1
        while first <= last and cleaned_lines[first] == "":
            first += 1
        while last >= first and cleaned_lines[last] == "":
            last -= 1
        if first > last:
            return ""
        return "\n".join(cleaned_lines[first:last + 1])

    parts = []
    for key in ("Arth", "Chhand", "Bhav"):
        block_text = _clean_join(blocks.get(key, []))
        if block_text:
            parts.append(block_text)

    if not parts:
        return original

    return "\n\n".join(parts)

# ------------------------------------------------------------------
# Tracker helpers: 'Assess by Word' Excel (1.1.6 XXXX.xlsx)
# - create, load, append
# - enforce: no nested writers; ExcelFile closed before writing
# - preserve and rewrite non-spec sheets in original order
# ------------------------------------------------------------------

TRACKER_WORDS_SHEET = "Words"
TRACKER_PROGRESS_SHEET = "Progress"

_WORDS_COLUMNS = [
    "word",
    "word_key_norm",
    "listed_by_user",
    "listed_at",
    "selected_for_analysis",
    "selected_at",
    "analysis_started",
    "analysis_started_at",
    "analysis_completed",
    "analysis_completed_at",
    "sequence_index",
    "notes",
]

_PROGRESS_COLUMNS = [
    "word",
    "word_key_norm",
    "word_index",
    "verse",
    "page_number",
    "selected_for_analysis",
    "selected_at",
    "status",
    "completed_at",
    "reanalyzed_count",
    "last_reanalyzed_at",
]


def _empty_tracker_frames():
    words_df = pd.DataFrame(columns=_WORDS_COLUMNS)
    progress_df = pd.DataFrame(columns=_PROGRESS_COLUMNS)
    return words_df, progress_df


def _coerce_dt(val):
    """Best-effort parse for datetime-like values with Excel-safe semantics.

    Rules:
    - None/empty -> None
    - If datetime-like and tz-aware -> convert to UTC and drop tzinfo (Excel can't store tz).
    - If datetime-like and naive -> return as-is.
    - Else try pd.to_datetime; if tz-aware result, convert to UTC and drop tz; return python datetime.
    - On failure, return original value.
    """
    # Empty
    if val is None or val == "":
        return None

    # pandas Timestamp handling
    if isinstance(val, pd.Timestamp):
        if val.tz is not None:
            try:
                return val.tz_convert('UTC').tz_localize(None).to_pydatetime()
            except Exception:
                try:
                    # Fallback: drop tz without convert (may shift interpretation)
                    return val.tz_localize(None).to_pydatetime()
                except Exception:
                    return val
        # Naive Timestamp -> python datetime
        try:
            return val.to_pydatetime()
        except Exception:
            return val

    # Python datetime handling
    if isinstance(val, datetime):
        try:
            aware = (val.tzinfo is not None) and (val.tzinfo.utcoffset(val) is not None)
        except Exception:
            aware = val.tzinfo is not None
        if aware:
            try:
                return val.astimezone(timezone.utc).replace(tzinfo=None)
            except Exception:
                try:
                    return val.replace(tzinfo=None)
                except Exception:
                    return val
        return val

    # Generic parse
    try:
        dt = pd.to_datetime(val, utc=False)
        if isinstance(dt, pd.Timestamp):
            if dt.tz is not None:
                try:
                    dt = dt.tz_convert('UTC').tz_localize(None)
                except Exception:
                    try:
                        dt = dt.tz_localize(None)
                    except Exception:
                        return val
            try:
                return dt.to_pydatetime()
            except Exception:
                return dt
        return dt
    except Exception:
        return val


def _ensure_columns(df: pd.DataFrame, required: list[str]) -> pd.DataFrame:
    # Add missing columns as empty; keep existing order for present columns
    for col in required:
        if col not in df.columns:
            df[col] = pd.Series([None] * len(df))
    # Reorder to required first, then any extras in original relative order
    extras = [c for c in df.columns if c not in required]
    return df[required + extras]


def ensure_word_tracker(
    tracker_path: str,
    words_sheet: str = TRACKER_WORDS_SHEET,
    progress_sheet: str = TRACKER_PROGRESS_SHEET,
):
    """Ensure a tracker workbook exists with 'Words' and 'Progress' sheets.

    Non-spec sheets are NEVER reserialized. If the file exists, we only
    modify the Words/Progress sheets and leave all other sheets intact.

    Constraints:
    - Single ExcelWriter per save.
    - Close pd.ExcelFile before writing.
    - Preserve original sheet order (spec sheets keep their position; new ones append).
    """
    def _is_xlsm(path: str) -> bool:
        try:
            return str(path).lower().endswith(".xlsm")
        except Exception:
            return False

    try:
        if not os.path.exists(tracker_path):
            # Create fresh workbook with the two spec sheets only (atomic write)
            words_df, prog_df = _empty_tracker_frames()
            # Coerce datetime-like columns
            for c in ["listed_at", "selected_at", "analysis_started_at", "analysis_completed_at"]:
                if c in words_df.columns:
                    words_df[c] = words_df[c].map(_coerce_dt)
            for c in ["selected_at", "completed_at", "last_reanalyzed_at"]:
                if c in prog_df.columns:
                    prog_df[c] = prog_df[c].map(_coerce_dt)

            base_dir = os.path.dirname(os.path.abspath(tracker_path)) or os.getcwd()
            ext = os.path.splitext(tracker_path)[1] or ".xlsx"
            with tempfile.NamedTemporaryFile(prefix=".tmp_tracker_", suffix=ext, dir=base_dir, delete=False) as tf:
                tmp_path = tf.name
            try:
                # Do NOT pass keep_vba when creating a new workbook; openpyxl does not support this.
                engine_kwargs = None
                with pd.ExcelWriter(tmp_path, engine="openpyxl", mode="w", engine_kwargs=engine_kwargs) as writer:
                    words_df.to_excel(writer, index=False, sheet_name=words_sheet)
                    prog_df.to_excel(writer, index=False, sheet_name=progress_sheet)
                # Validate temp file before replacing
                if not zipfile.is_zipfile(tmp_path):
                    raise RuntimeError("Tracker temp file is not a valid XLSX/XL file (zip)")
                _ = load_workbook(tmp_path, keep_vba=_is_xlsm(tracker_path))
                os.replace(tmp_path, tracker_path)
                return True
            except Exception as e:
                try:
                    os.unlink(tmp_path)
                except Exception:
                    pass
                print(f"ensure_word_tracker: aborted create due to error: {e}")
                return False

        # Load existing data for spec sheets only, using a context-managed ExcelFile
        with pd.ExcelFile(tracker_path, engine="openpyxl") as xls:
            names = list(xls.sheet_names)
            if words_sheet in names:
                words_df = pd.read_excel(xls, sheet_name=words_sheet)
            else:
                words_df = pd.DataFrame(columns=_WORDS_COLUMNS)
            if progress_sheet in names:
                prog_df = pd.read_excel(xls, sheet_name=progress_sheet)
            else:
                prog_df = pd.DataFrame(columns=_PROGRESS_COLUMNS)

        words_df = _ensure_columns(words_df, _WORDS_COLUMNS)
        prog_df = _ensure_columns(prog_df, _PROGRESS_COLUMNS)

        # Normalize datetime-like columns prior to write
        for c in ["listed_at", "selected_at", "analysis_started_at", "analysis_completed_at"]:
            if c in words_df.columns:
                words_df[c] = words_df[c].map(_coerce_dt)
        for c in ["selected_at", "completed_at", "last_reanalyzed_at"]:
            if c in prog_df.columns:
                prog_df[c] = prog_df[c].map(_coerce_dt)

        # Atomic in-place update: copy original, refresh spec sheets in-place without recreating them, then replace
        base_dir = os.path.dirname(os.path.abspath(tracker_path)) or os.getcwd()
        ext = os.path.splitext(tracker_path)[1] or ".xlsx"
        with tempfile.NamedTemporaryFile(prefix=".tmp_tracker_", suffix=ext, dir=base_dir, delete=False) as tf:
            tmp_path = tf.name
        try:
            shutil.copyfile(tracker_path, tmp_path)
            # Pre-clear existing sheet contents without recreating sheets to preserve formatting/VBA
            try:
                wb = load_workbook(tmp_path, keep_vba=_is_xlsm(tracker_path))
                if words_sheet not in wb.sheetnames:
                    wb.create_sheet(title=words_sheet)
                if progress_sheet not in wb.sheetnames:
                    wb.create_sheet(title=progress_sheet)
                try:
                    ws = wb[words_sheet]
                    ws.delete_rows(1, ws.max_row or 1)
                except Exception:
                    pass
                try:
                    ws = wb[progress_sheet]
                    ws.delete_rows(1, ws.max_row or 1)
                except Exception:
                    pass
                wb.save(tmp_path)
            except Exception:
                # If clearing fails, continue; overlay will still update data
                pass

            engine_kwargs = {"keep_vba": True} if _is_xlsm(tracker_path) else None
            # Overlay into existing sheets to avoid replacing sheet objects
            with pd.ExcelWriter(
                tmp_path,
                engine="openpyxl",
                mode="a",
                if_sheet_exists="overlay",
                engine_kwargs=engine_kwargs,
            ) as writer:
                words_df.to_excel(writer, index=False, sheet_name=words_sheet)
                prog_df.to_excel(writer, index=False, sheet_name=progress_sheet)

            if not zipfile.is_zipfile(tmp_path):
                raise RuntimeError("Tracker temp file is not a valid XLSX/XL file (zip)")
            _ = load_workbook(tmp_path, keep_vba=_is_xlsm(tracker_path))
            os.replace(tmp_path, tracker_path)
            return True
        except Exception as e:
            try:
                os.unlink(tmp_path)
            except Exception:
                pass
            print(f"ensure_word_tracker: aborted update due to error: {e}")
            return False
    except Exception as e:
        print(f"ensure_word_tracker: unexpected error: {e}")
        return False


def load_word_tracker(
    tracker_path: str,
    words_sheet: str = TRACKER_WORDS_SHEET,
    progress_sheet: str = TRACKER_PROGRESS_SHEET,
):
    """Load tracker workbook and return (words_df, progress_df, other_sheets_ordered).

    Uses a context-managed ExcelFile to ensure clean close before any write.
    """
    if not os.path.exists(tracker_path):
        # Create a new in-memory structure if missing
        words_df, prog_df = _empty_tracker_frames()
        return words_df, prog_df, []

    with pd.ExcelFile(tracker_path, engine="openpyxl") as xls:
        names = list(xls.sheet_names)
        words_df = (
            pd.read_excel(xls, sheet_name=words_sheet)
            if words_sheet in names else pd.DataFrame(columns=_WORDS_COLUMNS)
        )
        words_df = _ensure_columns(words_df, _WORDS_COLUMNS)
        prog_df = (
            pd.read_excel(xls, sheet_name=progress_sheet)
            if progress_sheet in names else pd.DataFrame(columns=_PROGRESS_COLUMNS)
        )
        prog_df = _ensure_columns(prog_df, _PROGRESS_COLUMNS)
        others = [n for n in names if n not in {words_sheet, progress_sheet}]
    return words_df, prog_df, others


def _save_tracker(
    tracker_path: str,
    words_df: pd.DataFrame,
    progress_df: pd.DataFrame,
    others: list[str],
    words_sheet: str = TRACKER_WORDS_SHEET,
    progress_sheet: str = TRACKER_PROGRESS_SHEET,
):
    """Save tracker sheets without touching non-spec sheets.

    - Uses a single ExcelWriter bound to an openpyxl workbook.
    - Does not reserialize other sheets; only updates Words/Progress.
    - Preserves sheet order by clearing content within existing spec sheets
      or appending new spec sheets to the end if missing.
    """
    # Coerce datetime-like columns just before writing
    for c in ["listed_at", "selected_at", "analysis_started_at", "analysis_completed_at"]:
        if c in words_df.columns:
            words_df[c] = words_df[c].map(_coerce_dt)
    for c in ["selected_at", "completed_at", "last_reanalyzed_at"]:
        if c in progress_df.columns:
            progress_df[c] = progress_df[c].map(_coerce_dt)

    def _is_xlsm(path: str) -> bool:
        try:
            return str(path).lower().endswith(".xlsm")
        except Exception:
            return False

    # Atomic write: work on a temp file in the same directory, then replace
    base_dir = os.path.dirname(os.path.abspath(tracker_path)) or os.getcwd()
    ext = os.path.splitext(tracker_path)[1] or ".xlsx"
    with tempfile.NamedTemporaryFile(prefix=".tmp_tracker_", suffix=ext, dir=base_dir, delete=False) as tf:
        tmp_path = tf.name

    try:
        if os.path.exists(tracker_path):
            shutil.copyfile(tracker_path, tmp_path)
            mode = "a"
        else:
            mode = "w"

        if mode == "a":
            engine_kwargs = {"keep_vba": True} if _is_xlsm(tracker_path) else None
            # Clear existing content without recreating sheets to preserve formatting/VBA
            try:
                wb = load_workbook(tmp_path, keep_vba=_is_xlsm(tracker_path))
                if words_sheet not in wb.sheetnames:
                    wb.create_sheet(title=words_sheet)
                if progress_sheet not in wb.sheetnames:
                    wb.create_sheet(title=progress_sheet)
                try:
                    ws = wb[words_sheet]
                    ws.delete_rows(1, ws.max_row or 1)
                except Exception:
                    pass
                try:
                    ws = wb[progress_sheet]
                    ws.delete_rows(1, ws.max_row or 1)
                except Exception:
                    pass
                wb.save(tmp_path)
            except Exception:
                pass

            # Overlay into existing sheets so sheet objects remain intact
            with pd.ExcelWriter(
                tmp_path,
                engine="openpyxl",
                mode="a",
                if_sheet_exists="overlay",
                engine_kwargs=engine_kwargs,
            ) as writer:
                words_df.to_excel(writer, index=False, sheet_name=words_sheet)
                progress_df.to_excel(writer, index=False, sheet_name=progress_sheet)
        else:
            # New workbook: do NOT pass keep_vba; openpyxl cannot create macros from scratch.
            with pd.ExcelWriter(
                tmp_path,
                engine="openpyxl",
                mode="w",
            ) as writer:
                words_df.to_excel(writer, index=False, sheet_name=words_sheet)
                progress_df.to_excel(writer, index=False, sheet_name=progress_sheet)

        if not zipfile.is_zipfile(tmp_path):
            raise RuntimeError("Tracker temp file is not a valid XLSX/XL file (zip)")
        _ = load_workbook(tmp_path, keep_vba=_is_xlsm(tracker_path))
        os.replace(tmp_path, tracker_path)
    except Exception as e:
        try:
            os.unlink(tmp_path)
        except Exception:
            pass
        print(f"_save_tracker: aborted save due to error: {e}")
        return


def append_to_word_tracker(
    tracker_path: str,
    words_rows: list[dict] | pd.DataFrame | None = None,
    progress_rows: list[dict] | pd.DataFrame | None = None,
    words_sheet: str = TRACKER_WORDS_SHEET,
    progress_sheet: str = TRACKER_PROGRESS_SHEET,
):
    """Append rows to Words and/or Progress sheets.

    - Ensures the workbook and sheets exist (creating them if necessary).
    - Loads the current frames using a context-managed ExcelFile.
    - Appends the provided rows and writes back using a single ExcelWriter.
    - Preserves other sheets by re-writing them in original order.
    """
    # Ensure base file and sheets exist
    ensure_word_tracker(tracker_path, words_sheet, progress_sheet)

    words_df, prog_df, others = load_word_tracker(tracker_path, words_sheet, progress_sheet)

    if words_rows is not None:
        w_add = pd.DataFrame(words_rows) if not isinstance(words_rows, pd.DataFrame) else words_rows.copy()
        w_add = _ensure_columns(w_add, _WORDS_COLUMNS)
        # Normalize booleans where applicable to avoid stringy truth values
        for c in ["listed_by_user", "selected_for_analysis", "analysis_started", "analysis_completed"]:
            if c in w_add.columns:
                try:
                    w_add[c] = w_add[c].astype("boolean")
                except Exception:
                    pass
        # Drop all-NA columns to avoid concat FutureWarnings
        w_add = w_add.dropna(axis=1, how="all")
        words_df = pd.concat([words_df, w_add], ignore_index=True)

    if progress_rows is not None:
        p_add = pd.DataFrame(progress_rows) if not isinstance(progress_rows, pd.DataFrame) else progress_rows.copy()
        p_add = _ensure_columns(p_add, _PROGRESS_COLUMNS)
        for c in ["selected_for_analysis"]:
            if c in p_add.columns:
                try:
                    p_add[c] = p_add[c].astype("boolean")
                except Exception:
                    pass
        for c in ["reanalyzed_count", "word_index"]:
            if c in p_add.columns:
                try:
                    # Avoid 'errors="ignore"' to tame FutureWarning; coerce invalid to NaN
                    p_add[c] = pd.to_numeric(p_add[c], errors="coerce")
                except Exception:
                    pass
        # Drop all-NA columns to avoid concat FutureWarnings
        p_add = p_add.dropna(axis=1, how="all")
        prog_df = pd.concat([prog_df, p_add], ignore_index=True)

    _save_tracker(tracker_path, words_df, prog_df, others, words_sheet, progress_sheet)
    return True

# Structured Darpan sources: JSON/CSV loader helpers
def _normalize_simple(text: str) -> str:
    try:
        s = str(text)
    except Exception:
        return ""
    try:
        s = unicodedata.normalize('NFC', s)
    except Exception:
        pass
    s = s.strip().lower()
    return " ".join(s.split())

def _normalize_verse_key(text: str) -> str:
    """Robust comparable key for verse text.
    - Unicode normalize, lowercase
    - Remove danda marks (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¥Ãƒâ€šÃ‚Â¤, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¥Ãƒâ€šÃ‚Â¥) and digits (ASCII + Gurmukhi)
    - Collapse whitespace
    """
    try:
        s = str(text)
    except Exception:
        return ""
    try:
        s = unicodedata.normalize('NFC', s)
    except Exception:
        pass
    s = s.lower()
    s = s.replace('ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¥Ãƒâ€šÃ‚Â¥', ' ').replace('ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¥Ãƒâ€šÃ‚Â¤', ' ')
    s = re.sub(r"[\u0A66-\u0A6F0-9]+", " ", s)
    s = " ".join(s.split())
    return s

def _parse_page_value(val):
    try:
        if val is None:
            return None
        if isinstance(val, int):
            return str(val)
        if isinstance(val, float):
            if math.isnan(val):
                return None
            return str(int(val))
        s = str(val).strip()
        # Find first ASCII digit group anywhere, e.g., "171-172", "[171]"
        m = re.search(r"(\d+)", s)
        if m:
            return m.group(1)
        return s or None
    except Exception:
        return None

def _normalize_record(rec: dict) -> dict:
    try:
        items = rec.items()
    except Exception:
        try:
            items = dict(rec).items()
        except Exception:
            items = []
    keys = {str(k).lower(): v for k, v in items}
    out = {
        'verse': keys.get('verse', ''),
        'padarth': keys.get('padarth', ''),
        'arth': keys.get('arth', ''),
        'chhand': keys.get('chhand', ''),
        'bhav': keys.get('bhav', ''),
        'excel_verses': keys.get('excel_verses') if 'excel_verses' in keys else keys.get('excel_verse'),
        'pages': keys.get('pages') if 'pages' in keys else keys.get('page'),
    }
    # Preserve whether source explicitly provided fragment list
    has_excel_verses_field = 'excel_verses' in keys
    ex = out.get('excel_verses') or out.get('verse') or ''
    norm_parts = []
    norm_key_parts = []
    parts_source = []
    if isinstance(ex, (list, tuple)):
        parts_source = [p for p in ex]
    else:
        # Attempt to parse JSON-style list encoded as a string
        parsed = None
        if isinstance(ex, str):
            s_ex = ex.strip()
            if s_ex.startswith('[') and s_ex.endswith(']'):
                try:
                    j = json.loads(s_ex)
                    if isinstance(j, list):
                        parsed = j
                except Exception:
                    parsed = None
        if isinstance(parsed, list):
            parts_source = parsed
        else:
            # Split multiline text into fragments; conservatively allow semicolon as delimiter when excel_verses present
            try:
                s = str(ex)
            except Exception:
                s = ""
            split_lines = [seg.strip() for seg in re.split(r"[\r\n]+", s) if str(seg).strip()]
            if len(split_lines) >= 2:
                parts_source = split_lines
            else:
                if has_excel_verses_field and (';' in s):
                    semi = [seg.strip() for seg in s.split(';') if seg and seg.strip()]
                    if len(semi) >= 2:
                        parts_source = semi
                if not parts_source:
                    parts_source = [s]

    for part in parts_source:
        try:
            part_str = str(part)
        except Exception:
            part_str = ""
        norm_parts.append(_normalize_simple(part_str))
        norm_key_parts.append(_normalize_verse_key(part_str))

    # For an overall verse key, join fragments into a single string
    ex_join = " ".join([str(p) for p in parts_source]) if parts_source else str(ex)
    out['norm_excel'] = _normalize_simple(ex_join)
    out['norm_excel_key'] = _normalize_verse_key(ex_join)
    out['norm_excel_parts'] = norm_parts
    out['norm_excel_key_parts'] = norm_key_parts
    out['norm_page'] = _parse_page_value(out.get('pages'))
    return out

def _load_arth_sources_once(self):
    if getattr(self, '_arth_loaded', False):
        return
    self._arth_loaded = True
    records = []
    # JSON sources (try multiple known filenames)
    for json_path in [
        '1.1.3 sggs_extracted_with_page_numbers.json',
        '1.1.4 Verse_Padarth_Arth_with_pages.json',
    ]:
        try:
            with open(json_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
                items = list(data.values())[0] if isinstance(data, dict) else data
                if isinstance(items, list):
                    for rec in items:
                        if isinstance(rec, dict):
                            records.append(_normalize_record(rec))
        except Exception:
            continue
    # CSV sources (try multiple known filenames)
    for csv_path in [
        '1.1.4 Verse_Padarth_Arth_with_pages.csv',
        '1.1.5 Verse_Padarth_Arth_with_pages.csv',
    ]:
        try:
            df = pd.read_csv(csv_path)
            for _, row in df.iterrows():
                records.append(_normalize_record(row.to_dict()))
        except Exception:
            continue
    # Deduplicate by (norm_excel, norm_page)
    seen = {}
    for rec in records:
        key = (rec.get('norm_excel') or '', rec.get('norm_page') or '')
        if key not in seen:
            seen[key] = rec
    self._arth_records = list(seen.values())
    _dbg_autofill(self, f"[AutoFill] Loaded {len(records)} records, dedup -> {len(self._arth_records)} by (norm_excel, norm_page)")

def _find_arth_for(self, verse_text: str, page_num):
    try:
        _load_arth_sources_once(self)
    except Exception:
        return None
    if not getattr(self, '_arth_records', None):
        return None
    target_norm_verse = _normalize_simple(verse_text)
    target_norm_key = _normalize_verse_key(verse_text)
    target_page = _parse_page_value(page_num)
    # Pass 1: strict match on normalized text or key + page
    for rec in self._arth_records:
        parts = rec.get('norm_excel_parts', [])
        key_parts = rec.get('norm_excel_key_parts', [])
        matched_by = None
        matched_fragment = None
        if target_norm_verse == rec.get('norm_excel'):
            matched_by = 'norm_excel'
        elif target_norm_verse in parts:
            matched_by = 'norm_excel_parts'
            try:
                idx = parts.index(target_norm_verse)
                matched_fragment = rec.get('excel_verses', [None])[idx] if isinstance(rec.get('excel_verses'), list) else None
            except Exception:
                matched_fragment = None
        elif target_norm_key == rec.get('norm_excel_key'):
            matched_by = 'norm_excel_key'
        elif target_norm_key in key_parts:
            matched_by = 'norm_excel_key_parts'
            try:
                idx = key_parts.index(target_norm_key)
                matched_fragment = rec.get('excel_verses', [None])[idx] if isinstance(rec.get('excel_verses'), list) else None
            except Exception:
                matched_fragment = None
        if matched_by:
            if target_page is None or rec.get('norm_page') == target_page:
                _dbg_autofill(self, f"[AutoFill] Pass=1 strict match by {matched_by} (page_ok={target_page is None or rec.get('norm_page') == target_page}); fragment={matched_fragment!r}")
                return rec
    # Pass 2: verse-only strict match
    for rec in self._arth_records:
        parts = rec.get('norm_excel_parts', [])
        key_parts = rec.get('norm_excel_key_parts', [])
        matched_by = None
        matched_fragment = None
        if target_norm_verse == rec.get('norm_excel'):
            matched_by = 'norm_excel'
        elif target_norm_verse in parts:
            matched_by = 'norm_excel_parts'
            try:
                idx = parts.index(target_norm_verse)
                matched_fragment = rec.get('excel_verses', [None])[idx] if isinstance(rec.get('excel_verses'), list) else None
            except Exception:
                matched_fragment = None
        elif target_norm_key == rec.get('norm_excel_key'):
            matched_by = 'norm_excel_key'
        elif target_norm_key in key_parts:
            matched_by = 'norm_excel_key_parts'
            try:
                idx = key_parts.index(target_norm_key)
                matched_fragment = rec.get('excel_verses', [None])[idx] if isinstance(rec.get('excel_verses'), list) else None
            except Exception:
                matched_fragment = None
        if matched_by:
            _dbg_autofill(self, f"[AutoFill] Pass=2 strict (no page) match by {matched_by}; fragment={matched_fragment!r}")
            return rec
    # Pass 3: fuzzy match on verse key (prefer page matches)
    try:
        best = None
        best_score = 0
        best_detail = None
        for rec in self._arth_records:
            keys = rec.get('norm_excel_key_parts', []) + [rec.get('norm_excel_key') or '']
            frag_scores = []
            for k in keys:
                try:
                    frag_scores.append((k, fuzz.partial_ratio(target_norm_key, k)))
                except Exception:
                    frag_scores.append((k, 0))
            if frag_scores:
                k_best, score = max(frag_scores, key=lambda t: t[1])
            else:
                k_best, score = ('', 0)
            if target_page is not None and rec.get('norm_page') == target_page:
                score += 5
            if score > best_score:
                best_score = score
                best = rec
                best_detail = (k_best, score)
        if best and best_score >= 85:
            _dbg_autofill(self, f"[AutoFill] Pass=3 fuzzy match score={best_score}; page_bonus={(target_page is not None and best.get('norm_page') == target_page)}; fragment_key_used={best_detail[0]!r}")
            return best
    except Exception as e:
        _dbg_autofill(self, f"[AutoFill] Fuzzy pass error: {e}")
    return None

# Helper to determine whether a given string is a full Punjabi word
def is_full_word(s: str) -> bool:
    """Return ``True`` if *s* looks like a complete Punjabi word."""
    s = str(s).strip()
    # Words starting with a vowel matra are generally suffixes
    return len(s) > 1 and not ("\u0A3E" <= s[0] <= "\u0A4C")

# ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ Canonical ending-class labels for the dropdown ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬
CANONICAL_ENDINGS = [
    "NA",
    "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ Ending",      # bare consonant
    "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ Ending",       # ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾
    "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ Ending",     # ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿
    "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ Ending",     # ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬
    "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ Ending",       # ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ / ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ poetic
    "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â° Ending",          # ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â
    "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ Ending",          # ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡
]

# ------------------------------------------------------------------
#  FULL-WORD EXEMPLARS FOR EACH ENDING-CLASS
#  (trim / extend these lists whenever you like)
# ------------------------------------------------------------------

# ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ Canonical ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“keepÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â vowel for each ending-class ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬
KEEP_CHAR = {
    "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ Ending": "",
    "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ Ending": ("ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â "),
    "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ Ending": "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿",
    "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ Ending": "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬",
    "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ Ending": "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹",
    "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â° Ending": "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â",
    "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ Ending": "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡",
}

ENDING_EXAMPLES = {
    "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ Ending": [
        "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â ","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°",
        "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â§ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â ",
        "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¥ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹",
        "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â§ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¼ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®"
    ],

    "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ Ending": [
        "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â§ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾",
        "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â­ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã¢â‚¬Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ",
        "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â­ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã¢â‚¬Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã¢â‚¬Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂºÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ",
        "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â§ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã¢â‚¬Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã¢â‚¬Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¥ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â "
    ],

    "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ Ending": [
        "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â­ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿",
        "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿",
        "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¼ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿",
        "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â§ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿"
    ],

    "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ Ending": [
        "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â­ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬",
        "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬",
        "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã¢â‚¬Â ","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¼ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â§ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬",
        "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã¢â‚¬Â ","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬"
    ],

    "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ Ending": [
        "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã¢â‚¬Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã¢â‚¬Â ","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â"
    ],

    "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â° Ending": [
        "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂºÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â",
        "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â­ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â",
        "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¯ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â"
    ],

    "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ Ending": [
        "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â­ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¶ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â§ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡",
        "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡"
    ],
}

# ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ Function that turns ENDING_EXAMPLES into (Full, Base, Suffix) tuples ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬

def build_example_bases(
    csv_path: str = "1.1.1_birha.csv",
    ending_examples: dict[str, list[str]] = None,
    keep_char: dict[str, str] = None,
) -> dict[str, list[tuple[str, str, str]]]:
    if ending_examples is None or keep_char is None:
        raise ValueError("Pass ENDING_EXAMPLES and KEEP_CHAR")


    df = (pd.read_csv(csv_path).rename(columns={"Vowel Ending": "\ufeffVowel Ending", "Word Type": "Type"}).fillna("")
            .assign(**{
                "Word Root": lambda d: (
                    d["Word Root"]
                      .str.replace("ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â± Ending","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ Ending", regex=False)
                      .str.replace("ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ Ending","ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ Ending", regex=False)
                )
            }))

    # map: same 5-feature key ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ list of 1-glyph endings
    suffix_lookup = {}
    small = df[~df["\ufeffVowel Ending"].apply(is_full_word)]
    for _, r in small.iterrows():
        k = (r["Word Root"], r["Type"], r["Grammar / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¯ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£"],
             r["Gender / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â"], r["Number / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨"])
        suffix_lookup.setdefault(k, []).append(r["\ufeffVowel Ending"].strip())

    result = {}
    for label, wordlist in ending_examples.items():
        canon = keep_char.get(label, "")
        canon_set = set(canon) if isinstance(canon, (list, tuple, set)) else {canon}
        triples = []
        for full in wordlist:
            row = df[(df["\ufeffVowel Ending"].str.strip() == full) &
                     (df["Word Root"] == label)]
            if row.empty:
                triples.append((full, full, ""))
                continue
            r = row.iloc[0]
            k = (r["Word Root"], r["Type"], r["Grammar / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¯ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£"],
                 r["Gender / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â"], r["Number / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨"])
            base, suf = full, ""
            for cand in suffix_lookup.get(k, []):
                cand = cand.strip()
                if cand in canon_set or cand == "":
                    continue
                if full.endswith(cand):
                    base = full[:-len(cand)]
                    suf = cand
                    break
               
            if label == "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ Ending" and base == full and len(full) > 1:
                last = full[-1]
                # Unicode range for Gurmukhi matras (U+0A3EÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œU+0A4C)
                if "\u0A3E" <= last <= "\u0A4C":
                    # strip that final matra as a true detachment
                    base, suf = full[:-1], last
            
            triples.append((full, base, suf))
        result[label] = triples
    return result

EXAMPLE_BASES = build_example_bases(
    csv_path="1.1.1_birha.csv",
    ending_examples=ENDING_EXAMPLES,
    keep_char=KEEP_CHAR,
)


@lru_cache(maxsize=1)
def build_noun_map(csv_path="1.1.1_birha.csv"):
    """
    Returns a nested dict:
        noun_map[ending][gender][number][case] -> [list of attested forms]
    The loader also normalises stray spaces & typo-variants so look-ups never fail
    due to invisible characters.
    """
    raw = pd.read_csv(csv_path)
    raw.columns = raw.columns.str.replace("\ufeff", "").str.strip()
    df = (
        raw
          .loc[raw["Type"].astype(str).str.startswith("Noun", na=False)]
          .fillna("NA")
          .rename(columns={
              "Vowel Ending"        : "ending",
              "Number / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨"         : "num",
              "Grammar / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¯ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£"     : "case",
              "Gender / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â"         : "gender",
              "Word Root"           : "root",
          })
    )

    # --- normalise whitespace & common misspellings -----------------
    for c in ("ending", "gender", "num", "case", "root"):
        df[c] = (
            df[c].astype(str)
                 .str.replace(r"\s+", " ", regex=True)   # collapse weird spaces
                 .str.strip()                            # trim front/back
        )

    # unify the Kanna spelling
    df["root"] = df["root"].str.replace("ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â± Ending", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ Ending")

    # --- build nested dictionary ------------------------------------
    by_end = {}
    for ending, g1 in df.groupby("ending"):
        g_dict = {}
        for gender, g2 in g1.groupby("gender"):
            n_dict = {}
            for num, g3 in g2.groupby("num"):
                case_dict = (
                    g3.groupby("case")["ending"]     # store the surface form
                      .apply(list)                   # list of forms
                      .to_dict()
                )
                n_dict[num] = case_dict
            g_dict[gender] = n_dict
        by_end[ending] = g_dict
    return by_end


class GrammarApp:
    def __init__(self, root):
        """
        Initialize the application and display the dashboard as the main window.
        """
        # ------------------------------------------------------------------
        # ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ 1.  BASIC ROOTÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‹Å“WINDOW SETUP ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬
        # ------------------------------------------------------------------
        self.root = root
        self.root.title("Dashboard")
        self.root.configure(bg="light gray")
        try:
            self._wm_for(self.root).maximize()
        except Exception:
            pass
      
        # ------------------------------------------------------------------
        # ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ 2.  APPÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‹Å“WIDE STATE VARIABLES ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬
        # ------------------------------------------------------------------
        self.number_var  = tk.StringVar(value="NA")
        self.gender_var  = tk.StringVar(value="NA")
        self.pos_var     = tk.StringVar(value="NA")

        self.new_entries                   = []
        self.accumulated_pankti            = ""
        self.accumulated_meanings          = []
        self.accumulated_grammar_matches   = []
        self.accumulated_finalized_matches = []
        self.current_pankti                = ""
        self.match_vars                    = []
        self.all_matches                   = []
        self.all_new_entries               = []   # global accumulator

        # Lexicon index cache (built lazily from 1.1.3 Excel)
        self._lexicon_index = None
        self._lexicon_index_path = "1.1.3_lexicon_index.json"

        # wordÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‹Å“byÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‹Å“word navigation
        self.current_word_index = 0
        self.pankti_words       = []

        # per-verse repeat-word note tracking
        self._repeat_note_shown = set()
        self._suppress_repeat_notes_for_verse = False
        self._use_inline_literal_banner = True
        self._always_show_literal_banner_frame = False
        self._last_literal_verse_key = None
        self._first_repeat_token = None
        self._last_dropdown_verse_key = None

        self._LITERAL_NOTE_TEXT = (
            "In literal analysis: This word appears multiple times in this verse. "
            "The highlighted grammar options reflect your past selections for this word "
            "(or close matches) to encourage consistency. TheyÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â‚¬Å¾Ã‚Â¢re suggestions, not mandatesÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â"
            "adjust if the current context differs."
        )

        # ------------------------------------------------------------------
        # ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ 3.  DATA LOAD ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬
        # ------------------------------------------------------------------
        self.grammar_data   = self.load_grammar_data("1.1.1_birha.csv")
        self.dictionary_data = pd.read_csv(
            "1.1.2 Grammatical Meanings Dictionary.csv",
            encoding="utf-8"
        )

        # ------------------------------------------------------------------
        # ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ 4.  LAUNCH DASHBOARD ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬
        # ------------------------------------------------------------------
        self.show_dashboard()
    def _norm_get(self, d, key):
        """Unified getter that tolerates legacy field names."""
        if key == "\ufeffVowel Ending" or key == "Vowel Ending":
            return d.get("\ufeffVowel Ending") or d.get("Vowel Ending")
        if key == "Type" or key == "Word Type":
            return d.get("Type") or d.get("Word Type")
        return d.get(key)

    # TODO: Reuse in user_input(...) and prompt_save_results(...) for consistent comparisons.
    def _norm_tok(self, t: str) -> str:
        """Normalize token via NFC; drop dandas, zero-width spaces, ZWJ/ZWNJ, trailing digits & punctuation."""
        t = unicodedata.normalize("NFC", t.strip())
        t = re.sub(r"[ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¥Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¥Ãƒâ€šÃ‚Â¥]", "", t)  # danda/double-danda
        # remove ZERO WIDTH SPACE, ZWNJ, ZWJ
        t = t.replace("\u200b", "").replace("\u200c", "").replace("\u200d", "")
        t = re.sub(r"[\d\u0A66-\u0A6F.,;:!?\"'ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ-]+$", "", t)  # trailing digits (Latin+Gurmukhi) & punct
        return t

    def _verse_key(self, verse_text: str) -> str:
        """NFC + collapse spaces + remove danda variations; used for verse-scoped de-dupe keys."""
        cleaned = re.sub(r"[ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¥Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¥Ãƒâ€šÃ‚Â¥]", "", verse_text).strip()
        cleaned = re.sub(r"\s+", " ", cleaned)
        return unicodedata.normalize("NFC", cleaned)

    # ------------------------------------------------------------------
    # Lexicon index: build from 1.1.3 Excel (tokenize, normalize, aggregate counts)
    # ------------------------------------------------------------------
    def _tokenize_and_normalize(self, text: str) -> list[str]:
        # Guard against pandas NaN becoming literal 'nan' tokens
        try:
            if pd.isna(text):
                return []
        except Exception:
            pass
        try:
            s = str(text or "")
        except Exception:
            s = ""
        # strip danda/double-danda before split to avoid them as separate tokens
        s = re.sub(r"[\u0964\u0965]", "", s)
        toks = []
        for raw in s.split():
            t = self._norm_tok(raw)
            if t:
                toks.append(t)
        return toks

    def build_lexicon_index(self, force_rebuild: bool = False) -> dict:
        """Build or load a word->count index from 1.1.3 sggs_extracted_with_page_numbers.xlsx.

        - Tokenize verses, normalize each token via _norm_tok, and aggregate counts.
        - Caches to JSON on disk for faster subsequent loads.
        - Returns a dict {token: count}.
        """
        existing = getattr(self, "_lexicon_index", None)
        if (not force_rebuild) and isinstance(existing, dict) and len(existing) > 0:
            return existing

        cache_path = getattr(self, "_lexicon_index_path", None)
        excel_path = "1.1.3 sggs_extracted_with_page_numbers.xlsx"

        # Try to load cache if present and not forcing rebuild
        if not force_rebuild and cache_path and os.path.exists(cache_path):
            try:
                with open(cache_path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                if isinstance(data, dict) and len(data) > 0:
                    self._lexicon_index = {str(k): int(v) for k, v in data.items()}
                    return self._lexicon_index
            except Exception:
                pass

        # Build fresh from Excel
        try:
            df = pd.read_excel(excel_path, engine="openpyxl")
        except Exception as e:
            try:
                messagebox.showerror("Lexicon Error", f"Failed to read '{excel_path}':\n{e}")
            except Exception:
                pass
            # Do not cache an empty/failed build; allow future retries
            return {}

        counts: dict[str, int] = {}
        verse_col = "Verse"
        if verse_col not in df.columns:
            candidates = [c for c in df.columns if str(c).strip().lower() == "verse"]
            if candidates:
                verse_col = candidates[0]
            else:
                try:
                    messagebox.showerror("Lexicon Error", "Excel is missing 'Verse' column for lexicon build.")
                except Exception:
                    pass
                # Do not cache an empty/failed build; allow future retries
                return {}

        for _, row in df.iterrows():
            verse = row.get(verse_col, "")
            for tok in self._tokenize_and_normalize(verse):
                counts[tok] = counts.get(tok, 0) + 1

        self._lexicon_index = counts

        # Write cache (best-effort)
        try:
            if cache_path:
                with open(cache_path, "w", encoding="utf-8") as f:
                    json.dump(self._lexicon_index, f, ensure_ascii=False)
        except Exception:
            pass

        return self._lexicon_index

    def search_lexicon(self, query: str, limit: int = 50, min_score: int = 55) -> list[dict]:
        """Fuzzy-search the lexicon for closest matches to query.

        - Normalizes the query similar to _norm_tok (drop dandas/ZW chars).
        - Returns up to `limit` unique tokens sorted by score desc, then count desc.
        - Each result is {"token": str, "count": int, "score": int}.
        """
        index = self.build_lexicon_index()
        if not index:
            return []

        q = self._norm_tok(query or "")
        if not q:
            return []

        candidates = list(index.keys())
        try:
            scored = process.extract(q, candidates, scorer=fuzz.WRatio, limit=limit * 3)
        except Exception:
            scored = [(cand, fuzz.WRatio(q, cand), None) for cand in candidates]

        seen = set()
        results = []
        for cand, score, _ in scored:
            if score < min_score:
                continue
            if cand in seen:
                continue
            seen.add(cand)
            results.append({
                "token": cand,
                "count": int(index.get(cand, 0)),
                "score": int(score),
            })
            if len(results) >= (limit * 2):
                break

        results.sort(key=lambda r: (r["score"], r["count"]), reverse=True)
        return results[:limit]

    def show_word_search_modal(self):
        """Modal: search lexicon tokens with fuzzy matching and multi-select tick-list."""
        # Ensure index is built (may show an error and yield empty)
        self.build_lexicon_index()

        win = tk.Toplevel(self.root)
        win.title("Lexicon Word Search")
        win.configure(bg='light gray')
        win.transient(self.root)
        try:
            win.grab_set()
        except Exception:
            pass
        # Attach WindowManager and defer exact maximize (no margin) after layout
        # Windows: move to work-area origin + state('zoomed'); else: per-monitor work area
        self._wm_apply(win, margin_px=0, defer=True)

        header = tk.Label(
            win,
            text="Search Words (Lexicon)",
            font=("Arial", 16, "bold"),
            bg='dark slate gray', fg='white', pady=8
        )
        header.pack(fill=tk.X)

        body = tk.Frame(win, bg='light gray')
        body.pack(fill=tk.BOTH, expand=True, padx=20, pady=16)

        # Search row
        sr = tk.Frame(body, bg='light gray')
        sr.pack(fill=tk.X, pady=(0, 10))
        tk.Label(sr, text="Query:", font=("Arial", 12), bg='light gray').pack(side=tk.LEFT)
        q_var = tk.StringVar(value="")
        q_entry = tk.Entry(sr, textvariable=q_var, font=("Arial", 12), width=40)
        q_entry.pack(side=tk.LEFT, padx=(8, 8))

        status_var = tk.StringVar(value="Type to search; press Enter to refresh")
        tk.Label(sr, textvariable=status_var, font=("Arial", 10, "italic"),
                 bg='light gray', fg='#333').pack(side=tk.LEFT)

        # Toolbar actions on a row ABOVE the results list (keeps controls always visible)
        def _copy_selected():
            chosen = [w for var, w in getattr(self, "_lex_chk_vars", []) if var.get()]
            if not chosen:
                messagebox.showinfo("No Selection", "Please select one or more words to copy.")
                return
            try:
                pyperclip.copy("\n".join(chosen))
                messagebox.showinfo("Copied", f"Copied {len(chosen)} word(s) to clipboard.")
            except Exception as e:
                messagebox.showerror("Copy Failed", str(e))

        def _proceed_to_hits():
            chosen = [w for var, w in getattr(self, "_lex_chk_vars", []) if var.get()]
            if len(chosen) != 1:
                try:
                    messagebox.showwarning("Pick One Word", "Please select exactly one word to proceed to verse hits.")
                except Exception:
                    pass
                return
            self.show_word_verse_hits_modal(chosen[0], parent=win)

        # Selection row: keep only the checkbox above results
        sel_row = tk.Frame(body, bg='light gray')
        sel_row.pack(fill=tk.X, pady=(8, 6))
        sel_all_var = tk.BooleanVar(value=False)
        def _toggle_all():
            for var, _ in getattr(self, "_lex_chk_vars", []):
                var.set(bool(sel_all_var.get()))
        tk.Checkbutton(sel_row, text="Select/Deselect All", variable=sel_all_var,
                       bg='light gray', command=_toggle_all).pack(side=tk.LEFT)

        # Results area
        res_frame = tk.Frame(body, bg='light gray')
        res_frame.pack(fill=tk.BOTH, expand=True)
        canvas = tk.Canvas(res_frame, bg='light gray', highlightthickness=0)
        vsb = tk.Scrollbar(res_frame, orient="vertical", command=canvas.yview)
        inner = tk.Frame(canvas, bg='light gray')
        inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=inner, anchor='nw')
        canvas.configure(yscrollcommand=vsb.set)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)

        # Bottom action bar: compute bottom gap from body's external bottom padding (16)
        btns = tk.Frame(body, bg='light gray')
        btns.pack(side=tk.BOTTOM, fill=tk.X, pady=(6, max(0, BOTTOM_PAD - 16)))
        tk.Button(btns, text="Copy Selected", bg='teal', fg='white', font=("Arial", 11),
                  command=_copy_selected).pack(side=tk.LEFT)
        tk.Button(btns, text="Next: Verse Hits", bg='navy', fg='white', font=("Arial", 11, 'bold'),
                  command=_proceed_to_hits).pack(side=tk.LEFT, padx=(8, 0))
        tk.Button(btns, text="Back to Dashboard", bg='#2f4f4f', fg='white', font=("Arial", 11),
                  command=lambda: self._go_back_to_dashboard(win)).pack(side=tk.RIGHT, padx=(8,0))
        tk.Button(btns, text="Close", bg='gray', fg='white', font=("Arial", 11),
                  command=win.destroy).pack(side=tk.RIGHT)

        self._lex_chk_vars = []

        def _render_results(items: list[dict]):
            for w in list(inner.winfo_children()):
                w.destroy()
            self._lex_chk_vars = []
            # header row
            header_row = tk.Frame(inner, bg='light gray')
            header_row.grid(row=0, column=0, sticky='ew', padx=2, pady=(0, 6))
            tk.Label(header_row, text="Select", font=("Arial", 11, "bold"), width=8, anchor='w', bg='light gray').pack(side=tk.LEFT)
            tk.Label(header_row, text="Word", font=("Arial", 11, "bold"), width=28, anchor='w', bg='light gray').pack(side=tk.LEFT)
            tk.Label(header_row, text="Count", font=("Arial", 11, "bold"), width=8, anchor='e', bg='light gray').pack(side=tk.LEFT)
            tk.Label(header_row, text="Score", font=("Arial", 11, "bold"), width=8, anchor='e', bg='light gray').pack(side=tk.LEFT)

            for i, r in enumerate(items, start=1):
                rowf = tk.Frame(inner, bg='light gray')
                rowf.grid(row=i, column=0, sticky='ew', padx=2, pady=2)
                v = tk.BooleanVar(value=False)
                tk.Checkbutton(rowf, variable=v, bg='light gray').pack(side=tk.LEFT, padx=(0, 8))
                tk.Label(rowf, text=r["token"], font=("Arial", 12), width=28, anchor='w', bg='light gray').pack(side=tk.LEFT)
                tk.Label(rowf, text=str(r["count"]), font=("Arial", 12), width=8, anchor='e', bg='light gray').pack(side=tk.LEFT)
                tk.Label(rowf, text=str(r["score"]), font=("Arial", 12), width=8, anchor='e', bg='light gray').pack(side=tk.LEFT)
                self._lex_chk_vars.append((v, r["token"]))

        def _do_search(event=None):
            q = q_var.get()
            if not q.strip():
                status_var.set("Type to search; press Enter to refresh")
                _render_results([])
                return
            results = self.search_lexicon(q)
            status_var.set(f"{len(results)} result(s)")
            _render_results(results)

        tk.Button(sr, text="Search", command=_do_search, bg='teal', fg='white').pack(side=tk.LEFT, padx=(8,0))
        q_entry.bind("<Return>", _do_search)
        q_entry.focus_set()

    # ---------------------------
    # Assess-by-Word: Verse hits UI
    # ---------------------------
    def _get_word_tracker_path(self) -> str:
        """Return the workbook path for the Assess-by-Word tracker.

        Kept simple: a project-local XLSX alongside datasets.
        """
        try:
            return os.path.join(os.getcwd(), "1.1.6_assess_by_word.xlsx")
        except Exception:
            return "1.1.6_assess_by_word.xlsx"

    def _compute_verse_hits_for_token(self, token: str) -> list[dict]:
        """Compute verse hits for an exact token (normalized) from SGGS Excel.

        Uses the existing tokenization/normalization pipeline and returns a list of
        dicts with keys: verse, page_number.
        """
        # Ensure SGGS data is loaded (reused by other flows)
        self.load_sggs_data()

        norm_target = self._norm_tok(token or "")
        if not norm_target:
            return []

        hits = []
        df = getattr(self, 'sggs_data', None)
        if df is None or df.empty:
            return []
        verse_col = 'Verse' if 'Verse' in df.columns else list(df.columns)[0]
        page_col = 'Page Number' if 'Page Number' in df.columns else 'Page'

        for _, row in df.iterrows():
            verse_txt = row.get(verse_col, "")
            toks = self._tokenize_and_normalize(verse_txt)
            if norm_target in toks:
                page_val = row.get(page_col, "")
                try:
                    page_str = str(int(page_val)) if isinstance(page_val, (int, float)) and not pd.isna(page_val) else str(page_val)
                except Exception:
                    page_str = str(page_val)
                hits.append({
                    'verse': str(verse_txt),
                    'page_number': page_str,
                })
        return hits

    def show_word_verse_hits_modal(self, word: str, parent=None):
        """Modal window: show verse hits for a chosen word with checkbox selection.

        - Scrollable list with pagination when large.
        - Select All / Clear All buttons.
        - Add Selected appends rows to Progress via tracker helpers.
        - Back uses _go_back_to_dashboard(parent_win).
        """
        hits = self._compute_verse_hits_for_token(word)

        win = tk.Toplevel(self.root)
        win.title(f"Verse Hits for: {word}")
        win.configure(bg='light gray')
        try:
            win.transient(self.root)
            win.grab_set()
        except Exception:
            pass
        # Attach WindowManager and defer exact maximize (no margin) after layout
        # Windows: move to work-area origin + state('zoomed'); else: per-monitor work area
        self._wm_apply(win, margin_px=0, defer=True)

        header = tk.Label(
            win,
            text=f"Verse Hits for: {word}",
            font=("Arial", 16, "bold"),
            bg='dark slate gray', fg='white', pady=8
        )
        header.pack(fill=tk.X)

        body = tk.Frame(win, bg='light gray')
        body.pack(fill=tk.BOTH, expand=True, padx=20, pady=12)

        # Pagination controls
        ctrl = tk.Frame(body, bg='light gray')
        ctrl.pack(fill=tk.X, pady=(0, 6))
        tk.Label(ctrl, text=f"Total hits: {len(hits)}", bg='light gray', font=("Arial", 11)).pack(side=tk.LEFT)

        page_size_var = tk.IntVar(value=50)
        cur_page_var = tk.IntVar(value=1)

        def _page_count():
            ps = max(1, int(page_size_var.get() or 50))
            return max(1, (len(hits) + ps - 1)//ps)

        tk.Label(ctrl, text=" | Page size:", bg='light gray').pack(side=tk.LEFT, padx=(6,0))
        ttk.Combobox(ctrl, values=[25, 50, 100, 200], width=5, textvariable=page_size_var, state='readonly').pack(side=tk.LEFT, padx=(3, 8))

        nav = tk.Frame(ctrl, bg='light gray')
        nav.pack(side=tk.RIGHT)
        page_label_var = tk.StringVar(value=f"Page 1 / {_page_count()}")
        tk.Label(nav, textvariable=page_label_var, bg='light gray').pack(side=tk.RIGHT, padx=(6,0))

        def _change_page(delta):
            newp = cur_page_var.get() + delta
            newp = max(1, min(_page_count(), newp))
            if newp != cur_page_var.get():
                cur_page_var.set(newp)
                _render_rows()

        tk.Button(nav, text="Prev", command=lambda: _change_page(-1), bg='gray', fg='white').pack(side=tk.RIGHT, padx=(4,2))
        tk.Button(nav, text="Next", command=lambda: _change_page(1),  bg='gray', fg='white').pack(side=tk.RIGHT, padx=(2,4))

        # Prepare dynamic add button label (used by selection bar)
        add_btn_text = tk.StringVar(value="Add Selected (0)")

        # Selection bar above the results list (keeps toggles visible)
        selbar = tk.Frame(body, bg='light gray')
        selbar.pack(fill=tk.X, pady=(6, 0))
        def _select_all_page(val: bool):
            ps = max(1, int(page_size_var.get() or 50))
            p  = max(1, int(cur_page_var.get() or 1))
            start = (p-1)*ps
            end   = min(len(hits), start+ps)
            for bv, idx in hit_vars[start:end]:
                bv.set(val)
            _update_add_button()
        tk.Button(selbar, text="Select All", bg='teal', fg='white', command=lambda: _select_all_page(True)).pack(side=tk.LEFT)
        tk.Button(selbar, text="Clear All", bg='gray', fg='white', command=lambda: _select_all_page(False)).pack(side=tk.LEFT, padx=(6,0))
        # Note: action buttons are moved to a bottom bar for consistent spacing

        # Scrollable area
        list_frame = tk.Frame(body, bg='light gray')
        list_frame.pack(fill=tk.BOTH, expand=True)
        
        # Bottom action bar - compute bottom gap from body's external bottom padding (12)
        btns = tk.Frame(body, bg='light gray')
        btns.pack(side=tk.BOTTOM, fill=tk.X, pady=(6, max(0, BOTTOM_PAD - 12)))
        add_btn_bottom = tk.Button(btns, textvariable=add_btn_text, bg='navy', fg='white', font=("Arial", 12, 'bold'))
        add_btn_bottom.pack(side=tk.LEFT)
        tk.Button(btns, text="Back to Dashboard", bg='#2f4f4f', fg='white', font=("Arial", 11),
                  command=lambda: self._go_back_to_dashboard(win)).pack(side=tk.RIGHT, padx=(8,0))
        tk.Button(btns, text="Close", bg='gray', fg='white', font=("Arial", 11), command=win.destroy).pack(side=tk.RIGHT)
        canvas = tk.Canvas(list_frame, bg='light gray', highlightthickness=0)
        vsb = tk.Scrollbar(list_frame, orient='vertical', command=canvas.yview)
        inner = tk.Frame(canvas, bg='light gray')
        inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox('all')))
        canvas.create_window((0,0), window=inner, anchor='nw')
        canvas.configure(yscrollcommand=vsb.set)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)

        # Selection state across pages (local to this modal)
        hit_vars = []  # list[(BooleanVar, index)] parallel to hits
        for i in range(len(hits)):
            hit_vars.append((tk.BooleanVar(value=False), i))

        # Row rendering
        def _update_add_button():
            total = sum(1 for bv, _ in hit_vars if bv.get())
            add_btn_text.set(f"Add Selected ({total})")
            try:
                add_btn_bottom.configure(state=tk.NORMAL if total > 0 else tk.DISABLED)
            except Exception:
                pass

        def _render_rows():
            for wdg in list(inner.winfo_children()):
                wdg.destroy()
            ps = max(1, int(page_size_var.get() or 50))
            p  = max(1, int(cur_page_var.get() or 1))
            start = (p-1)*ps
            end   = min(len(hits), start+ps)
            # header
            hdr = tk.Frame(inner, bg='light gray')
            hdr.grid(row=0, column=0, sticky='ew', padx=2, pady=(0, 6))
            tk.Label(hdr, text="Select", font=("Arial", 11, 'bold'), width=8, anchor='w', bg='light gray').pack(side=tk.LEFT)
            tk.Label(hdr, text="Verse", font=("Arial", 11, 'bold'), width=64, anchor='w', bg='light gray').pack(side=tk.LEFT)
            tk.Label(hdr, text="Page", font=("Arial", 11, 'bold'), width=8, anchor='e', bg='light gray').pack(side=tk.LEFT)

            for i, idx in enumerate(range(start, end), start=1):
                row = hits[idx]
                rf = tk.Frame(inner, bg='light gray')
                rf.grid(row=i, column=0, sticky='ew', padx=2, pady=2)
                bv, _ = hit_vars[idx]
                def _mk_toggle(var):
                    return lambda: (_update_add_button())
                tk.Checkbutton(rf, variable=bv, bg='light gray', command=_mk_toggle(bv)).pack(side=tk.LEFT, padx=(0, 8))
                tk.Label(rf, text=row['verse'], font=("Arial", 12), width=64, anchor='w', bg='light gray', justify='left', wraplength=900).pack(side=tk.LEFT)
                tk.Label(rf, text=str(row['page_number']), font=("Arial", 12), width=8, anchor='e', bg='light gray').pack(side=tk.LEFT)

            page_label_var.set(f"Page {p} / {_page_count()}")
            list_frame.update_idletasks()
            canvas.configure(scrollregion=canvas.bbox('all'))
            _update_add_button()

        def _on_page_size_change(event=None):
            cur_page_var.set(1)
            _render_rows()

        try:
            page_size_var.trace_add('write', lambda *args: _on_page_size_change())
        except Exception:
            pass

        def _do_add_selected():
            now = datetime.now()
            norm = self._norm_tok(word)
            rows = []
            for bv, idx in hit_vars:
                if not bv.get():
                    continue
                rec = hits[idx]
                rows.append({
                    'word': word,
                    'word_key_norm': norm,
                    'verse': rec.get('verse', ''),
                    'page_number': rec.get('page_number', ''),
                    'selected_for_analysis': True,
                    'selected_at': now,
                    'status': 'not started',
                })
            if not rows:
                return
            try:
                tracker_path = self._get_word_tracker_path()
                # Ensure a Words row exists and mark selected_for_analysis
                words_df, prog_df, _others = load_word_tracker(tracker_path, TRACKER_WORDS_SHEET, TRACKER_PROGRESS_SHEET)
                have_row = False
                try:
                    if not words_df.empty and 'word_key_norm' in words_df.columns:
                        have_row = any((_normalize_simple(str(x)) == _normalize_simple(norm)) for x in words_df['word_key_norm'])
                except Exception:
                    have_row = False
                words_rows = None
                if not have_row:
                    words_rows = [{
                        'word': word,
                        'word_key_norm': norm,
                        'listed_by_user': True,
                        'listed_at': now,
                        'selected_for_analysis': True,
                        'selected_at': now,
                        'analysis_started': False,
                        'analysis_started_at': None,
                        'analysis_completed': False,
                        'analysis_completed_at': None,
                        'sequence_index': 0,
                        'notes': '',
                    }]
                append_to_word_tracker(tracker_path, words_rows=words_rows, progress_rows=rows)
                try:
                    messagebox.showinfo("Added", f"Appended {len(rows)} row(s) to Progress:\n{tracker_path}")
                except Exception:
                    pass
            except Exception as e:
                try:
                    messagebox.showerror("Tracker Error", str(e))
                except Exception:
                    pass
            # Offer to start driver immediately
            try:
                start_now = messagebox.askyesno("Start Analysis", "Start analyzing selected verses for this word now?")
            except Exception:
                start_now = True
            if start_now:
                try:
                    # Just close this modal; do NOT set suppression here.
                    # Suppression is only used around ABW translation submit.
                    win.destroy()
                except Exception:
                    pass
                self.start_word_driver_for(word)

        # Wire the bottom 'Add Selected' button now that it's defined
        try:
            add_btn_bottom.configure(command=_do_add_selected)
        except Exception:
            pass
        _render_rows()

    # ---------------------------
    # Assess-by-Word: Driver (sequential verse opener)
    # ---------------------------
    def start_word_driver_for(self, word: str):
        """Start the Assess-by-Word driver for the given word.

        - Builds the verse queue from the tracker Progress sheet where status != 'completed'.
        - Marks Words.analysis_started on first run.
        - Opens a small controller with Next Verse and Pause/Resume.
        - Opens the first verse in the Assess-by-Verse analyzer.
        """
        try:
            norm = self._norm_tok(word or "")
            tracker_path = self._get_word_tracker_path()
            ensure_word_tracker(tracker_path, TRACKER_WORDS_SHEET, TRACKER_PROGRESS_SHEET)
            words_df, prog_df, others = load_word_tracker(tracker_path, TRACKER_WORDS_SHEET, TRACKER_PROGRESS_SHEET)

            # Build queue
            self._word_driver_queue = []
            if not prog_df.empty:
                try:
                    df = prog_df.copy()
                    # normalize key
                    if 'word_key_norm' in df.columns:
                        mask = df['word_key_norm'].astype(str).map(lambda s: _normalize_simple(s)) == _normalize_simple(norm)
                        df = df.loc[mask]
                    # status filter
                    if 'status' in df.columns:
                        df = df.loc[df['status'].astype(str).str.lower() != 'completed']
                    # stabilize order by selected_at then index
                    if 'selected_at' in df.columns:
                        try:
                            df['_sel'] = pd.to_datetime(df['selected_at'], errors='coerce')
                        except Exception:
                            df['_sel'] = pd.Series([None]*len(df))
                        df = df.sort_values(by=['_sel']).drop(columns=['_sel'])
                    self._word_driver_queue = df.to_dict('records')
                except Exception:
                    self._word_driver_queue = []

            if not self._word_driver_queue:
                try:
                    messagebox.showinfo("No Verses", "No pending verses found in Progress for this word.")
                except Exception:
                    pass
                return

            # Mark analysis_started in Words if not already
            if not words_df.empty:
                try:
                    mask = (words_df.get('word_key_norm', pd.Series(dtype=str)).astype(str)
                            .map(lambda s: _normalize_simple(s)) == _normalize_simple(norm))
                    if mask.any():
                        # set flags
                        if 'analysis_started' in words_df.columns:
                            words_df.loc[mask, 'analysis_started'] = True
                        if 'analysis_started_at' in words_df.columns:
                            # only set if empty
                            def _maybe_set(dt):
                                return dt if pd.notna(dt) and str(dt).strip() != '' else datetime.now()
                            words_df.loc[mask, 'analysis_started_at'] = words_df.loc[mask, 'analysis_started_at'].apply(_maybe_set)
                        _save_tracker(tracker_path, words_df, prog_df, others, TRACKER_WORDS_SHEET, TRACKER_PROGRESS_SHEET)
                except Exception:
                    pass

            # Init state
            self._word_driver_active = True
            self._word_driver_paused = False
            self._word_driver_in_progress = False
            self._word_driver_word = word
            self._word_driver_norm = norm
            self._word_driver_index = 0
            self._word_driver_current_verse = None

            # Integrated driver controls are part of the ABW windows; no floating controller.
            # Open first verse immediately
            self._word_driver_open_current_verse()
        except Exception as e:
            try:
                messagebox.showerror("Driver Error", f"Failed to start driver: {e}")
            except Exception:
                pass

    def _word_driver_open_controller(self):
        try:
            if hasattr(self, '_word_driver_win') and self._word_driver_win and self._word_driver_win.winfo_exists():
                # reuse
                self._word_driver_update_ui()
                return
        except Exception:
            pass
        win = tk.Toplevel(self.root)
        self._word_driver_win = win
        win.title("Assess by Word - Driver")
        try:
            win.attributes('-topmost', True)
        except Exception:
            pass
        win.configure(bg='light gray')
        frm = tk.Frame(win, bg='light gray')
        frm.pack(padx=12, pady=10)
        self._driver_lbl = tk.Label(frm, text="", font=('Arial', 12), bg='light gray')
        self._driver_lbl.pack(anchor='w')
        btns = tk.Frame(frm, bg='light gray')
        btns.pack(fill=tk.X, pady=(8, 0))
        self._driver_next_btn = tk.Button(btns, text="Next Verse ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Å“Ãƒâ€šÃ‚Â¶", bg='navy', fg='white', font=('Arial', 11, 'bold'), command=self._word_driver_open_current_verse)
        self._driver_next_btn.pack(side=tk.LEFT)
        self._driver_pause_btn = tk.Button(btns, text="Pause", bg='gray', fg='white', font=('Arial', 11), command=self._word_driver_toggle_pause)
        self._driver_pause_btn.pack(side=tk.LEFT, padx=(6, 0))
        tk.Button(btns, text="Close", bg='#2f4f4f', fg='white', font=('Arial', 11), command=win.destroy).pack(side=tk.RIGHT)
        try:
            win.bind('<Escape>', lambda e: win.destroy())
        except Exception:
            pass
        self._word_driver_update_ui()

    def _word_driver_update_ui(self):
        try:
            total = len(getattr(self, '_word_driver_queue', []) or [])
            idx = getattr(self, '_word_driver_index', 0)
            paused = bool(getattr(self, '_word_driver_paused', False))
            word = getattr(self, '_word_driver_word', '')
            cur = min(idx + 1, total) if total else 0
            status = 'Paused' if paused else 'Running'
            txt = f"Word: {word}   |   {status}   |   Verse {cur}/{total}"
            if hasattr(self, '_driver_lbl') and self._driver_lbl.winfo_exists():
                self._driver_lbl.config(text=txt)
            if hasattr(self, '_driver_next_btn') and self._driver_next_btn.winfo_exists():
                # Disable Next while in-progress
                busy = bool(getattr(self, '_word_driver_in_progress', False))
                try:
                    self._driver_next_btn.config(state=tk.DISABLED if busy or paused else tk.NORMAL)
                except Exception:
                    pass
            if hasattr(self, '_driver_pause_btn') and self._driver_pause_btn.winfo_exists():
                self._driver_pause_btn.config(text=("Resume" if paused else "Pause"))
            # Update integrated ABW header controls if present
            if hasattr(self, '_abw_driver_status_var') and isinstance(self._abw_driver_status_var, tk.StringVar):
                try:
                    self._abw_driver_status_var.set(txt)
                except Exception:
                    pass
            busy = bool(getattr(self, '_word_driver_in_progress', False))
            if hasattr(self, '_abw_driver_next_btn') and self._abw_driver_next_btn and self._abw_driver_next_btn.winfo_exists():
                try:
                    self._abw_driver_next_btn.config(state=tk.DISABLED if busy or paused or (total and (idx >= total)) else tk.NORMAL)
                except Exception:
                    pass
            if hasattr(self, '_abw_driver_prev_btn') and self._abw_driver_prev_btn and self._abw_driver_prev_btn.winfo_exists():
                try:
                    self._abw_driver_prev_btn.config(state=tk.DISABLED if busy or paused or idx <= 0 else tk.NORMAL)
                except Exception:
                    pass
            if hasattr(self, '_abw_driver_pause_btn') and self._abw_driver_pause_btn and self._abw_driver_pause_btn.winfo_exists():
                try:
                    self._abw_driver_pause_btn.config(text=("Resume" if paused else "Pause"))
                except Exception:
                    pass
        except Exception:
            pass

    def _word_driver_toggle_pause(self):
        try:
            self._word_driver_paused = not bool(getattr(self, '_word_driver_paused', False))
            self._word_driver_update_ui()
            # If resuming and not busy, open next verse
            if not self._word_driver_paused and not getattr(self, '_word_driver_in_progress', False):
                self._word_driver_open_current_verse()
        except Exception:
            pass

    def _word_driver_open_current_verse(self):
        """Open the current verse in the ABW analyzer (dedicated window)."""
        try:
            if getattr(self, '_word_driver_paused', False):
                return
            queue = getattr(self, '_word_driver_queue', []) or []
            idx = getattr(self, '_word_driver_index', 0)
            if idx >= len(queue):
                # done
                self._word_driver_complete_word_if_done()
                return
            rec = queue[idx]
            verse = str(rec.get('verse', ''))
            if not verse.strip():
                # Skip invalid
                self._word_driver_index += 1
                self._word_driver_update_ui()
                self._word_driver_open_current_verse()
                return
            # flag busy and remember current
            self._word_driver_in_progress = True
            self._word_driver_current_verse = verse
            # feed into Assess-by-Word analyzer
            self.selected_verse_text = verse
            self.show_word_translation_input()
            self._word_driver_update_ui()
        except Exception as e:
            try:
                messagebox.showerror("Driver Error", f"Failed to open verse: {e}")
            except Exception:
                pass

    def _word_driver_cancel_current(self):
        """Cancel the current verse if driver is in-progress.

        - Clears in-progress flag and current verse pointer.
        - Advances the queue index to skip this verse.
        - Updates the controller UI so Next is enabled and the driver can continue.
        """
        try:
            # clear busy state
            self._word_driver_in_progress = False
            # skip the current verse if any
            try:
                total = len(getattr(self, '_word_driver_queue', []) or [])
            except Exception:
                total = 0
            try:
                idx = int(getattr(self, '_word_driver_index', 0))
            except Exception:
                idx = 0
            if idx < total:
                self._word_driver_index = idx + 1
            self._word_driver_current_verse = None
            # refresh controller UI
            self._word_driver_update_ui()
        except Exception:
            pass

    def _word_driver_after_verse_finished(self):
        """Called when a verse finishes its assessment flow to update tracker and advance."""
        try:
            tracker_path = self._get_word_tracker_path()
            words_df, prog_df, others = load_word_tracker(tracker_path, TRACKER_WORDS_SHEET, TRACKER_PROGRESS_SHEET)
            norm = getattr(self, '_word_driver_norm', '')
            verse = getattr(self, '_word_driver_current_verse', '')
            if not prog_df.empty and verse:
                try:
                    mask_word = (prog_df.get('word_key_norm', pd.Series(dtype=str)).astype(str)
                                 .map(lambda s: _normalize_simple(s)) == _normalize_simple(norm))
                    mask_verse = (prog_df.get('verse', pd.Series(dtype=str)).astype(str)
                                  .map(lambda s: _normalize_simple(s)) == _normalize_simple(verse))
                    mask = mask_word & mask_verse
                    if mask.any():
                        if 'status' in prog_df.columns:
                            prog_df.loc[mask, 'status'] = 'completed'
                        if 'completed_at' in prog_df.columns:
                            prog_df.loc[mask, 'completed_at'] = datetime.now()
                        _save_tracker(tracker_path, words_df, prog_df, others, TRACKER_WORDS_SHEET, TRACKER_PROGRESS_SHEET)
                except Exception:
                    pass
            # advance index
            self._word_driver_in_progress = False
            self._word_driver_index += 1
            self._word_driver_update_ui()
            # if more verses and not paused, open next
            if self._word_driver_index < len(getattr(self, '_word_driver_queue', []) or []) and not getattr(self, '_word_driver_paused', False):
                self._word_driver_open_current_verse()
            else:
                # possibly complete word
                self._word_driver_complete_word_if_done()
        except Exception:
            # avoid breaking the core flow
            self._word_driver_in_progress = False
            self._word_driver_update_ui()

    def _word_driver_complete_word_if_done(self):
        try:
            tracker_path = self._get_word_tracker_path()
            words_df, prog_df, others = load_word_tracker(tracker_path, TRACKER_WORDS_SHEET, TRACKER_PROGRESS_SHEET)
            norm = getattr(self, '_word_driver_norm', '')
            # if no pending verses remain for this word, mark completed
            pending = 0
            if not prog_df.empty:
                try:
                    df = prog_df.copy()
                    mask = (df.get('word_key_norm', pd.Series(dtype=str)).astype(str)
                            .map(lambda s: _normalize_simple(s)) == _normalize_simple(norm))
                    df = df.loc[mask]
                    if not df.empty and 'status' in df.columns:
                        pending = int((df['status'].astype(str).str.lower() != 'completed').sum())
                except Exception:
                    pending = 0
            if pending == 0 and not words_df.empty:
                try:
                    maskw = (words_df.get('word_key_norm', pd.Series(dtype=str)).astype(str)
                             .map(lambda s: _normalize_simple(s)) == _normalize_simple(norm))
                    if maskw.any():
                        if 'analysis_completed' in words_df.columns:
                            words_df.loc[maskw, 'analysis_completed'] = True
                        if 'analysis_completed_at' in words_df.columns:
                            words_df.loc[maskw, 'analysis_completed_at'] = datetime.now()
                        _save_tracker(tracker_path, words_df, prog_df, others, TRACKER_WORDS_SHEET, TRACKER_PROGRESS_SHEET)
                        try:
                            messagebox.showinfo("Done", f"Completed analysis for word: {getattr(self, '_word_driver_word', '')}")
                        except Exception:
                            pass
                except Exception:
                    pass
        except Exception:
            pass

    def show_continue_incomplete(self):
        """List incomplete words (Words.analysis_completed == False) and allow resume."""
        try:
            tracker_path = self._get_word_tracker_path()
            ensure_word_tracker(tracker_path, TRACKER_WORDS_SHEET, TRACKER_PROGRESS_SHEET)
            words_df, prog_df, _ = load_word_tracker(tracker_path, TRACKER_WORDS_SHEET, TRACKER_PROGRESS_SHEET)
            if words_df.empty:
                messagebox.showinfo("None", "No words found in tracker.")
                return
            df = words_df.copy()
            # Compute set of words with any pending (status != 'completed') rows in Progress
            pending_norms = set()
            try:
                if prog_df is not None and not prog_df.empty:
                    p = prog_df.copy()
                    if 'status' in p.columns:
                        p = p.loc[p['status'].astype(str).str.lower() != 'completed']
                    if 'word_key_norm' in p.columns:
                        norms = p['word_key_norm'].astype(str).map(lambda s: _normalize_simple(s)).tolist()
                        pending_norms = set(norms)
            except Exception:
                pending_norms = set()

            # Base filter: Words where analysis_completed is not True
            if 'analysis_completed' in df.columns:
                incom = df.loc[df['analysis_completed'] != True]
            else:
                incom = df

            # Strengthen filter: include also any words with pending rows even if analysis_completed flag wasn't set
            try:
                if not df.empty and 'word_key_norm' in df.columns and pending_norms:
                    df['_norm'] = df['word_key_norm'].astype(str).map(lambda s: _normalize_simple(s))
                    extra = df.loc[df['_norm'].isin(pending_norms)]
                    incom = pd.concat([incom, extra], ignore_index=True).drop_duplicates(subset=['word_key_norm'])
            except Exception:
                pass
            if incom.empty:
                messagebox.showinfo("All Done", "No incomplete words remain.")
                return
            # Build simple chooser
            win = tk.Toplevel(self.root)
            win.title("Continue Incomplete - Choose Word")
            win.configure(bg='light gray')
            try:
                self._wm_for(win)
            except Exception:
                pass
            tk.Label(win, text="Select a word to resume:", font=('Arial', 12, 'bold'), bg='light gray').pack(padx=12, pady=(10,6), anchor='w')
            lb = tk.Listbox(win, height=10)
            items = []
            try:
                # If Progress pending counts are available, show them in the list for context
                pending_counts = {}
                if not prog_df.empty:
                    pp = prog_df.copy()
                    if 'status' in pp.columns:
                        pp['_is_pending'] = pp['status'].astype(str).str.lower() != 'completed'
                    else:
                        pp['_is_pending'] = True
                    if 'word_key_norm' in pp.columns:
                        pp['_norm'] = pp['word_key_norm'].astype(str).map(lambda s: _normalize_simple(s))
                        grp = pp.groupby('_norm')['_is_pending'].sum()
                        pending_counts = {k: int(v) for k, v in grp.to_dict().items()}
                for _, r in incom.iterrows():
                    w = str(r.get('word', ''))
                    n = str(r.get('word_key_norm', ''))
                    norm = _normalize_simple(n)
                    cnt = pending_counts.get(norm)
                    label = f"{w}" if not cnt else f"{w}  (pending: {cnt})"
                    items.append((w, n))
                    lb.insert(tk.END, label)
            except Exception:
                for _, r in incom.iterrows():
                    w = str(r.get('word', ''))
                    n = str(r.get('word_key_norm', ''))
                    items.append((w, n))
                    lb.insert(tk.END, w)
            lb.pack(fill=tk.BOTH, expand=True, padx=12, pady=(0,8))
            def _resume():
                try:
                    sel = lb.curselection()
                    if not sel:
                        return
                    w = items[sel[0]][0]
                    win.destroy()
                    self.start_word_driver_for(w)
                except Exception:
                    pass
            btns = tk.Frame(win, bg='light gray')
            btns.pack(fill=tk.X, padx=12, pady=(0, 10))
            tk.Button(btns, text="Resume", bg='navy', fg='white', font=('Arial', 11, 'bold'), command=_resume).pack(side=tk.RIGHT)
            tk.Button(btns, text="Cancel", bg='gray', fg='white', font=('Arial', 11), command=win.destroy).pack(side=tk.RIGHT, padx=(0,6))
        except Exception as e:
            try:
                messagebox.showerror("Tracker Error", f"Failed to load tracker: {e}")
            except Exception:
                pass

    def show_view_completed(self):
        """List completed words and allow re-analyze actions (all or selected verses)."""
        try:
            tracker_path = self._get_word_tracker_path()
            ensure_word_tracker(tracker_path, TRACKER_WORDS_SHEET, TRACKER_PROGRESS_SHEET)
            words_df, prog_df, _ = load_word_tracker(tracker_path, TRACKER_WORDS_SHEET, TRACKER_PROGRESS_SHEET)
            if words_df.empty and (prog_df is None or prog_df.empty):
                messagebox.showinfo("None", "No tracker data found.")
                return

            df = (words_df.copy() if words_df is not None else pd.DataFrame(columns=_WORDS_COLUMNS))
            completed = pd.DataFrame(columns=df.columns)
            # 1) Prefer Words.analysis_completed flag
            try:
                if not df.empty and 'analysis_completed' in df.columns:
                    completed = df.loc[df['analysis_completed'] == True]
            except Exception:
                completed = pd.DataFrame(columns=df.columns)

            # 2) Also include words for which all Progress rows are completed
            try:
                if prog_df is not None and not prog_df.empty and 'word_key_norm' in prog_df.columns:
                    p = prog_df.copy()
                    p['_norm'] = p['word_key_norm'].astype(str).map(lambda s: _normalize_simple(s))
                    if 'status' in p.columns:
                        grp = p.groupby('_norm')['status'].apply(lambda s: (s.astype(str).str.lower() == 'completed').all())
                    else:
                        grp = p.groupby('_norm').size() * 0  # no statuses implies nothing to add
                    norms_all_done = {k for k, v in grp.to_dict().items() if v is True}
                    if not df.empty and 'word_key_norm' in df.columns:
                        df['_norm'] = df['word_key_norm'].astype(str).map(lambda s: _normalize_simple(s))
                        extra = df.loc[df['_norm'].isin(norms_all_done)]
                        if not extra.empty:
                            completed = pd.concat([completed, extra], ignore_index=True).drop_duplicates(subset=['word_key_norm'])
            except Exception:
                pass

            if completed is None or completed.empty:
                messagebox.showinfo("None", "No completed words found.")
                return

            # Build chooser
            win = tk.Toplevel(self.root)
            win.title("Completed Words - Review / Re-Analyze")
            win.configure(bg='light gray')
            try:
                self._wm_for(win)
            except Exception:
                pass
            tk.Label(win, text="Select a word to re-analyze:", font=('Arial', 12, 'bold'), bg='light gray').pack(padx=12, pady=(10,6), anchor='w')
            lb = tk.Listbox(win, height=10)
            items = []
            try:
                # Optionally show count of completed verses per word
                comp_counts = {}
                if prog_df is not None and not prog_df.empty and 'status' in prog_df.columns and 'word_key_norm' in prog_df.columns:
                    pp = prog_df.copy()
                    pp['_is_completed'] = pp['status'].astype(str).str.lower() == 'completed'
                    pp['_norm'] = pp['word_key_norm'].astype(str).map(lambda s: _normalize_simple(s))
                    grp = pp[pp['_is_completed']].groupby('_norm').size()
                    comp_counts = {k: int(v) for k, v in grp.to_dict().items()}
                for _, r in completed.iterrows():
                    w = str(r.get('word', ''))
                    n = str(r.get('word_key_norm', ''))
                    norm = _normalize_simple(n)
                    cnt = comp_counts.get(norm)
                    label = f"{w}" if not cnt else f"{w}  (completed verses: {cnt})"
                    items.append((w, n))
                    lb.insert(tk.END, label)
            except Exception:
                for _, r in completed.iterrows():
                    w = str(r.get('word', ''))
                    n = str(r.get('word_key_norm', ''))
                    items.append((w, n))
                    lb.insert(tk.END, w)
            lb.pack(fill=tk.BOTH, expand=True, padx=12, pady=(0,8))

            def _get_selected_word():
                sel = lb.curselection()
                if not sel:
                    return None
                return items[sel[0]][0]

            def _re_analyze_all():
                try:
                    word = _get_selected_word()
                    if not word:
                        return
                    self._reset_progress_for_word_and_restart(word, only_selected=None)
                    try:
                        win.destroy()
                    except Exception:
                        pass
                except Exception:
                    pass

            def _open_selector():
                try:
                    word = _get_selected_word()
                    if not word:
                        return
                    self._open_reanalyze_selector(word)
                except Exception:
                    pass

            btns = tk.Frame(win, bg='light gray')
            btns.pack(fill=tk.X, padx=12, pady=(0, 10))
            tk.Button(btns, text="Re-analyze...", bg='navy', fg='white', font=('Arial', 11, 'bold'), command=_open_selector).pack(side=tk.RIGHT)
            tk.Button(btns, text="Re-analyze All", bg='teal', fg='white', font=('Arial', 11, 'bold'), command=_re_analyze_all).pack(side=tk.RIGHT, padx=(0,6))
            tk.Button(btns, text="Cancel", bg='gray', fg='white', font=('Arial', 11), command=win.destroy).pack(side=tk.RIGHT, padx=(0,6))
        except Exception as e:
            try:
                messagebox.showerror("Tracker Error", f"Failed to load tracker: {e}")
            except Exception:
                pass

    def _open_reanalyze_selector(self, word: str):
        """Open a modal to pick completed verses for a word to re-analyze (reset status and start driver)."""
        try:
            tracker_path = self._get_word_tracker_path()
            words_df, prog_df, others = load_word_tracker(tracker_path, TRACKER_WORDS_SHEET, TRACKER_PROGRESS_SHEET)
            norm = self._norm_tok(word)
            df = prog_df.copy() if prog_df is not None else pd.DataFrame(columns=_PROGRESS_COLUMNS)
            if not df.empty:
                if 'word_key_norm' in df.columns:
                    df = df.loc[df['word_key_norm'].astype(str).map(lambda s: _normalize_simple(s)) == _normalize_simple(norm)]
                if 'status' in df.columns:
                    df = df.loc[df['status'].astype(str).str.lower() == 'completed']
            if df is None or df.empty:
                try:
                    messagebox.showinfo("None", f"No completed verses found for '{word}'.")
                except Exception:
                    pass
                return

            win = tk.Toplevel(self.root)
            win.title(f"Re-Analyze Verses - {word}")
            win.configure(bg='light gray')
            try:
                self._wm_for(win)
            except Exception:
                pass
            tk.Label(win, text=f"Select completed verses to re-analyze for '{word}':", font=('Arial', 12, 'bold'), bg='light gray').pack(padx=12, pady=(10,6), anchor='w')
            outer = tk.Frame(win, bg='light gray')
            outer.pack(fill=tk.BOTH, expand=True, padx=12, pady=8)
            canvas = tk.Canvas(outer, bg='light gray', highlightthickness=0)
            vsb = tk.Scrollbar(outer, orient='vertical', command=canvas.yview)
            canvas.configure(yscrollcommand=vsb.set)
            vsb.pack(side='right', fill='y')
            canvas.pack(side='left', fill='both', expand=True)
            list_frame = tk.Frame(canvas, bg='light gray')
            canvas.create_window((0, 0), window=list_frame, anchor='nw')

            sel_vars = []
            rows = df.reset_index(drop=True).to_dict('records')

            def _on_configure(event=None):
                try:
                    canvas.configure(scrollregion=canvas.bbox('all'))
                except Exception:
                    pass
            list_frame.bind('<Configure>', _on_configure)

            # Top toolbar: Select All / Clear All
            toolbar = tk.Frame(win, bg='light gray')
            toolbar.pack(fill=tk.X, padx=12, pady=(0, 8))
            def _select_all():
                for v, _ in sel_vars:
                    try:
                        v.set(True)
                    except Exception:
                        pass
            def _clear_all():
                for v, _ in sel_vars:
                    try:
                        v.set(False)
                    except Exception:
                        pass
            tk.Button(toolbar, text="Select All", bg='teal', fg='white', command=_select_all).pack(side=tk.LEFT)
            tk.Button(toolbar, text="Clear All", bg='gray', fg='white', command=_clear_all).pack(side=tk.LEFT, padx=(6,0))

            for i, rec in enumerate(rows):
                rf = tk.Frame(list_frame, bg='light gray')
                rf.pack(fill=tk.X, pady=2)
                v = tk.BooleanVar(value=False)
                sel_vars.append((v, i))
                tk.Checkbutton(rf, variable=v, bg='light gray').pack(side=tk.LEFT)
                verse = str(rec.get('verse', ''))
                page = str(rec.get('page_number', ''))
                tk.Label(rf, text=verse, font=('Arial', 12), width=64, anchor='w', bg='light gray', justify='left', wraplength=900).pack(side=tk.LEFT)
                tk.Label(rf, text=page, font=('Arial', 12), width=8, anchor='e', bg='light gray').pack(side=tk.LEFT)

            btns = tk.Frame(win, bg='light gray')
            btns.pack(fill=tk.X, padx=12, pady=(8, 10))

            def _apply_and_start(selected_only=True):
                try:
                    indices = [i for v, i in sel_vars if v.get()] if selected_only else list(range(len(rows)))
                    if not indices:
                        return
                    # Reset chosen Progress rows to 'not started' and mark Words as incomplete
                    p2 = prog_df.copy()
                    if 'word_key_norm' in p2.columns:
                        maskw = p2['word_key_norm'].astype(str).map(lambda s: _normalize_simple(s)) == _normalize_simple(norm)
                    else:
                        maskw = pd.Series([False] * len(p2))
                    # Narrow to only completed rows selected
                    if 'status' in p2.columns:
                        maskw = maskw & (p2['status'].astype(str).str.lower() == 'completed')
                    subset = p2.loc[maskw].reset_index()
                    keep_idx = set(subset.loc[indices, 'index'].tolist())
                    sel_mask = p2.index.isin(list(keep_idx))
                    if 'status' in p2.columns:
                        p2.loc[sel_mask, 'status'] = 'not started'
                    if 'completed_at' in p2.columns:
                        try:
                            p2.loc[sel_mask, 'completed_at'] = None
                        except Exception:
                            pass

                    w2 = words_df.copy()
                    if 'word_key_norm' in w2.columns:
                        m2 = w2['word_key_norm'].astype(str).map(lambda s: _normalize_simple(s)) == _normalize_simple(norm)
                        if 'analysis_completed' in w2.columns:
                            w2.loc[m2, 'analysis_completed'] = False
                        if 'analysis_completed_at' in w2.columns:
                            try:
                                w2.loc[m2, 'analysis_completed_at'] = None
                            except Exception:
                                pass

                    _save_tracker(tracker_path, w2, p2, others, TRACKER_WORDS_SHEET, TRACKER_PROGRESS_SHEET)
                    try:
                        win.destroy()
                    except Exception:
                        pass
                    self.start_word_driver_for(word)
                except Exception:
                    pass

            tk.Button(btns, text="Re-analyze Selected", bg='navy', fg='white', font=('Arial', 11, 'bold'), command=lambda: _apply_and_start(True)).pack(side=tk.RIGHT)
            tk.Button(btns, text="Re-analyze All", bg='teal', fg='white', font=('Arial', 11, 'bold'), command=lambda: _apply_and_start(False)).pack(side=tk.RIGHT, padx=(0,6))
            tk.Button(btns, text="Cancel", bg='gray', fg='white', font=('Arial', 11), command=win.destroy).pack(side=tk.RIGHT, padx=(0,6))
        except Exception as e:
            try:
                messagebox.showerror("Tracker Error", f"Failed to load tracker: {e}")
            except Exception:
                pass

    def _reset_progress_for_word_and_restart(self, word: str, only_selected: list[int] | None = None):
        """Reset Progress rows to 'not started' for the given word (optionally by row indices),
        mark Words incomplete, save, then start the driver."""
        try:
            tracker_path = self._get_word_tracker_path()
            words_df, prog_df, others = load_word_tracker(tracker_path, TRACKER_WORDS_SHEET, TRACKER_PROGRESS_SHEET)
            norm = self._norm_tok(word)
            p2 = prog_df.copy()
            # Build mask for this word
            if 'word_key_norm' in p2.columns:
                maskw = p2['word_key_norm'].astype(str).map(lambda s: _normalize_simple(s)) == _normalize_simple(norm)
            else:
                maskw = pd.Series([False] * len(p2))
            if 'status' in p2.columns:
                maskw = maskw & (p2['status'].astype(str).str.lower() == 'completed')
            # If only_selected provided, map them to source indices
            if only_selected is not None:
                subset = p2.loc[maskw].reset_index()
                keep_idx = set(subset.loc[subset.index.isin(only_selected), 'index'].tolist())
                sel_mask = p2.index.isin(list(keep_idx))
            else:
                sel_mask = maskw
            if 'status' in p2.columns:
                p2.loc[sel_mask, 'status'] = 'not started'
            if 'completed_at' in p2.columns:
                try:
                    p2.loc[sel_mask, 'completed_at'] = None
                except Exception:
                    pass

            w2 = words_df.copy()
            if 'word_key_norm' in w2.columns:
                m2 = w2['word_key_norm'].astype(str).map(lambda s: _normalize_simple(s)) == _normalize_simple(norm)
                if 'analysis_completed' in w2.columns:
                    w2.loc[m2, 'analysis_completed'] = False
                if 'analysis_completed_at' in w2.columns:
                    try:
                        w2.loc[m2, 'analysis_completed_at'] = None
                    except Exception:
                        pass

            _save_tracker(tracker_path, w2, p2, others, TRACKER_WORDS_SHEET, TRACKER_PROGRESS_SHEET)
            self.start_word_driver_for(word)
        except Exception:
            pass

    def _banner_wraplength(self, win=None) -> int:
        """Return a wraplength tuned to the window width (clamped 600ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ900)."""
        try:
            target = win or (self.match_window if hasattr(self, "match_window") else None)
            if target and target.winfo_exists():
                target.update_idletasks()
                w = target.winfo_width()
                return max(600, min(900, w - 120))
        except Exception:
            pass
        return 900

    def _modal_wraplength(self, win=None) -> int:
        """Return a wraplength tuned for the small modal (clamped 360ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ520)."""
        try:
            target = win or (self.root if hasattr(self, "root") else getattr(self, "match_window", None))
            if target and target.winfo_exists():
                target.update_idletasks()
                w = target.winfo_width()
                return max(360, min(520, w - 200))
        except Exception:
            pass
        return 400

    def _on_match_window_resize(self, event=None):
        """Resize handler to reflow the inline banner text, if present."""
        try:
            if hasattr(self, "literal_note_body") and self.literal_note_body and self.literal_note_body.winfo_exists():
                self.literal_note_body.config(wraplength=self._banner_wraplength(self.match_window))
        except Exception:
            pass

    def _ensure_literal_banner(self, text: str):
        """Create/reuse and render the inline literal-analysis banner."""
        reuse_ok = (
            hasattr(self, "literal_note_frame")
            and self.literal_note_frame
            and self.literal_note_frame.winfo_exists()
            and self.literal_note_frame.master is self.match_window
        )
        if not reuse_ok:
            if hasattr(self, "literal_note_frame") and getattr(self, "literal_note_frame", None):
                try:
                    if self.literal_note_frame.winfo_exists():
                        self.literal_note_frame.destroy()
                except Exception:
                    pass
            self.literal_note_frame = tk.Frame(
                self.match_window, bg="AntiqueWhite", relief="groove", bd=2
            )
            self.literal_note_title = tk.Label(
                self.literal_note_frame,
                text="Important Note ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â Literal Analysis",
                bg="AntiqueWhite",
                font=("Arial", 14, "bold"),
            )
            self.literal_note_title.pack(anchor="w", padx=10, pady=(5, 0))
            self.literal_note_body = tk.Label(
                self.literal_note_frame,
                bg="AntiqueWhite",
                wraplength=self._banner_wraplength(self.match_window),
                justify=tk.LEFT,
                font=("Arial", 12),
            )
        if not self.literal_note_frame.winfo_ismapped():
            self.literal_note_frame.pack(fill=tk.X, padx=20, pady=(5, 10))
        if not self.literal_note_body.winfo_ismapped():
            self.literal_note_body.pack(anchor="w", padx=10, pady=(0, 5))
        self.literal_note_body.config(
            text=text, wraplength=self._banner_wraplength(self.match_window)
        )
        try:
            if not getattr(self, "_inline_resize_bound", False):
                self.match_window.bind("<Configure>", self._on_match_window_resize, add="+")
                self._inline_resize_bound = True
        except Exception:
            pass

    def _has_repeat(self, norm_words, norm_target: str) -> bool:
        """Return True iff *norm_target* appears at least twice in ``norm_words``.

        ``norm_words`` is expected to be a list of pre-normalized tokens so this
        helper can operate without re-normalizing each time it is called.
        """
        if not norm_target:
            return False
        return norm_words.count(norm_target) >= 2

    def _maybe_show_repeat_important_note(self, word, occurrence_idx, verse_norm):
        """Show an explanatory note for repeated words within a verse."""
        if occurrence_idx < 1 or self._suppress_repeat_notes_for_verse:
            return
        norm_word = self._norm_tok(word)
        if not norm_word:
            return
        norm_verse = self._verse_key(verse_norm)
        key = (norm_verse, norm_word, "second")
        if key in self._repeat_note_shown:
            return
        self._repeat_note_shown.add(key)

        top = tk.Toplevel(self.root)
        top.title("Important Note - Literal Analysis")
        top.configure(bg='AntiqueWhite')
        top.transient(self.root)
        top.grab_set()
        try:
            self._wm_for(top)
        except Exception:
            pass

        body_lbl = tk.Label(
            top,
            text=self._LITERAL_NOTE_TEXT,
            bg='AntiqueWhite',
            wraplength=self._modal_wraplength(top),
            justify=tk.LEFT,
            font=('Arial', 12)
        )
        body_lbl.pack(padx=20, pady=(15,5))
        # Reflow text on modal resize
        try:
            top.bind("<Configure>", lambda e: body_lbl.config(wraplength=self._modal_wraplength(top)))
        except Exception:
            pass

        dont_var = tk.BooleanVar(value=False)
        tk.Checkbutton(
            top,
            text="Don't show again for this verse",
            variable=dont_var,
            bg='AntiqueWhite',
            font=('Arial', 11)
        ).pack(pady=(0,10))

        def _commit_and_close():
            try:
                if dont_var.get():
                    self._suppress_repeat_notes_for_verse = True
            except Exception:
                pass
            top.destroy()

        ok_btn = tk.Button(
            top,
            text="OK",
            command=_commit_and_close,
            font=('Arial', 12, 'bold'),
            bg='navy',
            fg='white',
            padx=10,
            pady=5
        )
        ok_btn.pack(pady=(0,15))
        try:
            ok_btn.focus_set()
            top.bind("<Return>", lambda e: _commit_and_close(), add="+")
            top.bind("<Escape>", lambda e: _commit_and_close(), add="+")
        except Exception:
            pass

        def _on_close():
            _commit_and_close()
        top.protocol("WM_DELETE_WINDOW", _on_close)

        top.update_idletasks()
        w, h = top.winfo_width(), top.winfo_height()
        x = self.root.winfo_x() + (self.root.winfo_width() - w)//2
        y = self.root.winfo_y() + (self.root.winfo_height() - h)//2
        top.geometry(f"{w}x{h}+{x}+{y}")
        top.wait_window()

    def show_dashboard(self):
        """Creates the dashboard interface directly in the main root window."""
        # Clear any existing widgets from the root
        for widget in self.root.winfo_children():
            widget.destroy()

        # Set up the dashboard appearance in the root window
        self.root.title("Dashboard")
        self.root.configure(bg='light gray')
        try:
            self._wm_for(self.root).maximize()
        except Exception:
            pass

        # Dashboard header label
        header = tk.Label(
            self.root,
            text="Welcome to Gurbani Software Dashboard",
            font=('Arial', 18, 'bold'),
            bg='dark slate gray',
            fg='white'
        )
        header.pack(fill=tk.X, pady=20)

        # Create a frame to hold dashboard buttons
        button_frame = tk.Frame(self.root, bg='light gray')
        button_frame.pack(expand=True)

        # New Button to open the Verse Analysis Dashboard
        verse_analysis_btn = tk.Button(
            button_frame,
            text="Verse Analysis Dashboard",
            font=('Arial', 14, 'bold'),
            bg='dark cyan',
            fg='white',
            padx=20,
            pady=10,
            command=self.launch_verse_analysis_dashboard
        )
        verse_analysis_btn.pack(pady=10)

        # Button to open the GrammarÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‹Å“DB Update window
        grammar_update_btn = tk.Button(
            button_frame,
            text="Grammar DB Update",
            font=('Arial', 14, 'bold'),
            bg='teal',
            fg='white',
            padx=20,
            pady=10,
            command=self.launch_grammar_update_dashboard
        )
        grammar_update_btn.pack(pady=10)

        # Placeholder for future features (e.g., Grammar Correction)
        future_btn = tk.Button(
            button_frame,
            text="Upcoming Feature: Grammar Correction",
            font=('Arial', 14, 'bold'),
            bg='gray',
            fg='white',
            padx=20,
            pady=10,
            state=tk.DISABLED
        )
        future_btn.pack(pady=10)

        # What's New / Releases button
        whats_new_btn = tk.Button(
            button_frame,
            text="What's New",
            font=('Arial', 12, 'bold'),
            bg='#2f4f4f',
            fg='white',
            padx=16,
            pady=8,
            command=self.show_whats_new
        )
        whats_new_btn.pack(pady=(20, 10))

        # One-time prompt to announce recent UI changes
        try:
            self.root.after(800, self.maybe_prompt_whats_new)
        except Exception:
            pass

    def show_whats_new(self):
        """Display a dialog with recent UI updates and release links."""
        win = tk.Toplevel(self.root)
        win.title("What's New")
        win.configure(bg='light gray')
        win.transient(self.root)
        try:
            win.grab_set()
        except Exception:
            pass
        try:
            self._wm_for(win)
        except Exception:
            pass

        header = tk.Label(
            win,
            text="What's New",
            font=('Arial', 16, 'bold'),
            bg='dark slate gray',
            fg='white',
            pady=8
        )
        header.pack(fill=tk.X)

        body = tk.Frame(win, bg='light gray')
        body.pack(fill=tk.BOTH, expand=True, padx=20, pady=16)

        tk.Label(
            body,
            text=(
                "Recent UI improvements: two-column card parity in matches view, "
                "centered layout, equal column widths, radios never overlap text, and a centered final "
                "card without stretching when there's an odd number of results."
            ),
            font=('Arial', 12),
            bg='light gray',
            wraplength=800,
            justify='left'
        ).pack(anchor='w', pady=(0, 10))

        links = tk.Frame(body, bg='light gray')
        links.pack(anchor='w', pady=(0, 8))

        def link(label_parent, text, url):
            lbl = tk.Label(
                label_parent,
                text=text,
                font=('Arial', 12, 'underline'),
                fg='blue',
                bg='light gray',
                cursor='hand2'
            )
            lbl.pack(anchor='w', pady=2)
            lbl.bind('<Button-1>', lambda e: webbrowser.open(url))

        link(links, f"View UI tag: {WHATS_NEW_ID}",
             f"https://github.com/vije1711/Birha/tree/{WHATS_NEW_ID}")
        link(links, 'All Releases', 'https://github.com/vije1711/Birha/releases')

        btns = tk.Frame(win, bg='light gray')
        btns.pack(fill=tk.X, padx=20, pady=(0, 16))
        tk.Button(
            btns,
            text='Close',
            font=('Arial', 12, 'bold'),
            bg='gray', fg='white',
            padx=16, pady=6,
            command=win.destroy
        ).pack(side=tk.RIGHT)

        # Center over root
        win.update_idletasks()
        try:
            w, h = win.winfo_width(), win.winfo_height()
            x = self.root.winfo_x() + (self.root.winfo_width() - w)//2
            y = self.root.winfo_y() + (self.root.winfo_height() - h)//2
            win.geometry(f"{w}x{h}+{x}+{y}")
        except Exception:
            pass

    def _go_back_to_dashboard(self, win=None):
        """Minimal helper: return to main dashboard in the root window.

        Optional 'win' is ignored for API compatibility with other back buttons.
        """
        try:
            if win is not None and hasattr(win, 'destroy'):
                # If a child window was passed accidentally, close it safely
                win.destroy()
        except Exception:
            pass
        self.show_dashboard()

    def launch_word_assessment_dashboard(self):
        """Root-rendered UI for 'Assess by Word' dashboard (UI-only, no logic)."""
        # Clear any existing widgets from the root and prep the surface
        for widget in self.root.winfo_children():
            try:
                widget.destroy()
            except Exception:
                pass

        self.root.title("Assess by Word")
        try:
            self.root.configure(bg='light gray')
            try:
                self._wm_for(self.root).maximize()
            except Exception:
                pass
        except Exception:
            pass

        # Header
        header = tk.Label(
            self.root,
            text="Assess by Word - Dashboard",
            font=('Arial', 18, 'bold'),
            bg='dark slate gray',
            fg='white',
            pady=10
        )
        header.pack(fill=tk.X, pady=(0, 10))

        # Central buttons container
        body = tk.Frame(self.root, bg='light gray')
        body.pack(expand=True)

        # New Assessment menu housing the Word Search (Lexicon) launcher
        def open_new_word_assessment_menu():
            win = tk.Toplevel(self.root)
            win.title("New Assessment ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ Assess by Word")
            win.configure(bg='light gray')
            try:
                mgr = self._wm_for(win)
                if mgr:
                    mgr.maximize()
            except Exception:
                # Some platforms/window managers may not support maximize; ignore silently
                pass

            header = tk.Label(
                win,
                text="Start a New Assessment",
                font=('Arial', 18, 'bold'),
                bg='dark slate gray', fg='white', pady=10
            )
            header.pack(fill=tk.X)

            body_inner = tk.Frame(win, bg='light gray')
            body_inner.pack(expand=True, padx=24, pady=20)

            # Launcher: Word Search (Lexicon) ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ opens existing modal
            tk.Button(
                body_inner,
                text="Word Search (Lexicon)",
                font=('Arial', 14, 'bold'),
                bg='#4b8bbe', fg='white',
                padx=20, pady=10,
                command=self.show_word_search_modal
            ).pack(pady=10)

            footer = tk.Frame(win, bg='light gray')
            footer.pack(fill=tk.X, padx=24, pady=(0, 16))
            tk.Button(
                footer,
                text="< Back",
                font=('Arial', 12, 'bold'),
                bg='#2f4f4f', fg='white',
                padx=16, pady=6,
                command=lambda: self._go_back_to_dashboard(win)
            ).pack(side=tk.LEFT)
            tk.Button(
                footer,
                text="Close",
                font=('Arial', 12, 'bold'),
                bg='gray', fg='white',
                padx=16, pady=6,
                command=win.destroy
            ).pack(side=tk.RIGHT)

        tk.Button(
            body,
            text="New Assessment",
            font=('Arial', 14, 'bold'),
            bg='dark cyan', fg='white',
            padx=20, pady=10,
            command=open_new_word_assessment_menu
        ).pack(pady=8)

        tk.Button(
            body,
            text="Continue Incomplete",
            font=('Arial', 14, 'bold'),
            bg='teal', fg='white',
            padx=20, pady=10,
            command=self.show_continue_incomplete
        ).pack(pady=8)

        tk.Button(
            body,
            text="View Completed",
            font=('Arial', 14, 'bold'),
            bg='#2f4f4f', fg='white',
            padx=20, pady=10,
            command=self.show_view_completed
        ).pack(pady=8)

        # Bottom bar with Back
        bottom = tk.Frame(self.root, bg='light gray')
        bottom.pack(fill=tk.X, padx=20, pady=15)
        tk.Button(
            bottom,
            text="< Back",
            font=('Arial', 12, 'bold'),
            bg='gray', fg='white',
            padx=14, pady=6,
            command=self._go_back_to_dashboard  # no 'win' arg
        ).pack(side=tk.LEFT)

    # ---------------------------
    # One-time "What's New" helpers
    # ---------------------------
    def _state_file_path(self):
        try:
            base = os.path.join(os.path.expanduser("~"), ".birha")
            os.makedirs(base, exist_ok=True)
            return os.path.join(base, "state.json")
        except Exception:
            # Fallback to current working directory if home is not accessible
            return os.path.join(os.getcwd(), ".birha_state.json")

    def _load_state(self):
        path = self._state_file_path()
        try:
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {}

    def _save_state(self, state: dict):
        path = self._state_file_path()
        try:
            with open(path, "w", encoding="utf-8") as f:
                json.dump(state, f)
        except Exception:
            pass

    def maybe_prompt_whats_new(self):
        # Avoid multiple prompts per session
        if getattr(self, "_whats_new_checked", False):
            return
        self._whats_new_checked = True

        state = self._load_state()
        if state.get("last_whats_new") != WHATS_NEW_ID:
            # Ask the user if they'd like to view details, then record as shown
            self.root.after(100, lambda: self._do_prompt_whats_new_fixed(state))

    # ---------------------------
    # Window manager helper
    # ---------------------------
    def _wm_for(self, win):
        """Return a WindowManager bound to 'win', creating and binding <F11> if needed."""
        try:
            mgr = getattr(win, "_wmgr", None)
            if mgr is None:
                mgr = WindowManager(win)
                setattr(win, "_wmgr", mgr)
            return mgr
        except Exception:
            return None

    def _wm_apply(self, win, margin_px: int | None = None, defer: bool = True):
        """Attach WindowManager to 'win' and optionally maximize.

        - margin_px=None: only bind F11 (no sizing).
        - margin_px>0: maximize_client(margin_px) for taskbar-safe sizing.
        - margin_px==0: normal maximize().
        If defer is True, schedule sizing after idle to let geometry realize first.
        Returns the WindowManager or None.
        """
        mgr = self._wm_for(win)
        if not mgr:
            return None
        def _do_size():
            try:
                if margin_px is None:
                    return
                if isinstance(margin_px, int) and margin_px > 0:
                    mgr.maximize_client(bottom_margin_px=margin_px)
                else:
                    mgr.maximize()
            except Exception:
                pass
        try:
            if margin_px is not None:
                if defer:
                    try:
                        win.after_idle(_do_size)
                    except Exception:
                        _do_size()
                else:
                    _do_size()
        except Exception:
            pass
        return mgr

    def _do_prompt_whats_new(self, state: dict):
        try:
            if messagebox.askyesno(
                "WhatÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â‚¬Å¾Ã‚Â¢s New",
                (
                    "WeÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â‚¬Å¾Ã‚Â¢ve improved verse selection cards: centered layout, equal column widths, "
                    "and radios no longer overlap text. View details now?"
                ),
            ):
                self.show_whats_new()
        finally:
            try:
                state["last_whats_new"] = WHATS_NEW_ID
                self._save_state(state)
            except Exception:
                pass
        return

        header = tk.Label(
            win,
            text="What's New",
            font=('Arial', 16, 'bold'),
            bg='dark slate gray',
            fg='white',
            pady=8
        )
        header.pack(fill=tk.X)

        body = tk.Frame(win, bg='light gray')
        body.pack(fill=tk.BOTH, expand=True, padx=20, pady=16)

        tk.Label(
            body,
            text=(
                "Recent UI improvements: two-column card parity in matches view, "
                "centered layout, equal column widths, radios never overlap text, and a centered final "
                "card without stretching when thereÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â‚¬Å¾Ã‚Â¢s an odd number of results."
            ),
            font=('Arial', 12),
            bg='light gray',
            wraplength=800,
            justify='left'
        ).pack(anchor='w', pady=(0, 10))

        links = tk.Frame(body, bg='light gray')
        links.pack(anchor='w', pady=(0, 8))

        def link(label_parent, text, url):
            lbl = tk.Label(
                label_parent,
                text=text,
                font=('Arial', 12, 'underline'),
                fg='blue',
                bg='light gray',
                cursor='hand2'
            )
            lbl.pack(anchor='w', pady=2)
            lbl.bind('<Button-1>', lambda e: webbrowser.open(url))

        link(links, f"View UI tag: {WHATS_NEW_ID}",
             f"https://github.com/vije1711/Birha/tree/{WHATS_NEW_ID}")
        link(links, 'All Releases', 'https://github.com/vije1711/Birha/releases')

        btns = tk.Frame(win, bg='light gray')
        btns.pack(fill=tk.X, padx=20, pady=(0, 16))
        tk.Button(
            btns,
            text='Close',
            font=('Arial', 12, 'bold'),
            bg='gray', fg='white',
            padx=16, pady=6,
            command=win.destroy
        ).pack(side=tk.RIGHT)

        # Center over root
        win.update_idletasks()
        try:
            w, h = win.winfo_width(), win.winfo_height()
            x = self.root.winfo_x() + (self.root.winfo_width() - w)//2
            y = self.root.winfo_y() + (self.root.winfo_height() - h)//2
            win.geometry(f"{w}x{h}+{x}+{y}")
        except Exception:
            pass

    def launch_grammar_update_dashboard(self):
        win = tk.Toplevel(self.root)
        win.title("Grammar Database Update")
        win.configure(bg='#e0e0e0')  # light neutral background
        # Taskbar-safe sizing
        self._wm_apply(win, margin_px=0, defer=True)

        # ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â Header Bar ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â
        header = tk.Frame(win, bg='#2f4f4f', height=60)
        header.pack(fill=tk.X)
        tk.Label(
            header,
            text="Grammar Database Update",
            font=('Arial', 20, 'bold'),
            bg='#2f4f4f',
            fg='white'
        ).place(relx=0.5, rely=0.5, anchor='center')

        # ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â Separator ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â
        sep = tk.Frame(win, bg='#cccccc', height=2)
        sep.pack(fill=tk.X)

        # ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â Navigation Buttons ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â
        nav = tk.Frame(win, bg='#e0e0e0')
        nav.pack(pady=30)
        btn_kwargs = dict(
            font=('Arial', 14, 'bold'),
            width=20,
            padx=10, pady=10,
            relief='flat',
            activebackground='#007d7d',
            bg='#008c8c', fg='white'
        )

        btn_verse = tk.Button(
            nav, text="Assess by Verse", **btn_kwargs,
            command=self.launch_verse_assessment
        )
        btn_verse.grid(row=0, column=0, padx=20)

        btn_word = tk.Button(
            nav, text="Assess by Word", **btn_kwargs,
            command=lambda: (win.destroy(), self.launch_word_assessment_dashboard())
        )
        btn_word.grid(row=0, column=1, padx=20)

        # ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â Instruction / Description ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â
        instr = (
            "Choose ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“Assess by VerseÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â to look up verses and refine grammar entries.\n"
            "The ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“Assess by WordÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â workflow is coming in the next release."
        )
        tk.Label(
            win, text=instr,
            font=('Arial', 16),
            bg='#e0e0e0', fg='#333333',
            justify='center', wraplength=800
        ).pack(pady=20)

        # ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â Bottom Back Button ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â
        bottom = tk.Frame(win, bg='#e0e0e0')
        bottom.pack(side=tk.BOTTOM, pady=30)
        back_btn = tk.Button(
            bottom,
            text="ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â Ãƒâ€šÃ‚Â Back to Dashboard",
            font=('Arial', 14),
            bg='#2f4f4f', fg='white',
            activebackground='#3f6f6f',
            padx=20, pady=10,
            command=lambda: self._go_back_to_dashboard(win)
        )
        back_btn.pack()

        # Optional: make ESC key close this window
        win.bind("<Escape>", lambda e: win.destroy())

    def launch_verse_assessment(self):
        """Window for searching & selecting verses to assess grammar using a 2-column card layout."""
        win = tk.Toplevel(self.root)
        win.title("Assess by Verse")
        win.configure(bg='light gray')
        # Taskbar-safe sizing
        self._wm_apply(win, margin_px=0, defer=True)
        
        # ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â Optional pageÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Âwide heading ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â
        tk.Label(
            win,
            text="Select a Verse to Refine Grammar Entries",
            font=("Arial", 20, "bold"),
            bg="dark slate gray",
            fg="white",
            pady=10
        ).pack(fill=tk.X)

        # keep track of which card is selected
        self._selected_verse_idx = tk.IntVar(value=-1)
        # ensure safe defaults so Next can't crash before a search
        self._last_filtered = []
        try:
            self._selected_verse_idx.trace_add("write", lambda *args: self._update_next_button_state())
        except Exception:
            pass

        # ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â Top frame: entry + Search button ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â
        top = tk.Frame(win, bg='light gray')
        top.pack(fill=tk.X, padx=20, pady=15)
        tk.Label(top, text="Enter Verse:", font=("Arial", 16), bg='light gray').pack(side=tk.LEFT)
        self._verse_var = tk.StringVar()
        tk.Entry(top, textvariable=self._verse_var, font=("Arial", 16))\
        .pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(10,10))
        tk.Button(
            top, text="Search", font=("Arial", 16, "bold"),
            bg='dark cyan', fg='white',
            command=self._populate_cards
        ).pack(side=tk.LEFT)

        # ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â Middle frame: scrollable canvas + 2ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‹Å“column grid of ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“cardsÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â
        middle = tk.Frame(win, bg='light gray')
        middle.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)

        canvas = tk.Canvas(middle, bg='light gray', highlightthickness=0)
        vsb    = tk.Scrollbar(middle, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        # This frame goes inside the canvas and will hold our cards
        self._cards_frame = tk.Frame(canvas, bg='light gray')

        # create_window with anchor="n" so its x coordinate is the top-center of cards_frame
        cards_window = canvas.create_window((0, 0), window=self._cards_frame, anchor="n")

        # configure two equalÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‹Å“weight columns for 2ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‹Å“column layout
        self._cards_frame.grid_columnconfigure(0, weight=1, minsize=450)
        self._cards_frame.grid_columnconfigure(1, weight=1, minsize=450)

        # keep scrollregion up to date
        def _on_cards_configure(event):
            canvas.configure(scrollregion=canvas.bbox("all"))
        self._cards_frame.bind("<Configure>", _on_cards_configure)

        # **New:** whenever the canvas resizes, recenter cards_frame horizontally
        def _on_canvas_resize(event):
            canvas.coords(cards_window, event.width // 2, 0)
        canvas.bind("<Configure>", _on_canvas_resize)

        # ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â Bottom frame: navigation buttons ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â
        bottom = tk.Frame(win, bg='light gray')
        bottom.pack(fill=tk.X, padx=20, pady=15)
        tk.Button(
            bottom, text="ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¹ Back", font=("Arial", 14),
            bg='gray', fg='white', command=win.destroy
        ).pack(side=tk.LEFT)
        tk.Button(
            bottom, text="Back to Dashboard", font=("Arial", 14),
            bg='gray', fg='white', command=lambda: self._go_back_to_dashboard(win)
        ).pack(side=tk.LEFT, padx=5)
        tk.Button(
            bottom, text="Next ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢", font=("Arial", 14, "bold"),
            bg='dark cyan', fg='white',
            command=lambda: self.proceed_to_word_assessment(self._selected_verse_idx.get())
        ).pack(side=tk.RIGHT)
        # Keep a reference to the Next button and disable it until selection
        try:
            self._next_btn = bottom.winfo_children()[-1]
            self._next_btn.configure(state=tk.DISABLED)
        except Exception:
            pass

    def _update_next_button_state(self, *args):
        """Enable Next only when there are results and a valid selection."""
        try:
            results = getattr(self, "_last_filtered", [])
            idx = self._selected_verse_idx.get() if hasattr(self, "_selected_verse_idx") else -1
            ok = bool(results) and (0 <= idx < len(results))
            if hasattr(self, "_next_btn") and self._next_btn and self._next_btn.winfo_exists():
                self._next_btn.configure(state=(tk.NORMAL if ok else tk.DISABLED))
        except Exception:
            try:
                if hasattr(self, "_next_btn") and self._next_btn and self._next_btn.winfo_exists():
                    self._next_btn.configure(state=tk.DISABLED)
            except Exception:
                pass

    def _populate_cards(self):
        """Perform the verse search, filter & then render up to 10 cards in two columns."""
        # first, clear any existing cards
        for w in self._cards_frame.winfo_children():
            w.destroy()
        # disable Next and clear stale results on refresh
        try:
            self._last_filtered = []
            self._selected_verse_idx.set(-1)
            self._update_next_button_state()
        except Exception:
            pass

        # Ensure equal column widths using a uniform group to avoid asymmetry
        try:
            self._cards_frame.grid_columnconfigure(0, weight=1, minsize=450, uniform='cards')
            self._cards_frame.grid_columnconfigure(1, weight=1, minsize=450, uniform='cards')
        except Exception:
            # Fallback in case uniform isn't supported, keep prior settings
            self._cards_frame.grid_columnconfigure(0, weight=1, minsize=450)
            self._cards_frame.grid_columnconfigure(1, weight=1, minsize=450)

        # 1) run search & filter
        query = self._verse_var.get().strip()
        headers, all_matches = self.match_sggs_verse(query)
        filtered = [m for m in all_matches if m.get("Score",0) >= 25.0][:10]
        # remember these for the ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“Next ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â step
        self._last_filtered = filtered

        # reset selection
        self._selected_verse_idx.set(-1)

        # 2) render each card
        total_cards = len(filtered)
        for idx, m in enumerate(filtered):
            row, col = divmod(idx, 2)
            card = tk.Frame(
                self._cards_frame,
                bd=1,
                relief="solid",
                bg="white",
                padx=8,
                pady=8
            )
            # If odd number of cards and this is the last one, span both columns for visual centering
            if (total_cards % 2 == 1) and (idx == total_cards - 1):
                # Do not stretch the final full-width card; keep it centered with natural width
                card.grid(row=row, column=0, columnspan=2, padx=10, pady=10, sticky="n")
            else:
                card.grid(row=row, column=col, padx=10, pady=10, sticky="nsew")

            # the verse itself, wrapped
            tk.Label(
                card,
                text=m.get("Verse","").strip(),
                font=("Arial", 14, "bold"),
                wraplength=500,
                justify="center",
                bg="white"
            ).pack(pady=(14,4), padx=(28,8))

            # a little radiobutton at top-left for selection (created after the label and raised)
            rb = tk.Radiobutton(
                card,
                variable=self._selected_verse_idx,
                value=idx,
                bg="white",
                activebackground="white",
                command=self._update_next_button_state
            )
            rb.place(x=6, y=6)
            try:
                rb.lift()
            except Exception:
                pass

            # metadata line
            # build a list of (label, key) pairs
            fields = [
                ("Raag",   "Raag (Fixed)"),
                ("Writer", "Writer (Fixed)"),
                ("Bani",   "Bani Name"),
                ("Page",   "Page Number"),
            ]

            meta_parts = []
            for label, key in fields:
                v = m.get(key)
                # skip if missing or NaN
                if v is None or (isinstance(v, float) and math.isnan(v)):
                    continue
                meta_parts.append(f"{label}: {v}")

            # always include the match%
            meta_parts.append(f"Match: {m.get('Score',0):.1f}%")

            # join with separators
            meta = "   |   ".join(meta_parts)

            tk.Label(
                card,
                text=meta,
                font=("Arial", 12),
                bg="white"
            ).pack()

        # 3) force a canvas update of its scroll region
        self._cards_frame.update_idletasks()
        self._cards_frame.master.configure(
            scrollregion=self._cards_frame.master.bbox("all")
        )

    def show_translation_input(self):
        win = tk.Toplevel(self.root)
        win.title("Paste Darpan Translation")
        win.configure(bg='light gray')
        # Taskbar-safe sizing: exact work-area maximize (no extra bottom gap)
        self._wm_apply(win, margin_px=0, defer=True)
        win.transient(self.root)
        win.grab_set()

        # Ensure any close action (X or Back button wired to win.destroy) cancels driver state safely
        try:
            _orig_destroy = win.destroy

            def _driver_safe_destroy():
                try:
                    if getattr(self, '_abw_suppress_driver_cancel_once', False):
                        try:
                            self._abw_suppress_driver_cancel_once = False
                        except Exception:
                            pass
                    elif getattr(self, '_word_driver_active', False) and getattr(self, '_word_driver_in_progress', False):
                        try:
                            self._word_driver_cancel_current()
                        except Exception:
                            pass
                finally:
                    try:
                        _orig_destroy()
                    except Exception:
                        pass
            win.destroy = _driver_safe_destroy
            try:
                win.protocol("WM_DELETE_WINDOW", win.destroy)
            except Exception:
                pass
        except Exception:
            pass

        # Prefer a Gurmukhi-safe font to avoid clipping of shirorekha and matras
        try:
            if not hasattr(self, '_gurmukhi_font_family'):
                families = set(map(str, tkfont.families()))
                candidates = [
                    'Nirmala UI', 'Raavi', 'Noto Sans Gurmukhi', 'Noto Serif Gurmukhi',
                    'GurbaniAkhar', 'GurbaniAkhar-Thick', 'AnmolLipi', 'AnmolUni',
                    'Lohit Gurmukhi', 'Mukta Mahee', 'Saab', 'Gurmukhi MN'
                ]
                chosen = None
                for name in candidates:
                    if name in families:
                        chosen = name
                        break
                if not chosen:
                    # fall back to Tk's default family rather than Arial to keep system fallback working
                    chosen = tkfont.nametofont('TkDefaultFont').cget('family')
                self._gurmukhi_font_family = chosen
        except Exception:
            if not hasattr(self, '_gurmukhi_font_family'):
                # last resort: don't force a specific family; rely on Tk's default
                try:
                    self._gurmukhi_font_family = tkfont.nametofont('TkDefaultFont').cget('family')
                except Exception:
                    self._gurmukhi_font_family = 'TkDefaultFont'

        # - Heading -
        tk.Label(
            win,
            text=self.selected_verse_text,
            # Use a font with proper ascent for Gurmukhi + extra top padding
            font=(self._gurmukhi_font_family, 20, "bold"),
            bg="light gray",
            wraplength=900,
            justify="center",
            pady=12
        ).pack(fill=tk.X, padx=20, pady=(15,10))

        # Make a content container so bottom buttons are anchored reliably
        content = tk.Frame(win, bg='light gray')
        content.pack(side=tk.TOP, fill=tk.BOTH, expand=True)

        # - Translation area -
        tf = tk.LabelFrame(
            content,
            text="Established Darpan Translation",
            font=("Arial", 14, "bold"),
            bg='light gray',
            fg='black',
            padx=10, pady=5
        )
        # allow translation area to take the extra space
        tf.pack(fill=tk.BOTH, expand=True, padx=20, pady=(0,15))

        # Text area with vertical scrollbar so content scrolls instead of pushing layout
        text_container = tk.Frame(tf, bg='light gray')
        text_container.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
        scroll_y = tk.Scrollbar(text_container, orient="vertical")
        self._translation_text = tk.Text(
            text_container, wrap=tk.WORD, font=(self._gurmukhi_font_family, 13),
            height=10, padx=8, pady=12
        )
        self._translation_text.configure(yscrollcommand=scroll_y.set)
        scroll_y.configure(command=self._translation_text.yview)
        self._translation_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scroll_y.pack(side=tk.RIGHT, fill=tk.Y)

        # Status + Refresh row under the translation box
        status_row = tk.Frame(tf, bg='light gray')
        # keep status row pinned to the bottom of the translation frame
        status_row.pack(side=tk.BOTTOM, fill=tk.X, pady=(6, 0))
        # grid inside the row: label left (expands), button right
        status_row.columnconfigure(0, weight=1)
        status_row.columnconfigure(1, weight=0)
        self._translation_status_var = tk.StringVar(value="")
        tk.Label(
            status_row,
            textvariable=self._translation_status_var,
            font=("Arial", 10, "italic"),
            bg='light gray', fg='#333333',
            anchor='w'
        ).grid(row=0, column=0, sticky='w', padx=(0,8))
        tk.Button(
            status_row,
            text="Refresh from data files",
            font=("Arial", 10),
            bg='gray', fg='white',
            command=self._refresh_translation_from_data
        ).grid(row=0, column=1, sticky='e')

        # Try to auto-populate translation from structured sources
        filled, status = self._populate_translation_from_structured()
        self._translation_status_var.set(status)

        # (Removed one-off height cap; scrollable text keeps buttons visible)

        # ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â WordÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Âselection area ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â
        wf = tk.LabelFrame(
            content,
            text="Select Words to Assess Grammar",
            font=("Arial", 14, "bold"),
            bg='light gray',
            fg='black',
            padx=10, pady=10
        )
        # reduce the vertical footprint of the word-selection area
        wf.pack(fill=tk.X, expand=False, padx=20, pady=(0,15))

        # select/deselect all
        self._select_all_words_var = tk.BooleanVar(value=False)
        tk.Checkbutton(
            wf,
            text="Select/Deselect All Words",
            variable=self._select_all_words_var,
            bg="light gray",
            font=("Arial", 12, "italic"),
            command=self._toggle_all_word_selection
        ).pack(anchor="w", pady=(0,10))

        # scrollable word row (single line, horizontal scroll)
        canvas = tk.Canvas(wf, bg='light gray', highlightthickness=0)
        # limit the canvas height to roughly a single checkbox row (~75% reduction)
        canvas.configure(height=48)
        scrollbar = tk.Scrollbar(wf, orient="horizontal", command=canvas.xview)
        word_frame = tk.Frame(canvas, bg='light gray')
        canvas.configure(xscrollcommand=scrollbar.set)

        canvas.pack(side="top", fill="x", expand=False)
        scrollbar.pack(side="top", fill="x")
        canvas.create_window((0,0), window=word_frame, anchor="nw")

        def _on_wf_resize(event):
            canvas.configure(scrollregion=canvas.bbox("all"))
        word_frame.bind("<Configure>", _on_wf_resize)

        # lay out each word
        self._word_selection_vars = []

        # 1) grab the verse text, remove any trailing danda symbols:
        verse_text = self.selected_verse_text.split('ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¥Ãƒâ€šÃ‚Â¥', 1)[0].strip()

        # 2) split into words (now ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¥Ãƒâ€šÃ‚Â¥ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â wonÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â‚¬Å¾Ã‚Â¢t appear as its own token)
        words = verse_text.split()

        # 3) build your checkboxes off `words` instead of the raw text:
        for i, w in enumerate(words):
            var = tk.BooleanVar(value=False)
            chk = tk.Checkbutton(
                word_frame,
                text=w,
                variable=var,
                bg='light gray',
                font=('Arial', 12),
                anchor='w',
                justify='left'
            )
            # arrange all checkboxes in a single row; scroll horizontally if needed
            chk.grid(row=0, column=i, sticky='w', padx=5, pady=3)
            self._word_selection_vars.append((var, w))

        # - Bottom buttons -
        btn_frame = tk.Frame(win, bg="light gray")
        # Anchor to bottom; keep visual gap equal to BOTTOM_PAD
        btn_frame.pack(side=tk.BOTTOM, fill=tk.X, padx=20, pady=(6, BOTTOM_PAD))

        tk.Button(
            btn_frame,
            text="ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â Ãƒâ€šÃ‚Â Back to Verse Search",
            font=("Arial", 12),
            bg="gray",
            fg="white",
            command=win.destroy,
            padx=15, pady=8
        ).pack(side=tk.LEFT)

        tk.Button(
            btn_frame,
            text="Submit Translation ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢",
            font=("Arial", 12, "bold"),
            bg="dark cyan",
            fg="white",
            command=lambda: self._on_translation_submitted(win),
            padx=15, pady=8
        ).pack(side=tk.RIGHT)

    def _populate_translation_from_structured(self):
        """Attempt to fill the translation Text from structured JSON/CSV.
        Returns ``(filled, status_message)``.
        """
        try:
            meta = getattr(self, 'selected_verse_meta', {}) or {}
            verse_text = self.selected_verse_text if hasattr(self, 'selected_verse_text') else ''
            page_num = meta.get('Page Number')
            record = _find_arth_for(self, verse_text, page_num)
            if not record:
                return False, "Manual input"
            v = record.get('verse') or record.get('Verse') or ''
            p = record.get('padarth') or record.get('Padarth') or ''
            a = record.get('arth') or record.get('Arth') or ''
            ch = record.get('chhand') or record.get('Chhand') or ''
            bh = record.get('bhav') or record.get('Bhav') or ''

            if not any([v, p, a, ch, bh]):
                return False, "Manual input"

            # Write into Text widget using a bold header tag for labels
            self._translation_text.delete('1.0', tk.END)
            try:
                self._translation_text.tag_configure('hdr', font=('Arial', 12, 'bold'))
            except Exception:
                # If tag config fails, continue without styling
                pass

            def _ins(label, value):
                if value:
                    self._translation_text.insert('end', f"{label}:\n", 'hdr')
                    self._translation_text.insert('end', str(value).strip() + "\n\n")

            _ins('Verse', v)
            _ins('Padarth', p)
            _ins('Arth', a)
            _ins('Chhand', ch)
            _ins('Bhav', bh)
            return True, "Auto-filled from structured data"
        except Exception:
            return False, "Manual input"

    def _refresh_translation_from_data(self):
        """Handler for the Refresh button to try loading from data files again."""
        filled, status = self._populate_translation_from_structured()
        if hasattr(self, '_translation_status_var'):
            self._translation_status_var.set(status if filled else "No structured data match found")

    # ---------------------------
    # Assess-by-Word: Dedicated translation/selection UI (no Verse deps)
    # ---------------------------
    def show_word_translation_input(self):
        win = tk.Toplevel(self.root)
        win.title("Assess by Word - Translation & Selection")
        win.configure(bg='light gray')
        # Taskbar-safe exact work-area maximize (deferred) and F11 toggle
        self._wm_apply(win, margin_px=0, defer=True)
        win.transient(self.root)
        try:
            win.grab_set()
        except Exception:
            pass

        # Safe destroy cancels driver state if in-progress
        try:
            _orig_destroy = win.destroy
            def _driver_safe_destroy():
                try:
                    if getattr(self, '_abw_suppress_driver_cancel_once', False):
                        try:
                            self._abw_suppress_driver_cancel_once = False
                        except Exception:
                            pass
                    elif getattr(self, '_word_driver_active', False) and getattr(self, '_word_driver_in_progress', False):
                        try:
                            self._word_driver_cancel_current()
                        except Exception:
                            pass
                finally:
                    try:
                        _orig_destroy()
                    except Exception:
                        pass
            win.destroy = _driver_safe_destroy
            win.protocol("WM_DELETE_WINDOW", win.destroy)
        except Exception:
            pass

        # Header with integrated driver controls
        hdr = tk.Frame(win, bg='dark slate gray')
        hdr.pack(fill=tk.X)
        self._abw_driver_status_var = tk.StringVar(value="")
        tk.Label(hdr, textvariable=self._abw_driver_status_var, font=("Arial", 12, 'bold'),
                 bg='dark slate gray', fg='white').pack(side=tk.LEFT, padx=12, pady=8)
        ctrls = tk.Frame(hdr, bg='dark slate gray')
        ctrls.pack(side=tk.RIGHT, padx=8, pady=6)
        # Prev/Next/Pause/Close
        def _prev():
            if getattr(self, '_word_driver_in_progress', False):
                return
            try:
                if getattr(self, '_word_driver_index', 0) > 0:
                    self._word_driver_index -= 1
                    self._word_driver_update_ui()
                    self._word_driver_open_current_verse()
            except Exception:
                pass
        def _next():
            if getattr(self, '_word_driver_in_progress', False):
                return
            try:
                self._word_driver_open_current_verse()
            except Exception:
                pass
        def _pause_toggle():
            try:
                self._word_driver_toggle_pause()
            except Exception:
                pass
        self._abw_driver_prev_btn = tk.Button(ctrls, text="ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ Prev", bg='gray', fg='white', command=_prev)
        self._abw_driver_prev_btn.pack(side=tk.LEFT, padx=(0,6))
        self._abw_driver_next_btn = tk.Button(ctrls, text="Next ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Å“Ãƒâ€šÃ‚Â¶", bg='navy', fg='white', command=_next)
        self._abw_driver_next_btn.pack(side=tk.LEFT, padx=(0,6))
        self._abw_driver_pause_btn = tk.Button(ctrls, text="Pause", bg='#444', fg='white', command=_pause_toggle)
        self._abw_driver_pause_btn.pack(side=tk.LEFT, padx=(0,6))
        tk.Button(ctrls, text="Close", bg='#2f4f4f', fg='white', command=win.destroy).pack(side=tk.LEFT)
        # Initialize status/enablement
        try:
            self._word_driver_update_ui()
        except Exception:
            pass

        # Prefer a Gurmukhi-safe font for heading
        try:
            if not hasattr(self, '_gurmukhi_font_family'):
                families = set(map(str, tkfont.families()))
                for name in ['Nirmala UI','Raavi','Noto Sans Gurmukhi','Noto Serif Gurmukhi','GurbaniAkhar','AnmolLipi','Lohit Gurmukhi','Mukta Mahee','Saab','Gurmukhi MN']:
                    if name in families:
                        self._gurmukhi_font_family = name
                        break
                if not hasattr(self, '_gurmukhi_font_family'):
                    self._gurmukhi_font_family = tkfont.nametofont('TkDefaultFont').cget('family')
        except Exception:
            if not hasattr(self, '_gurmukhi_font_family'):
                self._gurmukhi_font_family = 'TkDefaultFont'

        # Verse heading
        tk.Label(
            win,
            text=self.selected_verse_text,
            font=(self._gurmukhi_font_family, 20, 'bold'),
            bg='light gray', wraplength=900, justify='center', pady=12
        ).pack(fill=tk.X, padx=20, pady=(12,10))

        content = tk.Frame(win, bg='light gray')
        content.pack(side=tk.TOP, fill=tk.BOTH, expand=True)

        # Translation area
        tf = tk.LabelFrame(content, text="Established Darpan Translation",
                           font=("Arial", 14, 'bold'), bg='light gray', fg='black', padx=10, pady=5)
        tf.pack(fill=tk.BOTH, expand=True, padx=20, pady=(0,15))

        text_container = tk.Frame(tf, bg='light gray')
        text_container.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
        scroll_y = tk.Scrollbar(text_container, orient='vertical')
        self._translation_text = tk.Text(text_container, wrap=tk.WORD,
                                         font=(self._gurmukhi_font_family, 13),
                                         height=10, padx=8, pady=12)
        self._translation_text.configure(yscrollcommand=scroll_y.set)
        scroll_y.configure(command=self._translation_text.yview)
        self._translation_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scroll_y.pack(side=tk.RIGHT, fill=tk.Y)

        status_row = tk.Frame(tf, bg='light gray')
        status_row.pack(side=tk.BOTTOM, fill=tk.X, pady=(6,0))
        self._translation_status_var = tk.StringVar(value="Manual input")
        tk.Label(status_row, textvariable=self._translation_status_var, bg='light gray', fg='#333', font=("Arial", 10, 'italic')).pack(side=tk.LEFT)
        tk.Button(status_row, text="Refresh from Data", command=self._refresh_translation_from_data,
                  bg='teal', fg='white', font=("Arial", 10)).pack(side=tk.RIGHT)

        # Word selection area
        wf = tk.LabelFrame(content, text="Select Words to Assess Grammar",
                           font=("Arial", 14, 'bold'), bg='light gray', fg='black', padx=10, pady=10)
        wf.pack(fill=tk.X, expand=False, padx=20, pady=(0,15))
        self._select_all_words_var = tk.BooleanVar(value=False)
        tk.Checkbutton(wf, text="Select/Deselect All Words", variable=self._select_all_words_var,
                       bg='light gray', font=("Arial", 12, 'italic'), command=self._toggle_all_word_selection).pack(anchor='w', pady=(0,10))
        canvas = tk.Canvas(wf, bg='light gray', highlightthickness=0)
        canvas.configure(height=48)
        scrollbar = tk.Scrollbar(wf, orient='horizontal', command=canvas.xview)
        word_frame = tk.Frame(canvas, bg='light gray')
        canvas.configure(xscrollcommand=scrollbar.set)
        canvas.pack(side='top', fill='x', expand=False)
        scrollbar.pack(side='top', fill='x')
        canvas.create_window((0,0), window=word_frame, anchor='nw')
        word_frame.bind('<Configure>', lambda e: canvas.configure(scrollregion=canvas.bbox('all')))

        self._word_selection_vars = []
        verse_text = (self.selected_verse_text or '').split('\u0965', 1)[0].strip()
        words = verse_text.split()
        for i, w in enumerate(words):
            var = tk.BooleanVar(value=False)
            chk = tk.Checkbutton(word_frame, text=w, variable=var, bg='light gray', font=('Arial', 12), anchor='w', justify='left')
            chk.grid(row=0, column=i, sticky='w', padx=5, pady=3)
            self._word_selection_vars.append((var, w))

        # Bottom buttons (consistent bottom spacing)
        btn_frame = tk.Frame(win, bg='light gray')
        # Visual bottom gap equals BOTTOM_PAD
        btn_frame.pack(side=tk.BOTTOM, fill=tk.X, padx=20, pady=(6, BOTTOM_PAD))
        tk.Button(btn_frame, text="Close", font=('Arial', 12), bg='gray', fg='white', command=win.destroy, padx=15, pady=8).pack(side=tk.LEFT)
        tk.Button(btn_frame, text="Submit Translation ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢", font=('Arial', 12, 'bold'), bg='dark cyan', fg='white',
                  command=lambda: self._abw_on_translation_submitted(win), padx=15, pady=8).pack(side=tk.RIGHT)

        # Prefill translation if possible
        try:
            filled, status = self._populate_translation_from_structured()
            self._translation_status_var.set(status if filled else "Manual input")
        except Exception:
            pass

        # Update driver controls state once laid out
        try:
            self._word_driver_update_ui()
        except Exception:
            pass

    def _abw_on_translation_submitted(self, win):
        # Equivalent to verse handler but ABW-only
        text = self._translation_text.get("1.0", tk.END).strip()
        if not text:
            messagebox.showwarning("No Translation", "Please paste a translation before submitting.")
            return
        self.current_translation = text
        selected_idxs = [idx for idx, (var, _) in enumerate(self._word_selection_vars) if var.get()]
        if not selected_idxs:
            messagebox.showwarning("Nothing Selected", "Please select at least one word to assess.")
            return
        self._selected_word_indices = selected_idxs
        try:
            self._abw_suppress_driver_cancel_once = True
            win.destroy()
        except Exception:
            pass
        self.abw_initialize_grammar_queue()

    def proceed_to_word_assessment(self, idx):
        """Proceed only if there are results and a valid selected index; otherwise prompt the user."""
        results = getattr(self, "_last_filtered", [])
        try:
            if not results or idx is None or idx < 0 or idx >= len(results):
                messagebox.showinfo("Select a Verse", "Please search and select a verse first")
                return
        except Exception:
            # Even if messagebox fails for any reason, do not crash
            return

        # grab the metadata dict from the last search
        self.selected_verse_meta = results[idx]
        self.selected_verse_text = self.selected_verse_meta["Verse"]
        # now pop up the translation-paste window
        self.show_translation_input()

    def process_next_word_assessment(self):
        if self.current_queue_pos >= len(self.grammar_queue):
            return self.finish_and_prompt_save()

        idx, word = self.grammar_queue[self.current_queue_pos]
        self.current_word_index = idx
        self.user_input_grammar(word, self.current_translation, idx)

    def finish_and_prompt_save(self):
        """Finalize grammar assessment and prompt to save results."""
        try:
            if hasattr(self, "save_results_btn") and self.save_results_btn.winfo_exists():
                self.save_results_btn.config(state=tk.NORMAL)
        except Exception:
            pass

        if self.all_new_entries:
            try:
                self.prompt_save_results(self.all_new_entries)
            except Exception as e:
                messagebox.showerror("Error", f"An error occurred while saving: {e}")
        else:
            try:
                messagebox.showinfo("No Entries", "No grammar assessments were recorded.")
            except Exception:
                pass

        # Word-driver hook: after a verse completes, advance and update tracker
        try:
            if getattr(self, "_word_driver_active", False) and getattr(self, "_word_driver_in_progress", False):
                self._word_driver_after_verse_finished()
        except Exception:
            pass

    def _on_translation_submitted(self, win):
        # 1) grab and validate the translation itself
        text = self._translation_text.get("1.0", tk.END).strip()
        if not text:
            messagebox.showwarning("No Translation", "Please paste a translation before submitting.")
            return
        self.current_translation = text

        # 2) capture exactly which indices the user checked
        #    (you built self._word_selection_vars = [(var, word), ...] in show_translation_input)
        selected_idxs = [
            idx for idx, (var, _) in enumerate(self._word_selection_vars)
            if var.get()
        ]
        if not selected_idxs:
            messagebox.showwarning("Nothing Selected", "Please select at least one word to assess.")
            return
        self._selected_word_indices = selected_idxs

        # 3) tear down and hand off to your queue initializer
        win.destroy()
        self.initialize_grammar_queue()
        # ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â Ãƒâ€šÃ‚Â NO MORE direct call to process_next_word_assessment() here,
        #     initialize_grammar_queue() will immediately invoke it.

    def initialize_grammar_queue(self):
        """
        After the user has pasted their translation, split the verse
        into words and build the queue of those words the user selected
        for grammar assessment. Then immediately start the first word.
        """
        # split the verse text
        words = self.selected_verse_text.strip().split()

        # collect exactly those indices the user checked in show_translation_input()
        # (you should have created a tk.BooleanVar list self._word_vars there)
        selected_indices = self._selected_word_indices

        # build the queue
        self.grammar_queue = [
            (i, words[i]) for i in selected_indices
        ]
        self.grammar_meanings = []        # ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â Ãƒâ€šÃ‚Â NEW: clear out any old entries
        self.current_queue_pos = 0

        if not self.grammar_queue:
            messagebox.showinfo("Nothing Selected",
                "You didnÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â‚¬Å¾Ã‚Â¢t select any words for grammar assessment.")
            return

        # **IMMEDIATELY** start your per-word flow
        self.process_next_word_assessment()

    # ABW-only: queue initializer and driver
    def abw_initialize_grammar_queue(self):
        """ABW: Build grammar queue from selected word indices and start first word (ABW UI only)."""
        verse_text = (self.selected_verse_text or '').split('ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¥Ãƒâ€šÃ‚Â¥', 1)[0].strip()
        words = verse_text.split()
        selected_indices = getattr(self, '_selected_word_indices', [])
        self.grammar_queue = [(i, words[i]) for i in selected_indices if 0 <= i < len(words)]
        self.grammar_meanings = []
        self.current_queue_pos = 0
        if not self.grammar_queue:
            messagebox.showinfo("Nothing Selected", "You didn't select any words for grammar assessment.")
            return
        self.abw_process_next_word_assessment()

    def abw_process_next_word_assessment(self):
        if self.current_queue_pos >= len(self.grammar_queue):
            return self.finish_and_prompt_save()
        idx, word = self.grammar_queue[self.current_queue_pos]
        self.current_word_index = idx
        self.user_input_grammar_for_word(word, self.current_translation, idx)

    def _toggle_all_word_selection(self):
        """Called by the top ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¹Ã…â€œSelect/Deselect All WordsÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â‚¬Å¾Ã‚Â¢ checkbox."""
        val = self._select_all_words_var.get()
        for var, _ in getattr(self, "_word_selection_vars", []):
            var.set(val)

    def _user_input_grammar_impl(self, word, translation, index):
        """
        Pop up a window to collect grammar info for one word:
        - shows full verse with the `index`th word highlighted
        - shows the Darpan translation
        - left pane: dictionary meanings
        - right pane: Number/Gender/POS radio buttons + Expert-Prompt button
        - bottom row: Back / Skip / Submit
        """
        win = tk.Toplevel(self.root)
        win.title(f"Assess Grammar: {word}")
        win.configure(bg='light gray')
        # Taskbar-safe exact maximize to the work area (deferred)
        self._wm_apply(win, margin_px=0, defer=True)
        win.resizable(True, True)

        # 1) Verse display + highlight (use GurmukhiÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‹Å“safe font + metrics padding)
        vf = tk.Frame(win, bg='light gray')
        vf.pack(fill=tk.X, padx=20, pady=(20,10))

        # Reuse or compute a Gurmukhi-safe font family (shared with translation input)
        try:
            if not hasattr(self, '_gurmukhi_font_family'):
                families = set(map(str, tkfont.families()))
                candidates = [
                    'Nirmala UI', 'Raavi', 'Noto Sans Gurmukhi', 'Noto Serif Gurmukhi',
                    'GurbaniAkhar', 'GurbaniAkhar-Thick', 'AnmolLipi', 'AnmolUni',
                    'Lohit Gurmukhi', 'Mukta Mahee', 'Saab', 'Gurmukhi MN'
                ]
                chosen = None
                for name in candidates:
                    if name in families:
                        chosen = name
                        break
                if not chosen:
                    chosen = tkfont.nametofont('TkDefaultFont').cget('family')
                self._gurmukhi_font_family = chosen
        except Exception:
            if not hasattr(self, '_gurmukhi_font_family'):
                try:
                    self._gurmukhi_font_family = tkfont.nametofont('TkDefaultFont').cget('family')
                except Exception:
                    self._gurmukhi_font_family = 'TkDefaultFont'

        # Construct fonts for the verse and highlight (persist to avoid GC fallback)
        if not hasattr(self, '_verse_font'):
            self._verse_font = tkfont.Font(family=self._gurmukhi_font_family, size=24)
        if not hasattr(self, '_verse_font_bold'):
            self._verse_font_bold = tkfont.Font(family=self._gurmukhi_font_family, size=24, weight='bold')

        # Compute top/bottom padding from font metrics to avoid clipping of shirorekha/matras
        try:
            ascent = int(self._verse_font.metrics('ascent') or 0)
            descent = int(self._verse_font.metrics('descent') or 0)
        except Exception:
            ascent = descent = 0
        pad_top = int(math.ceil(ascent * 0.25))
        pad_bottom = int(math.ceil(descent * 0.35))

        # Text widget for the verse (centered, word-wrap), with internal padding and external pady
        td = tk.Text(
            vf,
            wrap=tk.WORD,
            bg='light gray',
            font=self._verse_font,
            height=1,
            bd=0,
            padx=4,
            pady=max(2, int(math.ceil(max(ascent, descent) * 0.15)))  # internal cushion
        )
        td.pack(fill=tk.X, pady=(pad_top, pad_bottom))
        td.insert('1.0', self.selected_verse_text)
        td.tag_add('center', '1.0', 'end')
        td.tag_configure('center', justify='center')

        # highlight the word (keep blue styling, but use Gurmukhi font variant)
        words = self.selected_verse_text.split()
        start = sum(len(w)+1 for w in words[:index])
        end   = start + len(words[index])
        td.tag_add('highlight', f'1.{start}', f'1.{end}')
        td.tag_configure('highlight', font=self._verse_font_bold, foreground='blue')
        td.config(state=tk.DISABLED)

        # Keep text wrapping responsive to window width; adjust char width on resize (approximation)
        def _sync_text_width(evt):
            try:
                # Estimate characters that fit in available width (minus frame padding)
                avg_px = max(1, int(self._verse_font.measure('0') or 8))
                width_chars = max(20, int((evt.width - 40) / avg_px))
                td.configure(width=width_chars)
            except Exception:
                pass
        try:
            win.bind('<Configure>', _sync_text_width, add='+')
        except Exception:
            pass

        # 2) Translation LabelFrame (taller; expands with window)
        tf = tk.LabelFrame(win, text="Darpan Translation",
                           font=('Arial',16,'bold'),
                           bg='light gray', fg='black',
                           padx=10, pady=10)
        # Allow translation area to grow and use available vertical space
        tf.pack(fill=tk.BOTH, expand=True, padx=20, pady=(0,15))
        trans = tk.Text(tf, wrap=tk.WORD, font=('Arial',14),
                        height=6, bd=0)
        trans.insert('1.0', translation)
        trans.config(state=tk.DISABLED)
        # Let the Text grow within the frame
        trans.pack(fill=tk.BOTH, expand=True)

        # Prepare vars for grammar options
        # Default to ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“UnknownÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â (NA)
        self.number_var = tk.StringVar(value="NA")
        self.gender_var = tk.StringVar(value="NA")
        self.pos_var    = tk.StringVar(value="NA")

        # 3+4) Stack: meanings (wide) above options (wide), both near bottom
        stack = tk.Frame(win, bg='light gray')
        # Pack as a normal/top sibling so the bottom-anchored button row remains the last element.
        # This ensures the stack sits directly above the action buttons.
        stack.pack(fill=tk.X, padx=20, pady=(0,15))

        # ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â Meanings (wide, dense columns) ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â
        left = tk.LabelFrame(
            stack,
            text=f"Meanings for \u201c{word}\u201d",
            font=('Arial',16,'bold'),
            bg='light gray', fg='black',
            padx=10, pady=10
        )
        left.pack(fill=tk.X, expand=False)

        self.meanings_canvas = tk.Canvas(left, bg='light gray', borderwidth=0, height=200)
        scrollbar = tk.Scrollbar(left, orient=tk.VERTICAL, command=self.meanings_canvas.yview)
        self.meanings_canvas.configure(yscrollcommand=scrollbar.set)

        scrollbar.pack(side='right', fill='y')
        self.meanings_canvas.pack(side='left', fill='both', expand=True)
        self.meanings_inner_frame = tk.Frame(self.meanings_canvas, bg='light gray')
        self.meanings_canvas.create_window((0,0), window=self.meanings_inner_frame, anchor='nw')

        def _on_meanings_configure(evt):
            self.meanings_canvas.configure(scrollregion=self.meanings_canvas.bbox("all"))
        self.meanings_inner_frame.bind("<Configure>", _on_meanings_configure)

        self.current_word = word   # ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â Ãƒâ€šÃ‚Â NEW: remember which word weÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â‚¬Å¾Ã‚Â¢re looking up
        threading.Thread(
            target=lambda: self.lookup_grammar_meanings_thread(word),
            daemon=True
        ).start()


        # ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â Grammar Options (wide row of groups) ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â
        right = tk.LabelFrame(
            stack,
            text="Select Grammar Options",
            font=("Arial", 16, "bold"),
            bg="light gray", fg="black",
            padx=10, pady=10
        )
        right.pack(fill=tk.X, expand=False, pady=(8, 0))

        # prepare your choices
        nums = [
            ("Singular", "Singular / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢"),
            ("Plural",   "Plural / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â"),
            ("Unknown",  "NA")
        ]
        gends = [
            ("Masculine", "Masculine / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â"),
            ("Feminine",  "Feminine / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬"),
            ("Neuter",    "Trans / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢"),
            ("Unknown",   "NA")
        ]
        pos_choices = [
            ("Noun",        "Noun / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Âµ"),
            ("Adjective",   "Adjectives / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¶ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¶ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£"),
            ("Adverb",      "Adverb / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â  ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¶ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£"),
            ("Verb",        "Verb / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â "),
            ("Pronoun",     "Pronoun / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Âµ"),
            ("Postposition","Postposition / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â§ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢"),
            ("Conjunction", "Conjunction / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¯ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢"),
            ("Interjection", "Interjection / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢"),
            ("Unknown",     "NA")
        ]

        # Horizontal row of three groups: Number | Gender | Part of Speech
        # Inside 'right', use grid for all children to avoid pack/grid mix conflicts
        right.grid_columnconfigure(0, weight=1)
        grp_row = tk.Frame(right, bg="light gray")
        grp_row.grid(row=0, column=0, sticky="nsew")
        # Give more weight to POS group to use width better
        for c in range(3):
            grp_row.grid_columnconfigure(c, weight=(2 if c == 2 else 1))

        # Number
        num_frame = tk.LabelFrame(grp_row, text="Number",
                                  font=("Arial", 14, "bold"),
                                  bg="light gray", padx=8, pady=8)
        num_frame.grid(row=0, column=0, sticky="nsew", padx=5)
        # Arrange Number radios in 2 columns / 2 rows (Singular | Plural on row 0; Unknown on row 1)
        for i, (txt, val) in enumerate(nums):
            r = 0 if i < 2 else 1
            c = i if i < 2 else 0
            rb = tk.Radiobutton(
                num_frame, text=txt, variable=self.number_var, value=val,
                bg="light gray", font=("Arial", 12), anchor="w", justify="left")
            rb.grid(row=r, column=c, sticky='w', padx=2, pady=2)
        num_frame.grid_columnconfigure(0, weight=1)
        num_frame.grid_columnconfigure(1, weight=1)

        # Gender
        gend_frame = tk.LabelFrame(grp_row, text="Gender",
                                   font=("Arial", 14, "bold"),
                                   bg="light gray", padx=8, pady=8)
        gend_frame.grid(row=0, column=1, sticky="nsew", padx=5)

        # two sub-frames for the two columns
        gf_col1 = tk.Frame(gend_frame, bg="light gray")
        gf_col2 = tk.Frame(gend_frame, bg="light gray")
        try:
            gap = COL_GAP
        except Exception:
            gap = 24
        # Equalize the inter-column gap in Gender to match Number's 2-column gap
        gf_col1.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, max(0, gap//2)))
        gf_col2.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(max(0, gap//2), 0))

        # split the list in half
        half = (len(gends)+1)//2
        for i, (txt, val) in enumerate(gends):
            parent = gf_col1 if i < half else gf_col2
            tk.Radiobutton(
                parent, text=txt, variable=self.gender_var, value=val,
                bg="light gray", font=("Arial", 12),
                anchor="w", justify="left"
            ).pack(anchor="w", pady=2)

        # Part of Speech (compact: 3 columns inside the group)
        pos_frame = tk.LabelFrame(grp_row, text="Part of Speech",
                                  font=("Arial", 14, "bold"),
                                  bg="light gray", padx=8, pady=8)
        pos_frame.grid(row=0, column=2, sticky="nsew", padx=5)

        # Arrange POS radios in two rows using multiple columns to minimize height
        pos_rows = 2
        pos_cols = -(-len(pos_choices) // pos_rows)
        try:
            gap = COL_GAP
        except Exception:
            gap = 24
        for i, (txt, val) in enumerate(pos_choices):
            r = i % pos_rows
            c = i // pos_rows
            cc = 2 * c
            rb = tk.Radiobutton(
                pos_frame, text=txt, variable=self.pos_var, value=val,
                bg="light gray", font=("Arial", 12), anchor="w", justify="left"
            )
            if c == 0:
                pad = (0, max(0, gap//2))
            elif c == pos_cols - 1:
                pad = (max(0, gap//2), 0)
            else:
                pad = (max(0, gap//2), max(0, gap//2))
            rb.grid(row=r, column=cc, sticky='w', padx=pad, pady=2)
        # configure spacer columns (odd) to expand, content columns (even) to fixed
        _total_cols = 2 * pos_cols - 1
        for _cc in range(_total_cols):
            if _cc % 2 == 0:
                pos_frame.grid_columnconfigure(_cc, weight=0)
            else:
                pos_frame.grid_columnconfigure(_cc, weight=1, minsize=gap)

        # B & D) Increase Translation height using space reclaimed from Meanings (ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°Ãƒâ€¹Ã¢â‚¬Â 10% shrink; guard for small)
        try:
            # Current translation height (text lines)
            t_h = int(trans.cget('height'))
            # Approximate meanings height in lines via canvas height and label line spacing
            m_h_px = int(self.meanings_canvas.cget('height') or 200)
            label_font = tkfont.Font(family='Arial', size=12)
            line_px = max(12, int(label_font.metrics('linespace') or 16))
            m_h_lines = max(1, int(round(m_h_px / float(line_px))))
            # Target meanings lines: ~10% shorter, min 8 lines; if very small, keep at least m_h-1
            new_m_lines = max(8, int(round(m_h_lines * 0.90)))
            if m_h_lines <= 10:
                new_m_lines = max(8, m_h_lines - 1)
            delta_lines = max(0, m_h_lines - new_m_lines)
            # Clamp translation to a reasonable max
            t_max = 20
            gain = max(1, delta_lines)
            trans.config(height=min(t_max, t_h + gain))
            # Apply new meanings height in pixels
            new_m_px = max(8 * line_px, int(round(new_m_lines * line_px)))
            self.meanings_canvas.config(height=new_m_px)
        except Exception:
            pass

        # Expert-prompt builder
        def ask_suggestion():
            verse = self.selected_verse_text
            trans = self.current_translation
            word  = self.current_word
            num   = self.number_var.get() or "ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ"
            gen   = self.gender_var.get() or "ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ"
            pos   = self.pos_var.get()    or "ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ"

            # --------------------------------------------------------------
            # Surface existing matches from the grammar database.  We mimic
            # perform_search_and_finish_reanalysis() to determine whether to
            # call search_by_criteria() or search_by_inflections().
            # --------------------------------------------------------------
            try:
                search_num = self.number_var.get()
                search_gen = self.gender_var.get()
                search_pos = self.pos_var.get()

                if (
                    search_num == "NA" and
                    search_gen == "NA" and
                    search_pos == "NA"
                ):
                    matches = self.search_by_inflections(word)
                else:
                    matches = self.search_by_criteria(
                        word, search_num, search_gen, search_pos
                    )
                    if not matches:
                        matches = self.search_by_inflections(word)

                # Precompute keyset for Evaluation == "Predefined" rows from CSV
                pre_keys = load_predefined_keyset("1.1.1_birha.csv")
                rows = []
                for result, _count, _perc in matches:
                    parts = [p.strip() for p in result.split("|")]
                    if len(parts) < 7:
                        parts += [""] * (7 - len(parts))

                    # Filter out rows not marked Predefined in CSV (by core feature key)
                    try:
                        key = tuple((parts[i] or "").strip() for i in (2, 3, 4, 5, 6))
                    except Exception:
                        key = ("") * 5
                    if key not in pre_keys:
                        continue

                    highlight = parts[0] == parts[1] and is_full_word(parts[0])
                    if highlight:
                        parts = [f"**{p}**" for p in parts]
                        parts[0] = "ÃƒÆ’Ã‚Â¢Ãƒâ€¦Ã¢â‚¬Å“ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ " + parts[0]

                    rows.append(
                        "| "
                        + " | ".join(parts + [str(_count), f"{_perc:.1f}%"])
                        + " |"
                    )
                    if len(rows) >= 5:
                        break

                if rows:
                    headers = [
                        "Word under Analysis",
                        "Vowel Ending / Word Matches",
                        "Number / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨",
                        "Grammar / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¯ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£",
                        "Gender / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â",
                        "Word Root",
                        "Type",
                        "Match Count",
                        "Match %",
                    ]
                    table_lines = [
                        "**Top Grammar Matches**",
                        "| " + " | ".join(headers) + " |",
                        "| " + " | ".join(["---"] * len(headers)) + " |",
                        *rows,
                    ]
                    matches_block = "\n".join(table_lines)
                else:
                    matches_block = "**Top Grammar Matches**\nNo predefined examples found"
            except Exception as exc:
                print(f"search for matches failed: {exc}")
                matches_block = ""

            # pull the meanings we stored for this word
            meanings = next(
                (e["meanings"] for e in self.grammar_meanings if e["word"] == word),
                []
            )
            meanings_block = "\n".join(f"- {m}" for m in meanings) or "- (no dictionary meanings found)"

            prompt = textwrap.dedent(f"""
            You are a Punjabi grammar expert trained in the grammatical framework of Sri Guru Granth Sahib (SGGS). I will provide:

            1. **Verse** (in Gurmukhi)  
            2. **Established Darpan Translation** (by Prof. Sahib Singh)  
            3. **Word under scrutiny**, along with my selected values for Number, Gender, and Part of Speech  
            4. **Dictionary Meanings** of that word (as a secondary reference)

            Your job is to confirm or correct my selections based on the **Darpan Translation and its contextual meaning**, which is the **primary reference**. Override my input only if the Darpan explanation makes it grammatically, semantically, or functionally incorrect within the SGGS grammatical framework.

           ---

            ## ÃƒÆ’Ã‚Â°Ãƒâ€¦Ã‚Â¸ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â€šÂ¬Ã…Â¾ Two-Pass Analysis Workflow
            **Phase 1 ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ Functional Tagging**  
            1 a. Locate every occurrence of the stem in the verse.  
            1 b. Assign provisional POS to each occurrence from context.  

            **Phase 2 ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ Morphological Reconciliation**  
            2 a. Compare endings of all identical stems found in 1 a.  
            2 b. If endings differ ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ mark the stem **declinable** and align each form with its noun/pronoun.  
            2 c. If endings never differ ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ note ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“No declension detected.ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â  

            If Phase 2 detects a declinable pattern but any token fails to agree with its noun/pronoun, **STOP** and return ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“Agreement Error ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ Review Needed.ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â

            ---

            ## ÃƒÆ’Ã‚Â°Ãƒâ€¦Ã‚Â¸ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒâ€¹Ã…â€œ Reference Framework ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ SGGS Grammar Definitions

            ### ÃƒÆ’Ã‚Â°Ãƒâ€¦Ã‚Â¸Ãƒâ€šÃ‚Â§Ãƒâ€šÃ‚Â© Implicit Case Logic in Gurbani Grammar
            Many case roles in SGGS are conveyed through **inflection or contextual meaning**, not modern postpositions. Refer to the gloss clues (ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“ofÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â, ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“byÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â, ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“withÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â, etc.) to infer case correctly.

            ### 1. **Noun (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Âµ)**  
            A noun is a word that names a person, place, thing, quality, or idea.

            #### ÃƒÆ’Ã‚Â°Ãƒâ€¦Ã‚Â¸ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒâ€šÃ‚Â¹ Types:
            - **Proper Noun (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¼ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¼ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Âµ)** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ e.g., ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢
            - **Common Noun (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â§ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Âµ)** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ e.g., ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬
            - **Abstract Noun (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â­ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Âµ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Âµ)** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ e.g., ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨
            - **Material Noun (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Âµ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Âµ)** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ e.g., ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²
            - **Collective Noun (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Âµ)** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ e.g., ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â«ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¦Ã¢â‚¬â„¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“

            #### ÃƒÆ’Ã‚Â°Ãƒâ€¦Ã‚Â¸ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒâ€šÃ‚Â¹ Cases in Gurbani Grammar:
            Nouns in Gurbani may appear in the following **grammatical cases** (*vibhakti*), sometimes **without explicit post-positions**:

            | Case         | Helper (Gloss Clue)             | Modern Marker    | When to Use                                                       |
            |--------------|----------------------------------|------------------|-------------------------------------------------------------------|
            | **Nominative**     | No helper, subject role         | None             | Default when noun is subject of verb                              |
            | **Accusative**     | No helper, object role          | None             | Default when noun is object of verb                               |
            | **Genitive**       | ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“ofÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â, ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â                | `ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡`, `ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬`, `ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾` | Use when gloss adds ownership/association                         |
            | **Instrumental**   | ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“byÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â, ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“withÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â, ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“underÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â           | `ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²`, `ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â§ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨`     | Use when gloss suggests means/manner (even if unstated in verse)  |
            | **Dative**         | ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“toÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â, ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“forÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â                     | `ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°`, `ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã¢â‚¬Â `       | When gloss implies recipient/beneficiary                          |
            | **Locative**       | ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“inÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â, ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“onÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â, ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“atÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â                | `ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â±ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡`, `ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡`      | When gloss places noun in space/context                           |
            | **Ablative**       | ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“fromÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â, ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“out ofÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â                | `ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡`, `ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡`      | When gloss implies source                                         |
            | **Vocative**       | ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“OÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â, ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“HeyÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â                      | *(address)*       | Used for direct address (e.g., *ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â­ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã¢â‚¬Â !*)                          |

            > ÃƒÆ’Ã‚Â°Ãƒâ€¦Ã‚Â¸ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒâ€šÃ‚Â¸ **Implicit Post-Positions:** If Darpan adds ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â², ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â±ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â etc., treat it as a **helper** for inferring the nounÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â‚¬Å¾Ã‚Â¢s **grammatical case**, even if the verse lacks a marker.
            >
            > ÃƒÆ’Ã‚Â°Ãƒâ€¦Ã‚Â¸ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒâ€šÃ‚Â¸ **Indeclinable Loan Nouns:** Sanskrit-based nouns (like *ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â§ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿*, *ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬*) may not show visible inflection. Their case must be inferred from semantic role and Darpan gloss, not suffix alone.

            > ÃƒÆ’Ã‚Â°Ãƒâ€¦Ã‚Â¸ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒâ€šÃ‚Â¹ **Fallback Rule:**  
            > When the gloss offers no helper and the noun does not visibly decline, default to **Nominative or Accusative**, then refine based on sentence structure and implied role in the Darpan explanation.

            ### 2. **Pronoun (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Âµ)**  
            Used in place of nouns. Types include:  
            - **Personal**, **Demonstrative**, **Reflexive**, **Possessive**, **Relative**, **Indefinite**, **Interrogative**

            ### 3. **Adjective (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¼ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¼ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£) ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ Agreement Framework**
            Describes or qualifies a noun or pronoun only. Must be directly linked to one.  
            Adjectives include: **Qualitative**, **Demonstrative**, **Indefinite**, **Pronominal**, **Numeral**, and **Interrogative**.
            Examples include: ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Å“, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â , ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã¢â‚¬Â  ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â±ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Å“

            ÃƒÆ’Ã‚Â°Ãƒâ€¦Ã‚Â¸ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒâ€šÃ‚Â´ **GURBANI RULE (STRICT)**  
            ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Å“Ãƒâ€šÃ‚Â¶ÃƒÆ’Ã‚Â¯Ãƒâ€šÃ‚Â¸Ãƒâ€šÃ‚Â **All adjectives in Gurbani MUST agree in Number and Gender with the noun or pronoun they qualify.**  
            This is a **non-negotiable rule** confirmed by both **Sikh Research Institute (SikhRi)** and **Prof. Sahib SinghÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â‚¬Å¾Ã‚Â¢s Gurbani Vyakaran**.  
            The agreement must be:
            - **Semantic** (referring to the correct noun/pronoun)
            - **Morphological** (adjective form visibly matches Number & Gender)

            ÃƒÆ’Ã‚Â°Ãƒâ€¦Ã‚Â¸ÃƒÂ¢Ã¢â€šÂ¬Ã‹Å“ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â° *In Gurbani, adjectives are always **declined** to match the Number and Gender of the noun or pronoun they describe. This means adjectives **change form** based on their grammatical role. They are not fixed or invariable by default.*

            If the adjectiveÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â‚¬Å¾Ã‚Â¢s form appears fixed (e.g., ending in ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¹Ã…â€œÃƒÆ’Ã¢â‚¬Â¦Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â‚¬Å¾Ã‚Â¢ or ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¹Ã…â€œauÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â‚¬Å¾Ã‚Â¢), consult its grammatical root ending (MuktÃƒÆ’Ã¢â‚¬Å¾Ãƒâ€šÃ‚Â, KannÃƒÆ’Ã¢â‚¬Å¾Ãƒâ€šÃ‚Â, AunkÃƒÆ’Ã¢â‚¬Å¾Ãƒâ€šÃ‚Âr, HorÃƒÆ’Ã¢â‚¬Å¾Ãƒâ€šÃ‚Â, BihÃƒÆ’Ã¢â‚¬Å¾Ãƒâ€šÃ‚ÂrÃƒÆ’Ã¢â‚¬Å¾Ãƒâ€šÃ‚Â«) to verify its role and alignment.

            ÃƒÆ’Ã‚Â°Ãƒâ€¦Ã‚Â¸ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒâ€šÃ‚Â *Do not assume that any adjective is morphologically invariable unless **Gurbani Vyakaran** explicitly identifies it as a poetic variant that still maintains grammatical agreement.* **Do not conclude invariance merely because the same form appears with multiple nouns.**
            **Many adjectives follow internal paradigms that are consistent across different contexts, even if they *look* fixed.**

            ÃƒÆ’Ã‚Â°Ãƒâ€¦Ã‚Â¸Ãƒâ€šÃ‚Â§Ãƒâ€šÃ‚Â  *If the adjectiveÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â‚¬Å¾Ã‚Â¢s ending appears unchanged, it must still be evaluated against known adjective paradigms (e.g., hÃƒÆ’Ã¢â‚¬Â¦Ãƒâ€šÃ‚ÂrÃƒÆ’Ã¢â‚¬Å¾Ãƒâ€šÃ‚Â-ending, kannÃƒÆ’Ã¢â‚¬Å¾Ãƒâ€šÃ‚Â-ending). Only when those forms confirm invariance through grammatical structureÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‚Ânot intuitionÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‚Âshould it be marked as ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¹Ã…â€œinvariableÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â‚¬Å¾Ã‚Â¢ in the agreement table.*

            > **Cross-token check ** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ If the same stem re-appears with a different ending in the *line*, treat that as conclusive evidence it is **declinable**; do not invoke ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“indeclinableÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â unless all tokens are identical in form *and* no paradigm lists inflected endings.

            ---

            **ÃƒÆ’Ã‚Â°Ãƒâ€¦Ã‚Â¸ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂºÃƒÂ¢Ã¢â€šÂ¬Ã‹Å“ Mandatory Adjective Agreement Table**
            ÃƒÆ’Ã‚Â¢Ãƒâ€¦Ã‚Â¡Ãƒâ€šÃ‚Â ÃƒÆ’Ã‚Â¯Ãƒâ€šÃ‚Â¸Ãƒâ€šÃ‚Â **Caution:**  
            Do **not** classify a word as an Adjective merely because it appears near a noun.  
            Carefully check whether the word is:
            - Acting as the **object of a postposition** (e.g., "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â§ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â±ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â±ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡"), in which case it is a **noun**, not an adjective.
            - Part of an **oblique noun phrase** and not qualifying the noun directly.
            - Functioning as a **noun in instrumental case** (e.g., ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â§ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ by/with threefold means); these may **appear** descriptive but are **semantically instrumental nouns**, not adjectives.
            
            These constructions often create **false links**. Always confirm grammatical agreement and functional relationship before assigning Adjective.

            If a word is confirmed as an adjective, this table is required:

            | Step | Requirement | Observation | Result |
            |------|-------------|-------------|--------|
            | 1 | Identify the qualified noun/pronoun | (e.g., ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ masculine singular) | ... |
            | 2 | Show matching Number & Gender in adjective form | (e.g., ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ = masculine singular form of ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¦Ã¢â‚¬â„¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾-ending adjective) | ÃƒÆ’Ã‚Â¢Ãƒâ€¦Ã¢â‚¬Å“ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ / ÃƒÆ’Ã‚Â¢Ãƒâ€šÃ‚ÂÃƒâ€¦Ã¢â‚¬â„¢ |
            | 3 | Stem-variation observed? | e.g. ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â«ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¦Ã¢â‚¬Å“ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â«ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â | ÃƒÆ’Ã‚Â¢Ãƒâ€¦Ã¢â‚¬Å“ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ / ÃƒÆ’Ã‚Â¢Ãƒâ€šÃ‚ÂÃƒâ€¦Ã¢â‚¬â„¢ |

            ÃƒÆ’Ã‚Â¢Ãƒâ€šÃ‚ÂÃƒâ€¦Ã¢â‚¬â„¢ *Responses that skip this table or assume invariable adjectives will be treated as incomplete.*
            *(skip the table entirely if final POS ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°Ãƒâ€šÃ‚Â  Adjective)*

            ### 4. **Verb (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â )**  
            Expresses an action, state, or condition. Includes forms like transitive/intransitive, passive, causative, auxiliary, etc.

            ### 5. **Adverb (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â  ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¼ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¼ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£)**  
            Modifies verbs only. Never nouns. Categories include Time, Place, Manner, Degree, Frequency, etc.

            ### 6. **Postposition (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â§ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢)** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ e.g., ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â², ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â±ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â±ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡  
            ### 7. **Conjunction (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¯ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢)** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ e.g., ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°  
            ### 8. **Interjection (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢)** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ e.g., ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹!, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â!

            ---

            ## ÃƒÆ’Ã‚Â°Ãƒâ€¦Ã‚Â¸Ãƒâ€¦Ã‚Â½Ãƒâ€šÃ‚Â¯ Evaluation Guidelines

            1. Use **Darpan Translation** to determine the wordÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â‚¬Å¾Ã‚Â¢s semantic role.  
            2. Confirm **Part of Speech**:  
            - Modifies noun/pronoun ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ Adjective (**triggers the agreement check**)  
            - Modifies verb/adjective/adverb ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ Adverb  
            - If noun/pronoun ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ classify accordingly  
            3. For Adjectives:
            - Confirm Number & Gender based on the noun/pronoun the adjective qualifies. If the adjective form appears fixed, verify its grammatical alignment using its root ending.
            - If adjective doesnÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â‚¬Å¾Ã‚Â¢t change form (invariable), still list target noun and declare this explicitly 
            - ÃƒÆ’Ã‚Â¢Ãƒâ€¦Ã‚Â¡Ãƒâ€šÃ‚Â ÃƒÆ’Ã‚Â¯Ãƒâ€šÃ‚Â¸Ãƒâ€šÃ‚Â The **nounÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â‚¬Å¾Ã‚Â¢s gender and number** must be derived from **Gurbani Grammar definitions** (as per Darpan and Vyakaran), not from modern Punjabi intuition or pronunciation. For example, abstract nouns like **ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾** are feminine singular by SGGS convention.
            ÃƒÆ’Ã‚Â¢Ãƒâ€¦Ã¢â‚¬Å“ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ *Trigger Adjective Agreement Table only if:*  
            - Word semantically modifies a noun/pronoun (confirmed in Darpan gloss)  
            - Is not the subject/object of a helper-preposition  
            - Does not serve as the head of a noun phrase or abstract concept (e.g., ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â§ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ = by/through threefold mode)  
            4. Do not guess based on spelling or intuitionÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â**rely on function and context from translation**  
            5. Output is **incomplete** if POS = Adjective and Adjective Agreement Table is missing

            ---

            ## ÃƒÆ’Ã‚Â°Ãƒâ€¦Ã‚Â¸ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒâ€šÃ‚Â¥ Inputs

            **Verse (Gurmukhi):**  
            {verse}

            **Darpan Translation:**  
            {trans}

            **Word under scrutiny:**  
            {word}

            **My Selections:**  
            - Number: {num}  
            - Gender: {gen}  
            - Part of Speech: {pos}

            **Dictionary Meanings (Secondary Aid):**
            {meanings_block}

            {matches_block}

            ---

            ## ÃƒÆ’Ã‚Â°Ãƒâ€¦Ã‚Â¸ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ Response Format (Follow exactly)

            1. **Feature Confirmation**  
            - Number: (Correct / Incorrect) ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ based on Darpan gloss and noun agreement  
            - Gender: (Correct / Incorrect) ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ based on noun gender  
            - Part of Speech: (Correct / Incorrect) ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ based on function and Darpan context  

            2. **Corrections (if needed)**  
            - Number: <correct value> ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ with rationale  
            - Gender: <correct value> ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ with rationale  
            - Part of Speech: <correct value> ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ with rationale  

            3. **Commentary**  
            - Explain briefly how the Darpan translation and noun/pronoun connection led to your decision  
            - If adjective form is invariable, name the adjective group (e.g., **Horaa** ending or **Poetic variation**)

            4. **Adjective-Agreement Table (REQUIRED if POS = Adjective)**  
            | Step | Requirement              | Observation                    | Result        |
            |------|--------------------------|--------------------------------|---------------|
            | 1    | Qualified noun/pronoun   | (e.g., ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ masculine-singular) | (Identified) |
            | 2    | Number & Gender match    | (e.g., adjective ends with -ÃƒÆ’Ã¢â‚¬Â¦Ãƒâ€šÃ‚Â, matches masculine singular noun; or declare as invariable) | ÃƒÆ’Ã‚Â¢Ãƒâ€¦Ã¢â‚¬Å“ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦/ÃƒÆ’Ã‚Â¢Ãƒâ€šÃ‚ÂÃƒâ€¦Ã¢â‚¬â„¢ |
            
            ---

            ÃƒÆ’Ã‚Â°Ãƒâ€¦Ã‚Â¸ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒâ€¹Ã…â€œ **Quick Reference: Common Adjective Endings in Gurbani**

            | Ending      | Number & Gender         | Example           |
            |-------------|--------------------------|-------------------|
            | **-ÃƒÆ’Ã¢â‚¬Â¦Ãƒâ€šÃ‚Â**      | Masculine singular        | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹       |
            | **-ÃƒÆ’Ã¢â‚¬Å¾ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â**  | Masculine plural          | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡         |
            | **-ÃƒÆ’Ã¢â‚¬Å¾Ãƒâ€šÃ‚Â«**      | Feminine singular         | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬         |
            | **-ÃƒÆ’Ã¢â‚¬Å¾Ãƒâ€šÃ‚Â«ÃƒÆ’Ã¢â‚¬Å¾Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â¡Ãƒâ€šÃ‚Â¹Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡** | Feminine plural         | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡      |

            These endings are drawn from adjective groups described in Prof. Sahib SinghÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â‚¬Å¾Ã‚Â¢s *Gurbani Vyakaran*, e.g., hÃƒÆ’Ã¢â‚¬Â¦Ãƒâ€šÃ‚ÂrÃƒÆ’Ã¢â‚¬Å¾Ãƒâ€šÃ‚Â-samÃƒÆ’Ã¢â‚¬Å¾Ãƒâ€šÃ‚Âpt adjectives. Always match these with the gender and number of the qualified noun.
            ÃƒÆ’Ã‚Â°Ãƒâ€¦Ã‚Â¸ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒâ€šÃ‚Â¹ *Tatsam Words (Sanskrit-Derived)*:  
            Many Sanskrit-origin words in GurbaniÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‚Âsuch as **ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â§ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿**, **ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“**, **ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤**ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‚Âoften appear morphologically fixed and may superficially resemble adjectives. However, they frequently function as **abstract nouns** or appear in **instrumental** or other oblique grammatical cases.

            > ÃƒÆ’Ã‚Â°Ãƒâ€¦Ã‚Â¸ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒâ€šÃ‚Â¸ **Tatsam Adjectives vs Indeclinable Nouns:**  
            > Do **not** classify such words as adjectives unless the **Darpan gloss clearly shows them qualifying a noun**, with **visible agreement in Number and Gender**.  
            > ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Å“Ãƒâ€šÃ‚Â¶ÃƒÆ’Ã‚Â¯Ãƒâ€šÃ‚Â¸Ãƒâ€šÃ‚Â If the gloss inserts a helper like *ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“by,ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“with,ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“in,ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â or ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“ofÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â*, this usually signals a **noun in an oblique case**ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‚Ânot an adjective.  
            > ÃƒÆ’Ã‚Â¢Ãƒâ€¦Ã‚Â¾ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ For example, **ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â§ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿** may mean *ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“by threefold meansÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â* or *ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“through the three qualitiesÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â*, serving a **functional role** rather than describing a noun.

            ÃƒÆ’Ã‚Â°Ãƒâ€¦Ã‚Â¸ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒâ€šÃ‚Â *Key Insight:*  
            Words like **ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â§ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿**, despite their descriptive appearance, often act as **instrumental-case nouns** or form part of a **compound abstract expression** (e.g., *ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â *). Always validate their role against the **Darpan translation** and **Gurbani grammar definitions**, not surface resemblance.

            ---

            ### ÃƒÆ’Ã‚Â°Ãƒâ€¦Ã‚Â¸ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÂ¢Ã¢â€šÂ¬Ã‹Å“ Stem-Variation Check ÃƒÆ’Ã‚Â°Ãƒâ€¦Ã‚Â¸ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢
            *(Fill this mini-grid during Phase 2 if you detected more than one token of the same stem)*  
            | Token | Ending | Nearby noun/pronoun | Expected agreement | Matches? |
            |-------|--------|---------------------|--------------------|----------|

            ---

            ÃƒÆ’Ã‚Â°Ãƒâ€¦Ã‚Â¸ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂºÃƒâ€šÃ‚Â  **Debug Trace** ÃƒÆ’Ã‚Â°Ãƒâ€¦Ã‚Â¸ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ (single line at the very end):  
            `[TokensChecked:X | Declined:Yes/No | FinalPOS:___ | AgreementOK:Yes/No]`

            """).strip()

            # copy to clipboard
            self.root.clipboard_clear()
            self.root.clipboard_append(prompt)
            messagebox.showinfo(
                "Prompt Ready",
                "Expert-level prompt (with secondary dictionary meanings) has been copied to your clipboard.\n"
                "Paste it into ChatGPT for its recommendation."
            )

        # (Button moved to the bottom action bar)

        # 5) Bottom separator + buttons
        sep = tk.Frame(win, bg='#cccccc', height=2)
        sep.pack(side=tk.BOTTOM, fill=tk.X, padx=20, pady=(0,5))

        btns = tk.Frame(win, bg='light gray')
        btns.pack(side=tk.BOTTOM, fill=tk.X, padx=20, pady=(6, BOTTOM_PAD))
        tk.Button(btns, text="ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¹ Back to Translation",
                  font=('Arial',12), bg='gray', fg='white',
                  padx=20, pady=8,
                  command=lambda: [win.destroy(), self.show_translation_input()]
        ).pack(side=tk.LEFT)
        tk.Button(btns, text="Skip Word",
                  font=('Arial',12), bg='orange', fg='white',
                  padx=20, pady=8,
                  command=lambda: self.skip_word_grammar(win)
        ).pack(side=tk.LEFT, padx=10)
        # Moved here from options frame: Build Expert Prompt
        tk.Button(btns,
                  text="ÃƒÆ’Ã‚Â°Ãƒâ€¦Ã‚Â¸ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ Build Expert Prompt",
                  font=("Arial", 14, "italic"),
                  bg='white', fg='dark cyan',
                  padx=6, pady=4,
                  command=ask_suggestion
        ).pack(side=tk.LEFT, padx=10)
        tk.Button(btns, text="Submit",
                  font=('Arial',12,'bold'),
                  bg='dark cyan', fg='white',
                  padx=20, pady=8,
                  command=lambda: self.submit_input_grammar(word, index)
        ).pack(side=tk.RIGHT)

        # Modal
        win.transient(self.root)
        win.grab_set()
        self.root.wait_window(win)

    def _build_user_input_grammar(self, win, *, word, translation, index, mode):
        """Shared builder for the per-word Grammar UI. Does not start a mainloop.

        Returns (win, ctx) where ctx exposes key widgets for callers to tweak.
        mode in {"verse", "word"} controls small call-site differences (e.g., back button).
        """
        # Stash current mode so skip/advance chooses correct queue handler
        try:
            self._current_grammar_mode = mode
        except Exception:
            pass
        win.configure(bg='light gray')
        win.resizable(True, True)

        # 1) Verse display + highlight (use Gurmukhi-safe font + metrics padding)
        vf = tk.Frame(win, bg='light gray')
        vf.pack(fill=tk.X, padx=20, pady=(20,10))

        # Reuse or compute a Gurmukhi-safe font family (shared with translation input)
        try:
            if not hasattr(self, '_gurmukhi_font_family'):
                families = set(map(str, tkfont.families()))
                candidates = [
                    'Nirmala UI', 'Raavi', 'Noto Sans Gurmukhi', 'Noto Serif Gurmukhi',
                    'GurbaniAkhar', 'GurbaniAkhar-Thick', 'AnmolLipi', 'AnmolUni',
                    'Lohit Gurmukhi', 'Mukta Mahee', 'Saab', 'Gurmukhi MN'
                ]
                chosen = None
                for name in candidates:
                    if name in families:
                        chosen = name
                        break
                if not chosen:
                    chosen = tkfont.nametofont('TkDefaultFont').cget('family')
                self._gurmukhi_font_family = chosen
        except Exception:
            if not hasattr(self, '_gurmukhi_font_family'):
                try:
                    self._gurmukhi_font_family = tkfont.nametofont('TkDefaultFont').cget('family')
                except Exception:
                    self._gurmukhi_font_family = 'TkDefaultFont'

        # Construct fonts for the verse and highlight (persist to avoid GC fallback)
        if not hasattr(self, '_verse_font'):
            self._verse_font = tkfont.Font(family=self._gurmukhi_font_family, size=24)
        if not hasattr(self, '_verse_font_bold'):
            self._verse_font_bold = tkfont.Font(family=self._gurmukhi_font_family, size=24, weight='bold')

        # Compute top/bottom padding from font metrics to avoid clipping of shirorekha/matras
        try:
            ascent = int(self._verse_font.metrics('ascent') or 0)
            descent = int(self._verse_font.metrics('descent') or 0)
        except Exception:
            ascent = descent = 0
        pad_top = int(math.ceil(ascent * 0.25))
        pad_bottom = int(math.ceil(descent * 0.35))

        # Text widget for the verse (centered, word-wrap), with internal padding and external pady
        td = tk.Text(
            vf,
            wrap=tk.WORD,
            bg='light gray',
            font=self._verse_font,
            height=1,
            bd=0,
            padx=4,
            pady=max(2, int(math.ceil(max(ascent, descent) * 0.15)))
        )
        td.pack(fill=tk.X, pady=(pad_top, pad_bottom))
        td.insert('1.0', self.selected_verse_text)
        td.tag_add('center', '1.0', 'end')
        td.tag_configure('center', justify='center')

        # highlight the word
        try:
            words = self.selected_verse_text.split()
            start = sum(len(w)+1 for w in words[:index])
            end   = start + len(words[index])
            td.tag_add('highlight', f'1.{start}', f'1.{end}')
            td.tag_configure('highlight', font=self._verse_font_bold, foreground='blue')
        except Exception:
            pass
        td.config(state=tk.DISABLED)

        # Keep text wrapping responsive to window width
        def _sync_text_width(evt):
            try:
                avg_px = max(1, int(self._verse_font.measure('0') or 8))
                width_chars = max(20, int((evt.width - 40) / avg_px))
                td.configure(width=width_chars)
            except Exception:
                pass
        try:
            win.bind('<Configure>', _sync_text_width, add='+')
        except Exception:
            pass

        # 2) Translation LabelFrame (taller; expands with window)
        tf = tk.LabelFrame(win, text="Darpan Translation",
                           font=('Arial',16,'bold'),
                           bg='light gray', fg='black',
                           padx=10, pady=10)
        tf.pack(fill=tk.BOTH, expand=True, padx=20, pady=(0,15))
        trans = tk.Text(tf, wrap=tk.WORD, font=('Arial',14), height=6, bd=0)
        trans.insert('1.0', translation)
        trans.config(state=tk.DISABLED)
        trans.pack(fill=tk.BOTH, expand=True)

        # Prepare vars for grammar options
        self.number_var = tk.StringVar(value="NA")
        self.gender_var = tk.StringVar(value="NA")
        self.pos_var    = tk.StringVar(value="NA")

        # 3+4) Stack: meanings (wide) above options (wide)
        stack = tk.Frame(win, bg='light gray')
        stack.pack(fill=tk.X, padx=20, pady=(0,15))

        # Meanings
        left = tk.LabelFrame(
            stack,
            text=f"Meanings for \u201c{word}\u201d",
            font=('Arial',16,'bold'),
            bg='light gray', fg='black',
            padx=10, pady=10
        )
        left.pack(fill=tk.X, expand=False)
        self.meanings_canvas = tk.Canvas(left, bg='light gray', borderwidth=0, height=200)
        scrollbar = tk.Scrollbar(left, orient=tk.VERTICAL, command=self.meanings_canvas.yview)
        self.meanings_canvas.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side='right', fill='y')
        self.meanings_canvas.pack(side='left', fill='both', expand=True)
        self.meanings_inner_frame = tk.Frame(self.meanings_canvas, bg='light gray')
        self.meanings_canvas.create_window((0,0), window=self.meanings_inner_frame, anchor='nw')
        self.meanings_inner_frame.bind("<Configure>", lambda e: self.meanings_canvas.configure(scrollregion=self.meanings_canvas.bbox("all")))
        self.current_word = word
        threading.Thread(target=lambda: self.lookup_grammar_meanings_thread(word), daemon=True).start()

        # Options
        right = tk.LabelFrame(
            stack,
            text="Select Grammar Options",
            font=("Arial", 16, "bold"),
            bg="light gray", fg="black",
            padx=10, pady=10
        )
        right.pack(fill=tk.X, expand=False, pady=(8, 0))
        grp_row = tk.Frame(right, bg="light gray")
        grp_row.pack(fill=tk.X)
        # Use a 9-column grid and span frames to enforce a visible 2:2:5 ratio
        for c in range(9):
            grp_row.grid_columnconfigure(c, weight=1, uniform='grpcols')

        # Number
        num_frame = tk.LabelFrame(grp_row, text="Number",
                                  font=("Arial", 14, "bold"),
                                  bg="light gray", padx=8, pady=8)
        num_frame.grid(row=0, column=0, columnspan=2, sticky="nsew", padx=(0,5))
        for txt, val in [("Singular","Singular / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢"),("Plural","Plural / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â"),("Unknown","NA")]:
            tk.Radiobutton(
                num_frame, text=txt, variable=self.number_var, value=val,
                bg="light gray", font=("Arial", 12), anchor="w", justify="left"
            ).pack(anchor="w", pady=2)

        # A) Re-layout Number radios into 2x2 grid: row0 (Singular, Plural), row1 (Unknown, spacer)
        try:
            for child in list(num_frame.winfo_children()):
                try:
                    child.pack_forget()
                except Exception:
                    pass
            try:
                num_frame.grid_columnconfigure(0, weight=0)
                num_frame.grid_columnconfigure(1, weight=1, minsize=COL_GAP)
                num_frame.grid_columnconfigure(2, weight=0)
            except Exception:
                pass
            COL_GAP = 24
            tk.Radiobutton(
                num_frame, text="Singular", variable=self.number_var, value="Singular / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢",
                bg="light gray", font=("Arial", 12), anchor="w", justify="left"
            ).grid(row=0, column=0, sticky='w', padx=0, pady=2)
            tk.Radiobutton(
                num_frame, text="Plural", variable=self.number_var, value="Plural / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â",
                bg="light gray", font=("Arial", 12), anchor="w", justify="left"
            ) .grid(row=0, column=2, sticky='w', padx=0, pady=2)
            tk.Radiobutton(
                num_frame, text="Unknown", variable=self.number_var, value="NA",
                bg="light gray", font=("Arial", 12), anchor="w", justify="left"
            ).grid(row=1, column=0, sticky='w', padx=0, pady=2)
        except Exception:
            pass

        # Gender (two columns)
        gend_frame = tk.LabelFrame(grp_row, text="Gender",
                                   font=("Arial", 14, "bold"),
                                   bg="light gray", padx=8, pady=8)
        gend_frame.grid(row=0, column=2, columnspan=2, sticky="nsew", padx=5)
        gends = [("Masculine","Masculine / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â"),("Feminine","Feminine / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬"),("Neuter","Trans / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢"),("Unknown","NA")]
        gf_col1 = tk.Frame(gend_frame, bg="light gray")
        gf_col2 = tk.Frame(gend_frame, bg="light gray")
        try:
            gap = COL_GAP
        except Exception:
            gap = 24
        gf_col1.pack(side=tk.LEFT, fill=tk.BOTH, expand=False, padx=(0,0))
        gf_gap = tk.Frame(gend_frame, bg="light gray", width=gap)
        gf_gap.pack(side=tk.LEFT, fill=tk.X, expand=True)
        gf_col2.pack(side=tk.LEFT, fill=tk.BOTH, expand=False, padx=(0,0))
        
        half = (len(gends)+1)//2
        for i, (txt, val) in enumerate(gends):
            parent = gf_col1 if i < half else gf_col2
            tk.Radiobutton(
                parent, text=txt, variable=self.gender_var, value=val,
                bg="light gray", font=("Arial", 12), anchor="w", justify="left"
            ).pack(anchor="w", pady=2)

        # POS
        pos_frame = tk.LabelFrame(grp_row, text="Part of Speech",
                                  font=("Arial", 14, "bold"),
                                  bg="light gray", padx=8, pady=8)
        pos_frame.grid(row=0, column=4, columnspan=5, sticky="nsew", padx=(5,0))
        pos_choices = [("Noun","Noun / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Âµ"),("Adjective","Adjectives / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¶ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¶ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£"),("Adverb","Adverb / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â  ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¶ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£"),
                       ("Verb","Verb / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â "),("Pronoun","Pronoun / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Âµ"),("Postposition","Postposition / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â§ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢"),
                       ("Conjunction","Conjunction / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¯ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢"),("Interjection","Interjection / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢"),("Unknown","NA")]
        pos_rows = 2
        pos_cols = -(-len(pos_choices) // pos_rows)
        try:
            gap = COL_GAP
        except Exception:
            gap = 24
        for i, (txt, val) in enumerate(pos_choices):
            r = i % pos_rows
            c = i // pos_rows
            cc = 2 * c
            rb = tk.Radiobutton(
                pos_frame, text=txt, variable=self.pos_var, value=val,
                bg="light gray", font=("Arial", 12), anchor="w", justify="left"
            )
            if c == 0:
                pad = (0, max(0, gap//2))
            elif c == pos_cols - 1:
                pad = (max(0, gap//2), 0)
            else:
                pad = (max(0, gap//2), max(0, gap//2))
            rb.grid(row=r, column=cc, sticky='w', padx=pad, pady=2)
        total_cols = 2*pos_cols - 1
        for cc in range(total_cols):
            if cc % 2 == 0:
                pos_frame.grid_columnconfigure(cc, weight=0)
            else:
                pos_frame.grid_columnconfigure(cc, weight=1, minsize=gap)

        # Column spanning above enforces the 2:2:5 proportion; no extra handling needed

        # Expert-prompt builder
        def ask_suggestion():
            verse = self.selected_verse_text
            trans = self.current_translation
            w     = self.current_word
            num   = self.number_var.get() or "-"
            gen   = self.gender_var.get() or "-"
            pos   = self.pos_var.get()    or "-"
            prior = ""
            try:
                search_num = self.number_var.get(); search_gen = self.gender_var.get(); search_pos = self.pos_var.get()
                if search_num == "NA" and search_gen == "NA" and search_pos == "NA":
                    matches = self.search_by_inflections(w)
                else:
                    matches = self.search_by_criteria(w, search_num, search_gen, search_pos) or self.search_by_inflections(w)
                pre_keys = load_predefined_keyset("1.1.1_birha.csv")
                rows = []
                for result, _count, _perc in matches:
                    parts = [p.strip() for p in result.split("|")]
                    if len(parts) < 7: parts += [""] * (7 - len(parts))
                    try:
                        key = tuple((parts[i] or "").strip() for i in (2,3,4,5,6))
                    except Exception:
                        key = ("") * 5
                    if key not in pre_keys: continue
                    highlight = parts[0] == parts[1] and is_full_word(parts[0])
                    if highlight:
                        parts = [f"**{p}**" for p in parts]
                        parts[0] = "? " + parts[0]
                    rows.append(
                        "| " + " | ".join(parts + [str(_count), f"{_perc:.1f}%"]) + " |"
                    )
                    if len(rows) >= 5: break
                if rows:
                    prior = (
                        "\n\n**Top Predefined Grammar Matches (from Birha CSV):**\n\n"
                        "| Word | Matched Token | Number | Gender | Grammar | Root | Type | Count | Frequency |\n"
                        "|------|----------------|--------|--------|---------|------|------|-------:|----------:|\n"
                        + "\n".join(rows) + "\n\n"
                    )
            except Exception:
                prior = ""
            prompt = f"""
            You are a Gurbani grammar expert. Given the verse and Darpan translation, help decide the grammar for the highlighted word.

            ### Verse
            {verse}

            ### Darpan Translation (condensed)
            {extract_darpan_translation(trans)}

            ### Target Word
            - Word: {w}
            - Number: {num}
            - Gender: {gen}
            - Part of Speech: {pos}

            {prior}

            ### Guidance
            1. Consider whether the word is a Noun, Pronoun, Adjective, Adverb, Verb, Postposition, Conjunction, or Interjection.
            2. If Noun/Pronoun: infer the correct case (e.g., genitive, instrumental, dative, locative, ablative) from gloss clues like "of", "by/with", "to/for", "in/on/at", "from".
            3. If Adjective: ensure it matches the qualified noun/pronoun in Number and Gender; avoid misclassification of indeclinable nouns.
            4. Keep output concise and evidence-based from the gloss.
            """.strip()
            self.root.clipboard_clear(); self.root.clipboard_append(prompt)
            messagebox.showinfo(
                "Prompt Ready",
                "Expert-level prompt (with secondary dictionary meanings) has been copied to your clipboard.\n"
                "Paste it into ChatGPT for its recommendation."
            )

        # Bottom separator + buttons
        sep = tk.Frame(win, bg='#cccccc', height=2)
        sep.pack(side=tk.BOTTOM, fill=tk.X, padx=20, pady=(0,5))
        btns = tk.Frame(win, bg='light gray')
        btns.pack(side=tk.BOTTOM, fill=tk.X, padx=20, pady=(6, BOTTOM_PAD))
        back_btn_text = "< Back to Translation" if mode == "verse" else "< Back to ABW Translation"
        back_cmd = (lambda: [win.destroy(), self.show_translation_input()]) if mode == "verse" else (lambda: [win.destroy(), self.show_word_translation_input()])
        back_btn = tk.Button(btns, text=back_btn_text, font=('Arial',12), bg='gray', fg='white', padx=20, pady=8, command=back_cmd)
        back_btn.pack(side=tk.LEFT)
        skip_btn = tk.Button(btns, text="Skip Word", font=('Arial',12), bg='orange', fg='white', padx=20, pady=8, command=lambda: self.skip_word_grammar(win))
        skip_btn.pack(side=tk.LEFT, padx=10)
        prompt_btn = tk.Button(btns, text="ÃƒÆ’Ã‚Â°Ãƒâ€¦Ã‚Â¸ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ Build Expert Prompt", font=("Arial",14,'italic'), bg='white', fg='dark cyan', padx=6, pady=4, command=ask_suggestion)
        prompt_btn.pack(side=tk.LEFT, padx=10)
        # Route submit to the appropriate handler per mode to avoid Verse dependencies in ABW
        submit_handler = (lambda: self.submit_input_grammar(word, index)) if mode == "verse" else (lambda: self.submit_input_grammar_for_word(word, index))
        submit_btn = tk.Button(btns, text="Submit", font=('Arial',12,'bold'), bg='dark cyan', fg='white', padx=20, pady=8, command=submit_handler)
        submit_btn.pack(side=tk.RIGHT)

        ctx = {"verse_text": td, "translation_text": trans, "back_button": back_btn, "skip_button": skip_btn,
               "prompt_button": prompt_btn, "submit_button": submit_btn, "mode": mode}
        return win, ctx

    def user_input_grammar(self, word, translation, index):
        """Verse entry: build shared UI and run modally (unchanged behavior)."""
        win = tk.Toplevel(self.root)
        win.title(f"Assess Grammar: {word}")
        self._wm_apply(win, margin_px=0, defer=True)
        try:
            self.current_translation = translation
        except Exception:
            pass
        self._build_user_input_grammar(win, word=word, translation=translation, index=index, mode="verse")
        win.transient(self.root)
        try:
            win.grab_set()
        except Exception:
            pass
        self.root.wait_window(win)

    def user_input_grammar_for_word(self, word, translation, index):
        """ABW entry point: build and show per-word grammar UI as a modal.

        - Sets ABW-specific title
        - Taskbar-safe maximize + F11 toggle
        - Modal via transient + grab_set
        - Self-contained; does not call Verse UI handlers
        """
        win = tk.Toplevel(self.root)
        win.title(f"Assess by Word ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ Grammar: {word}")
        self._wm_apply(win, margin_px=0, defer=True)
        try:
            self.current_translation = translation
        except Exception:
            pass
        self._build_user_input_grammar(win, word=word, translation=translation, index=index, mode="word")
        win.transient(self.root)
        try:
            win.grab_set()
        except Exception:
            pass
        def _on_close():
            try:
                win.grab_release()
            except Exception:
                pass
            try:
                win.destroy()
            except Exception:
                pass
        try:
            win.protocol('WM_DELETE_WINDOW', _on_close)
        except Exception:
            pass
        self.root.wait_window(win)

    def skip_word_grammar(self, win):
        """Skip grammar assessment for the current word and advance to the next."""
        try:
            confirm_skip = messagebox.askyesno("Confirm Skip", "Are you sure you want to skip this word?")
            if not confirm_skip:
                return
        except Exception:
            # If the confirmation dialog fails for any reason, proceed with skipping.
            pass

        try:
            if win and win.winfo_exists():
                win.destroy()
        except Exception:
            pass

        try:
            self.current_queue_pos += 1
            is_abw = (getattr(self, '_current_grammar_mode', 'verse') == 'word') or (getattr(self, '_current_detailed_mode', 'verse') == 'word')
            if is_abw:
                self.abw_process_next_word_assessment()
            else:
                self.process_next_word_assessment()
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred while skipping: {e}")

    def lookup_grammar_meanings_thread(self, word):
        """
        Look up dictionary meanings for ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¹Ã…â€œwordÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â‚¬Å¾Ã‚Â¢ on a background thread,
        then schedule update into the grammar UI.
        """
        meanings = self.lookup_word_in_dictionary(word)
        # schedule into mainloop
        self.root.after(0, lambda: self.update_grammar_meanings_ui(meanings))

    def update_grammar_meanings_ui(self, meanings):
        """
        Populate the meanings_inner_frame into N columns (now 7).
        """
        # 1) Clear any old widgets
        for w in self.meanings_inner_frame.winfo_children():
            w.destroy()

        # 2) Decide on how many columns (wider, denser layout across the canvas width)
        try:
            # Try to infer available width for meaning columns
            self.meanings_canvas.update_idletasks()
            avail_w = max(800, int(self.meanings_canvas.winfo_width() or 0))
        except Exception:
            avail_w = 1000
        col_w = 240  # approximate width per column incl. padding (slightly narrower to fit more)
        num_cols = max(6, min(12, avail_w // col_w))
        total   = len(meanings)
        # Ceil division so each column has at most ceil(total/num_cols) entries
        per_col = -(-total // num_cols)

        # 3) Grid each meaning into (row, column)
        for idx, m in enumerate(meanings):
            col = idx // per_col
            row = idx % per_col
            tk.Label(
                self.meanings_inner_frame,
                text=f"ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¢ {m}",
                bg='light gray',
                font=('Arial', 12),
                wraplength=max(140, col_w - 20),
                justify='left'
            ).grid(
                row=row,
                column=col,
                sticky='nw',
                padx=8, pady=2
            )
        
        # 4) NEW: stash into a growing list of dicts:
        entry = {
            "word": getattr(self, "current_word", None),
            "meanings": meanings
        }
        self.grammar_meanings.append(entry)

    def submit_input_grammar(self, word, index):
        """
        Collects grammar input and transitions to the dropdown step.
        """
        # 1) Extract the basic Number/Gender/POS the user just picked:
        number = self.number_var.get()
        gender = self.gender_var.get()
        pos    = self.pos_var.get()

        # 2) Gather verse + translation context:
        verse            = self.selected_verse_text
        raw_translation  = self.current_translation
        translation      = extract_darpan_translation(raw_translation)

        # 3) Pull the previously lookedÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Âup meanings out of self.grammar_meanings:
        meanings = next(
            (e["meanings"] for e in self.grammar_meanings if e["word"] == word),
            []
        )

        # 4) Build the initial "detailed" entry dict:
        entry = {
            "Vowel Ending":       word,
            "Number / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨":       number,
            "Grammar / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¯ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£":    "",   # to be filled in dropdown step
            "Gender / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â":       gender,
            "Word Root":           "",   # to be filled next
            "Type":                pos,
            "Evaluation":          "Derived",
            "Reference Verse":     verse,
            "Darpan Translation":  translation,
            "Darpan Meaning":      "| ".join(m.strip() for m in meanings),
            "ChatGPT Commentary":  "",    # to be pasted later
            # Ensure each saved word within the same verse is uniquely identified
            "Word Index":          int(index) if index is not None else ""
        }

        # 5) Store it so the next window can read & update it:
        self.current_detailed_entry = entry

        # 6) Hand off to your dropdown-UI (Verse wrapper):
        self.open_final_grammar_dropdown(word, entry["Type"], index)

    def submit_input_grammar_for_word(self, word, index):
        """ABW: Collect grammar input and transition to ABW dropdown (no Verse dependency)."""
        number = self.number_var.get()
        gender = self.gender_var.get()
        pos    = self.pos_var.get()
        verse            = self.selected_verse_text
        raw_translation  = self.current_translation
        translation      = extract_darpan_translation(raw_translation)
        meanings = next((e["meanings"] for e in self.grammar_meanings if e["word"] == word), [])
        entry = {
            "Vowel Ending":       word,
            COL_NUMBER:            number,
            COL_GRAMMAR:           "",
            COL_GENDER:            gender,
            "Word Root":           "",
            "Type":                pos,
            "Evaluation":          "Derived",
            "Reference Verse":     verse,
            "Darpan Translation":  translation,
            "Darpan Meaning":      "| ".join(m.strip() for m in meanings),
            "ChatGPT Commentary":  "",
            "Word Index":          int(index) if index is not None else ""
        }
        self.current_detailed_entry = entry
        self.open_final_grammar_dropdown_for_word(word, entry["Type"], index)

    # ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬
    # MAIN METHOD  ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ  drop-in replacement
    # ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬
    def _open_final_grammar_dropdown_common(self, word, pos, index, mode: str = "verse"):
        """
        After the user has chosen a Part-of-Speech, pop up a Toplevel
        with dropdowns for the detailed grammar fields _and_ a place
        to paste ChatGPT's commentary.
        """
        # Stash mode for save-phase routing (ABW vs Verse)
        try:
            self._current_detailed_mode = mode
        except Exception:
            pass

        # 1) --------------  Load & filter your CSV  -----------------
        self.grammar_db = pd.read_csv("1.1.1_birha.csv")
        df = self.grammar_db[self.grammar_db["Type"] == pos]

        # option lists
        # option lists (resolve actual headers robustly)
        _num_col  = _resolve_col(df, COL_NUMBER, 'Number / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨', 'Number')
        _gram_col = _resolve_col(df, COL_GRAMMAR, 'Grammar / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¯ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£', 'Grammar Case / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¯ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£', 'Grammar')
        _gen_col  = _resolve_col(df, COL_GENDER, 'Gender / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â', 'Gender')
        num_opts  = sorted(df[_num_col].dropna().unique().tolist()) if _num_col else []
        gram_opts = sorted(df[_gram_col].dropna().unique().tolist()) if _gram_col else []
        gen_opts  = sorted(df[_gen_col].dropna().unique().tolist()) if _gen_col else []
        entry = getattr(self, "current_detailed_entry", {}) or {}
        pos_type = str((entry.get("Type") if isinstance(entry, dict) else None) or pos or "")

        # Choose how to build root_opts based on whether it's a Noun
        if pos_type == "Noun / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Âµ":
            # Option-1: For Nouns, use hard-wired canonical endings
            root_opts = CANONICAL_ENDINGS.copy()
            for lst in (num_opts, gram_opts, gen_opts):
                if "NA" not in lst:
                    lst.insert(0, "NA")
        else:
            # Option-2: For all other types (e.g., Pronoun), use actual values from database
            root_opts = sorted(df["Word Root"].dropna().unique().tolist())
            for lst in (num_opts, gram_opts, gen_opts, root_opts):
                if "NA" not in lst:
                    lst.insert(0, "NA")

        # ---- Repeat-word banner bookkeeping ----
        verse_text = getattr(self, "current_pankti", "")
        verse_key = self._verse_key(verse_text)
        if getattr(self, "_last_dropdown_verse_key", None) != verse_key:
            self._repeat_note_shown = set()
            self._first_repeat_token = None
            self._last_dropdown_verse_key = verse_key
            self._suppress_repeat_notes_for_verse = False
            # kill any stale frame
            if hasattr(self, "literal_note_frame") and self.literal_note_frame:
                try:
                    if self.literal_note_frame.winfo_exists():
                        self.literal_note_frame.destroy()
                except Exception:
                    pass
                self.literal_note_frame = None
                self.literal_note_title = None
                self.literal_note_body = None

        words = list(getattr(self, "pankti_words", []))
        if not words and verse_text:
            words = verse_text.split()
        cached_words = getattr(self, "_raw_words_cache", None)
        if cached_words != words:
            norm_words = [self._norm_tok(w) for w in words]
            self._norm_words_cache = norm_words
            self._raw_words_cache = words
        else:
            norm_words = getattr(self, "_norm_words_cache", []) or [self._norm_tok(w) for w in words]

        idx = index
        display_word = words[idx] if idx < len(words) else word
        word_norm = norm_words[idx] if idx < len(norm_words) else self._norm_tok(display_word)
        # guard empty tokens; ignore vanished punctuation/ZW chars
        has_repeat = bool(word_norm) and norm_words.count(word_norm) >= 2
        if has_repeat and self._first_repeat_token is None and word_norm:
            self._first_repeat_token = word_norm
        seen_before = norm_words[:idx].count(word_norm) if word_norm else 0
        key = (verse_key, word_norm, "second")
        is_special_hit = (
            has_repeat
            and word_norm == self._first_repeat_token
            and seen_before == 1
            and key not in self._repeat_note_shown
        )

        inline_allowed = (
            getattr(self, "_use_inline_literal_banner", True)
            and not getattr(self, "_suppress_repeat_notes_for_verse", False)
        )

        suppress_first_occurrence_of_first_token = (
            has_repeat
            and self._first_repeat_token is not None
            and word_norm == self._first_repeat_token
            and seen_before == 0
        )

        if inline_allowed and is_special_hit:
            self._repeat_note_shown.add(key)
            special_text = (
                f"In literal analysis: The word ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“{display_word}ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â appears multiple times in this verse. "
                "The highlighted grammar options reflect your past selections for this word (or close matches) "
                "to encourage consistency. TheyÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â‚¬Å¾Ã‚Â¢re suggestions, not mandatesÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‚Âadjust if the current context differs."
            )
            self._ensure_literal_banner(special_text)
        elif inline_allowed and has_repeat and not suppress_first_occurrence_of_first_token:
            self._ensure_literal_banner(self._LITERAL_NOTE_TEXT)
        else:
            if not inline_allowed or not has_repeat:
                if hasattr(self, "literal_note_frame") and self.literal_note_frame:
                    try:
                        if self.literal_note_frame.winfo_exists():
                            self.literal_note_frame.destroy()
                    except Exception:
                        pass
                    self.literal_note_frame = None
                    self.literal_note_title = None
                    self.literal_note_body = None

        # 2) --------------  Build the window  -----------------------
        win = tk.Toplevel(self.root)
        win.title(f"Detail Grammar for ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¹Ã…â€œ{word}ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â‚¬Å¾Ã‚Â¢")
        win.configure(bg="light gray")
        # Exact per-monitor work-area maximize, deferred to avoid 1x1 restores
        try:
            self._wm_apply(win, margin_px=0, defer=True)
        except Exception:
            pass

        frm = tk.LabelFrame(
            win, text="Finalize Detailed Grammar",
            font=("Arial", 16, "bold"), bg="light gray",
            padx=10, pady=10
        )
        frm.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
        frm.grid_columnconfigure(1, weight=1)

        def _add_dropdown(row, label, var, options, colspan=1):
            ttk.Label(frm, text=label, font=("Arial", 12),
                    background="light gray").grid(
                row=row, column=0, sticky="w", padx=5, pady=8)
            cb = ttk.Combobox(
                frm, textvariable=var, values=options,
                state="readonly", font=("Arial", 12))
            cb.grid(row=row, column=1, columnspan=colspan,
                    sticky="ew", padx=5, pady=8)
            return cb

        # 3) --------------  Five dropdowns  -------------------------
        self.detailed_ve_var      = tk.StringVar(value=self._norm_get(entry, "\ufeffVowel Ending"))
        self.detailed_number_var  = tk.StringVar(value=(entry.get(COL_NUMBER)  if isinstance(entry, dict) else None) or "NA")
        self.detailed_grammar_var = tk.StringVar(value=(entry.get(COL_GRAMMAR) if isinstance(entry, dict) else None) or "")
        self.detailed_gender_var  = tk.StringVar(value=(entry.get(COL_GENDER)  if isinstance(entry, dict) else None) or "NA")
        self.detailed_root_var    = tk.StringVar(value=entry["Word Root"])

        _add_dropdown(0, "Word Under Analysis:", self.detailed_ve_var, [word], colspan=2)
        _add_dropdown(1, "Number / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨:",        self.detailed_number_var,  num_opts)
        _add_dropdown(2, "Grammar Case / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¯ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£:", self.detailed_grammar_var, gram_opts)
        _add_dropdown(3, "Gender / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â:",        self.detailed_gender_var,   gen_opts)
        _add_dropdown(4, "Word Root:",            self.detailed_root_var,     root_opts)

        # 4) --------------  Commentary box  -------------------------
        cm_frame = tk.LabelFrame(
            frm, text="ChatGPT Commentary", font=("Arial", 14, "bold"),
            bg="light gray", padx=8, pady=8
        )
        cm_frame.grid(row=5, column=0, columnspan=2,
                    sticky="nsew", padx=5, pady=(10, 0))
        self.detailed_commentary = tk.Text(
            cm_frame, wrap=tk.WORD, font=("Arial", 12),
            height=6, bd=1, relief="sunken", padx=5, pady=5
        )
        self.detailed_commentary.pack(fill=tk.BOTH, expand=True)

        # ---------- dynamic noun-map in self ----------
        if not hasattr(self, "noun_map"):
            self.noun_map = build_noun_map()

        def build_examples_footer():
            """
            Return a Markdown block that lists every ending-class with its full word,
            base form, and detachable suffix, taken from EXAMPLE_BASES.
            """
            lines = ["*Ending-class examples*"]
            for label in CANONICAL_ENDINGS:
                if label == "NA":
                    continue

                triples = EXAMPLE_BASES.get(label, [])
                if not triples:
                    continue

                # Build ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â  ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ + ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â style strings
                rendered = [
                    f"{full} ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ {base}{' + ' + suf if suf else ''}"
                    for full, base, suf in triples
                ]
                lines.append(f"- **{label}** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ " + ", ".join(rendered))

            return "\n".join(lines)

        # helper ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ build cheat-sheet table from noun_map
        def make_cheat_sheet(word: str, gender: str, number: str) -> str:
            """
            Progressive right-edge matcher, now bounded by len(word):
            ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¢ For L = 1 ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¦ len(word):
                    slice_w = word[-L:]
                    for every ending key E in noun_map:
                        if E[-L:] == slice_w  ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ collect E
            ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¢ Merge all collected endingsÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â‚¬Å¾Ã‚Â¢ case tables (deduped), build Markdown.
            """

            word_len = len(word)                              # new upper bound
            matched: list[str] = []

            # 1) -------- gather every ending with the same right-edge ------------
            for L in range(1, word_len + 1):                  # 1 ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¦ len(word)
                slice_w = word[-L:]
                for ending in self.noun_map:
                    if ending[-L:] == slice_w and ending not in matched:
                        matched.append(ending)

            if not matched:
                return ""                                     # nothing found

            # 2) -------- merge case ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ suffix lists for gender & number ----------
            merged: dict[str, list[str]] = {}
            for end in matched:
                cases = (
                    self.noun_map[end]
                        .get(gender or "NA", {})
                        .get(number or "NA", {})
                )
                for case, forms in cases.items():
                    merged.setdefault(case, []).extend(forms)

            if not merged:
                return ""                                     # no data for this combo

            # Deduplicate each list while preserving order
            for case, forms in merged.items():
                seen = set()
                merged[case] = [f for f in forms if not (f in seen or seen.add(f))]

            # 3) -------- build the mini-table -----------------------------------
            rows = [
                f"| {case:11} | {', '.join(forms)} |"
                for case, forms in merged.items()
            ]
            ending_list = ", ".join(matched)

            # build the core table but DONÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â‚¬Å¾Ã‚Â¢T return yet
            table_rows = "\n".join(rows)
            table_markdown = textwrap.dedent(f"""
                **Morphology map ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ endings matched: {ending_list}
                ({gender.split()[0]}/{number.split()[0]})**
                | Case         | Attested suffix(es) |
                |--------------|----------------------|
                {table_rows}
                _Table shows **attested** suffixes.
                If you need an unlisted case, propose a plausible form and justify._
            """).strip()

            # --- build a footer that shows EVERY ending-class with examples -------------
            footer = "\n" + build_examples_footer()
            return table_markdown + footer + "\n\n"

        # 5) --------------  Prompt-builder button  ------------------
        def build_detailed_prompt(num_opts=num_opts,
                                gram_opts=gram_opts,
                                gen_opts=gen_opts,
                                root_opts=root_opts):

            ve    = self.detailed_ve_var.get()      or "(please choose)"
            num   = self.detailed_number_var.get()  or "(please choose)"
            gram  = self.detailed_grammar_var.get() or "(please choose)"
            gen   = self.detailed_gender_var.get()  or "(please choose)"
            root  = self.detailed_root_var.get()    or "(please choose)"
            verse = entry["Reference Verse"]
            trans = entry["Darpan Translation"]
            dm    = entry["Darpan Meaning"]

            def make_block(title, items):
                lines = [f"- **{title}**"]
                for it in items:
                    lines.append(f"  ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ {it}")
                return "\n".join(lines)

            # ------------------------------------------------------------------
            # Use the existing search_by_criteria helper to surface any grammar
            # matches that align with the current selections.  This gives the
            # language model extra context about how the form appears in the
            # database.  If no match is found or the search fails, we simply
            # omit the block from the prompt.
            # ------------------------------------------------------------------
            try:
                crit_num = num if num != "(please choose)" else "NA"
                crit_gen = gen if gen != "(please choose)" else "NA"
                crit_matches = self.search_by_criteria(word, crit_num, crit_gen, pos)

                # Predefined-only filter keyset from CSV
                pre_keys = load_predefined_keyset("1.1.1_birha.csv")
                rows = []
                for result, _count, _perc in crit_matches:
                    parts = [p.strip() for p in result.split("|")]
                    if len(parts) < 7:
                        parts += [""] * (7 - len(parts))

                    # Skip if the match does not map to a Predefined row in CSV
                    try:
                        key = tuple((parts[i] or "").strip() for i in (2, 3, 4, 5, 6))
                    except Exception:
                        key = ("") * 5
                    if key not in pre_keys:
                        continue

                    highlight = parts[0] == parts[1] and is_full_word(parts[0])
                    if highlight:
                        parts = [f"**{p}**" for p in parts]
                        parts[0] = "ÃƒÆ’Ã‚Â¢Ãƒâ€¦Ã¢â‚¬Å“ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ " + parts[0]

                    rows.append("| " + " | ".join(
                        parts + [str(_count), f"{_perc:.1f}%"]
                    ) + " |")
                    if len(rows) >= 5:
                        break

                if rows:
                    headers = [
                        "Word under Analysis",
                        "Vowel Ending / Word Matches",
                        "Number / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨",
                        "Grammar / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¯ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£",
                        "Gender / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â",
                        "Word Root",
                        "Type",
                        "Match Count",
                        "Match %",
                    ]
                    table_lines = [
                        "**Top Grammar Matches**",
                        "| " + " | ".join(headers) + " |",
                        "| " + " | ".join(["---"] * len(headers)) + " |",
                        *rows,
                    ]
                    matches_block = "\n".join(table_lines)
                else:
                    matches_block = "**Top Grammar Matches**\nNo predefined examples found"
            except Exception as exc:
                print(f"search_by_criteria failed: {exc}")
                matches_block = ""

            opts_block = "\n\n".join([
                make_block("Word Under Analysis", [ve]),
                make_block("Number / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ options",   num_opts),
                make_block("Grammar Case / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¯ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£ options", gram_opts),
                make_block("Gender / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â options",  gen_opts),
                make_block("Word-Root options",      root_opts),
            ])

            # noun-specific notes
            ending_cheat_sheet = ""
            implicit_note      = ""
            common_sense_note  = ""

            if entry["Type"] == "Noun / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Âµ":
                ending_cheat_sheet = make_cheat_sheet(ve, gen, num)

                implicit_note = textwrap.dedent("""\
                    **IMPLICIT POST-POSITIONS & CASE DECLENSIONS**  
                    In GurbÃƒÆ’Ã¢â‚¬Å¾Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â¡Ãƒâ€šÃ‚Â¹ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã¢â‚¬Å¾Ãƒâ€šÃ‚Â«, relationships such as *to, from, with, of, in* are conveyed
                    by **inflected endings** rather than modern post-positions (`ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°`, `ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²`
                    ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¦). A noun may appear unmarked while the Darpan gloss supplies a helper.

                    **How to read the gloss**  
                    ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¢ If the gloss inserts **to / for / of / by / with / from / in / on / at / O / Hey**
                    that is absent in the verse, treat it as an **implicit post-position**
                    and pick the matching **case**.  
                    ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¢ If the gloss repeats the word without a helper, default to
                    **Nominative / Accusative** and let context refine the choice.

                    | Helper | Punjabi marker | Case |
                    |--------|----------------|------|
                    | to / for   | `ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°`, `ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã¢â‚¬Â `     | **Dative** |
                    | of         | `ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬`      | **Genitive** |
                    | by / with  | `ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²`, `ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡`  | **Instrumental** |
                    | from / out of | `ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡`, `ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡` | **Ablative** |
                    | in / on / at | `ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â±ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡`, `ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â±ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡`, `ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡` | **Locative** |
                    | O / Hey    | *(address)*     | **Vocative** |

                    _Endings overlap: NomÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°Ãƒâ€¹Ã¢â‚¬Â Acc, GenÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°Ãƒâ€¹Ã¢â‚¬Â Dat, InstÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°Ãƒâ€¹Ã¢â‚¬Â Loc ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ use semantics to decide._
                """).strip() + "\n\n"

                common_sense_note = textwrap.dedent("""\
                    **SEMANTIC SANITY CHECK ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ DOES THE LABEL REALLY FIT?**  
                    Match the case to the *role* the noun plays.

                    **Quick Meanings**  Nom=subject | Acc=object | Inst=by/with | Dat=to/for |
                    Gen=of | Abl=from | Loc=in/on | Voc=address

                    ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¢ Instrumental ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ means, agency, tool  
                    ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¢ Locative     ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ spatial/temporal setting  
                    ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¢ Dative       ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ recipient, purpose  
                    ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¢ Genitive     ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ ownership, relation  
                    ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¢ Ablative     ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ source, cause  
                    ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¢ Nom / Acc    ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ subject vs. direct object (no helper)  
                    ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¢ Vocative     ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ direct address

                    **Ambiguity reminder** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ If **one suffix stands for two cases**
                    (e.g., ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã¢â‚¬Â  = Nom *and* Acc), *explain your semantic reason* for choosing.

                    **Oblique + Post-position lines** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ GurbÃƒÆ’Ã¢â‚¬Å¾Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â¡Ãƒâ€šÃ‚Â¹ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã¢â‚¬Å¾Ãƒâ€šÃ‚Â« occasionally stacks a
                    post-position **after** an oblique form **and** after a direct form
                    (see examples with *ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡*, *ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â *).  Either is validÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‚Âchoose the case
                    that best reflects the combined meaning.
                """).strip() + "\n\n"
                
            elif entry["Type"] == "Pronoun / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Âµ":
                # ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ Pronoun block with enriched cross-category logic ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬
                implicit_note = textwrap.dedent("""\
                    **PRONOUNS ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ INFLECTIONS, IDENTITY & IMPLIED MEANINGS**  
                    In GurbÃƒÆ’Ã¢â‚¬Å¾Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â¡Ãƒâ€šÃ‚Â¹ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã¢â‚¬Å¾Ãƒâ€šÃ‚Â«, pronouns diverge from noun patterns and inflect by **person, number, and gender**.  
                    Their meaning is sometimes explicit (like ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ = I), but often **derived from Darpan's gloss**.

                    **Core Steps to Identify the Case**  
                    1. **Read the gloss literally.**  
                    If it adds a helper like *to, from, with, in*, this signals an **implicit post-position**.  
                    Match it with:  
                    ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¢ `ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°`, `ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã¢â‚¬Â ` ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ Dative  
                    ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¢ `ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡`, `ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡` ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ Genitive  
                    ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¢ `ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡`, `ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡`, `ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡`, `ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡` ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ Ablative  
                    ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¢ `ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²`, `ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â±ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡`, `ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â±ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡`, `ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²`, `ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°`, etc. ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ Instrumental / Locative  
                    ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¢ `O`, `Hey` ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ Vocative

                    2. **Check form compatibility.**  
                    Every person/gender/number has a finite set of endings (see below).  
                    Match the surface form to a standard **canonical pronoun**.

                    3. **For Relative / Interrogative / Reflexive / Indefinite types**,  
                    blend case logic with **semantic roles**: e.g.,  
                    ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¢ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â° ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“to whomÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ Dative  
                    ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¢ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“on whomÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ Locative  
                    ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¢ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Âª ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ Reflexive emphatic  
                    ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¢ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ Genitive relative

                    _Postpositions are often absent but impliedÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‚Âyour judgment is key._  
                    Also note: **GurbÃƒÆ’Ã¢â‚¬Å¾Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â¡Ãƒâ€šÃ‚Â¹ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã¢â‚¬Å¾Ãƒâ€šÃ‚Â« often uses plural pronouns to show respect.**
                """).strip() + "\n\n"

                common_sense_note = textwrap.dedent("""\
                    **PRONOUN SEMANTIC CHECK ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ ROLE IN MEANINGFUL CONTEXT**  
                    Pronouns are **not just replacements for nouns**ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‚Âthey carry personhood, humility, or divinity.

                    ÃƒÆ’Ã‚Â¢Ãƒâ€¦Ã¢â‚¬Å“ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ Use this test logic:  
                    - **Is the pronoun the subject?** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ Nom  
                    - **Receiving the action?** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ Acc  
                    - **Belonging to someone?** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ Gen  
                    - **Given to someone?** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ Dat  
                    - **Means or tool or ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“withÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â sense?** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ Inst  
                    - **Place or inner state?** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ Loc  
                    - **Directly addressed?** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ Voc  

                    ÃƒÆ’Ã‚Â¢Ãƒâ€¦Ã‚Â¡Ãƒâ€šÃ‚Â ÃƒÆ’Ã‚Â¯Ãƒâ€šÃ‚Â¸Ãƒâ€šÃ‚Â For overlapping forms:  
                    - Use the Darpan helper (e.g., "to me", "from them", "by whom")  
                    - Ask what semantic role the pronoun plays **in that line**  
                    - e.g., ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â may be Nom or Acc depending on meaning

                    **Special Guidance per Category**  
                    - **Reflexive** (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Âª, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡): Self-reference or emphasis  
                    - **Relative/Correlative** (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹...ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹): Link two ideas (doer/result, condition/result)  
                    - **Interrogative** (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¦Ã¢â‚¬â„¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸): Structure question  
                    - **Indefinite** (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã¢â‚¬Â , ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â­): Ambiguous subject  
                    - **Honorific 2nd Person** (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®): May appear plural but refer to one Divine

                    **Final Tip**: Plural/oblique/abstract usage may reflect poetic or spiritual nuance more than grammar. Follow meaning.
                """).strip() + "\n\n"

                ending_cheat_sheet = textwrap.dedent("""\
                    **PRONOUN CASE ENDINGS ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ EXAMPLES ACROSS CATEGORIES**

                    ÃƒÆ’Ã‚Â°Ãƒâ€¦Ã‚Â¸ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒâ€šÃ‚Â¹ **Valid Number / Gender Combinations per Category**  
                    *(Use this to cross-check if your feature choices are logically possible)*

                    - **1st Person / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â±ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â® ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Å“**  
                    ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ Number: Singular / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢, Plural / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â  
                    ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ Gender: Trans / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢

                    - **2nd Person / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â§ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â® ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Å“**  
                    ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ Number: Singular / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢, Plural / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â  
                    ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ Gender: Trans / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢

                    - **3rd Person / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¯ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Å“**  
                    ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ Number: Singular / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢, Plural / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â  
                    ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ Gender: Masculine / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â, Feminine / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬, Trans / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢

                    - **CoRelative / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â§**  
                    ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ Number: Singular / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢, Plural / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â  
                    ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ Gender: Masculine / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â, Feminine / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬, Trans / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢

                    - **Relative / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â§**  
                    ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ Number: Singular / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢, Plural / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â  
                    ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ Gender: Masculine / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â, Feminine / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬, Trans / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢

                    - **Interrogative / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¶ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢**  
                    ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ Number: Singular / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢, Plural / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â  
                    ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ Gender: Masculine / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â, Feminine / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬, Trans / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢

                    - **Reflexive / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢**  
                    ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ Number: Singular / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢, Plural / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â  
                    ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ Gender: Masculine / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â, Feminine / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬, Trans / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢

                    - **Indefinite / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢**  
                    ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ Number: Singular / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢, Plural / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â  
                    ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ Gender: Masculine / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â, Feminine / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬, Trans / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢

                    _ÃƒÆ’Ã‚Â¢Ãƒâ€¦Ã¢â‚¬Å“Ãƒâ€šÃ‚Â³ Note: ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“TransÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢) appears for most categories due to universal/neutral references or poetic plurality._

                    **1st Person / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â±ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â® ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Å“ Pronouns ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ Case Examples**
                    - Ablative ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨: ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â  / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â® ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡
                    - Accusative ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®: ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â  / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â° / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â° / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â® / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿
                    - Dative ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨: ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â  / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â  / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â  / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â° / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â® (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°) / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â 
                    - Genitive ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â§: ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â° / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â  / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â° / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â  / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾
                    - Locative ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â§ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£: ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â  / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â  / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â  ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â  ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â® / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â  / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â 
                    - Nominative ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾: ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â  / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â° / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â® / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â

                    **2nd Person / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â§ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â® ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Å“ Pronouns ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ Case Examples**
                    - Ablative ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨: ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â  / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â  ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â  ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â§ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â§ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â  ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â® ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡
                    - Accusative ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®: ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â° / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â  / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â§ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â§ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â§ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â§ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â§ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â  / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â® / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â° / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬
                    - Dative ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨: ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â° / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â  / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â§ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â§ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â® / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â® ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â° / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â  / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â  ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â° / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¥ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¥ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡
                    - Genitive ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â§: ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â° / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Âµ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â  / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â° / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â  / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â° / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¥ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¥ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡
                    - Locative ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â§ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£: ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â  / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â  ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â§ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â§ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â  / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â® / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿
                    - Nominative ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾: ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â° / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â  / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â§ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â§ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â  / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â® / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â® ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â  / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â  / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â  / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â  / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿

                    **3rd Person / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¯ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Å“ Pronouns ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ Case Examples**
                    - Ablative ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨: ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡) / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â  / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡) / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â
                    - Accusative ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®: ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã¢â‚¬Â  / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â 
                    - Dative ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨: ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â  / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬â„¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°) / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â
                    - Genitive ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â§: ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â  / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â  (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾) / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬) / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾) / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾) (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â ) (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡)
                    - Instrumental ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£: ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿)
                    - Locative ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â§ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£: ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â ) / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂºÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â ) / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡
                    - Nominative ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾: ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â  / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã¢â‚¬Â 

                    **CoRelative / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â§ Pronouns ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ Case Examples**
                    - Ablative ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨: ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡)
                    - Accusative ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®: ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°) / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â  / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â 
                    - Dative ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨: ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°) / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°) / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°) / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹) / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â  / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â 
                    - Genitive ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â§: ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹) / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬) / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾) / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬) / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡) / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿) / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬) / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â  / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡)
                    - Instrumental ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£: ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â
                    - Locative ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â§ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£: ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â ) / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬) / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿) / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿
                    - Nominative ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾: ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã¢â‚¬Â  / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã¢â‚¬Â  / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã¢â‚¬Â  / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â 

                    **Indefinite / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ Pronouns ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ Case Examples**
                    - Ablative ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨: ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â­ (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡) / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â­ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â° / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ
                    - Accusative ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®: ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â° / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹) / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â) / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â  / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂºÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â  / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â  (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂºÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â) / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂºÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹) / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â  / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã¢â‚¬Â  / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â  / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¥ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¥ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â² / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â­ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â­ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â­ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â­ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â  (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹) / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â­ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â­ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂºÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â) / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â­ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â) / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â­ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â  / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â­ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â­ (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂºÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â) / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹) / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â
                    - Dative ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨: ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â  / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬) / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â  / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â­ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â­ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â 
                    - Genitive ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â§: ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â° / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â  / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â­ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â­ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â 
                    - Instrumental ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£: ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â  / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â
                    - Locative ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â§ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£: ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿)
                    - Nominative ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾: (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°) ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â° / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹) / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â­ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿) / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂºÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â) / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â  / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â  / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂºÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã¢â‚¬Â  / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â  / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂºÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬) / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â  / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã¢â‚¬Â  / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã¢â‚¬Â  / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â  / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â² / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â  / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â­ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â­ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â­ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â­ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â­ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â­ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â­ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂºÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â) / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â­ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹) / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â­ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡) / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â­ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã¢â‚¬Â ) / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â­ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â­ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â­ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â

                    **Interrogative / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¶ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ Pronouns ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ Case Examples**
                    - Accusative ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®: ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â  / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â
                    - Dative ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨: ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°) / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°) / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â 
                    - Genitive ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â§: ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â
                    - Locative ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â§ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£: ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿) / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°) / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿) / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â  (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿)
                    - Nominative ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾: ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â  / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹

                    **Reflexive / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ Pronouns ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ Case Examples**
                    - Ablative ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨: ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡) / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¦Ã¢â‚¬â„¢
                    - Accusative ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®: ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â
                    - Dative ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨: ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°) / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â  (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹)
                    - Genitive ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â§: ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Âª / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â  / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Âª / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â  / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾
                    - Instrumental ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£: ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â  (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿)
                    - Locative ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â§ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£: ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â 
                    - Nominative ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾: ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Âª (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬) / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬) / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â 

                    **Relative / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â§ Pronouns ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ Case Examples**
                    - Ablative ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨: ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡) / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡)
                    - Accusative ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®: ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°) / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°) / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã¢â‚¬Â  ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã¢â‚¬Â  / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¯ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â
                    - Dative ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨: ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â 
                    - Genitive ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â§: ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â ) / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿) / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡) / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬) / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾) / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬) / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡) / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹
                    - Instrumental ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£: ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹
                    - Locative ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â§ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£: ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹
                    - Nominative ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾: ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã¢â‚¬Â 

                    _Ending note: **ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°** is often **omitted** before postpositions like ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡.  
                    e.g., **ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¥ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿** instead of **ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¥ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿**_
                """).strip() + "\n\n"

            elif entry["Type"] == "Adjectives / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¶ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¶ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£":
                # ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬
                # 3-B  IMPLICIT-NOTE  ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ how to ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“readÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â the gloss
                # ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬
                implicit_note = textwrap.dedent("""
                    **ADJECTIVES IN GURBÃƒÆ’Ã¢â‚¬Å¾ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¡Ãƒâ€šÃ‚Â¹ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã¢â‚¬Å¾Ãƒâ€šÃ‚Âª ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ AGREEMENT & HINTS FROM THE DARPAN GLOSS**

                    ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¢ An adjective always **agrees in gender & number** with the noun /
                    pronoun it qualifies.  Case is *not* tagged independently for adjectives;
                    if a noun shifts to an oblique form (due to post-positions like
                    `ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¦`) the adjective may simply copy that *ending*.

                    ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¢ **Look at the helper words the Darpan adds**:
                    - If the gloss inserts a post-position after the noun
                        (*e.g.* ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“to the **good** oneÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â, ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“in the **other** realmÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â), the adjective
                        will mirror whatever oblique ending the noun shows ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ **but you still
                        classify the adjective only by Gender / Number / Class**.
                    - If the gloss repeats the adjective without a helper,
                        treat the form you see in the verse as the **direct** (base) form.

                    _Quick reminder ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ common agreement endings_  
                    | Ending-class | Masc.Sg | Fem.Sg | Plural | Notes |
                    |--------------|---------|--------|--------|-------|
                    | **Mukta**    | ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦      | ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã¢â‚¬â€ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ **ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦** dropped for fem./pl. |
                    | **KannÃƒÆ’Ã¢â‚¬Å¾Ãƒâ€šÃ‚Â**    | ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â       | ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã¢â‚¬Â      | ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â     | |
                    | **SihÃƒÆ’Ã¢â‚¬Å¾Ãƒâ€šÃ‚ÂrÃƒÆ’Ã¢â‚¬Å¾Ãƒâ€šÃ‚Â«**   | ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿      | ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿      | ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡      | |
                    | **BihÃƒÆ’Ã¢â‚¬Å¾Ãƒâ€šÃ‚ÂrÃƒÆ’Ã¢â‚¬Å¾Ãƒâ€šÃ‚Â«**   | ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬      | ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã¢â‚¬Â      | ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â/ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã¢â‚¬Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡| |

                    _When in doubt: match what the noun is doing rather than forcing
                    a new inflection on the adjective._
                """).strip() + "\n\n"

                # ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬
                # 3-C  COMMON-SENSE-NOTE  ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ semantic & class sanity
                # ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬
                common_sense_note = textwrap.dedent("""
                    **SEMANTIC CHECK ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ DOES THE LABEL FIT THIS ADJECTIVE?**

                    ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‹Å“Ãƒâ€šÃ‚Â  **Identify the class** (use the column ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“Adjective Class / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¶ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¶ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â):  
                    ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¢ **Qualitative / Descriptive (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢)** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ *ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾*  
                    ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¢ **Demonstrative (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¼ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢)** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ *ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â*  
                    ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¢ **Indefinite (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â°Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â°Ãƒâ€šÃ‚Â¶ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢)** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ *ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã¢â‚¬Â , ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â , ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â­*  
                    ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¢ **Pronominal**  
                        ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ *ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ (possessive) / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â , ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â° (relative)*  
                    ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¢ **Interrogative (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¼ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢)** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ *ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â³, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡*  
                    ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¢ **Numeral (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â  ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢)**  
                        ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ **Cardinal** *ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹* | **Ordinal** *ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¦*

                    ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‹Å“Ãƒâ€šÃ‚Â¡ **Verify agreement** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ does the ending you see match the gender &
                    number of the noun in the gloss?  Typical pitfalls:  
                    ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¢ plural nouns paired with singular adjective forms,  
                    ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¢ masculine endings left on a feminine noun after emendation.

                    ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‹Å“Ãƒâ€šÃ‚Â¢ **Ambiguity guardrails**  
                    ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¢ Many demonstratives (*ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¦*) double as pronouns ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ keep them
                        in **Adjective** only when they *modify* a following noun.  
                    ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¢ Some numerals can work adverbially (*ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â­ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡*, ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“ran a lotÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â) ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ do not
                        tag those as adjectives.

                    _If two classes seem possible, pick the one that best serves the
                    **function in that specific gloss line** and give one-line reasoning._
                """).strip() + "\n\n"

                ending_cheat_sheet = textwrap.dedent("""\
                **ADJECTIVE ENDINGS ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ QUICK REFERENCE (GurbÃƒÆ’Ã¢â‚¬Å¾Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â¡Ãƒâ€šÃ‚Â¹ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã¢â‚¬Å¾Ãƒâ€šÃ‚Â« corpus)**

                ÃƒÆ’Ã‚Â°Ãƒâ€¦Ã‚Â¸ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒâ€šÃ‚Â¹ **Agreement grid (what can legally combine)**  
                ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¢ **Number / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ Singular / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢, Plural / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â, NA  
                ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¢ **Gender / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ Masc / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â, Fem / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬, Neut / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢, NA  
                ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¢ **Surface ending-classes** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¦Ã¢â‚¬â„¢, NA  
                ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¢ **Sub-classes** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ Qualitative, Demonstrative, Indefinite, Possessive-pronom., Pronominal, Interrogative, Numeral (Card & Ord), Diminutive, Negation, Tat-sam, Compound, NA  

                <sub>Adjectives never carry an independent ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“caseÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â; if the noun is oblique, the adjective just copies that ending.</sub>

                ---

                ### A ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â· Canonical ending patterns  

                | Ending-class | Masc Sg | Fem Sg | Plural | Tiny sample from text |
                |--------------|---------|--------|--------|-----------------------|
                | **ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾**    | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡**ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾** | ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡**ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡** | **ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¥ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â**, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ |
                | **ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾**     | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â**ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾** | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â**ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬** | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â**ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡** | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ |
                | **ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬**   | ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â | ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²**ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡** | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¼ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ |
                | **ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬**   | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°**ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬** | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°**ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬** | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°**ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡** | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â­ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ |
                | **ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾**     | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â­**ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°** | ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â | ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â (rare) |
                | **ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¦Ã¢â‚¬â„¢** | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²**ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â** | ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â | ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â±ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¦Ã¢â‚¬â„¢ |

                ---

                ### B ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â· Sub-class snapshots  

                | Class / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â® | 2-4 high-frequency examples (agreement marked) |
                |--------------|-----------------------------------------------|
                | **Qualitative (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£)** | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ (M), ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ (F), ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ (Pl) ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¢ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¥ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â (M) ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¢ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â (M) |
                | **Demonstrative (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¼ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡)** | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â (M Sg), ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ (F Sg), ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â |
                | **Indefinite (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¼ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡)** | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã¢â‚¬Â , ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã¢â‚¬Â , ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â­, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ |
                | **Possessive-pronominal** | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ (M), ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ (F), ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ (Pl) ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¢ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ |
                | **Pronominal (relative etc.)** | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ (F/M), ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â |
                | **Interrogative (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¶ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨)** | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â (M Sg), ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â , ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â |
                | **Numeral ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ Cardinal** | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â° |
                | **Numeral ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ Ordinal** | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¥ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â  |
                | **Negation** | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ |
                | **Tat-sam (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ loan)** | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ |
                | **Diminutive** | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â |
                | **Compound** | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â§ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ |

                """).strip() + "\n\n"

            elif entry["Type"] == "Verb / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ":
                # ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬
                # 4-B  IMPLICIT-NOTE  ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ how to ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“readÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â the gloss
                # ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬
                implicit_note = textwrap.dedent("""\
                **VERBS IN GURBÃƒÆ’Ã¢â‚¬Å¾ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¡Ãƒâ€šÃ‚Â¹ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã¢â‚¬Å¾Ãƒâ€šÃ‚Âª ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ IMPLIED CLUES FROM THE GLOSS**

                Verbs in GurbÃƒÆ’Ã¢â‚¬Å¾Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â¡Ãƒâ€šÃ‚Â¹ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã¢â‚¬Å¾Ãƒâ€šÃ‚Â« span a wide linguistic spectrumÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂLahindÃƒÆ’Ã¢â‚¬Å¾Ãƒâ€šÃ‚Â«, Braj, HindustÃƒÆ’Ã¢â‚¬Å¾Ãƒâ€šÃ‚ÂnÃƒÆ’Ã¢â‚¬Å¾Ãƒâ€šÃ‚Â«, and archaic PanjÃƒÆ’Ã¢â‚¬Å¾Ãƒâ€šÃ‚ÂbÃƒÆ’Ã¢â‚¬Å¾Ãƒâ€šÃ‚Â«. The verse alone often omits explicit markers for **tense, voice, mood, or even subject**. Prof. SÃƒÆ’Ã¢â‚¬Å¾Ãƒâ€šÃ‚Âhib SiÃƒÆ’Ã‚Â¡Ãƒâ€šÃ‚Â¹ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ghÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â‚¬Å¾Ã‚Â¢s **Darpan gloss** therefore becomes our decoder ring: it regularly inserts the **hidden agent, auxiliary, or intent** that lets us recover the full verbal meaning.

                ---

                ### ÃƒÆ’Ã‚Â¢Ãƒâ€¦Ã¢â‚¬Å“ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â Step 1 ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â· Read the gloss literally
                Ask yourself:
                * Is the action **ongoing**, **completed**, or **yet to come**?
                * Is the subject **doing** the action or **receiving** it?
                * Is the clause a **command**, a **wish**, or a **hypothetical**?
                * Do helper words appearÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â*has, was, should, may, being, let*ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‚Âthat hint at aspect or mood?

                ---

                ### ÃƒÆ’Ã‚Â¢Ãƒâ€¦Ã¢â‚¬Å“ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â Step 2 ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â· Map the gloss cue to a grammatical category

                | Category            | Common cues in the gloss (Eng. gloss)            |
                |---------------------|--------------------------------------------------|
                | **Present**         | do, does, is, are, becomes, gives                |
                | **Past**            | did, was, were, had, gave, came                  |
                | **Future**          | will, shall, would                               |
                | **Imperative**      | (you) give, fall, listen ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â direct command forms  |
                | **Subjunctive**     | if ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¦ may / might / should / let us               |
                | **Passive**         | is called, was given ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â object promoted to subject |
                | **Participles**     | having done, while doing, upon going, imbued     |
                | **Compound/Aux**    | do come, has gone, may go ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â multi-verb chains    |

                ---

                ### ÃƒÆ’Ã‚Â°Ãƒâ€¦Ã‚Â¸Ãƒâ€šÃ‚Â§Ãƒâ€šÃ‚Â  Key heuristics from the Darpan gloss
                * **ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“was made / is givenÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ strong passive signal.  
                * **ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“has shown / had comeÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ perfect aspect; expect past-participle + auxiliary.  
                * If the gloss shows the subject **causing** another to act (*was made to go*) ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ tag the verb **causative**.

                ---

                ### ÃƒÆ’Ã‚Â°Ãƒâ€¦Ã‚Â¸ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒâ€¦Ã¢â‚¬â„¢ Postposition surrogates
                Gloss words like *to, by, with, for, from* often reveal an implied **shift in voice** or a **participial/causative chain** hidden in the surface form.

                ---

                ### ÃƒÆ’Ã‚Â°Ãƒâ€¦Ã‚Â¸ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â€šÂ¬Ã…Â¾ When in doubt
                * Subject absent, object prominent ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ suspect **passive**.  
                * Two verbs side-by-side (*will come go*, *has been given*) ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ parse for **compound** or **auxiliary** roles.  
                * Conditional tone (*if ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¦ may ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¦*, *let it be ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¦*) ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ test for **subjunctive**.

                ---

                ### ÃƒÆ’Ã‚Â°Ãƒâ€¦Ã‚Â¸Ãƒâ€šÃ‚Â§Ãƒâ€šÃ‚Â© Suffix hints  
                Endings like **ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°, ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬, ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®, ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦** (and LahindÃƒÆ’Ã¢â‚¬Å¾Ãƒâ€šÃ‚Â« ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°, ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â) can encode person or emphasis. Cross-check with the glossÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â‚¬Å¾Ã‚Â¢s subject reference.

                ---

                > **Rule of thumb**  
                > *If the gloss shows something **happening to** someone and the agent is missing ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ think passive.*  
                > *If multiple verbs are chained, the **right-most** verb usually carries tense/voice; earlier ones express the semantic action.*

                _Use the glossÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‚Âits hidden auxiliaries, agents, and helpersÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‚Âto uncover the verbÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â‚¬Å¾Ã‚Â¢s true grammatical load._\
                """).strip() + "\n\n"


                common_sense_note = textwrap.dedent("""\
                ### ÃƒÆ’Ã‚Â°Ãƒâ€¦Ã‚Â¸ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒâ€šÃ‚Â¹ `common_sense_note` ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ VERBS / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â  (semantic sanity layer)

                **Essence**ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€ Ã¢â‚¬â„¢A sieve that questions every verb label: *Does this person ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â number ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â tense truly fit what the verb is doing in the paÃƒÆ’Ã‚Â¡Ãƒâ€šÃ‚Â¹ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ktÃƒÆ’Ã¢â‚¬Å¾Ãƒâ€šÃ‚Â«?*

                **Vision**ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€ Ã¢â‚¬â„¢Fuse surface-form clues with syntactic/semantic roles so edge-cases (poetic plurals, ergative flips, auxiliary drop, LahindÃƒÆ’Ã¢â‚¬Å¾Ãƒâ€šÃ‚Â« quirks) are flagged, not rubber-stamped.

                ---

                ## 1 ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â· Finite vs Non-finite: cheat grid  

                | Tag you plan | Sanity checks (abort / relabel if violated) |
                |--------------|---------------------------------------------|
                | **Present / Future** | Ending shows **person+number; no gender**. If ending = ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ **without** auxiliary **ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â /ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨**, treat as participle (habitual/progressive) not finite. |
                | **Imperative** | Only 2nd-person. Command/request mood. If clause is conditional (*ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¦*) ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ **Subjunctive** not Imperative. |
                | **Subjunctive** | Expresses wish/suggestion; often with *ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡*. Never shows gender agreement. |
                | **Past / Perfective** | Built on past-participle endings **ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â  / ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã¢â‚¬Â  / ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â**. Transitive verbs agree with **object** (ergative); intransitives with **subject**. |
                | **Passive finite** | Look for **ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â  ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â** etc. Object promoted to subject; auxiliary **ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â** etc. present/past table (ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â§ passive pages). |
                | **Causative** | Endings ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾, ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â³, ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°, ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡, ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¦; semantics must show *caused* action. |
                | **Auxiliary-only token** | If root **ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹** form (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â , ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¦) appears **alone**, tag = **Auxiliary Verb** not main finite. |
                *If the Canonical row label is ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“Pronominal Suffixes ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â you **must tag Grammar Case = ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“Pronominal Suffixes ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â**, not plain Past/Present.*
                *For finite verbs, **Word-Root must record the person (1st / 2nd / 3rd)**; tense or aspect belongs in ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“Grammar Case / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¯ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£,ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â not in Word-Root.*

                ---

                ## 2 ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â· Past-participle agreement sanity  

                1. **Intransitive:** participle ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â subject.  
                2. **Transitive (ergative):** participle ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â object; subject in instrumental/obl.  
                3. **Pron.-suffix ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°/-ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â:** when object = **ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â /ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°**, endings like **ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â** act as clitics ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ tag ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“Pronominal-suffixÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â sub-type.  
                4. Gender/number mismatch with controller ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ flag for review.

                ---

                ## 2A ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â· When gender actually matters  

                * **Finite verbs** (Present, Future, Imperative, Subjunctive, Causative, Auxiliary)  
                  ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ **never carry masc/fem marks** in SGGS.  *Finite verbs must therefore be tagged **Gender = Trans / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢** (not NA).*

                * **Participles** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ the only verb forms that **do** mark gender:  
                  ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¢ Perfect / perfective: **Masc SG -ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â  / Fem SG -ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã¢â‚¬Â  / Masc PL -ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â / Fem PL -ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã¢â‚¬Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡**  
                  ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¢ Habitual / imperfective: **Masc SG -ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ / Fem SG -ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ / Masc PL -ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ / Fem PL -ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡**  
                  ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¢ Dialectal allomorphs (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ **-ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ**, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ **-ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¯ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹**, etc.) are **still Masc SG**.

                * **Controller rule**  
                  ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ **Intransitive** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ participle agrees with **subject**.  
                  ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ **Transitive perfective** (ergative) ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ participle agrees with **object**.

                * **Auxiliaries stay neuter.**  `ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â /ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¦` never add gender; only the participle does.

                ---

                ## 3 ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â· Auxiliary verbs & silent dropping  

                * Present auxiliaries: **ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ (1 sg), ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â  (2 sg), ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â  (3 sg), ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ (1 pl), ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â (2 pl respect), ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨/hin (3 pl)**.  
                * Past auxiliaries (rare): **ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â , ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾; 3 pl = ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾**.  
                * In GurbÃƒÆ’Ã¢â‚¬Å¾Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â¡Ãƒâ€šÃ‚Â¹ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã¢â‚¬Å¾Ãƒâ€šÃ‚Â« the auxiliary is **often absorbed** into a longer verb with pronominal suffix: *ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â , ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â­ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã¢â‚¬Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â*. If you canÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â‚¬Å¾Ã‚Â¢t locate a free auxiliary, confirm tense via surface ending first.

                ---

                ## 4 ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â· Imperative & Subjunctive overlap  

                | Ending cluster | True Imperative ifÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¦ | Else ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ likely Subjunctive |
                |----------------|---------------------|---------------------------|
                | **ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ / ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹** | Stand-alone command/request | Used inside conditional/wish |
                | **ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ / ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ / ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡** | Vocative context | Hypothetical clause |

                ---

                ## 5 ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â· Passive voice heuristics  

                * **Surface template:** participle (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ) + auxiliary **ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ** etc.  
                * Only 3rd-person shows full paradigm in tables; 1st/2nd are scarce ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ flag if you tag 1st-person finite passive without strong textual evidence.  
                * Present passive often masquerades as adjective; ensure a *patient-as-subject* reading is plausible.

                ---

                ## 6 ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â· Causative sanity  

                * First-person causatives: **ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ / ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾, ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾**. No object ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ verb likely **inchoative**, not causative.  
                * 3rd-person causatives: **ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â , ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â§ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â , ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â , ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã¢â‚¬Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¯ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â **: must show agent-causes-other scenario.  
                * If semantic agent = performer, drop ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“causativeÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â tag.

                ---

                ## 7 ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â· Compound verbs  

                * Earlier element -> conjunct ending **-ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ / -ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ / -ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â  / -ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡**.  
                * Last element holds tense/person.  
                * Tag first as ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“Conjunct Verb / GerundÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â, second as finite.

                ---

                ## 8 ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â· Auto-highlight (red flags)  

                | Pattern | Likely mis-label |
                |---------|------------------|
                | Ending **-ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡** but tag ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°Ãƒâ€šÃ‚Â  Future | Wrong tense |
                | Ending **-ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â/-ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡** tagged 1st/3rd person | Imperative bleed |
                | Ending **-ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡** with no **ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â /ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨** & tag = Present/Future | Participle, not finite |
                | Two consecutive finite-verb tags inside one clause | Probably compound verb ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ split roles |
                | Passive participle **ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â** but subjectÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Âagent reading given | Reverse voice |
                | Finite verb tagged Masc/Fem | Finite forms should be Trans ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ likely mis-tag |
                | Participial ending gender ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°Ãƒâ€šÃ‚Â  controller noun/pronoun | Agreement error (ergative or intransitive mix-up) |
                | Ending-tense combo not found in Canonical table | Illegal combination ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ override gloss |
                | Finite verb with Gender = NA | Should be Trans ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ fix label |

                ---

                <sub>Heuristics sourced from pages 5.1 ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ 5.12: Present, Past, Future, Imperative, Subjunctive, Participles, Compound, Passive, Causative, Auxiliary, Pron-suffix sections.</sub>\
                """).strip() + "\n\n"

                ending_cheat_sheet = textwrap.dedent("""\
                ÃƒÆ’Ã‚Â°Ãƒâ€¦Ã‚Â¸ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÂ¢Ã¢â€šÂ¬Ã‚Â **Authoritative workflow**

                1ÃƒÆ’Ã‚Â¯Ãƒâ€šÃ‚Â¸Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â¢Ãƒâ€ Ã¢â‚¬â„¢Ãƒâ€šÃ‚Â£ **Check legality** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ If a surface ending ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â person/number ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â tense combo is **absent** from the
                Canonical table below, reject or relabel.

                2ÃƒÆ’Ã‚Â¯Ãƒâ€šÃ‚Â¸Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â¢Ãƒâ€ Ã¢â‚¬â„¢Ãƒâ€šÃ‚Â£ **Decide meaning** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ Among the *legal* options, pick the tag that is **best supported by
                the Darpan Translation and Darpan Meanings** (Prof. SÃƒÆ’Ã¢â‚¬Å¾Ãƒâ€šÃ‚Âhib SiÃƒÆ’Ã‚Â¡Ãƒâ€šÃ‚Â¹ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦gh).  
                *Those glosses remain the primary key to tense, mood, voice, and agent/object choice.*

                3ÃƒÆ’Ã‚Â¯Ãƒâ€šÃ‚Â¸Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â¢Ãƒâ€ Ã¢â‚¬â„¢Ãƒâ€šÃ‚Â£ Apply common-sense sanity rules (ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â§ 1ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ8) for edge-case flags.

                ---

                **VERB / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â  ENDINGS ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ QUICK REFERENCE (GurbÃƒÆ’Ã¢â‚¬Å¾Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â¡Ãƒâ€šÃ‚Â¹ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã¢â‚¬Å¾Ãƒâ€šÃ‚Â« corpus, Sheet 1)**  

                ÃƒÆ’Ã‚Â°Ãƒâ€¦Ã‚Â¸ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒâ€šÃ‚Â¹ **Agreement grid (what can legally combine)**  
                ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¢ **Person / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Å“** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ 1st (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â±ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®) | 2nd (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â§ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®) | 3rd (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¯)  
                ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¢ **Number / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ Singular / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ | Plural / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â  
                ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¢ **Tense / Mood** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ Present / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ | Past / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â­ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ | Future / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â­ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â±ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ | Causative / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¥ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ | Pronominal suffix  
                <sub>*Finite verbs ignore noun-gender; ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾/ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬/ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ are participial*</sub>

                ---

                ### A ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â· Canonical ending patterns (+ three toy forms on **ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Âµ-**)

                | Person ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â· Number | Tense / Mood | Surface endings | Micro-examples |
                |-----------------|--------------|-----------------|---------------|
                | **1st Sg** | Present | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã¢â‚¬Â /ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â /ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã¢â‚¬Â , ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â° |
                |  | Past | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ |
                |  | Future | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â /ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â  ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â° | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â , ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ |
                |  | Causative | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã¢â‚¬Â /ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã¢â‚¬Â , ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ |
                |  | Pronominal | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â |
                | **1st Pl** | Present | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ |
                |  | Past | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ |
                |  | Future | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ |

                | Person ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â· Number | Tense / Mood | Surface endings | Micro-examples |
                |-----------------|--------------|-----------------|---------------|
                | **2nd Sg** | Present | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â /ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â , ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ |
                |  | Past | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â |
                |  | Future | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ |
                |  | Causative | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ |
                |  | Pronominal | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã¢â‚¬Â /ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã¢â‚¬Â  |
                | **2nd Pl** | Present | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¦Ã¢â‚¬â„¢/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã¢â‚¬Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¦Ã¢â‚¬â„¢ | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â° |
                |  | Past | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ |
                |  | Future | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â |

                | Person ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â· Number | Tense / Mood | Surface endings | Micro-examples |
                |-----------------|--------------|-----------------|---------------|
                | **3rd Sg** | Present | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã¢â‚¬Â /ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â /ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â  | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã¢â‚¬Â , ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ |
                |  | Past | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ |
                |  | Future | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã¢â‚¬Â /ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â /ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ |
                |  | Causative | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã¢â‚¬Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â /ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â  | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ |
                |  | Pronominal | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â |
                | **3rd Pl** | Present | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â /ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ |
                |  | Past | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ |
                |  | Future | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ |
                |  | Causative | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ |

                ---

                ### B ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â· How to use the dashboard  

                1. **Validate annotations** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ If you tag a form ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“2nd Pl FutureÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â but it ends in **ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾**, the table shows that combo never occurs ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ revisit the tag.  
                2. **Debug machine predictions** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ Surface ending not found under predicted role ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ flag for review.  
                3. **Handle sandhi** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ Remember silent ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â° can drop before postpositions (e.g. **ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°**).  

                _Export or further slicing on request._\
                """).strip() + "\n\n"

            elif entry["Type"] == "Adverb / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â  ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¶ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£":
                implicit_note = textwrap.dedent("""\
                ### ÃƒÆ’Ã‚Â°Ãƒâ€¦Ã‚Â¸ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒâ€šÃ‚Â¹ `implicit_note` ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ ADVERB / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â  ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¼ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¼ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£  
                *(SGGS-centric discovery guide)*  

                **Essence**ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡Teach the evaluator to recognise words that **modify the *action itself***ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‚Ânever the doer (noun) nor the qualityÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Âword (adjective).  

                **Vision**ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡Lean on *Prof. SÃƒÆ’Ã¢â‚¬Å¾Ãƒâ€šÃ‚Âhib SiÃƒÆ’Ã‚Â¡Ãƒâ€šÃ‚Â¹ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ghÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â‚¬Å¾Ã‚Â¢s* Darpan gloss to infer *how, when, where* the verb happensÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‚Âeven when SGGS omits explicit post-positions or auxiliaries.  

                ---

                ## 1 ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â· Adverb ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°Ãƒâ€šÃ‚Â  Adjective ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°Ãƒâ€šÃ‚Â  Noun ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â the litmus test ÃƒÆ’Ã‚Â°Ãƒâ€¦Ã‚Â¸Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Âº  

                | Ask this first | Pass ÃƒÆ’Ã‚Â¢Ãƒâ€¦Ã¢â‚¬Å“ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÆ’Ã‚Â¯Ãƒâ€šÃ‚Â¸Ãƒâ€šÃ‚Â ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ Adverb | Fail ÃƒÆ’Ã‚Â¢Ãƒâ€¦Ã¢â‚¬Å“ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Å“ÃƒÆ’Ã‚Â¯Ãƒâ€šÃ‚Â¸Ãƒâ€šÃ‚Â ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ something else |
                |----------------|------------------|--------------------------|
                | **Does the word alter the meaning of the verb?** <br>(time, place, manner, measureÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¦) | ÃƒÆ’Ã‚Â¢Ãƒâ€¦Ã¢â‚¬Å“ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÆ’Ã‚Â¯Ãƒâ€šÃ‚Â¸Ãƒâ€šÃ‚Â modifies *action* ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ keep testing | ÃƒÆ’Ã‚Â¢Ãƒâ€¦Ã¢â‚¬Å“ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Å“ÃƒÆ’Ã‚Â¯Ãƒâ€šÃ‚Â¸Ãƒâ€šÃ‚Â modifies noun ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ likely *Adjective* or *Noun* |
                | **Will the clause stay grammatical if the word is removed?** | ÃƒÆ’Ã‚Â¢Ãƒâ€¦Ã¢â‚¬Å“ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÆ’Ã‚Â¯Ãƒâ€šÃ‚Â¸Ãƒâ€šÃ‚Â sentence remains; nuance lost | ÃƒÆ’Ã‚Â¢Ãƒâ€¦Ã¢â‚¬Å“ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Å“ÃƒÆ’Ã‚Â¯Ãƒâ€šÃ‚Â¸Ãƒâ€šÃ‚Â structure breaks ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ maybe pronoun/helper |
                | **Can the word move freely in the clause?** | ÃƒÆ’Ã‚Â¢Ãƒâ€¦Ã¢â‚¬Å“ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÆ’Ã‚Â¯Ãƒâ€šÃ‚Â¸Ãƒâ€šÃ‚Â adverbs float (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â´ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â **ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿** ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ) | ÃƒÆ’Ã‚Â¢Ãƒâ€¦Ã¢â‚¬Å“ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Å“ÃƒÆ’Ã‚Â¯Ãƒâ€šÃ‚Â¸Ãƒâ€šÃ‚Â fixed next to noun ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ adjective/compound |
                | **Any number/gender inflection visible?** | ÃƒÆ’Ã‚Â¢Ãƒâ€¦Ã¢â‚¬Å“ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÆ’Ã‚Â¯Ãƒâ€šÃ‚Â¸Ãƒâ€šÃ‚Â none (adverbs are **indeclinable**) | ÃƒÆ’Ã‚Â¢Ãƒâ€¦Ã¢â‚¬Å“ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Å“ÃƒÆ’Ã‚Â¯Ãƒâ€šÃ‚Â¸Ãƒâ€šÃ‚Â ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â /ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã¢â‚¬Â /ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â etc. ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ participle/adjective |
                | **Darpan gloss clue** says: ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“now, then, quickly, here, twiceÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â | ÃƒÆ’Ã‚Â¢Ãƒâ€¦Ã¢â‚¬Å“ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÆ’Ã‚Â¯Ãƒâ€šÃ‚Â¸Ãƒâ€šÃ‚Â adopt adverb label | ÃƒÆ’Ã‚Â¢Ãƒâ€¦Ã¢â‚¬Å“ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Å“ÃƒÆ’Ã‚Â¯Ãƒâ€šÃ‚Â¸Ãƒâ€šÃ‚Â gloss uses ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“of, to, withÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ case marker |

                > **Rule:** In this framework an adverb may *expand* a phrase (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ **ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â­ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â **), but it still targets the action, **not** the noun.  

                ---

                ## 2 ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â· Functional buckets ÃƒÆ’Ã‚Â°Ãƒâ€¦Ã‚Â¸ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã‚Â¯Ãƒâ€šÃ‚Â¸Ãƒâ€šÃ‚Â  

                | Category (Punjabi) | Core semantic cue | Minimal examples* |
                |--------------------|-------------------|-------------------|
                | **ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ / Time**        | ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¹Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡? ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡?ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â‚¬Å¾Ã‚Â¢ | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ |
                | **ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¥ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ / Place**       | ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¹Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â±ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¥ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡?ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â‚¬Å¾Ã‚Â¢            | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â , ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â , ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ |
                | **ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â§ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ / Manner**     | ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¹Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡? ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²?ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â‚¬Å¾Ã‚Â¢ | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ |
                | **ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£ / Measure**   | ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¹Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾?ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â‚¬Å¾Ã‚Â¢            | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â­ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ |
                | **ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â  / Number**    | ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¹Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°?ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â‚¬Å¾Ã‚Â¢        | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â° ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â«ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â«ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ |
                | **ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â  / Decision**   | certainty / denial  | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â° |
                | **ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£ / Reason**     | causation           | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¯ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¥ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ |
                | **ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ / Stress**    | emphasis            | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â­ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ |

                * A full ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“high-freqÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â tableÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‚Âincluding **phrase, compound & iterative** idiomsÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‚Âfollows in *common_sense_note*.

                ---

                ## 3 ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â· Zero-inflection principle ÃƒÆ’Ã‚Â°Ãƒâ€¦Ã‚Â¸Ãƒâ€¦Ã‚Â¡Ãƒâ€šÃ‚Â«ÃƒÆ’Ã‚Â°Ãƒâ€¦Ã‚Â¸Ãƒâ€šÃ‚Â§Ãƒâ€šÃ‚Â¬  

                * Adverbs **never** show number (-ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â/-ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°), gender, person or case.  
                * If a token **does** decline, re-classify: participial verb (*-ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾/-ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬/-ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡*), adjective, or oblique noun.  

                ---

                ## 4 ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â· Typical gloss helpers ÃƒÆ’Ã‚Â°Ãƒâ€¦Ã‚Â¸ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒâ€šÃ‚Â  

                | Gloss clue | Likely adverb class | Illustration |
                |------------|--------------------|--------------|
                | ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“**now / today / always**ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â | Time | ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â |
                | ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“**here / everywhere / within**ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â | Place | ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â |
                | ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“**thus / quickly / secretly**ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â | Manner | ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â° ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â |
                | ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“**fully / a little**ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â | Measure | ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â­ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â |
                | ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“**again / twice**ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â | Number | ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â«ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â«ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â |

                ---

                ## 5 ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â· Quick detection workflow ÃƒÆ’Ã‚Â¢Ãƒâ€¦Ã‚Â¡Ãƒâ€šÃ‚Â¡  

                1. **Mark all gloss adverbials** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ scan Darpan for English adverbs.  
                2. **Map to Punjabi surface form** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ locate the SGGS token(s) that carry that nuance.  
                3. **Apply indeclinability test** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ no visible suffix change? keep as adverb.  
                4. **Check floating mobility** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ move token; if syntax survives, adverb confirmed.  
                5. **Edge alert** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ if token sits after a post-position (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¦), probably **oblique noun** not adverb.

                ---

                ## 6 ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â· Red-flag heuristics ÃƒÆ’Ã‚Â°Ãƒâ€¦Ã‚Â¸Ãƒâ€¦Ã‚Â¡Ãƒâ€šÃ‚Â©  

                * Word tagged *Adverb* but ends in **-ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾/-ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬/-ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ likely participial.  
                * Tagged *Adverb* but gloss shows possession (*of*) ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ test for Genitive noun.  
                * Compound form **ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿** mis-tagged as Time/Manner interchangeably ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ ensure Darpan intent.  
                * Form appears **twice with different endings** in same ÃƒÆ’Ã‚Â¡Ãƒâ€šÃ‚Â¹Ãƒâ€šÃ‚Â­uk ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ must be *declinable* ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ not adverb.  

                ---

                ### ÃƒÆ’Ã‚Â°Ãƒâ€¦Ã‚Â¸ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒâ€šÃ‚Â Footnote on spreadsheet codes  
                The Excel ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“AdverbsÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â sheet groups every token into **eight functional sets** above, plus **Compound / Phrase** and **Iterative** markers. These codes are referenced only for *high-freq tables* and require **no inflection logic**.

                _Use this guide, then apply the sanity layer in `common_sense_note` for mis-tag traps._
                """).strip() + "\n\n"
            
                common_sense_note = textwrap.dedent("""\
                ### ÃƒÆ’Ã‚Â°Ãƒâ€¦Ã‚Â¸ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒâ€šÃ‚Â¹ `common_sense_note` ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ ADVERBS / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â  ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¼ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¼ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£ (semantic sanity layer)

                **Essence**ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€ Ã¢â‚¬â„¢A quick triage: *Does this token truly act as an **adverb**ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‚Âi.e., modifies a verb (or a whole clause) and NEVER a noun/pronoun?*

                **Vision**ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€ Ã¢â‚¬â„¢Prevent false-positives caused by:
                * Post-positions or emphatic particles masquerading as adverbs  
                * Adjectival or nominal words that look ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“adverb-ishÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â but show agreement or case

                ---

                ## 1 ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â· Three-step sanity check ÃƒÆ’Ã‚Â°Ãƒâ€¦Ã‚Â¸Ãƒâ€šÃ‚Â§Ãƒâ€šÃ‚Âª  

                | Step | Ask yourself | Abort / Relabel ifÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¦ |
                |------|--------------|--------------------|
                | ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‹Å“Ãƒâ€šÃ‚Â  | **Function** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ Does the word modify a **verb or clause** (manner, time, place, degree)? | It directly qualifies a noun/pronoun ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ likely Adjective or Noun |
                | ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‹Å“Ãƒâ€šÃ‚Â¡ | **Morphology** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ No number / gender / person agreement & no case endings | You see ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â/ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â° etc. agreeing with noun ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ itÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â‚¬Å¾Ã‚Â¢s NOT an adverb |
                | ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‹Å“Ãƒâ€šÃ‚Â¢ | **Position / Helpers** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ Is it followed by a postposition (*ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²*)? | Token + post-position ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ treat token as **Noun in oblique**, PP = post-position |

                ---

                ## 2 ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â· Category reference with high-frequency SGGS tokens ÃƒÆ’Ã‚Â°Ãƒâ€¦Ã‚Â¸ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒâ€šÃ‚Â  

                | Category | Typical surface cues | SGGS high-freq examples |
                |----------|----------------------|-------------------------|
                | **Time / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡** | ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“when?ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â, duration, sequence | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â  |
                | **Place / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¥ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡** | ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“where?ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â, location, direction | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â , ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â , ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ |
                | **Manner / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â§ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬** | ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“how?ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â, style, attitude | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Âµ, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â |
                | **Measurement / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£** | quantity / degree | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â­ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â |
                | **Number / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ** | frequency / repetition | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â«ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â«ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â° ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Å“ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Å“, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â° |
                | **Decision / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â ** | negation / affirmation | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ |
                | **Reason / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£** | cause / purpose | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¯ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ |
                | **Stress / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦** | emphasis / focus | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â­ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â , ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ |
                
                ---

                ### ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Å“Ãƒâ€šÃ‚Â¸ Phrase / Compound & Iterative idioms (extended reference)

                | Sub-group | Token set ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ **all indeclinable adverbs** | Main category |
                |-----------|------------------------------------------|---------------|
                | **Time ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â Phrase** | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂºÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â  ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â  ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ | Time / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ |
                | **Place ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â Phrase** | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â­ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂºÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã¢â‚¬Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â§ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â° | Place / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¥ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ |
                | **Manner ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â Phrase** | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â­ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â , ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â­ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â° ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¥ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â§ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â  ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ | Manner / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â§ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ |
                | **Iterative (Time)** | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â«ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â«ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Å“ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Å“, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â° ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ | Time / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ |
                | **Iterative (Place)** | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ | Place / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¥ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ |
                | **Iterative (Manner)** | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â² ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â², ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â° ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ | Manner / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â§ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ |

                *(Duplicates collapsed; diacritics kept as in SGGS.)*

                ---

                ## 3 ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â· Red-flag heuristics ÃƒÆ’Ã‚Â°Ãƒâ€¦Ã‚Â¸Ãƒâ€¦Ã‚Â¡Ãƒâ€šÃ‚Â¨  

                | Pattern | Likely mis-tag |
                |---------|---------------|
                | Token shows **plural/oblique ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ / ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â / ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°** agreement | Probably a noun or adjective |
                | Token immediately followed by post-position (**ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â², ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡**) | Treat as noun + PP |
                | Token doubles as **auxiliary verb** (*ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â *) in context | Re-evaluate as Stress adverb OR auxiliary |
                | Same stem appears with changing endings inside verse | Likely **declinable adjective**, not adverb |
                | Gloss marks token as **object / subject** | Not an adverb |

                ---

                ## 4 ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â· Usage tips ÃƒÆ’Ã‚Â°Ãƒâ€¦Ã‚Â¸ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢Ãƒâ€šÃ‚Â¡  

                1. **No gender/number tags** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ Always set **Gender = NA** & **Number = NA** for adverbs.  
                2. **POS override wins** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ If sanity check fails, switch POS before finishing the task.  
                3. Quote at least one verb the adverb is modifying when you justify your choice.

                ---

                <sub>Source pages: Grammar book ch. 6 (pp. 6.1ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ6.2.6) & ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“AdverbsÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â sheet from 0.2 For Data to GPT.xlsx.</sub>\
                """).strip() + "\n\n"

                ending_cheat_sheet = (
                    "**ADVERBS:** Indeclinable in SGGS ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ no ending table required."
                )

            elif entry["Type"] == "Postposition / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â§ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢":
                implicit_note = textwrap.dedent("""\
                    **POSTPOSITIONS IN GURBÃƒÆ’Ã¢â‚¬Å¾ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¡Ãƒâ€šÃ‚Â¹ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã¢â‚¬Å¾Ãƒâ€šÃ‚Âª ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ SEEING THE HIDDEN LINKS**  

                    A postposition (_ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â§ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢_) expresses the *relationship* of a noun or pronoun to the
                    rest of the clause.  Think of it as a Punjabi sibling of the English preposition,
                    except it normally **follows** the word it governs.

                    ### 1 ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â· Why they matter in annotation  
                    ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¢ **Old case-endings ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ new helpers** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ Classical Punjabi often fused case endings
                    straight onto the noun (e.g. ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â , ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°).  Over centuries these endings began to act
                    like separate postpositionsÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‚Âand GurbÃƒÆ’Ã¢â‚¬Å¾Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â¡Ãƒâ€šÃ‚Â¹ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã¢â‚¬Å¾Ãƒâ€šÃ‚Â« preserves *both* layers.  
                    ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¢ **One helper ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°Ãƒâ€šÃ‚Â  one case** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ DonÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â‚¬Å¾Ã‚Â¢t map ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“each postposition to one caseÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â by reflex.
                    Many helpers (esp. ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¹Ã…â€œofÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â‚¬Å¾Ã‚Â¢, ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¹Ã…â€œfromÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â‚¬Å¾Ã‚Â¢, ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¹Ã…â€œwithÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â‚¬Å¾Ã‚Â¢) sit across **multiple traditional cases**.  
                    ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¢ **Pre-noun surprise** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ Forms such as **ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â ** can surface *before* the noun when
                    they co-occur with another postposition; still tag them as postpositions.

                    ### 2 ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â· How to read the Darpan gloss  
                    1. **Scan the English helper** inserted by Prof. SÃƒÆ’Ã¢â‚¬Å¾Ãƒâ€šÃ‚Âhib SiÃƒÆ’Ã‚Â¡Ãƒâ€šÃ‚Â¹ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦gh ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ _to, of, from,
                    with, without, in, on, before, after, near, farÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¦_  
                    2. **Locate the Punjabi token(s)** that deliver that meaning in the pÃƒÆ’Ã¢â‚¬Å¾Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â¡Ãƒâ€šÃ‚Â¹ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ktÃƒÆ’Ã¢â‚¬Å¾Ãƒâ€šÃ‚Â«.
                    They may be:  
                    ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¢ an **attached ending** (*ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â  ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤*),  
                    ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¢ a **stand-alone word** (*ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â², ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿*), or  
                    ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¢ an **archaic variant** (e.g. _ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡_).  
                    3. **Check the noun form** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ the governed noun should be in the **oblique** (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â§ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢)
                    if the language still marks one; otherwise, rely on meaning.

                    > **Rule of thumb** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ If the gloss supplies a relational word the verse omits,
                    > treat that English word as a flag that ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“a postposition is hiding here.ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â\
                    """).strip() + "\\n\\n"

                common_sense_note = textwrap.dedent("""\
                    **SEMANTIC SANITY CHECK ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ IS THIS *REALLY* A POSTPOSITION?**  

                    ### ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‹Å“Ãƒâ€šÃ‚Â   Function test  
                    ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¢ Does the candidate **link** its noun/pronoun to the verb or another noun?  
                    _Yes_ ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ proceed.  _No_ ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ it may be an **adverb**, **case-suffix**, or even
                    part of a **compound noun**.

                    ### ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‹Å“Ãƒâ€šÃ‚Â¡  Morphology test  
                    ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¢ Postpositions are **indeclinable** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ no gender/number/person endings of their
                    own.  If the token shows ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â /ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã¢â‚¬Â /ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â etc., suspect an *oblique noun* instead.  
                    ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¢ Possessive markers **ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬** *look* like adjectives but behave as
                    postpositions.  Tag them here only when they attach to another noun
                    (ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â® **ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾** ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â).  

                    ### ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‹Å“Ãƒâ€šÃ‚Â¢  Dependency test  
                    ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¢ A true postposition normally keeps a **dependent noun** close by.  If none
                    appears, ask whether the word is actually an **adverbial particle** (ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿,
                    ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â) or part of a **verb phrase**.

                    ### ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‹Å“Ãƒâ€šÃ‚Â£  Red-flag heuristics ÃƒÆ’Ã‚Â°Ãƒâ€¦Ã‚Â¸Ãƒâ€¦Ã‚Â¡Ãƒâ€šÃ‚Â©  
                    | Pattern | Likely mis-tag | Example cue |
                    |---------|---------------|-------------|
                    | Token plus **another postposition** with no noun in between | Missing oblique noun | ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â  **ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²**ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â |
                    | Token followed by *ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â /ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨* | Probably predicate adjective | ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ **ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿**ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â |
                    | Token appears twice with changing endings | Declining noun, not postposition | ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â |

                    ### ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‹Å“Ãƒâ€šÃ‚Â¤  Quick role alignment  
                    | Semantic role | Common helpers (non-exhaustive) |
                    |---------------|----------------------------------|
                    | **Genitive / OF** | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ |
                    | **Dative / TO, FOR** | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â , ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã¢â‚¬Â  |
                    | **Ablative / FROM** | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â° |
                    | **Instrumental / WITH** | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â², ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¥, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ |
                    | **Locative / IN, ON, AT** | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ |
                    | **Orientational / BEFORE, AFTER, NEAR, FAR** | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â , ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂºÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â , ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â², ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¸, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ |

                    _If a helper can sit in more than one row, choose the case that best matches the
                    **meaning of the clause**, and note the alternative in comments._\
                    """).strip() + "\\n\\n"
                
                ending_cheat_sheet = textwrap.dedent("""\
                    **POSTPOSITION QUICK-REFERENCE ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ SURFACE FORMS BY SEMANTIC GROUP**  

                    | Role (Eng.) | Core Punjabi forms* | Notes |
                    |-------------|---------------------|-------|
                    | **OF / Possessive** | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â· ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â· ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â , ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â° ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â· ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â  | Masculine/Feminine variants; decline with possessed noun, not with owner |
                    | **TO / FOR** | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â , ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â· ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â° ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â· ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã¢â‚¬Â  | Older endings (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¦) often fuse; **ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°** modern |
                    | **FROM / OUT OF** | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ | Ablative / separative sense; *ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾* also ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“withoutÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â |
                    | **WITH / BY / ALONG** | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â², ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¥, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ | Instrumental & associative; choice shaped by metre |
                    | **WITHOUT / THAN** | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â , ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¥ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ | Negative / comparative nuance |
                    | **IN / INSIDE / WITHIN** | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â¢Ãƒâ€šÃ‚Â¸Ãƒâ€šÃ‚Â±ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â  | Locative & internal |
                    | **ON / OVER / ABOVE** | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ | Spatial elevation; *ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡* doubles as generic PP |
                    | **UNDER / BELOW** | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¥ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â , ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â , ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ | Lower level |
                    | **BEFORE / FRONT** | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â , ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ | Temporal or spatial precedence |
                    | **AFTER / BEHIND** | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂºÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â , ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂºÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â , ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂºÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ | Temporal or spatial following |
                    | **TOWARDS / NEAR / FAR** | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â², ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â², ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¸, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ | Directional & proximity |

                    <sub>*Forms collated from pp. 1-7 of your textbook; diacritics left as printed.
                    The list is not exhaustiveÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‚Âadd dialectal or Braj variants as you meet them.</sub>

                    **Oblique rule** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ The governed noun normally appears in the **oblique**; the
                    postposition itself **never inflects**.

                    **Pre-noun exception** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ When **ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â ** precedes another PP, it may surface *before*
                    its noun (e.g. ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â® **ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â ** ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¥ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â) ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ still tag as postposition.

                    **Cross-case cautions**  
                    ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¢ Some helpers (esp. ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“withÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â, ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“inÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â, ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“fromÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â) can realise **Instrumental, Locative,
                    or Ablative** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ decide by semantics.  
                    ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¢ Genitive set **ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬** functions like an adjective in modern speech but
                    grammatically remains a postposition in SGGS.

                    _Use this sheet to *reject impossible guesses* and to **confirm legal surface
                    forms** before finalising your annotation._\
                    """).strip() + "\\n\\n"

            elif entry["Type"] == "Conjunction / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¯ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢":
                implicit_note = textwrap.dedent("""\
                    **CONJUNCTIONS IN GURBÃƒÆ’Ã¢â‚¬Å¾ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¡Ãƒâ€šÃ‚Â¹ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã¢â‚¬Å¾Ãƒâ€šÃ‚Âª ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ HOW TO HEAR THE HINGES**

                    A conjunction (_ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¯ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢_) links words, phrases, or entire clausesÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â*and, but, or,
                    if ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¦ then, even thoughÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¦. *  GurbÃƒÆ’Ã¢â‚¬Å¾Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â¡Ãƒâ€šÃ‚Â¹ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã¢â‚¬Å¾Ãƒâ€šÃ‚Â« uses a small core set, but the
                    multilingual texture of the text supplies many **variants** (ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â , ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°,
                    ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â«ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿; ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°; ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â­).

                    #### 1 ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â· Spotting them in the verse
                    1. **Look for clause boundaries** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ commas or the metrical ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“||ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â often signal the
                    join.  
                    2. **Map the gloss cue** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ Prof. SÃƒÆ’Ã¢â‚¬Å¾Ãƒâ€šÃ‚Âhib SiÃƒÆ’Ã‚Â¡Ãƒâ€šÃ‚Â¹ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦gh frequently inserts
                    *and / but / or / if / then / even*, etc.  Trace that helper back to a Punjabi
                    token (sometimes a tiny vowel like **ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡**).  
                    3. **Check the flow** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ removing a true conjunction should split the sentence
                    into two meaningful parts; if the sense collapses, the token may be an
                    **adverb** (*ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¦Ã¢â‚¬â„¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ = then* vs. *ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ = from*), **post-position**, or **particle**.

                    > **Rule of thumb** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ If the gloss supplies an English linker and the Punjabi
                    > token neither declines nor carries case, youÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â‚¬Å¾Ã‚Â¢ve found a conjunction.
                    """).strip() + "\\n\\n"
                
                common_sense_note = textwrap.dedent("""\
                    **SEMANTIC SANITY CHECK ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ DOES THIS REALLY JOIN THINGS?**

                    | Quick test | Keep as conjunction ÃƒÆ’Ã‚Â¢Ãƒâ€¦Ã¢â‚¬Å“ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÆ’Ã‚Â¯Ãƒâ€šÃ‚Â¸Ãƒâ€¦Ã‚Â½ | Rethink ÃƒÆ’Ã‚Â¢Ãƒâ€¦Ã¢â‚¬Å“Ãƒâ€¹Ã…â€œ |
                    |------------|------------------------|-----------|
                    | **Function** | Links two clauses / words of equal status | Adds a helper to a noun (*post-position*) |
                    | **Morphology** | Indeclinable; no gender/number | Ends -ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â /-ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã¢â‚¬Â /-ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ likely adjective/noun |
                    | **Mobility** | Can often move to clause edge without breaking grammar | Locked to noun it follows ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ PP/adjective |
                    | **Gloss cue** | gloss shows *and, but, or, if ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¦ then* | gloss shows *to, of, from* ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ case helper |

                    #### Red-flag patterns ÃƒÆ’Ã‚Â°Ãƒâ€¦Ã‚Â¸Ãƒâ€¦Ã‚Â¡Ãƒâ€šÃ‚Â©
                    * Token plus **post-position** (e.g. *ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹*): maybe *ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡* = ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“ifÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â (OK) but *ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹* =
                    Dative ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ label both separately.  
                    * **ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾** or **ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ might be emphatic repetition, not conjunction.  
                    * **ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡**: confirm rÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â´leÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â*ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾* = ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“thenÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â, *ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡* often Locative PP, *ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡* Ablative.
                    """).strip() + "\\n\\n"
                
                ending_cheat_sheet = textwrap.dedent("""\
                    **CONJUNCTION QUICK-REFERENCE ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ HIGH-FREQ FORMS IN SGGS**

                    | Logical role | Punjabi forms* | Example gloss cue |
                    |--------------|---------------|-------------------|
                    | **AND / THEN** | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â«ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ | ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“andÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â, ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“thenÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â, ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“alsoÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â |
                    | **OR** | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â , ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ | ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“or / whetherÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â |
                    | **BUT / HOWEVER** | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¸, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â«ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ | ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“butÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â, ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“yetÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â |
                    | **IF** | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ | ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“if / provided thatÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â |
                    | **IF ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¦ THEN** | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¦ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡/ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ | paired correlative |
                    | **EVEN IF / EVEN THEN** | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â­ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â° ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â­ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ | concessive |
                    | **NEITHER ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¦ NOR** | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¦ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ | correlative negative |
                    | **OTHERWISE** | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ | ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“otherwiseÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â |
                    | **THEREFORE / HENCE** | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ | result / inference |
                    | **AS / LIKE** | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ | comparative |
                    | **LEST** | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â | preventative |

                    <sub>*Forms taken from textbook pp. 8.1 ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ 8.4; diacritics preserved.</sub>

                    **Key reminders**

                    * **Indeclinable** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ conjunctions never carry case or agreement.
                    * **Dual tokens** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ Some forms (*ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡*) double as post-positions.
                    Decide by context: if it *links* clauses ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ conjunction; if it *marks* a noun
                    ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ post-position.
                    * **Correlative pairs** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ Tag both halves (e.g. **ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¦ **ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡**) as one
                    logical conjunction with a note ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“correlativeÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â.
                    """).strip() + "\\n\\n"
                
            elif entry["Type"] == "Interjection / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢":
                implicit_note = textwrap.dedent("""\
                    **INTERJECTIONS IN GURBÃƒÆ’Ã¢â‚¬Å¾ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¡Ãƒâ€šÃ‚Â¹ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã¢â‚¬Å¾Ãƒâ€šÃ‚Âª ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ PURE, UNINFLECTED EMOTION**

                    An interjection (_ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢_) erupts outside normal grammar to voice **feeling**:
                    surprise, pain, devotion, blessing, aweÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¦  Because they sit *outside* the clause
                    structure, they **never govern case, never inflect, never agree**.

                    #### 1 ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â· What to notice in a verse
                    1. **Standalone or comma-bound** tokens ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ often at the start, end, or mid-clause,
                    separated by a breve pause.  E.g. **ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â**, **ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â  ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â **, **ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿**.
                    2. **Gloss cue** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ Prof. SÃƒÆ’Ã¢â‚¬Å¾Ãƒâ€šÃ‚Âhib SiÃƒÆ’Ã‚Â¡Ãƒâ€šÃ‚Â¹ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦gh usually inserts an English exclamation
                    (*O!, Alas!, Wow!, Blessed!*) or italicises the Punjabi for emphasis.
                    3. **No syntactic load** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ if you remove the interjection, the grammar of the
                    sentence remains intact (though colour is lost).

                    #### 2 ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â· Ten broad emotional classes in SGGS
                    1. **Vocative** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ calling or invoking (*ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â , ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¦*).  
                    2. **Repulsive** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ aversion or disgust (*ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â«ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â*).  
                    3. **Painful** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ sorrow, lament (*ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â  ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â *).  
                    4. **Submission** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¹Ã…â€œDivine willingÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â‚¬Å¾Ã‚Â¢ (*ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹*).  
                    5. **Wondrous** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ ecstatic awe (*ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â­ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬*).  
                    6. **Caution / Warning** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ prudent cry (*ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡* used admonishingly).  
                    7. **Blessing** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ goodwill (*ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â*).  
                    8. **Curse** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ condemnation (*ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°*).  
                    9. **Sacrificial** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ self-offering (*ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿*).  
                    10. **Reverence** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ respectful welcome (*ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂºÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬*).

                    > **Rule of thumb** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ if the word communicates *only* emotion and detaches
                    > cleanly from clause syntax, tag it as Interjection; otherwise test Adverb,
                    > Vocative Noun, or Particle.
                    """).strip() + "\\n\\n"

                common_sense_note = textwrap.dedent("""\
                    **SEMANTIC SANITY CHECK ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ IS THIS TOKEN *JUST* AN EMOTION?**

                    | Quick probe | Keep as Interjection ÃƒÆ’Ã‚Â¢Ãƒâ€¦Ã¢â‚¬Å“ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â | Rethink ÃƒÆ’Ã‚Â¢Ãƒâ€¦Ã¢â‚¬Å“ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Å“ |
                    |-------------|-----------------------|-----------|
                    | **Function** | Adds emotional colour, no syntactic role | Performs grammatical work (case, link, inflection) |
                    | **Inflection** | Completely indeclinable | Shows ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â  / ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã¢â‚¬Â  / ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â endings ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ maybe adjective/noun |
                    | **Dependence** | Can float; removal leaves clause intact | Sentence breaks ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ probably verb/particle |
                    | **Gloss cue** | Gloss marks ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“O!ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â, ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“Alas!ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â, ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“Blessed!ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â etc. | Gloss gives ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“to, from, withÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ post-position |

                    #### Red-flag patterns ÃƒÆ’Ã‚Â°Ãƒâ€¦Ã‚Â¸Ãƒâ€¦Ã‚Â¡Ãƒâ€šÃ‚Â©
                    * **ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â** appears as noun/adjective elsewhere ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ decide per context.  
                    * **ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â  ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â , ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â­ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã¢â‚¬Â ** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ first token vocative interjection, second token noun;
                    split tags, donÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â‚¬Å¾Ã‚Â¢t bundle.  
                    * Repeated **ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿** could be mantra (noun) *or* caution interjection ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ
                    weigh meaning.

                    _For every interjection, fill **Number = NA** and **Gender = NA**; they never
                    agree with anything._
                    """).strip() + "\\n\\n"
                
                ending_cheat_sheet = textwrap.dedent("""\
                    **INTERJECTION QUICK-REFERENCE ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ FREQUENT FORMS BY EMOTIONAL CLASS**

                    | Class               | High-frequency tokens* (SGGS spelling)        |
                    |---------------------|----------------------------------------------|
                    | **Vocative**        | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â , ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â , ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â , ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ |
                    | **Repulsive**       | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â«ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â                                   |
                    | **Painful**         | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â  ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â , ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹           |
                    | **Submission**      | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹                                          |
                    | **Wondrous**        | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â­ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â , ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â |
                    | **Caution / Warning** | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡                       |
                    | **Blessing**        | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â               |
                    | **Curse**           | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â                  |
                    | **Sacrificial**     | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾    |
                    | **Reverence**       | ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â° ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬, ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂºÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬                       |

                    <sub>*Tokens taken from textbook pp. 9.1ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ9.4; diacritics preserved.  
                    Feel free to trim or expand as corpus stats evolve.</sub>

                    **Remember** ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ Interjections are **indeclinable** and **carry no grammatical
                    features**.  Therefore the spreadsheet needs **no ending table** beyond this
                    categorical list.
                    """).strip() + "\\n\\n"
                
            notes_block = ending_cheat_sheet + implicit_note + common_sense_note

            prompt = textwrap.dedent(f"""
                **You are a Punjabi grammar expert.**

                Below are the *allowed choices* for each feature of the highlighted word:

                {opts_block}
                {matches_block}

                {notes_block}

                **IMPORTANT:**  
                Base **all** confirmations or corrections **solely on the Darpan translation** below.  
                Do **not** consult any other translation or external context.

                **My Current Selections:**  
                - Word Under Analysis: **{ve}**  
                - Number / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨: **{num}**  
                - Grammar Case / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¯ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£: **{gram}**  
                - Gender / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â: **{gen}**  
                - Word Root: **{root}**

                **Context (use *only* the Darpan gloss):**  
                ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¢ **Verse:** {verse}  
                ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¢ **Darpan Translation:** {trans}  
                ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¢ **Darpan-Meanings:** {dm}

                **Task:**  
                1. **Confirm or correct** each featureÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‚Âif blank, **choose** the best option  
                (one-sentence rationale citing the inflection or usage).
                ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¢ For finite forms, choose **1st / 2nd / 3rd Person** in Word-Root (do not use Past/Perfect there). 
                2. **Corrections**, if any:  
                - Number ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¦  
                - Grammar Case ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¦  
                - Word Root ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¦  
                3. **Example Usage:**  
                Provide **one** new GurbÃƒÆ’Ã¢â‚¬Å¾Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â¡Ãƒâ€šÃ‚Â¹ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã¢â‚¬Å¾Ãƒâ€šÃ‚Â«-style sentence using **ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“{ve}ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â** with the
                confirmed ending, number, case, gender, and root.
                4. **Table citation:**  
                Quote the person ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â number ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â tense row header you matched in the Canonical table  
                (e.g., ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“1 Sg | PastÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â). **Use that rowÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â‚¬Å¾Ã‚Â¢s category name for ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“Grammar Case / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¯ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£,ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â unless a sanity rule forbids it.**
                5. **Ending ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¾ Case cross-check:**
                ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¢ If the cheat-sheet already lists a suffix for your chosen case, use it.  
                ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¢ If the case is **missing**, you may propose a likely form
                    (or say ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“uninflectedÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â) **but give one-line reasoning**.
                6. **Commentary:**  
                Please write 2ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ3 sentences as ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“ChatGPT Commentary:ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â explaining how you arrived at each feature choice.
            """).strip()

            self.root.clipboard_clear()
            self.root.clipboard_append(prompt)
            messagebox.showinfo(
                "Prompt Ready",
                "The detailed-grammar prompt has been copied to your clipboard.\n"
                "Paste it into ChatGPT, then paste its response back into the text box."
            )

        tk.Button(
            frm, text="ÃƒÆ’Ã‚Â°Ãƒâ€¦Ã‚Â¸ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ Build Detailed Grammar Prompt",
            font=("Arial", 12, "italic"),
            bg="white", fg="dark cyan",
            command=build_detailed_prompt
        ).grid(row=6, column=0, columnspan=2, pady=(10, 0))

        # 6) --------------  Bottom buttons (unchanged)  --------------
        sep = tk.Frame(win, bg="#cccccc", height=2)
        sep.pack(side=tk.BOTTOM, fill=tk.X, padx=20, pady=(5, 0))

        btns = tk.Frame(win, bg="light gray")
        btns.pack(side=tk.BOTTOM, fill=tk.X, padx=20, pady=(6, BOTTOM_PAD))

        tk.Button(
            btns, text="ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¹ Back",
            font=("Arial", 12), bg="gray", fg="white",
            command=lambda: [win.destroy(),
                            self.show_matches_grammar(self._last_matches, word, index)]
        ).pack(side=tk.LEFT)

        tk.Button(
            btns, text="Save & Finish ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢",
            font=("Arial", 12, "bold"), bg="dark cyan", fg="white",
            command=lambda: self.on_accept_detailed_grammar(win)
        ).pack(side=tk.RIGHT)

        win.transient(self.root)
        win.grab_set()
        self.root.wait_window(win)















    # Wrappers for Verse and ABW to avoid cross-calling
    def open_final_grammar_dropdown(self, word, pos, index):
        return self._open_final_grammar_dropdown_common(word, pos, index, mode='verse')

    def open_final_grammar_dropdown_for_word(self, word, pos, index):
        return self._open_final_grammar_dropdown_common(word, pos, index, mode='word')
    def on_accept_detailed_grammar(self, win):
        """Finalize the detailed grammar selection and append a row to 1.1.1_birha.csv.
        This is called by the 'Save & Finish' button in the detailed grammar dialog.
        """
        advance_ok = False
        try:
            # Pull selections from the dropdowns/commentary box
            ve    = (self.detailed_ve_var.get() if hasattr(self, 'detailed_ve_var') else "")
            num   = (self.detailed_number_var.get()  if hasattr(self, 'detailed_number_var')  else "")
            gram  = (self.detailed_grammar_var.get() if hasattr(self, 'detailed_grammar_var') else "")
            gen   = (self.detailed_gender_var.get()  if hasattr(self, 'detailed_gender_var')  else "")
            root  = (self.detailed_root_var.get()    if hasattr(self, 'detailed_root_var')    else "")
            comm  = ""
            try:
                if hasattr(self, 'detailed_commentary') and self.detailed_commentary is not None:
                    comm = self.detailed_commentary.get("1.0", tk.END).strip()
            except Exception:
                pass

            # Update the current detailed entry with finalized values
            entry = getattr(self, 'current_detailed_entry', {}) or {}
            if ve:
                entry["Vowel Ending"] = ve
            if num:
                entry["Number / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨"] = num
            if gram:
                entry["Grammar / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¯ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£"] = gram
            if gen:
                entry["Gender / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â"] = gen
            if root:
                entry["Word Root"] = root
            # Keep existing fields: Type, Evaluation, Reference Verse, Darpan Translation, Darpan Meaning
            # Store commentary in the UI key for internal use; CSV uses 'ChatGPT Commentry'
            entry["ChatGPT Commentary"] = comm

            # Append/Overwrite to CSV with composite key and confirm-overwrite UI
            saved, _row_no, _mode = self._append_birha_csv_row(
                entry,
                path="1.1.1_birha.csv",
                require_confirm=True,
                ui_parent=win
            )
            advance_ok = bool(saved)
        except Exception as e:
            try:
                messagebox.showwarning("Save Warning", f"Could not append CSV row: {e}")
            except Exception:
                pass
        finally:
            # Close and advance only if save succeeded or confirmed overwrite
            if advance_ok:
                try:
                    if win and win.winfo_exists():
                        win.destroy()
                except Exception:
                    pass
                try:
                    # advance in the selected-words queue if that flow is active
                    if hasattr(self, 'grammar_queue'):
                        self.current_queue_pos = getattr(self, 'current_queue_pos', 0) + 1
                        is_abw = (getattr(self, '_current_detailed_mode', 'verse') == 'word') or (getattr(self, '_current_grammar_mode', 'verse') == 'word')
                        if is_abw:
                            self.abw_process_next_word_assessment()
                        else:
                            self.process_next_word_assessment()
                except Exception:
                    pass

    def _append_birha_csv_row(self, entry: dict, path: str = "1.1.1_birha.csv", require_confirm: bool = False, ui_parent=None):
        """Append a UTF-8 row to 1.1.1_birha.csv, creating it with headers if missing.
        Expected headers: Vowel Ending, Number / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨, Grammar / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¯ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£, Gender / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â,
        Word Root, Type, Evaluation, Reference Verse, Darpan Translation, Darpan Meaning, ChatGPT Commentry
        """
        headers = [
            "\ufeffVowel Ending",  # keep BOM in first header for compatibility with existing readers
            "Number / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨",
            "Grammar / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¯ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£",
            "Gender / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â",
            "Word Root",
            "Type",
            "Evaluation",
            "Reference Verse",
            "Darpan Translation",
            "Darpan Meaning",
            "ChatGPT Commentry",   # note: CSV header uses 'Commentry'
        ]
        # Build a row mapping with safe defaults
        row = {h: "" for h in headers}
        # Map from internal keys (including 'ChatGPT Commentary') to CSV headers
        mapping = {
            "Vowel Ending": "Vowel Ending",
            "\ufeffVowel Ending": "Vowel Ending",
            "Number / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨": "Number / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨",
            "Grammar / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¯ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£": "Grammar / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¯ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£",
            "Gender / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â": "Gender / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â",
            "Word Root": "Word Root",
            "Type": "Type",
            "Evaluation": "Evaluation",
            "Reference Verse": "Reference Verse",
            "Darpan Translation": "Darpan Translation",
            "Darpan Meaning": "Darpan Meaning",
            # Internal UI key to CSV header
            "ChatGPT Commentary": "ChatGPT Commentry",
            "ChatGPT Commentry": "ChatGPT Commentry",
            "Word Index": "Word Index",
        }
        for k, v in (entry or {}).items():
            if k in mapping:
                row[mapping[k]] = v if v is not None else ""

        # Overwrite vs append based on composite key (Vowel Ending, Reference Verse, Word Index)
        key_vowel = row.get('Vowel Ending', row.get('\ufeffVowel Ending', ''))
        key_ref   = row.get('Reference Verse', '')
        key_widx  = str(row.get('Word Index', '') or '').strip()
        composite_key = (key_vowel, key_ref, key_widx)

        file_exists = os.path.exists(path)
        file_empty = False
        if file_exists:
            try:
                file_empty = os.path.getsize(path) == 0
            except Exception:
                file_empty = False

        existing_rows = []
        headers_from_file = None
        if file_exists and not file_empty:
            try:
                with open(path, 'r', encoding='utf-8-sig', newline='') as f:
                    reader = csv.DictReader(f)
                    headers_from_file = list(reader.fieldnames or [])
                    existing_rows = list(reader)
            except Exception:
                existing_rows = []

        # Normalize and preserve existing column order; ensure 'Word Index' exists at end
        def _norm_header_list(hdrs):
            out = []
            for h in (hdrs or []):
                s = str(h)
                if s.startswith('\ufeff'):
                    s = s[1:]
                out.append(s)
            return out

        if headers_from_file:
            headers = _norm_header_list(headers_from_file)
        else:
            headers = _norm_header_list(headers)
        if 'Word Index' not in headers:
            headers.append('Word Index')

        # Find a matching row by keys; treat missing/blank Word Index as wildcard for legacy rows
        match_idx = None
        for i, r in enumerate(existing_rows):
            ex_vowel = r.get('\ufeffVowel Ending', r.get('Vowel Ending', ''))
            ex_ref   = r.get('Reference Verse', '')
            ex_widx  = str(r.get('Word Index', '') or '').strip()
            if ex_vowel == key_vowel and ex_ref == key_ref:
                if (ex_widx == key_widx) or (ex_widx == '' or key_widx == ''):
                    match_idx = i
                    break

        if match_idx is not None:
            # Show confirm modal with diff if requested
            if require_confirm and ui_parent is not None:
                try:
                    if not self._confirm_overwrite_modal(ui_parent, headers, existing_rows[match_idx], row):
                        return False, None, None
                except Exception:
                    pass
            # Replace the existing row (store normalized keys so writer can map to BOM-safe headers)
            existing_rows[match_idx] = dict(row)
            with open(path, 'w', encoding='utf-8', newline='') as f:
                try:
                    f.write('\ufeff')
                except Exception:
                    pass
                writer = csv.DictWriter(f, fieldnames=headers)
                writer.writeheader()
                for r in existing_rows:
                    out = {h: '' for h in headers}
                    for k, v in (r or {}).items():
                        try:
                            nk = str(k)
                        except Exception:
                            nk = k
                        if isinstance(nk, str) and nk.startswith('\ufeff'):
                            nk = nk[1:]
                        if nk in out:
                            out[nk] = v
                    writer.writerow(out)
                try:
                    f.flush(); os.fsync(f.fileno())
                except Exception:
                    pass
                row_no = match_idx + 1
                try:
                    messagebox.showinfo("Saved", f"Saved to {path} (row #{row_no} overwritten)")
                except Exception:
                    pass
                print(f"[Save] overwrite row #{row_no}")
                return True, row_no, 'overwrite'
        else:
            # Create file with header if missing, then append
            if (not file_exists) or file_empty:
                with open(path, 'w', encoding='utf-8', newline='') as f:
                    try:
                        f.write('\ufeff')
                    except Exception:
                        pass
                    writer = csv.DictWriter(f, fieldnames=headers)
                    writer.writeheader()
                    try:
                        f.flush(); os.fsync(f.fileno())
                    except Exception:
                        pass
            else:
                # If the existing CSV lacks 'Word Index' in its header, upgrade the file by rewriting
                try:
                    existing_headers_norm = _norm_header_list(headers_from_file) if headers_from_file else []
                except Exception:
                    existing_headers_norm = list(headers_from_file or [])
                if headers_from_file and 'Word Index' not in existing_headers_norm:
                    try:
                        with open(path, 'w', encoding='utf-8', newline='') as f:
                            try:
                                f.write('\ufeff')
                            except Exception:
                                pass
                            writer = csv.DictWriter(f, fieldnames=headers)
                            writer.writeheader()
                            for r in existing_rows:
                                out = {h: '' for h in headers}
                                for k, v in (r or {}).items():
                                    try:
                                        nk = str(k)
                                    except Exception:
                                        nk = k
                                    if isinstance(nk, str) and nk.startswith('\ufeff'):
                                        nk = nk[1:]
                                    if nk in out:
                                        out[nk] = v
                                writer.writerow(out)
                            try:
                                f.flush(); os.fsync(f.fileno())
                            except Exception:
                                pass
                    except Exception:
                        # If upgrade fails, fall back to appending without Word Index to avoid misalignment
                        headers = existing_headers_norm
            with open(path, 'a', encoding='utf-8', newline='') as f:
                writer = csv.DictWriter(f, fieldnames=headers)
                # Build an output row aligned to headers (map normalized 'Vowel Ending' to BOM header if needed)
                out = {h: '' for h in headers}
                for k, v in (row or {}).items():
                    if k in out:
                        out[k] = v
                    elif k == 'Vowel Ending' and '\ufeffVowel Ending' in out:
                        out['\ufeffVowel Ending'] = v
                writer.writerow(out)
                try:
                    f.flush(); os.fsync(f.fileno())
                except Exception:
                    pass
            row_no = (len(existing_rows) if existing_rows else 0) + 1
            try:
                messagebox.showinfo("Saved", f"Saved to {path} (new row #{row_no})")
            except Exception:
                pass
            print(f"[Save] new row #{row_no}")
            return True, row_no, 'append'



























    def _confirm_overwrite_modal(self, parent, headers, existing_row, new_row) -> bool:
        try:
            win = tk.Toplevel(parent)
            win.title("Confirm Overwrite")
            win.configure(bg="light gray")
            win.transient(parent)
            win.grab_set()
            try:
                # Bind F11 toggle and enable user maximize/restore
                self._wm_for(win)
            except Exception:
                pass
            title = tk.Label(win, text="A matching entry exists. Review changes:",
                             bg="dark slate gray", fg="white", font=("Arial", 12, "bold"), padx=10, pady=6)
            title.pack(fill=tk.X)
            body = tk.Frame(win, bg="light gray")
            body.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            left = tk.Frame(body, bg="light gray"); right = tk.Frame(body, bg="light gray")
            left.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0,5))
            right.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(5,0))
            tk.Label(left, text="Existing", font=("Arial", 11, "bold"), bg="light gray").pack(anchor='w')
            tk.Label(right, text="New", font=("Arial", 11, "bold"), bg="light gray").pack(anchor='w')
            l_txt = tk.Text(left, wrap='word', height=18); r_txt = tk.Text(right, wrap='word', height=18)
            l_txt.pack(fill=tk.BOTH, expand=True); r_txt.pack(fill=tk.BOTH, expand=True)
            diff_lines = []
            for h in headers:
                # Normalize lookups to tolerate BOM-prefixed legacy headers
                try:
                    oldv = str(self._norm_get(existing_row, h) or '')
                except Exception:
                    oldv = str(existing_row.get(h, ''))
                try:
                    newv = str(self._norm_get(new_row, h) or '')
                except Exception:
                    newv = str(new_row.get(h, ''))
                l_line = f"{h}: {oldv}\n"; r_line = f"{h}: {newv}\n"
                l_txt.insert(tk.END, l_line); r_txt.insert(tk.END, r_line)
                if (oldv or "") != (newv or ""):
                    diff_lines.append(h)
            try:
                def _hl(txtw, header):
                    start = '1.0'
                    while True:
                        idx = txtw.search(f"{header}:", start, stopindex=tk.END)
                        if not idx:
                            break
                        txtw.tag_add('diff', idx, f"{idx} lineend")
                        start = f"{idx}+1c"
                for h in diff_lines:
                    _hl(l_txt, h); _hl(r_txt, h)
                l_txt.tag_config('diff', background='#fff3cd'); r_txt.tag_config('diff', background='#fff3cd')
            except Exception:
                pass
            btns = tk.Frame(win, bg="light gray"); btns.pack(fill=tk.X, pady=(8,6))
            def _do_overwrite():
                win._result = True; win.destroy()
            def _do_cancel():
                win._result = False; win.destroy()
            cancel = tk.Button(btns, text="Cancel", command=_do_cancel)
            ok = tk.Button(btns, text="Overwrite", command=_do_overwrite, bg='dark cyan', fg='white')
            cancel.pack(side=tk.RIGHT, padx=6); ok.pack(side=tk.RIGHT)
            cancel.focus_set()
            parent.wait_window(win)
            return bool(getattr(win, '_result', False))
        except Exception:
            return True

    def launch_verse_analysis_dashboard(self):
        """Clears the main dashboard and launches the Verse Analysis Dashboard."""
        for widget in self.root.winfo_children():
            widget.destroy()
        self.root.title("Verse Analysis Dashboard")
        self.setup_verse_analysis_dashboard()

    def setup_verse_analysis_dashboard(self):
        """Builds the Verse Analysis Dashboard interface."""
        for widget in self.root.winfo_children():
            widget.destroy()
        self.main_frame = tk.Frame(self.root, bg='light gray', padx=10, pady=10)
        self.main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)

        header_label = tk.Label(
            self.main_frame,
            text="Verse Analysis Dashboard",
            font=('Arial', 18, 'bold'),
            bg='dark slate gray',
            fg='white',
            pady=10
        )
        header_label.pack(fill=tk.X, pady=10)

        # Create a frame for the option buttons
        button_frame = tk.Frame(self.main_frame, bg='light gray')
        button_frame.pack(expand=True)

        # Literal Translation Button
        literal_btn = tk.Button(
            button_frame,
            text="Literal Translation",
            font=('Arial', 14, 'bold'),
            bg='dark cyan',
            fg='white',
            padx=20,
            pady=10,
            command=self.launch_literal_analysis
        )
        literal_btn.pack(pady=10)

        # New button for editing saved literal translation
        edit_saved_btn = tk.Button(
            button_frame,
            text="Edit Saved Literal Translation",
            font=('Arial', 14, 'bold'),
            bg='dark cyan',
            fg='white',
            padx=20,
            pady=10,
            command=self.launch_select_verse  # Updated command
        )
        edit_saved_btn.pack(pady=10)

        # Placeholder for Spiritual Translation (future implementation)
        spiritual_btn = tk.Button(
            button_frame,
            text="Spiritual Translation (Coming Soon)",
            font=('Arial', 14, 'bold'),
            bg='gray',
            fg='white',
            padx=20,
            pady=10,
            state=tk.DISABLED
        )
        spiritual_btn.pack(pady=10)

        # Placeholder for Translation Management (future implementation)
        management_btn = tk.Button(
            button_frame,
            text="Translation Management (Coming Soon)",
            font=('Arial', 14, 'bold'),
            bg='gray',
            fg='white',
            padx=20,
            pady=10,
            state=tk.DISABLED
        )
        management_btn.pack(pady=10)

        # Back button to return to the main dashboard
        back_btn = tk.Button(
            self.main_frame,
            text="Back to Main Dashboard",
            font=('Arial', 14, 'bold'),
            bg='red',
            fg='white',
            padx=20,
            pady=10,
            command=self.show_dashboard
        )
        back_btn.pack(pady=10)

    def launch_select_verse(self):
        """
        Creates a centered modal window to let the user select a verse by:
        1. Searching by verse content
        2. Filtering by metadata (Raag, Writer, Bani, Page)
        Displays only those verses that already exist in the assessment Excel.
        """

        # === Setup modal ===
        select_win = tk.Toplevel(self.root)
        select_win.title("Select Verse")
        select_win.geometry("800x600")
        try:
            try:
                self._wm_for(select_win).maximize()
            except Exception:
                pass
        except tk.TclError:
            pass
        select_win.configure(bg="light gray")

        # === Load Excel data ===
        file_path = "1.2.1 assessment_data.xlsx"
        df_existing = self.load_existing_assessment_data(file_path)

        # === Center content ===
        center_frame = tk.Frame(select_win, bg="light gray")
        center_frame.pack(expand=True)

        content_frame = tk.Frame(center_frame, bg="light gray", width=960)
        content_frame.pack()

        # === Header ===
        header_label = tk.Label(
            content_frame,
            text="Select Verse",
            bg="dark slate gray",
            fg="white",
            font=("Helvetica", 18, "bold"),
            padx=20, pady=10
        )
        header_label.pack(fill=tk.X, pady=(0, 10))

        # === Mode selection (Search or Filter) ===
        mode_var = tk.StringVar(value="search")
        mode_frame = tk.Frame(content_frame, bg="light gray")
        mode_frame.pack(pady=10)

        tk.Radiobutton(mode_frame, text="Search by Verse", variable=mode_var, value="search",
                    bg="light gray", font=("Arial", 12)).pack(side=tk.LEFT, padx=10)
        tk.Radiobutton(mode_frame, text="Filter by Metadata", variable=mode_var, value="filter",
                    bg="light gray", font=("Arial", 12)).pack(side=tk.LEFT, padx=10)

        # === Create container for dynamic frames ===
        main_content_area = tk.Frame(content_frame, bg="light gray")
        main_content_area.pack(fill=tk.BOTH, expand=True)

        # === Search Frame ===
        search_frame = tk.Frame(main_content_area, bg="light gray")
        tk.Label(search_frame, text="Enter or paste a verse to search:", bg="light gray",
                font=("Arial", 12, "bold")).pack(pady=5)

        search_entry = tk.Entry(search_frame, font=("Arial", 12), width=80)
        search_entry.pack(pady=5)

        search_button = tk.Button(
            search_frame, text="Search", font=("Arial", 12, "bold"),
            bg="navy", fg="white", width=12,
            command=lambda: perform_search(search_entry.get())
        )
        search_button.pack(pady=5)

        search_results_list = tk.Listbox(search_frame, font=("Arial", 12), width=80, height=10)
        search_results_list.pack(pady=10)

        # === Filter Frame ===
        filter_frame = tk.Frame(main_content_area, bg="light gray")
        tk.Label(filter_frame, text="Filter verses by metadata:", bg="light gray",
                font=("Arial", 12, "bold")).pack(pady=5)

        filter_controls = tk.Frame(filter_frame, bg="light gray")
        filter_controls.pack(pady=5)

        # === Dropdowns for filter ===
        df = df_existing.copy()
        raag_var, writer_var, bani_var, page_var = tk.StringVar(), tk.StringVar(), tk.StringVar(), tk.StringVar()
        # Sort Raag, Writer, Bani based on first page appearance
        initial_raag = df.dropna(subset=["Raag (Fixed)"]).drop_duplicates("Raag (Fixed)", keep="first").sort_values("Page Number")["Raag (Fixed)"].tolist()
        initial_writer = df.dropna(subset=["Writer (Fixed)"]).drop_duplicates("Writer (Fixed)", keep="first").sort_values("Page Number")["Writer (Fixed)"].tolist()
        initial_bani = df.dropna(subset=["Bani Name"]).drop_duplicates("Bani Name", keep="first").sort_values("Page Number")["Bani Name"].tolist()
        # Sort Page Numbers numerically
        initial_page = sorted(df["Page Number"].dropna().unique())
        initial_page = [str(p) for p in initial_page]  # Convert back to string for dropdown


        tk.Label(filter_controls, text="Raag:", bg="light gray", font=("Arial", 12)).grid(row=0, column=0, padx=5, pady=5)
        raag_dropdown = ttk.Combobox(filter_controls, textvariable=raag_var, values=initial_raag, width=15)
        raag_dropdown.grid(row=0, column=1, padx=5, pady=5)

        tk.Label(filter_controls, text="Writer:", bg="light gray", font=("Arial", 12)).grid(row=0, column=2, padx=5, pady=5)
        writer_dropdown = ttk.Combobox(filter_controls, textvariable=writer_var, values=initial_writer, width=15)
        writer_dropdown.grid(row=0, column=3, padx=5, pady=5)

        tk.Label(filter_controls, text="Bani:", bg="light gray", font=("Arial", 12)).grid(row=0, column=4, padx=5, pady=5)
        bani_dropdown = ttk.Combobox(filter_controls, textvariable=bani_var, values=initial_bani, width=15)
        bani_dropdown.grid(row=0, column=5, padx=5, pady=5)

        tk.Label(filter_controls, text="Page:", bg="light gray", font=("Arial", 12)).grid(row=0, column=6, padx=5, pady=5)
        page_dropdown = ttk.Combobox(filter_controls, textvariable=page_var, values=initial_page, width=10)
        page_dropdown.grid(row=0, column=7, padx=5, pady=5)

        # === Update filter dropdowns dynamically ===
        def update_dropdowns(*args):
            filtered_df = df.copy()
            if raag_var.get():
                filtered_df = filtered_df[filtered_df["Raag (Fixed)"] == raag_var.get()]
            if writer_var.get():
                filtered_df = filtered_df[filtered_df["Writer (Fixed)"] == writer_var.get()]
            if bani_var.get():
                filtered_df = filtered_df[filtered_df["Bani Name"] == bani_var.get()]
            if page_var.get():
                filtered_df = filtered_df[filtered_df["Page Number"].astype(str) == page_var.get()]
            raag_dropdown['values'] = (
                filtered_df.dropna(subset=["Raag (Fixed)"])
                .drop_duplicates("Raag (Fixed)", keep="first")
                .sort_values("Page Number")["Raag (Fixed)"].tolist()
            )

            writer_dropdown['values'] = (
                filtered_df.dropna(subset=["Writer (Fixed)"])
                .drop_duplicates("Writer (Fixed)", keep="first")
                .sort_values("Page Number")["Writer (Fixed)"].tolist()
            )

            bani_dropdown['values'] = (
                filtered_df.dropna(subset=["Bani Name"])
                .drop_duplicates("Bani Name", keep="first")
                .sort_values("Page Number")["Bani Name"].tolist()
            )

            sorted_pages = sorted(filtered_df["Page Number"].dropna().unique())
            page_dropdown['values'] = [str(p) for p in sorted_pages]

        for var in (raag_var, writer_var, bani_var, page_var):
            var.trace_add("write", update_dropdowns)

        # === Filter Button & Listbox ===
        filter_button = tk.Button(filter_frame, text="Apply Filter", font=("Arial", 12, "bold"),
                                bg="navy", fg="white", command=lambda: update_verse_list())
        filter_button.pack(pady=5)

        filter_results_list = tk.Listbox(filter_frame, font=("Arial", 12), width=80, height=10)
        filter_results_list.pack(pady=10)

        # === Toggle mode display ===
        def update_mode():
            filter_frame.pack_forget()
            search_frame.pack_forget()
            if mode_var.get() == "search":
                search_frame.pack(pady=10)
            else:
                filter_frame.pack(pady=10)

        mode_var.trace_add("write", lambda *args: update_mode())
        update_mode()

        # === Search Logic ===
        def perform_search(query):
            search_results_list.delete(0, tk.END)
            if not query:
                search_results_list.insert(tk.END, "No query entered.")
                return
            headers, candidate_matches = self.match_sggs_verse(query)
            candidate_verses = list(dict.fromkeys(candidate["Verse"] for candidate in candidate_matches))
            excel_verses = set(df_existing["Verse"].unique())
            unique_verses = [v for v in candidate_verses if v in excel_verses]
            if unique_verses:
                for verse in unique_verses:
                    search_results_list.insert(tk.END, verse)
            else:
                search_results_list.insert(tk.END, "No analyzed verse matches found.")

        # === Filter Logic ===
        def update_verse_list():
            filter_results_list.delete(0, tk.END)
            filtered_df = df_existing.copy()
            if raag_var.get():
                filtered_df = filtered_df[filtered_df["Raag (Fixed)"].str.contains(raag_var.get(), case=False, na=False)]
            if writer_var.get():
                filtered_df = filtered_df[filtered_df["Writer (Fixed)"].str.contains(writer_var.get(), case=False, na=False)]
            if bani_var.get():
                filtered_df = filtered_df[filtered_df["Bani Name"].str.contains(bani_var.get(), case=False, na=False)]
            if page_var.get():
                filtered_df = filtered_df[filtered_df["Page Number"].astype(str).str.contains(page_var.get(), case=False, na=False)]
            for verse in filtered_df["Verse"].unique():
                filter_results_list.insert(tk.END, verse)

        # === Finalize & Back buttons ===
        def finalize_selection():
            if mode_var.get() == "search":
                sel = search_results_list.curselection()
                final_verse = search_results_list.get(sel[0]) if sel else ""
            else:
                sel = filter_results_list.curselection()
                final_verse = filter_results_list.get(sel[0]) if sel else ""
            if final_verse:
                self.finalized_verse = final_verse
                select_win.destroy()
                self.launch_edit_saved_literal_translation()
            else:
                tk.messagebox.showerror("Error", "Please select a verse before proceeding.")

        bottom_frame = tk.Frame(main_content_area, bg="light gray")
        bottom_frame.pack(side=tk.BOTTOM, pady=10)

        tk.Button(bottom_frame, text="Finalize Selection", font=("Arial", 12, "bold"),
                bg="green", fg="white", padx=15, pady=5,
                command=finalize_selection).pack(side=tk.LEFT, padx=(10, 5))

        tk.Button(bottom_frame, text="Back to Dashboard", font=("Arial", 12, "bold"),
                bg="gray", fg="white", padx=15, pady=5,
                command=lambda: (select_win.destroy(), self.setup_verse_analysis_dashboard())).pack(side=tk.LEFT, padx=(5, 10))

        # === Modal behavior ===
        select_win.transient(self.root)
        select_win.grab_set()
        self.root.wait_window(select_win)

    def launch_edit_saved_literal_translation(self):
        """Launch a window to review and select words for re-analysis from a saved verse."""

        verse = self.finalized_verse
        df = self.load_existing_assessment_data("1.2.1 assessment_data.xlsx")

        # === Clear root ===
        for widget in self.root.winfo_children():
            widget.destroy()

        # === Setup main frame ===
        main_frame = tk.Frame(self.root, bg="light gray", padx=20, pady=20)
        main_frame.pack(fill=tk.BOTH, expand=True)

        # === Header ===
        header = tk.Label(
            main_frame,
            text="Edit Saved Literal Translation",
            font=("Helvetica", 18, "bold"),
            bg="dark slate gray",
            fg="white",
            pady=10
        )
        header.pack(fill=tk.X, pady=(0, 20))

        # === Display Verse ===
        verse_label = tk.Label(
            main_frame,
            text=f"Verse:\n  {verse}",
            font=("Arial", 14),
            bg="light gray",
            anchor="w",
            justify="left"
        )
        verse_label.pack(fill=tk.X, padx=10)

        # === Display Translation ===
        row_data = df[df['Verse'] == verse].iloc[0]
        translation = row_data.get('Translation', '')
        translation_label = tk.Label(
            main_frame,
            text=f"Translation:\n  {translation}",
            font=("Arial", 13, "italic"),
            bg="light gray",
            anchor="w",
            justify="left"
        )
        translation_label.pack(fill=tk.X, padx=10, pady=(10, 5))

        # Returns the real value if it isnÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â‚¬Å¾Ã‚Â¢t NaN; otherwise it returns a ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‚ÂÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â placeholder
        def safe(val):
            return val if pd.notna(val) else "ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â"

        # === Metadata (Raag, Writer, Bani, Page) ===
        metadata_frame = tk.Frame(main_frame, bg="light gray")
        metadata_frame.pack(fill=tk.X, padx=10, pady=(0, 15))

        for label, col in [
            ("Raag:", "Raag (Fixed)"),
            ("Writer:", "Writer (Fixed)"),
            ("Bani:", "Bani Name"),
            ("Page:", "Page Number"),
        ]:
            val = row_data.get(col)
            # only show non-missing values
            if val is None or pd.isna(val):
                continue

            meta = tk.Label(
                metadata_frame,
                text=f"{label} {val}",
                font=("Arial", 11, "bold"),
                bg="light gray",
                anchor="w",
                justify="left"
            )
            meta.pack(anchor="w")

        # === Editable Metadata for Entire Verse ===
        # Create a frame for editing verse-wide settings like "Framework?" and "Explicit?"
        edit_metadata_frame = tk.Frame(main_frame, bg="light gray")
        edit_metadata_frame.pack(fill=tk.X, padx=10, pady=(0, 15))

        # Helper: safely extract a boolean value (assuming Excel stores these as numbers 0/1 or booleans)
        def safe_bool(val):
            try:
                # If the value is a numpy integer, convert it to a normal int first
                if isinstance(val, np.integer):
                    val = int(val)
                # Now, if it's numeric, nonzero means True; otherwise, use its boolean conversion
                return bool(val) if isinstance(val, (int, float)) else val
            except Exception:
                return False

        # Get initial values from the Excel row data
        initial_framework = safe_bool(row_data.get("Framework?"))
        initial_explicit = safe_bool(row_data.get("Explicit?"))

        # Create checkboxes for Framework and Explicit metadata
        framework_var_edit = tk.BooleanVar(value=initial_framework)
        explicit_var_edit = tk.BooleanVar(value=initial_explicit)

        framework_cb_edit = tk.Checkbutton(
            edit_metadata_frame,
            text="Framework?",
            variable=framework_var_edit,
            font=("Arial", 11, "bold"),
            bg="light gray"
        )
        framework_cb_edit.pack(side=tk.LEFT, padx=10)

        explicit_cb_edit = tk.Checkbutton(
            edit_metadata_frame,
            text="Explicit?",
            variable=explicit_var_edit,
            font=("Arial", 11, "bold"),
            bg="light gray"
        )
        explicit_cb_edit.pack(side=tk.LEFT, padx=10)

        def update_verse_metadata(new_framework, new_explicit):
            file_path = "1.2.1 assessment_data.xlsx"
            try:
                # Load existing data
                df_existing = self.load_existing_assessment_data(file_path)
                # Create a mask for rows corresponding to the current verse
                verse_mask = df_existing["Verse"] == verse
                # Update the columns for the entire verse with the new values
                df_existing.loc[verse_mask, "Framework?"] = int(new_framework)
                df_existing.loc[verse_mask, "Explicit?"] = int(new_explicit)
                df_existing.to_excel(file_path, index=False)
                messagebox.showinfo("Updated", "Verse metadata updated successfully!")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to update verse metadata: {e}")

        update_btn = tk.Button(
            edit_metadata_frame,
            text="Update Verse Metadata",
            font=("Arial", 11, "bold"),
            bg="dark cyan", fg="white",
            padx=10, pady=5,
            command=lambda: update_verse_metadata(framework_var_edit.get(), explicit_var_edit.get())
        )
        update_btn.pack(side=tk.LEFT, padx=10)

        # === Word Table Frame ===
        word_frame = tk.LabelFrame(
            main_frame,
            text="Select words to re-analyze:",
            bg="light gray",
            font=("Arial", 12, "bold")
        )
        word_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # === Checkbox Simulation ===
        selected_items = set()

        columns = [
            'Select', 'Word', 'Vowel Ending', 'Number / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨',
            'Grammar / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¯ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£', 'Gender / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â', 'Word Type',
            'Word Root', 'Word Index'
        ]

        tree = ttk.Treeview(word_frame, columns=columns, show='headings', selectmode='none')
        for col in columns:
            tree.heading(col, text=col)
            tree.column(col, anchor=tk.CENTER, width=110 if col != 'Select' else 60)

        # Add scrollbar
        vsb = ttk.Scrollbar(word_frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        tree.pack(fill=tk.BOTH, expand=True)

        # Insert rows
        # 1) Configure tag styles for odd/even rows
        tree.tag_configure('oddrow', background='white')
        tree.tag_configure('evenrow', background='#E8E8E8')  # light gray

        # 2) Insert rows with alternating row tags
        rows = df[df['Verse'] == verse]
        for i, (_, row) in enumerate(rows.iterrows()):
            row_id = f"row{i}"
            # Your existing 'values' building
            values = [
                "",  # checkbox
                safe(self._norm_get(row, "Word")),
                safe(self._norm_get(row, "\ufeffVowel Ending")),
                safe(self._norm_get(row, "Number / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨")),
                safe(self._norm_get(row, "Grammar / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¯ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£")),
                safe(self._norm_get(row, "Gender / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â")),
                safe(self._norm_get(row, "Word Root")),
                safe(self._norm_get(row, "Type")),
                int(self._norm_get(row, "Word Index") or -1)
            ]

            # Determine odd/even row coloring
            if i % 2 == 0:
                tree.insert('', tk.END, iid=row_id, values=values, tags=('evenrow',))
            else:
                tree.insert('', tk.END, iid=row_id, values=values, tags=('oddrow',))

        # === Toggle ÃƒÆ’Ã‚Â¢Ãƒâ€¦Ã¢â‚¬Å“ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ in first column ===
        def on_tree_click(event):
            region = tree.identify_region(event.x, event.y)
            if region == 'cell':
                row_id = tree.identify_row(event.y)
                col = tree.identify_column(event.x)
                if row_id and col == '#1':
                    if row_id in selected_items:
                        selected_items.remove(row_id)
                        tree.set(row_id, 'Select', "")
                    else:
                        selected_items.add(row_id)
                        tree.set(row_id, 'Select', "ÃƒÆ’Ã‚Â¢Ãƒâ€¦Ã¢â‚¬Å“ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ")

        tree.bind('<Button-1>', on_tree_click)

        # === Action Buttons ===
        btn_frame = tk.Frame(main_frame, bg="light gray")
        btn_frame.pack(pady=20)

        def analyze_selected_words():
            if not hasattr(self, "results_text"):
                self.results_text = scrolledtext.ScrolledText(
                    self.root,
                    width=90,
                    height=20,
                    font=("Consolas", 11),
                    bd=3,
                    relief=tk.SUNKEN,
                    wrap=tk.WORD
                )
                self.results_text.pack_forget()  # DonÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â‚¬Å¾Ã‚Â¢t show it to the user during re-analysis

            # Get column names to dynamically determine index of "Word" and "Word Index"
            column_names = tree["columns"]
            word_col_index = column_names.index("Word")
            index_col_index = column_names.index("Word Index")

            # Build the list of selected words with indices
            selected_words_with_indices = [
                (
                    tree.item(rid)['values'][word_col_index],
                    int(tree.item(rid)['values'][index_col_index])
                )
                for rid in selected_items
            ]

            all_words_in_verse = [tree.item(rid)['values'][1] for rid in tree.get_children()]

            if not selected_words_with_indices:
                messagebox.showwarning(
                    "No Words Selected",
                    "You havenÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â‚¬Å¾Ã‚Â¢t selected any words for re-analysis.\n\n"
                    "Click the ÃƒÆ’Ã‚Â¢Ãƒâ€¦Ã¢â‚¬Å“ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ box beside the word(s) you wish to re-analyze, then press the button again."
                )
                return

            # Step 1: Set context before any processing
            self.current_pankti = verse
            self.accumulated_pankti = verse
            self.pankti_words = all_words_in_verse  # Keep 'ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¥Ãƒâ€šÃ‚Â¥' if part of original flow
            self.selected_verses = [verse]
            self.accumulated_meanings = [{} for _ in self.pankti_words]
            self.accumulated_finalized_matches = [[] for _ in self.pankti_words]
            self.all_new_entries = []
            self.current_reanalysis_index = []

            # Step 2: Load Excel and pre-fill all words from verse
            df = self.load_existing_assessment_data("1.2.1 assessment_data.xlsx")
            verse_rows = df[df["Verse"] == verse]

            for i, word in enumerate(self.pankti_words):
                word_rows = verse_rows[
                    (verse_rows["Word"] == word) &
                    (verse_rows["Verse"] == verse) &
                    (verse_rows["Word Index"] == i)
                ]
                if not word_rows.empty:
                    # Meanings
                    meanings = [
                        row.get("Selected Darpan Meaning", "")
                        for _, row in word_rows.iterrows()
                        if pd.notna(row.get("Selected Darpan Meaning"))
                    ]
                    self.accumulated_meanings[i] = {"word": word, "meanings": meanings}

                    # Grammar matches
                    finalized = word_rows.to_dict("records")
                    self.accumulated_finalized_matches[i] = finalized
                    self.all_new_entries.extend(finalized)

            # Step 3: Determine which indices to re-analyze
            selected_words_with_indices.sort(key=lambda x: x[1])  # Sort by Word Index (ascending)
            word_indices = [idx for (_, idx) in selected_words_with_indices]
            if not word_indices:
                messagebox.showinfo("Not Found", "Selected words not found in verse structure.")
                return

            # Populate past details (grammar and selected meanings) for each word in the verse.
            self.past_word_details = {}

            # Iterate over each tuple (word, idx) in selected_words_with_indices.
            for word, idx in selected_words_with_indices:
                # Filter the DataFrame for rows matching the word, verse, and unique word index.
                word_rows = verse_rows[
                    (verse_rows["Word"] == word) &
                    (verse_rows["Verse"] == verse) &
                    (verse_rows["Word Index"] == idx)
                ]
                
                if not word_rows.empty:
                    # Choose the representative row that has the highest Grammar Revision.
                    latest_idx = word_rows["Grammar Revision"].idxmax()
                    latest_row = word_rows.loc[latest_idx]
                    
                    # Gather all available past Darpan meanings from the filtered rows.
                    darpan_meanings = []
                    for _, row in word_rows.iterrows():
                        val = row.get("Selected Darpan Meaning")
                        if pd.notna(val):
                            # Split on comma and strip whitespace
                            split_meanings = [m.strip() for m in val.split("| ")]
                            darpan_meanings.extend(split_meanings)

                    # Store the past details in the dictionary using the word index as the key.
                    self.past_word_details[idx] = {
                        "Word": word,
                        "\ufeffVowel Ending": self._norm_get(latest_row, "\ufeffVowel Ending") or "",
                        "Number / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨": self._norm_get(latest_row, "Number / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨") or "",
                        "Grammar / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¯ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£": self._norm_get(latest_row, "Grammar / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¯ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£") or "",
                        "Gender / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â": self._norm_get(latest_row, "Gender / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â") or "",
                        "Type": self._norm_get(latest_row, "Type") or "",
                        "Word Root": self._norm_get(latest_row, "Word Root") or "",
                        "Word Index": idx,
                        "darpan_meanings": darpan_meanings
                    }

            # Step 4: Store queue and start
            self.reanalysis_queue = word_indices
            self.process_next_selected_word()

        def back_to_search():
            self.launch_select_verse()

        def back_to_dashboard():
            self.setup_verse_analysis_dashboard()

        tk.Button(
            btn_frame,
            text="Analyze Selected Words",
            bg="navy", fg="white",
            font=("Arial", 12, "bold"),
            padx=15, pady=5,
            command=analyze_selected_words
        ).pack(side=tk.LEFT, padx=10)

        tk.Button(
            btn_frame,
            text="Back to Search",
            bg="gray", fg="white",
            font=("Arial", 12, "bold"),
            padx=15, pady=5,
            command=back_to_search
        ).pack(side=tk.LEFT, padx=10)

        tk.Button(
            btn_frame,
            text="Back to Dashboard",
            bg="red", fg="white",
            font=("Arial", 12, "bold"),
            padx=15, pady=5,
            command=back_to_dashboard
        ).pack(side=tk.LEFT, padx=10)

        # === Optional: Customize style ===
        style = ttk.Style()
        style.configure("Treeview.Heading", font=('Arial', 11, 'bold'))
        style.configure("Treeview", rowheight=28, font=('Arial', 11))

    def process_next_selected_word(self):
        """Process the next word from re-analysis queue, or prompt to save if finished."""
        if not hasattr(self, 'reanalysis_queue') or not self.reanalysis_queue:
            messagebox.showinfo("Done", "Re-analysis completed for all selected words.")
            
            # Proceed to save reanalyzed results
            if hasattr(self, 'save_results_btn') and self.save_results_btn.winfo_exists():
                self.save_results_btn.config(state=tk.NORMAL)
            
            self.prompt_save_results_reanalysis(self.all_new_entries, skip_copy=False)  # Skip clipboard step for reanalysis
            return

        # Process the next word from queue
        idx = self.reanalysis_queue.pop(0)
        self.current_word_index = idx
        word = self.pankti_words[idx]
        self.ensure_meanings_slot_initialized(idx, word)

        self.fetch_data_for_reanalysis(word, self.accumulated_pankti, idx)

    def fetch_data_for_reanalysis(self, word, pankti, index):
        self.reset_input_variables()
        self.current_word_index = index
        self.user_input_reanalysis(word, pankti, index)

        if hasattr(self, 'input_window') and self.input_window.winfo_exists():
            self.root.wait_window(self.input_window)
        else:
            return

        if self.input_submitted:
            self.handle_submitted_input(word)
        else:
            print(f"Skipped: {word}")

        # Now move to next in reanalysis queue
        self.process_next_selected_word()

    def user_input_reanalysis(self, word, pankti, index):
        print(f"[Reanalysis] Opening input window for {word} (index {index})")
        self.input_submitted = False
        self.current_word_index = index  # Ensure correct word gets highlighted

        self.input_window = tk.Toplevel(self.root)
        self.input_window.title(f"[Edit Mode] Input for {word}")
        self.input_window.configure(bg='light gray')
        try:
            self._wm_for(self.input_window).maximize()
        except Exception:
            pass
        self.input_window.resizable(True, True)

        # Display the Pankti with word highlight
        pankti_frame = tk.Frame(self.input_window, bg='light gray')
        pankti_frame.pack(fill=tk.X, padx=20, pady=10)

        pankti_display = tk.Text(
            pankti_frame, wrap=tk.WORD, bg='light gray', font=('Arial', 32),
            height=2, padx=5, pady=5
        )
        pankti_display.pack(fill=tk.X, expand=False)
        pankti_display.insert(tk.END, pankti)
        pankti_display.tag_add("center", "1.0", "end")
        pankti_display.tag_configure("center", justify='center')

        # Highlight the word at the re-analysis index
        words = pankti.split()
        start_idx = 0
        for i, w in enumerate(words):
            if i == index:
                end_idx = start_idx + len(w)
                pankti_display.tag_add("highlight", f"1.{start_idx}", f"1.{end_idx}")
                pankti_display.tag_config("highlight", foreground="blue", font=('Arial', 32, 'bold'))
                break
            start_idx += len(w) + 1
        pankti_display.config(state=tk.DISABLED)

        # Create layout pane
        split_pane = tk.PanedWindow(self.input_window, orient=tk.HORIZONTAL, bg='light gray')
        split_pane.pack(fill=tk.BOTH, expand=True)

        # Left: Meanings
        self.left_pane = tk.Frame(split_pane, bg='light gray')
        split_pane.add(self.left_pane, stretch="always")
        tk.Label(self.left_pane, text=f"Re-analyze Meanings for {word}:", bg='light gray',
                font=('Arial', 14, 'bold')).pack(anchor='center', pady=(0, 10))
        self.meanings_scrollbar = tk.Scrollbar(self.left_pane, orient=tk.VERTICAL)
        self.meanings_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.meanings_canvas = tk.Canvas(self.left_pane, bg='light gray', borderwidth=0,
                                        yscrollcommand=self.meanings_scrollbar.set)
        self.meanings_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.meanings_scrollbar.config(command=self.meanings_canvas.yview)
        self.meanings_inner_frame = tk.Frame(self.meanings_canvas, bg='light gray')
        self.meanings_canvas.create_window((0, 0), window=self.meanings_inner_frame, anchor='nw')

        # Right: Grammar Options
        right_pane = tk.Frame(split_pane, bg='light gray')
        split_pane.add(right_pane, stretch="always")
        tk.Label(right_pane, text="Adjust Grammar Options:", bg='light gray',
                font=('Arial', 14, 'bold')).pack(pady=10)
        self.setup_options(
            right_pane,
            "Do you know the Number of the word?",
            [("Singular", "Singular / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢"), ("Plural", "Plural / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â"), ("Not Applicable", "NA")],
            self.number_var
        )
        self.setup_options(
            right_pane,
            "Do you know the Gender of the word?",
            [("Masculine", "Masculine / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â"), ("Feminine", "Feminine / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬"), ("Neutral", "Trans / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢")],
            self.gender_var
        )
        self.setup_options(
            right_pane,
            "Do you know the Part of Speech for the word?",
            [("Noun", "Noun / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Âµ"), ("Adjective", "Adjectives / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¶ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¶ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£"),
            ("Adverb", "Adverb / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â  ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¶ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£"), ("Verb", "Verb / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â "),
            ("Pronoun", "Pronoun / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Âµ"), ("Postposition", "Postposition / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â§ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢"),
            ("Conjunction", "Conjunction / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¯ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢"), ("Interjection", "Interjection / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢")],
            self.pos_var
        )

        # Submit / Skip
        button_frame = tk.Frame(self.input_window, bg='light gray')
        button_frame.pack(side=tk.BOTTOM, fill=tk.X, padx=20, pady=(6, 0))
        # Center the action buttons: use expanding spacers on both sides
        tk.Frame(button_frame, bg='light gray').pack(side=tk.LEFT, expand=True)
        tk.Button(button_frame, text="Submit", command=self.submit_input_reanalysis,
                font=('Arial', 12, 'bold'), bg='navy', fg='white',
                padx=30, pady=15).pack(side=tk.LEFT, padx=15)
        tk.Button(button_frame, text="Skip", command=self.skip_input,
                font=('Arial', 12, 'bold'), bg='navy', fg='white',
                padx=30, pady=15).pack(side=tk.LEFT, padx=15)
        tk.Frame(button_frame, bg='light gray').pack(side=tk.LEFT, expand=True)

        # Progress + Dictionary Lookup
        self.start_progress()
        threading.Thread(target=self.lookup_meanings_thread, args=(word,), daemon=True).start()

        self.input_window.transient(self.root)
        self.input_window.grab_set()
        self.root.wait_window(self.input_window)
        print(f"[Reanalysis] Closed input window for {word}")

    def submit_input_reanalysis(self):
        self.input_submitted = True
        if hasattr(self, 'input_window') and self.input_window.winfo_exists():
            self.input_window.destroy()

        # Start the progress bar
        self.start_progress()

        # Run the reanalysis search in a separate thread
        search_thread = threading.Thread(target=self.perform_search_and_finish_reanalysis)
        search_thread.start()

    def perform_search_and_finish_reanalysis(self):
        current_word = self.pankti_words[self.current_word_index]
        number = self.number_var.get()
        gender = self.gender_var.get()
        pos = self.pos_var.get()

        print(f"[Reanalysis] Processing word: {current_word}, Number: {number}, Gender: {gender}, POS: {pos}")
        
        if number == "NA" and gender == "NA" and pos == "NA":
            matches = self.search_by_inflections(current_word)
        else:
            matches = self.search_by_criteria(current_word, number, gender, pos)
            if not matches:
                messagebox.showinfo(
                    "No Matches Found",
                    "No matches were found as per the criteria. Now conducting a general search."
                )
                matches = self.search_by_inflections(current_word)

        # Meanings are guaranteed to be preloaded from the Excel into self.accumulated_meanings
        entry = self.accumulated_meanings[self.current_word_index]
        meanings = entry.get("meanings", []) if isinstance(entry, dict) else entry

        # Stop progress bar first
        self.root.after(0, self.stop_progress)

        if matches:
            print(f"[Reanalysis] Found matches for {current_word}: {matches}")
            self.root.after(0, lambda: self.show_matches_reanalysis(matches, self.current_pankti, meanings, self.current_word_index))
        else:
            self.root.after(0, lambda: messagebox.showinfo("No Matches", f"No matches found for the word: {current_word}"))
            self.current_word_index += 1
            self.root.after(0, self.process_next_selected_word)

    def show_matches_reanalysis(self, matches, pankti, meanings, index, max_display=30):
        # Destroy any existing match window
        if hasattr(self, 'match_window') and self.match_window.winfo_exists():
            self.match_window.destroy()

        self.match_window = tk.Toplevel(self.root)
        self.match_window.title("Re-analysis: Select Matches and Meanings")
        self.match_window.configure(bg='light gray')
        try:
            self._wm_for(self.match_window).maximize()
        except Exception:
            pass

        self.match_vars = []
        self.meaning_vars = []
        unique_matches = self.filter_unique_matches(matches)
        self.all_matches.append(unique_matches)

        self.current_reanalysis_index.append(index)
        
        # Display Pankti
        self.display_pankti_with_highlight(self.match_window, pankti, index)

        # --- Explanation Section ---
        explanation_frame = tk.Frame(self.match_window, bg='AntiqueWhite', 
                                    relief='groove', bd=2)  # A tinted frame with a grooved border
        explanation_frame.pack(fill=tk.X, padx=20, pady=(5, 10))

        heading_label = tk.Label(
            explanation_frame, 
            text="Important Note", 
            font=("Arial", 14, 'bold'),
            bg='AntiqueWhite'
        )
        heading_label.pack(pady=(5, 0))

        explanation_text = (
            "ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¢ Highlighted selections (displayed in MistyRose) indicate the meanings or grammar rules that "
            "were previously confirmed in your assessment.\n"
            "ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¢ This helps you quickly recognize which items reflect your earlier choices."
        )

        body_label = tk.Label(
            explanation_frame, 
            text=explanation_text,
            bg='AntiqueWhite', 
            fg='black', 
            font=('Arial', 12),
            wraplength=900,    # Adjust wrap length to your windowÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â‚¬Å¾Ã‚Â¢s width
            justify=tk.LEFT
        )
        body_label.pack(pady=(0, 10), padx=10)

        # Main layout
        main_frame = tk.Frame(self.match_window, bg='light gray')
        main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)

        # --- Left Pane: Meanings ---
        word = self.pankti_words[index]
        self.display_meanings_section_reanalysis(main_frame, word, index, meanings)

        # --- Right Pane: Matches ---
        self.display_matches_section_reanalysis(main_frame, unique_matches, index, max_display)

        # --- Bottom Buttons ---
        button_frame = tk.Frame(self.match_window, bg='light gray')
        button_frame.pack(pady=10)

        tk.Button(
            button_frame,
            text="Submit",
            command=self.submit_matches_reanalysis,
            font=('Arial', 12, 'bold'), bg='navy', fg='white',
            padx=20, pady=10
        ).pack(side=tk.LEFT, padx=5)

        tk.Button(
            button_frame,
            text="Back",
            command=lambda: self.back_to_user_input_reanalysis(pankti, index),
            font=('Arial', 12, 'bold'), bg='navy', fg='white',
            padx=20, pady=10
        ).pack(side=tk.LEFT, padx=5)

    def display_pankti_with_highlight(self, parent, pankti, index):
        """
        Displays the full pankti and highlights the word at the given index in a Text widget.
        """
        pankti_frame = tk.Frame(parent, bg='light gray')
        pankti_frame.pack(fill=tk.BOTH, padx=30, pady=20)

        display = tk.Text(
            pankti_frame,
            wrap=tk.WORD,
            bg='light gray',
            font=('Arial', 32),
            height=1,
            padx=5,
            pady=5
        )
        display.pack(fill=tk.BOTH, expand=False)
        display.insert(tk.END, pankti)
        display.tag_add("center", "1.0", "end")
        display.tag_configure("center", justify='center')

        words = pankti.split()
        start_idx = 0
        for i, w in enumerate(words):
            if i == index:
                break
            start_idx += len(w) + 1
        end_idx = start_idx + len(words[index])

        display.tag_add("highlight", f"1.{start_idx}", f"1.{end_idx}")
        display.tag_config("highlight", foreground="blue", font=('Arial', 32, 'bold'))
        display.config(state=tk.DISABLED)

    def display_meanings_section_reanalysis(self, parent_frame, word, index, meanings):
        """Display meanings as checkboxes for reanalysis with prior selection support."""
        meanings_frame = tk.Frame(parent_frame, bg='light gray')
        meanings_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=10)

        tk.Label(meanings_frame, text=f"Select Meanings for {word}:", bg='light gray',
                font=('Arial', 14, 'bold')).pack(pady=10)

        self.select_all_meanings_var = tk.BooleanVar(value=True)
        tk.Checkbutton(meanings_frame, text="Select/Deselect All Meanings",
                    variable=self.select_all_meanings_var, bg='light gray',
                    font=('Arial', 12), command=self.toggle_all_meanings).pack(pady=5)

        meanings_canvas = tk.Canvas(meanings_frame, bg='light gray', borderwidth=0)
        meanings_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar = tk.Scrollbar(meanings_frame, orient=tk.VERTICAL, command=meanings_canvas.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        meanings_canvas.config(yscrollcommand=scrollbar.set)

        inner_frame = tk.Frame(meanings_canvas, bg='light gray')
        meanings_canvas.create_window((0, 0), window=inner_frame, anchor='nw')

        # Merge past meanings
        first_index = next((i for i, w in enumerate(self.pankti_words) if w == word), index)
        merged_meanings = []
        for idx in range(first_index, index):
            if idx < len(self.accumulated_meanings):
                entry = self.accumulated_meanings[idx]
                if isinstance(entry, dict):
                    merged_meanings.extend(entry.get("meanings", []))
        prior_meanings = list(dict.fromkeys(merged_meanings))

        # Extract assessment-specific selected meanings from past_word_details.
        # These are the meanings the user had previously selected (from assessment).
        assessment_meanings = self.past_word_details.get(index, {}).get("darpan_meanings", [])

        # Reorder current meanings
        if isinstance(meanings, dict):
            current_meanings = meanings.get("meanings", [])
        else:
            current_meanings = meanings
            
        # Tier 1: assessment_meanings at very top
        reordered_assessment = [m for m in current_meanings if m in assessment_meanings]

        # Tier 2: prior_meanings next
        reordered_prior = [m for m in current_meanings 
                        if (m in prior_meanings and m not in assessment_meanings)]

        # Tier 3: everything else
        reordered_others = [m for m in current_meanings
                            if m not in prior_meanings and m not in assessment_meanings]

        reordered = reordered_assessment + reordered_prior + reordered_others

        split = self.split_meanings_for_display(reordered)
        self.meaning_vars = []

        for i, column in enumerate(split.values()):
            col_frame = tk.Frame(inner_frame, bg='light gray')
            col_frame.grid(row=0, column=i, padx=10, pady=10, sticky='nw')
            for meaning in column:
                # Determine whether this meaning was previously chosen during
                # the earlier assessment.  Those meanings should stand out in
                # MistyRose so that the user can easily recognise them when
                # reÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‹Å“analysing a word (mirroring the behaviour of grammar
                # rule highlighting).
                highlight = (meaning in assessment_meanings)

                # Default selection ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ for reÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‹Å“analysis we only preÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‹Å“select a
                # meaning if it was explicitly chosen earlier.  Previously the
                # first occurrence of a word had every meaning preÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‹Å“selected
                # which made it difficult to spot the assessed choice.  By
                # limiting the preÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‹Å“selection to the highlighted meanings we
                # keep the focus on what was actually picked before.
                if index != first_index:
                    preselect = (meaning in prior_meanings) or highlight
                else:
                    preselect = highlight

                # Apply the MistyRose background when the meaning was part of
                # the previous assessment; otherwise fall back to light gray.
                bg_color = "MistyRose" if highlight else "light gray"

                var = tk.BooleanVar(value=preselect)
                chk = tk.Checkbutton(
                    col_frame,
                    text=f"- {meaning}",
                    variable=var,
                    bg=bg_color,
                    font=('Arial', 12),
                    wraplength=325,
                    anchor='w',
                    justify=tk.LEFT,
                    selectcolor='light blue',
                )
                chk.pack(anchor='w', padx=15, pady=5)
                self.meaning_vars.append((var, meaning))

        inner_frame.update_idletasks()
        meanings_canvas.config(scrollregion=meanings_canvas.bbox("all"))

    def display_matches_section_reanalysis(self, parent_frame, unique_matches, index, max_display=30):
        """Display matching rule checkboxes in the reanalysis pane."""
        matches_frame = tk.Frame(parent_frame, bg='light gray')
        matches_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=10)

        tk.Label(matches_frame, text="Select the matching rules:",
                bg='light gray', font=('Arial', 14, 'bold')).pack(pady=10)

        canvas = tk.Canvas(matches_frame, bg='light gray', borderwidth=0)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar = tk.Scrollbar(matches_frame, orient=tk.VERTICAL, command=canvas.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        canvas.config(yscrollcommand=scrollbar.set)

        inner_frame = tk.Frame(canvas, bg='light gray')
        canvas.create_window((0, 0), window=inner_frame, anchor='nw')

        self.match_vars = []  # Reset match_vars in reanalysis flow

        # Retrieve the prior assessment grammar details for this word occurrence.
        grammar_assessment = self.past_word_details.get(index, {})
        assessment_fields = self.extract_grammar_fields(grammar_assessment)
        
        # For each match, we assume a tuple (label, value). In some cases,
        # label might be a composite string (e.g. with fields joined by " | ").
        # We reorder or highlight based on whether the extracted values match.
        reordered_matches = unique_matches[:max_display]  # (Assume prior reordering if needed)

        for match in reordered_matches:
            field_label, match_value = match[0], match[1]

            # If the field label is composite, parse it.
            if " | " in field_label:
                parsed = self.parse_composite(field_label)
                # Check for each target field whether the parsed value matches the assessment.
                highlight = True
                for key, expected in assessment_fields.items():
                    if key in parsed:
                        if not self.safe_equal_matches_reanalysis(parsed[key], expected):
                            highlight = False
                            break
                bg_color = "MistyRose" if highlight else "light gray"
            else:
                # Otherwise, if the label is a single field name, use a simple check.
                if field_label in assessment_fields and assessment_fields[field_label] == match_value:
                    bg_color = "MistyRose"
                else:
                    bg_color = "light gray"

            var = tk.BooleanVar()
            chk = tk.Checkbutton(inner_frame,
                                text=f"{field_label}: {match_value}",
                                variable=var,
                                bg=bg_color,
                                font=('Arial', 12),
                                selectcolor='light blue',
                                anchor='w')
            chk.pack(fill=tk.X, padx=10, pady=5)
            self.match_vars.append((var, match))

        inner_frame.update_idletasks()
        canvas.config(scrollregion=canvas.bbox("all"))

    def safe_equal_matches_reanalysis(self, val1, val2):
        def normalize(v):
            # Convert real NaN to empty
            if pd.isna(v):
                return ""
            # Turn into string, strip whitespace, and treat "NA" (any case) as empty
            s = str(v).strip()
            return "" if s.upper() == "NA" else s
        return normalize(val1) == normalize(val2)

    def extract_grammar_fields(self, grammar_assessment):
        """
        Extract only the fields we wish to highlight from grammar_assessment.
        Returns a dict with keys:
        - "Vowel Ending"
        - "Number / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨"
        - "Grammar / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¯ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£"
        - "Gender / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â"
        - "Word Root"
        - "Word Type"
        """
        target_keys = ["\ufeffVowel Ending", "Number / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨", "Grammar / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¯ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£",
                    "Gender / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â", "Word Root", "Type"]
        return {key: grammar_assessment.get(key) for key in target_keys}

    def parse_composite(self, label):
        """
        Assume a composite label is built by joining fields with " | ".
        This function splits the composite string into its individual parts
        and returns a dictionary mapping (in order) the following keys:
        "Word", "Vowel Ending", "Number / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨", "Grammar / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¯ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£",
        "Gender / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â", "Word Root", "Type"
        """
        parts = label.split(" | ")
        keys = ["Word", "\ufeffVowel Ending", "Number / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨", "Grammar / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¯ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£",
                "Gender / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â", "Word Root", "Type"]
        return dict(zip(keys, parts))

    def back_to_user_input_reanalysis(self, pankti, index):
        """
        Returns to the grammar input screen for the current word during re-analysis.
        """
        try:
            if hasattr(self, 'match_window') and self.match_window:
                self.match_window.destroy()

            if 0 <= index < len(self.pankti_words):
                word = self.pankti_words[index]
                self.current_word_index = index
                self.reset_input_variables()
                self.user_input_reanalysis(word, pankti, index)
            else:
                messagebox.showerror("Invalid Index", "Cannot return to word ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â index out of range.")

        except Exception as e:
            messagebox.showerror("Error", f"An error occurred while going back: {e}")

    def submit_matches_reanalysis(self):
        any_selection = False
        current_entries = []

        for var, match in self.match_vars:
            if var.get():
                match_string = match[0]
                data = match_string.split(" | ")
                new_entry = {
                    "Word": data[0],
                    "\ufeffVowel Ending": data[1],
                    "Number / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨": data[2],
                    "Grammar / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¯ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£": data[3],
                    "Gender / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â": data[4],
                    "Word Root": data[5],
                    "Type": data[6]
                }
                current_entries.append(new_entry)
                any_selection = True

        selected_meanings = [meaning for var, meaning in self.meaning_vars if var.get()]
        self.accumulate_meanings_data(selected_meanings)

        current_word = self.pankti_words[self.current_word_index]
        first_index = next((i for i, w in enumerate(self.pankti_words) if w == current_word), self.current_word_index)

        current_selected = selected_meanings
        if self.current_word_index != first_index:
            merged_meanings = []
            for idx in range(first_index, self.current_word_index):
                if idx < len(self.accumulated_meanings) and self.pankti_words[idx] == current_word:
                    entry = self.accumulated_meanings[idx]
                    if isinstance(entry, dict):
                        merged_meanings.extend(entry.get("meanings", []))
                    else:
                        merged_meanings.extend(entry)
            prior_meanings = list(dict.fromkeys(merged_meanings))

            if set(current_selected) != set(prior_meanings):
                update_prev = messagebox.askyesno(
                    "Update Previous Meanings",
                    f"You have selected different meanings for the word '{current_word}'.\n"
                    "Do you want to update the meanings for all previous occurrences of this word?"
                )
                if update_prev:
                    for idx in range(first_index, self.current_word_index):
                        if idx < len(self.accumulated_meanings) and self.pankti_words[idx] == current_word:
                            self.accumulated_meanings[idx] = {"word": current_word, "meanings": current_selected}

        if self.current_word_index < len(self.accumulated_meanings):
            self.accumulated_meanings[self.current_word_index] = {"word": current_word, "meanings": current_selected}
        else:
            self.accumulated_meanings.append({"word": current_word, "meanings": current_selected})

        # Assign the current verse using index-to-verse mapping
        verse_boundaries = []
        pointer = 0
        for verse in self.selected_verses:
            verse_words = verse.split()
            start = pointer
            end = pointer + len(verse_words)
            verse_boundaries.append((start, end))
            pointer = end

        current_verse = None
        for i, (start, end) in enumerate(verse_boundaries):
            if start <= self.current_word_index < end:
                current_verse = self.selected_verses[i]
                break

        for entry in current_entries:
            entry["Verse"] = current_verse
            entry["Word Index"] = self.current_word_index

        finalized_matches = []
        for var, match in self.match_vars:
            if var.get():
                match_word = match[0].split(" | ")[0]
                for entry in current_entries:
                    if entry["Word"] == match_word and entry not in finalized_matches:
                        finalized_matches.append(entry)
        self.accumulate_finalized_matches(finalized_matches)

        if not any_selection:
            messagebox.showwarning("No Selection", "No matches were selected. Please select at least one match.")
        else:
            self.match_window.destroy()
            self.all_new_entries.extend(current_entries)
            self.process_next_selected_word()

    def prompt_save_results_reanalysis(self, new_entries, skip_copy=False):
        file_path = "1.2.1 assessment_data.xlsx"
        existing_data = self.load_existing_assessment_data(file_path)
        original_accumulated_pankti = self.accumulated_pankti

        for verse in self.selected_verses:
            self.accumulated_pankti = verse
            current_verse_words = verse.replace('ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¥Ãƒâ€šÃ‚Â¥', '').split()
            selected_words = set(current_verse_words)

            # Filter grammar entries specific to this verse
            # now you can pick only the entries for that exact wordÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Âindex
            filtered_new_entries = [
                entry for entry in new_entries
                if entry.get("Verse", "").strip() == verse.strip()
                and entry.get("Word Index") in self.current_reanalysis_index
            ]

            # Silently remove exact duplicates based on your key fields
            seen = set()
            unique_entries = []

            keys = [
                "Word", "\ufeffVowel Ending", "Number / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨", "Grammar / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¯ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£",
                "Gender / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â", "Word Root", "Type", "Verse", 'Word Index'
            ]

            for entry in filtered_new_entries:
                # build a normalized tuple of comparison values
                key = tuple(self.normalize_save_results_reanalysis(entry.get(k)) for k in keys)
                if key not in seen:
                    seen.add(key)
                    unique_entries.append(entry)

            if not skip_copy:
                self.prompt_copy_to_clipboard_reanalysis()

            if unique_entries:
                save = messagebox.askyesno("Save Results", f"Would you like to save the new entries for the following verse?\n\n{verse}")
                if save:
                    assessment_data = self.prompt_for_assessment_once_reanalysis()

                    # Extract verse metadata from chosen_match if available
                    verse_to_match = self.accumulated_pankti.strip()
                    candidate = None
                    if hasattr(self, 'candidate_matches') and self.chosen_match:
                        for cand in self.chosen_match:
                            if cand.get("Verse", "").strip() == verse_to_match:
                                candidate = cand
                                break
                        if candidate is None:
                            candidate = self.chosen_match[0]

                        verse_metadata = {
                            "Verse": verse_to_match,
                            "S. No.": candidate.get("S. No.", ""),
                            "Verse No.": candidate.get("Verse No.", ""),
                            "Stanza No.": candidate.get("Stanza No.", ""),
                            "Text Set No.": candidate.get("Text Set No.", ""),
                            "Raag (Fixed)": candidate.get("Raag (Fixed)", ""),
                            "Sub-Raag": candidate.get("Sub-Raag", ""),
                            "Writer (Fixed)": candidate.get("Writer (Fixed)", ""),
                            "Verse Configuration (Optional)": candidate.get("Verse Configuration (Optional)", ""),
                            "Stanza Configuration (Optional)": candidate.get("Stanza Configuration (Optional)", ""),
                            "Bani Name": candidate.get("Bani Name", ""),
                            "Musical Note Configuration": candidate.get("Musical Note Configuration", ""),
                            "Special Type Demonstrator": candidate.get("Special Type Demonstrator", ""),
                            "Verse Type": candidate.get("Verse Type", ""),
                            "Page Number": candidate.get("Page Number", "")
                        }
                    else:
                        verse_metadata = {}

                    # Group grammar entries by word and partition by occurrence index
                    word_groups = {}
                    for entry in unique_entries:
                        word = entry["Word"]
                        word_groups.setdefault(word, []).append(entry)

                    final_entries = []
                    occurrence_mapping = {}
                    for word in set(current_verse_words):
                        count = current_verse_words.count(word)
                        if word not in word_groups:
                            continue
                        entries_list = word_groups[word]
                        n = len(entries_list)
                        k = count
                        group_size = n // k
                        remainder = n % k
                        start = 0
                        groups = []
                        for i in range(k):
                            size = group_size + (1 if i < remainder else 0)
                            group = entries_list[start:start+size]
                            groups.append(group)
                            start += size

                        occurrence_positions = [i for i, w in enumerate(current_verse_words) if w == word]
                        for occ, pos in zip(range(k), occurrence_positions):
                            occurrence_mapping[(word, pos)] = groups[occ]

                    for idx, word in enumerate(current_verse_words):
                        key = (word, idx)
                        entries = occurrence_mapping.get(key, [])
                        if not entries:
                            continue

                        seen = set()
                        dedup_entries = []
                        for entry in entries:
                            entry_tuple = tuple(sorted(entry.items()))
                            if entry_tuple not in seen:
                                seen.add(entry_tuple)
                                dedup_entries.append(entry)
                        entries = dedup_entries

                        if len(entries) > 1:
                            chosen_entry = self.prompt_for_final_grammar_reanalysis(entries)
                        else:
                            chosen_entry = entries[0]

                        chosen_entry['Word Index'] = idx

                        if len(self.accumulated_finalized_matches) <= idx:
                            self.accumulated_finalized_matches.extend([[]] * (idx - len(self.accumulated_finalized_matches) + 1))
                        self.accumulated_finalized_matches[idx] = [chosen_entry]
                        final_entries.append(chosen_entry)

                    for entry in final_entries:
                        entry.update(assessment_data)
                        entry.update(verse_metadata)
                        self.save_assessment_data_reanalysis(entry)

                    messagebox.showinfo("Saved", "Assessment data saved successfully for verse:\n" + verse)

        self.accumulated_pankti = original_accumulated_pankti

        if hasattr(self, 'copy_button') and self.copy_button.winfo_exists():
            self.copy_button.config(state=tk.NORMAL)

    def normalize_save_results_reanalysis(self, v):
        # Convert real NaN ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ ""
        if pd.isna(v):
            return ""
        # Convert None ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ ""
        if v is None:
            return ""
        # Convert the literal string "NA" (any case, with whitespace) ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ ""
        s = str(v).strip()
        return "" if s.upper() == "NA" else s

    def safe_equal_save_results_reanalysis(self, val1, val2):
        return self.normalize_save_results_reanalysis(val1) == self.normalize_save_results_reanalysis(val2)

    def prompt_copy_to_clipboard_reanalysis(self):
        print("Prompting to copy re-analysis to clipboard...")

        copy_prompt = messagebox.askyesno(
            "Copy to Clipboard",
            f"Would you like to copy the re-analysis for the verse '{self.accumulated_pankti}' to your clipboard?"
        )

        if not copy_prompt:
            return

        # Validation checks
        if not self.accumulated_pankti:
            messagebox.showerror("Error", "No verse (pankti) available to copy.")
            return

        if not self.accumulated_meanings or not self.accumulated_finalized_matches or not self.all_new_entries:
            messagebox.showerror("Error", "Missing re-analysis data to copy.")
            print("Error: Data is incomplete for clipboard copy.")
            return

        try:
            # Use the existing composing utility
            clipboard_text = self.compose_clipboard_text_for_chatgpt_reanalysis()
            pyperclip.copy(clipboard_text)
            messagebox.showinfo("Copied", "The re-analysis has been copied to the clipboard!")
            print("Clipboard content copied successfully.")

        except Exception as e:
            print(f"Unexpected error while copying reanalysis: {e}")
            messagebox.showerror("Error", f"Unexpected error occurred: {e}")

    def compose_clipboard_text_for_chatgpt_reanalysis(self):
        clipboard_text = "### Detailed Reanalysis & Literal Translation\n\n"
        clipboard_text += (
            f"The verse **'{self.accumulated_pankti}'** has undergone re-analysis. "
            "Below is a breakdown of each word with revised user-selected meanings and grammar details.\n\n"
        )

        # --- Preceding Verses & Translations ---
        existing_data = self.load_existing_assessment_data("1.2.1 assessment_data.xlsx")
        selected_columns = [
            'S. No.', 'Verse', 'Verse No.', 'Stanza No.', 'Text Set No.',
            'Raag (Fixed)', 'Sub-Raag', 'Writer (Fixed)', 'Verse Configuration (Optional)',
            'Stanza Configuration (Optional)', 'Bani Name', 'Musical Note Configuration',
            'Special Type Demonstrator', 'Verse Type', 'Page Number'
        ]
        df_filtered = existing_data[selected_columns]
        chosen_list = df_filtered[
            df_filtered["Verse"].str.strip() == self.accumulated_pankti.strip()
        ].drop_duplicates().to_dict(orient="records")
        current_candidate = chosen_list[0] if chosen_list else None

        preceding_verses_text = ""
        if current_candidate:
            text_set_no = current_candidate.get("Text Set No.")
            try:
                current_verse_no = int(current_candidate.get("Verse No."))
            except (ValueError, TypeError):
                current_verse_no = None

            if current_verse_no is not None:
                filtered_data = existing_data[existing_data["Text Set No."] == text_set_no]
                consecutive_verses = []
                target_verse_no = current_verse_no - 1
                while True:
                    row = filtered_data[filtered_data["Verse No."] == target_verse_no]
                    if row.empty:
                        break
                    row_data = row.iloc[0]
                    consecutive_verses.insert(0, row_data)
                    target_verse_no -= 1

                if consecutive_verses:
                    preceding_verses_text += "\n### Preceding Verses & Translations\n\n"
                    for row_data in consecutive_verses:
                        verse_no = row_data.get("Verse No.", "")
                        verse_text = row_data.get("Verse", "")
                        translation = row_data.get("Translation", "")
                        preceding_verses_text += f"**Verse {verse_no}:** {verse_text}\n"
                        preceding_verses_text += f"**Translation:** {translation}\n\n"

        clipboard_text += preceding_verses_text

        # --- Past Translation of the Current Verse ---
        past_translation_text = ""
        if current_candidate:
            text_set_no = current_candidate.get("Text Set No.")
            try:
                current_verse_no = int(current_candidate.get("Verse No."))
            except (ValueError, TypeError):
                current_verse_no = None

            if current_verse_no is not None:
                filtered_data = existing_data[existing_data["Text Set No."] == text_set_no]
                # Instead of iterating for preceding verses, we directly extract the row for the current verse.
                row = filtered_data[filtered_data["Verse No."] == current_verse_no]
                if not row.empty:
                    row_data = row.iloc[0]
                    verse_text = row_data.get("Verse", "")
                    translation = row_data.get("Translation", "")
                    past_translation_text += "\n### Past Translation of the Current Verse\n\n"
                    past_translation_text += f"**Verse {current_verse_no}:** {verse_text}\n"
                    past_translation_text += f"**Translation:** {translation}\n\n"

        clipboard_text += past_translation_text

        current_verse_words = self.accumulated_pankti.split()

        def find_sublist_index(haystack, needle):
            for i in range(len(haystack) - len(needle) + 1):
                if haystack[i:i + len(needle)] == needle:
                    return i
            return -1

        start_index = find_sublist_index(self.pankti_words, current_verse_words)
        if start_index == -1:
            start_index = 0

        for i, word in enumerate(current_verse_words):
            actual_index = start_index + i
            clipboard_text += f"**Word {i + 1}: {word}**\n"

            acc_entry = self.accumulated_meanings[actual_index] if actual_index < len(self.accumulated_meanings) else {}
            meanings_list = acc_entry.get("meanings", []) if isinstance(acc_entry, dict) else acc_entry
            meanings_str = ", ".join(meanings_list) if meanings_list else "No user-selected meanings available"
            clipboard_text += f"- **User-Selected Meanings:** {meanings_str}\n"

            # --- Past Assessment Details ---
            assessment_details = self.past_word_details.get(actual_index, {})
            if assessment_details:
                clipboard_text += "- **Past Assessment Details:**\n"
                # Display past meanings if available
                past_meanings = assessment_details.get("darpan_meanings", [])
                if past_meanings:
                    clipboard_text += f"   - **Past Meanings:** {', '.join(past_meanings)}\n"
                # Display grammar fields
                clipboard_text += f"   - **Vowel Ending:** {self._norm_get(assessment_details, '\\ufeffVowel Ending') or 'N/A'}\n"
                clipboard_text += f"   - **Number / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨:** {assessment_details.get('Number / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨', 'N/A')}\n"
                clipboard_text += f"   - **Grammar / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¯ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£:** {assessment_details.get('Grammar / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¯ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£', 'N/A')}\n"
                clipboard_text += f"   - **Gender / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â:** {assessment_details.get('Gender / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â', 'N/A')}\n"
                clipboard_text += f"   - **Word Root:** {assessment_details.get('Word Root', 'N/A')}\n"
                clipboard_text += f"   - **Word Type:** {self._norm_get(assessment_details, 'Type') or 'N/A'}\n"

            # --- Grammar Options ---
            clipboard_text += "- **Grammar Options:**\n"
            finalized_matches_list = self.accumulated_finalized_matches[actual_index] if actual_index < len(self.accumulated_finalized_matches) else []

            if finalized_matches_list:
                for option_idx, match in enumerate(finalized_matches_list, start=1):
                    clipboard_text += (
                        f"  - **Option {option_idx}:**\n"
                        f"      - **Word:** {self._norm_get(match, 'Word') or 'N/A'}\n"
                        f"      - **Vowel Ending:** {self._norm_get(match, '\\ufeffVowel Ending') or 'N/A'}\n"
                        f"      - **Number / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨:** {match.get('Number / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨', 'N/A')}\n"
                        f"      - **Grammar / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¯ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£:** {match.get('Grammar / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¯ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£', 'N/A')}\n"
                        f"      - **Gender / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â:** {match.get('Gender / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â', 'N/A')}\n"
                        f"      - **Word Root:** {match.get('Word Root', 'N/A')}\n"
                        f"      - **Type:** {self._norm_get(match, 'Type') or 'N/A'}\n"
                        f"      - **Literal Translation (Option {option_idx}):** The word '{word}' functions as a "
                        f"'{self._norm_get(match, 'Type') or 'N/A'}' with '{match.get('Grammar / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¯ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£', 'N/A')}' usage, "
                        f"in the '{match.get('Number / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨', 'N/A')}' form and '{match.get('Gender / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â', 'N/A')}' gender. Translation: ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¦\n"
                    )
            else:
                clipboard_text += "  - No finalized grammar options available\n"

            clipboard_text += "\n"

        if 'ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¥Ãƒâ€šÃ‚Â¥' in current_verse_words:
            clipboard_text += (
                "**Symbol:** ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¥Ãƒâ€šÃ‚Â¥\n"
                "- **Meaning:** End of verse or sentence\n"
                "- **Context:** Denotes the conclusion of the verse.\n\n"
            )

        clipboard_text += "\n### Literal Translation Prompt\n"
        clipboard_text += (
            f"Using the above user-selected meanings and grammar details for the verse '{self.accumulated_pankti}', "
            "please generate a literal translation that adheres strictly to the grammatical structure, "
            "capturing the tense, number, gender, and function accurately."
        )

        return clipboard_text

    def prompt_for_assessment_once_reanalysis(self):
        """Opens a modal prompt for re-analysis of the entire verse and returns the collected assessment data."""
        assessment_win = tk.Toplevel(self.root)
        assessment_win.title(f"Re-Assessment: '{self.accumulated_pankti}'")
        assessment_win.configure(bg='light gray')
        try:
            self._wm_for(assessment_win)
        except Exception:
            pass

        instruction_label = tk.Label(
            assessment_win,
            text="Paste the updated translation or assessment for this verse:",
            font=("Helvetica", 14), bg="light gray"
        )
        instruction_label.pack(pady=10)

        analysis_text = scrolledtext.ScrolledText(
            assessment_win, width=80, height=10,
            font=("Helvetica", 12), wrap=tk.WORD
        )
        analysis_text.pack(padx=20, pady=10)

        cb_frame = tk.Frame(assessment_win, bg="light gray")
        cb_frame.pack(pady=10)

        framework_var = tk.BooleanVar()
        explicit_var = tk.BooleanVar()

        framework_cb = tk.Checkbutton(cb_frame, text="Framework?", variable=framework_var,
                                    font=("Helvetica", 12), bg="light gray")
        framework_cb.pack(side=tk.LEFT, padx=10)

        explicit_cb = tk.Checkbutton(cb_frame, text="Explicit?", variable=explicit_var,
                                    font=("Helvetica", 12), bg="light gray")
        explicit_cb.pack(side=tk.LEFT, padx=10)

        assessment_data = {}

        def on_save():
            translation = analysis_text.get("1.0", tk.END).strip()
            if not translation:
                messagebox.showerror("Error", "Please provide the revised assessment or translation.")
                return
            assessment_data["Translation"] = translation
            assessment_data["Framework?"] = framework_var.get()
            assessment_data["Explicit?"] = explicit_var.get()
            assessment_win.destroy()

        save_btn = tk.Button(
            assessment_win, text="Save Re-Assessment",
            command=on_save, font=("Helvetica", 14, "bold"),
            bg="#2a7b39", fg="white", padx=20, pady=10
        )
        save_btn.pack(pady=20)

        assessment_win.transient(self.root)
        assessment_win.grab_set()
        self.root.wait_window(assessment_win)

        return assessment_data

    def prompt_for_final_grammar_reanalysis(self, word_entries):
        """
        Opens a modal window to finalize grammar during reanalysis.
        Displays a structured prompt for ChatGPT and allows selection from available grammar options.
        """
        final_choice = {}

        final_win = tk.Toplevel(self.root)
        final_win.title(f"Reanalysis: Finalize Grammar for '{word_entries[0]['Word']}'")
        final_win.configure(bg='light gray')
        try:
            self._wm_for(final_win)
        except Exception:
            pass

        # --- Build prompt for clipboard ---
        prompt_lines = [
            f"Finalize the applicable grammar for the word: {word_entries[0]['Word']}"
        ]
        prompt_lines.append("The following grammar options are available:")

        fields = [
            "\ufeffVowel Ending", "Number / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨", "Grammar / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¯ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£",
            "Gender / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â",   "Word Root", "Type"
        ]

        for idx, entry in enumerate(word_entries, start=1):
            # coerce each field to str, converting NaN ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ ""
            parts = []
            for f in fields:
                val = self._norm_get(entry, f) or ""
                if pd.isna(val):
                    val = ""
                parts.append(str(val))
            summary = " | ".join(parts)

            prompt_lines.append(f"Option {idx}: {summary}")

        prompt_text = "\n".join(prompt_lines)

        # --- Prompt display with copy ---
        prompt_frame = tk.Frame(final_win, bg="light gray")
        prompt_frame.pack(padx=20, pady=10, fill=tk.BOTH, expand=True)

        tk.Label(prompt_frame,
                text="ChatGPT Prompt for Grammar Reassessment:",
                font=("Helvetica", 14, "bold"),
                bg="light gray").pack(anchor="w", pady=(0, 5))

        prompt_text_widget = scrolledtext.ScrolledText(prompt_frame, width=80, height=6,
                                                    font=("Helvetica", 12), wrap=tk.WORD)
        prompt_text_widget.pack(fill=tk.BOTH, expand=True)
        prompt_text_widget.insert(tk.END, prompt_text)
        prompt_text_widget.config(state=tk.DISABLED)

        def copy_prompt():
            self.root.clipboard_clear()
            self.root.clipboard_append(prompt_text)
            messagebox.showinfo("Copied", "Prompt text copied to clipboard!")

        tk.Button(prompt_frame, text="Copy Prompt", command=copy_prompt,
                font=("Helvetica", 12, "bold"), bg="#007acc", fg="white", padx=10, pady=5
                ).pack(anchor="e", pady=5)

        # --- Instruction and option selection ---
        tk.Label(final_win,
                text="Please select the correct grammar from the following options:",
                font=("Helvetica", 14), bg="light gray").pack(pady=10)

        choice_var = tk.IntVar(value=0)
        options_container = tk.Frame(final_win, bg="light gray")
        options_container.pack(padx=20, pady=10, fill=tk.BOTH, expand=True)

        canvas = tk.Canvas(options_container, bg="light gray", highlightthickness=0)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vsb = tk.Scrollbar(options_container, orient="vertical", command=canvas.yview)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        canvas.configure(yscrollcommand=vsb.set)

        options_frame = tk.Frame(canvas, bg="light gray")
        canvas.create_window((0, 0), window=options_frame, anchor="nw")

        def on_frame_configure(event):
            canvas.configure(scrollregion=canvas.bbox("all"))
        options_frame.bind("<Configure>", on_frame_configure)

        def as_str(val):
            # turn real NaN ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ "" and everything else ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ string
            return "" if pd.isna(val) else str(val)

        fields = [
            "\ufeffVowel Ending", "Number / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨", "Grammar / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¯ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£",
            "Gender / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â",   "Word Root", "Type"
        ]

        for idx, entry in enumerate(word_entries):
            summary = " | ".join(
                as_str(self._norm_get(entry, f) or "") for f in fields
            )

            tk.Radiobutton(options_frame,
                        text=f"Option {idx+1}: {summary}",
                        variable=choice_var,
                        value=idx,
                        bg="light gray",
                        font=("Helvetica", 12),
                        anchor='w',
                        justify=tk.LEFT,
                        selectcolor='light blue'
                        ).pack(anchor="w", padx=10, pady=5)

        def on_save():
            selected_index = choice_var.get()
            nonlocal final_choice
            final_choice = word_entries[selected_index]
            final_win.destroy()

        tk.Button(final_win, text="Save Choice",
                command=on_save,
                font=("Helvetica", 14, "bold"),
                bg="#2a7b39", fg="white", padx=20, pady=10).pack(pady=20)

        final_win.transient(self.root)
        final_win.grab_set()
        self.root.wait_window(final_win)
        return final_choice

    def save_assessment_data_reanalysis(self, new_entry):
        """
        Saves a new re-analysis entry to the Excel file.
        Handles grammar and translation revision tracking for specific word occurrences.
        """
        file_path = "1.2.1 assessment_data.xlsx"
        df_existing = self.load_existing_assessment_data(file_path)

        grammar_keys = [
            '\ufeffVowel Ending', 'Number / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨', 'Grammar / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¯ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£',
            'Gender / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â', 'Word Root', 'Type'
        ]

        # Update translation for all entries of the same verse
        df_existing.loc[df_existing["Verse"] == new_entry["Verse"], "Translation"] = new_entry["Translation"]

        # Locate matching entries
        matching_rows = df_existing[
            (df_existing["Word"] == new_entry["Word"]) &
            (df_existing["Verse"] == new_entry["Verse"]) &
            (df_existing["Word Index"] == new_entry["Word Index"])
        ]

        if not matching_rows.empty:
            latest_idx = matching_rows["Grammar Revision"].idxmax()
            latest_row = df_existing.loc[latest_idx]
            differences = any(new_entry.get(key) != self._norm_get(latest_row, key) for key in grammar_keys)

            if differences:
                new_revision = matching_rows["Grammar Revision"].max() + 1
                new_entry["Grammar Revision"] = new_revision

                for idx in matching_rows.index:
                    for key in grammar_keys:
                        df_existing.at[idx, key] = new_entry.get(key)
                    df_existing.at[idx, "Grammar Revision"] = new_revision

                    for key, value in new_entry.items():
                        if key not in grammar_keys and key != "Translation":
                            if key in ("Framework?", "Explicit?"):
                                df_existing.at[idx, key] = int(value)
                            else:
                                df_existing.at[idx, key] = value

                # Update Selected Darpan Meaning
                global_index = new_entry.get("Word Index", 0)
                if len(self.accumulated_meanings) > global_index:
                    acc_entry = self.accumulated_meanings[global_index]
                    selected_meaning = "| ".join(acc_entry.get("meanings", [])) if isinstance(acc_entry, dict) else ", ".join(acc_entry)
                    for idx in matching_rows.index:
                        df_existing.at[idx, "Selected Darpan Meaning"] = selected_meaning

                # Update Translation Revision
                verse_mask = df_existing["Verse"] == new_entry["Verse"]
                latest_revision = df_existing.loc[verse_mask, "Grammar Revision"].max()
                df_existing.loc[verse_mask, "Translation Revision"] = latest_revision
            else:
                return  # No changes, skip saving
        else:
            new_entry["Grammar Revision"] = 1
            current_revision = df_existing[df_existing["Verse"] == new_entry["Verse"]]["Translation Revision"].max()
            new_entry["Translation Revision"] = (current_revision + 1) if not pd.isna(current_revision) else 1

            # Set Selected Darpan Meaning
            global_index = new_entry.get("Word Index", 0)
            if len(self.accumulated_meanings) > global_index:
                acc_entry = self.accumulated_meanings[global_index]
                selected_meaning = "| ".join(acc_entry.get("meanings", [])) if isinstance(acc_entry, dict) else ", ".join(acc_entry)
                new_entry["Selected Darpan Meaning"] = selected_meaning

            df_existing = pd.concat([df_existing, pd.DataFrame([new_entry])], ignore_index=True)

        try:
            df_existing.to_excel(file_path, index=False)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save reanalysis data: {e}")
            return

        # Also reflect reanalysis metadata in the Assess-by-Word tracker (increment count, set timestamp)
        try:
            tracker_path = self._get_word_tracker_path()
            words_df, prog_df, others = load_word_tracker(tracker_path, TRACKER_WORDS_SHEET, TRACKER_PROGRESS_SHEET)
            if not prog_df.empty:
                # Build a mask to identify the specific verse entry for this word
                try:
                    word_raw = new_entry.get("Word", "")
                    verse_raw = new_entry.get("Verse", "")
                    idx_raw = new_entry.get("Word Index", None)

                    def _norm(s):
                        try:
                            return _normalize_simple(str(s))
                        except Exception:
                            return str(s)

                    mask_word = False
                    if 'word_key_norm' in prog_df.columns:
                        mask_word = prog_df.get('word_key_norm', pd.Series(dtype=str)).astype(str).map(_norm) == _norm(word_raw)
                    else:
                        mask_word = prog_df.get('word', pd.Series(dtype=str)).astype(str).map(_norm) == _norm(word_raw)

                    mask_verse = prog_df.get('verse', pd.Series(dtype=str)).astype(str).map(_norm) == _norm(verse_raw)
                    mask = mask_word & mask_verse
                    # If we have an index, further narrow by word_index where present
                    if idx_raw is not None and 'word_index' in prog_df.columns:
                        try:
                            idx_num = pd.to_numeric(prog_df['word_index'], errors='coerce')
                            mask = mask & (idx_num == pd.to_numeric(idx_raw, errors='coerce'))
                        except Exception:
                            pass

                    if mask.any():
                        # Increment reanalyzed_count and set last_reanalyzed_at
                        if 'reanalyzed_count' in prog_df.columns:
                            try:
                                cur = pd.to_numeric(prog_df.loc[mask, 'reanalyzed_count'], errors='coerce').fillna(0).astype(int)
                                prog_df.loc[mask, 'reanalyzed_count'] = (cur + 1).astype(int)
                            except Exception:
                                # Fallback: simple add with coercion
                                prog_df.loc[mask, 'reanalyzed_count'] = prog_df.loc[mask, 'reanalyzed_count']
                        if 'last_reanalyzed_at' in prog_df.columns:
                            from datetime import datetime as _dt
                            prog_df.loc[mask, 'last_reanalyzed_at'] = _dt.now()
                        _save_tracker(tracker_path, words_df, prog_df, others, TRACKER_WORDS_SHEET, TRACKER_PROGRESS_SHEET)
                except Exception:
                    pass
        except Exception:
            # Non-fatal to the reanalysis save flow
            pass

    def ensure_meanings_slot_initialized(self, index, word):
        """Ensure self.accumulated_meanings[index] is initialized with a valid structure."""
        while len(self.accumulated_meanings) <= index:
            self.accumulated_meanings.append({"word": None, "meanings": []})

        if not isinstance(self.accumulated_meanings[index], dict) or "word" not in self.accumulated_meanings[index]:
            self.accumulated_meanings[index] = {"word": word, "meanings": []}

    def setup_main_analysis_interface(self):
        """Builds the main analysis interface for Literal Meaning Analysis."""
        # Define a consistent color scheme
        BG_COLOR = "#f0f0f0"        # Light gray background for main area
        HEADER_COLOR = "#2c3e50"    # Dark slate-like header
        HEADER_TEXT_COLOR = "white"
        BUTTON_COLOR = "#007acc"    # A pleasant blue for action buttons
        BUTTON_TEXT_COLOR = "white"
        NAV_BUTTON_COLOR = "#5f9ea0" # CadetBlue for navigation
        LABEL_TEXT_COLOR = "#333333"
        
        # Define fonts
        TITLE_FONT = ("Helvetica", 20, "bold")
        LABEL_FONT = ("Helvetica", 14, "bold")
        ENTRY_FONT = ("Helvetica", 14)
        BUTTON_FONT = ("Helvetica", 14, "bold")
        RESULT_FONT = ("Helvetica", 12)

        # Clear existing widgets in case we re-enter
        for widget in self.root.winfo_children():
            widget.destroy()

        self.root.configure(bg=BG_COLOR)

        # Main frame (entire interface)
        self.main_frame = tk.Frame(self.root, bg=BG_COLOR, padx=20, pady=20)
        self.main_frame.pack(fill=tk.BOTH, expand=True)

        # Header Label
        header_label = tk.Label(
            self.main_frame,
            text="Grammar Analyzer",
            bg=HEADER_COLOR,
            fg=HEADER_TEXT_COLOR,
            font=TITLE_FONT,
            pady=10
        )
        header_label.pack(fill=tk.X, pady=(0, 20))

        # Label and Entry for Pankti
        self.pankti_label = tk.Label(
            self.main_frame,
            text="Please share the Pankti:",
            bg=BG_COLOR,
            fg=LABEL_TEXT_COLOR,
            font=LABEL_FONT
        )
        self.pankti_label.pack(anchor="w", pady=(0, 5))

        self.pankti_entry = tk.Entry(
            self.main_frame,
            width=70,
            font=ENTRY_FONT,
            bd=3,
            relief=tk.GROOVE
        )
        self.pankti_entry.pack(pady=(0, 15))

        # Analyze Button
        self.analyze_button = tk.Button(
            self.main_frame,
            text="Analyze",
            command=self.analyze_pankti,
            font=BUTTON_FONT,
            bg=BUTTON_COLOR,
            fg=BUTTON_TEXT_COLOR,
            relief=tk.RAISED,
            padx=30,
            pady=10
        )
        self.analyze_button.pack(pady=(0, 20))

        # Results Text Area
        self.results_text = scrolledtext.ScrolledText(
            self.main_frame,
            width=90,
            height=20,
            font=RESULT_FONT,
            bd=3,
            relief=tk.SUNKEN,
            wrap=tk.WORD
        )
        self.results_text.pack(pady=(0, 20))

        # Navigation Frame
        nav_frame = tk.Frame(self.main_frame, bg=BG_COLOR)
        nav_frame.pack(fill=tk.X, pady=10)

        # --- New Word Navigation Panel ---
        # A label to display the current word
        self.word_label = tk.Label(nav_frame, text="", font=("Helvetica", 16, "bold"),
                                bg="white", fg="black", width=20, relief=tk.SUNKEN)
        self.word_label.pack(side=tk.LEFT, padx=10)
        self.update_current_word_label()  # Update this label based on self.current_word_index

        # Previous button for word navigation
        self.prev_button = tk.Button(
            nav_frame,
            text="Previous",
            command=self.prev_word,
            state=tk.DISABLED,
            font=BUTTON_FONT,
            bg=NAV_BUTTON_COLOR,
            fg="white",
            padx=20,
            pady=5
        )
        self.prev_button.pack(side=tk.LEFT, padx=(0, 10))

        # Next button for word navigation
        self.next_button = tk.Button(
            nav_frame,
            text="Next",
            command=self.next_word,
            state=tk.DISABLED,
            font=BUTTON_FONT,
            bg=NAV_BUTTON_COLOR,
            fg="white",
            padx=20,
            pady=5
        )
        self.next_button.pack(side=tk.LEFT, padx=(0, 10))

        # Select Word button to analyze the currently displayed word
        select_word_btn = tk.Button(
            nav_frame,
            text="Select Word",
            command=self.select_current_word,
            font=BUTTON_FONT,
            bg=BUTTON_COLOR,
            fg=BUTTON_TEXT_COLOR,
            padx=20,
            pady=5
        )
        select_word_btn.pack(side=tk.LEFT, padx=(0, 10))
        # --- End Word Navigation Panel ---

        # Copy Analysis Button (initially disabled)
        self.copy_button = tk.Button(
            nav_frame,
            text="Copy Analysis",
            command=self.prompt_copy_to_clipboard,
            font=BUTTON_FONT,
            bg="#1b95e0",   # A slightly different blue for variety
            fg="white",
            padx=20,
            pady=5,
            state=tk.DISABLED
        )
        self.copy_button.pack(side=tk.RIGHT, padx=(10, 0))

        # Back to Dashboard Button
        back_dashboard_btn = tk.Button(
            nav_frame,
            text="Back to Dashboard",
            command=self.back_to_dashboard,
            font=BUTTON_FONT,
            bg="red",
            fg="white",
            padx=20,
            pady=5
        )
        back_dashboard_btn.pack(side=tk.RIGHT, padx=(0, 10))
        
        # Create a new "Save Results" button:
        self.save_results_btn = tk.Button(
            nav_frame,
            text="Save Results",
            command=lambda: self.prompt_save_results(self.all_new_entries, skip_copy=True),
            font=BUTTON_FONT,
            bg="#1b95e0",   # Choose a color you like
            fg="white",
            padx=20,
            pady=5,
            state=tk.DISABLED  # Initially disabled
        )
        self.save_results_btn.pack(side=tk.RIGHT, padx=(0, 10))

    def back_to_dashboard(self):
        """Destroy the current main interface and return to the dashboard."""
        # Destroy the main analysis interface
        self.main_frame.destroy()
        # Optionally, reset any state if needed here.
        # Then, show the dashboard again:
        self.show_dashboard()

    def launch_literal_analysis(self):
        """Clears the dashboard and builds the literal meaning analysis interface."""
        # Clear the root window (removes the dashboard)
        for widget in self.root.winfo_children():
            widget.destroy()

        # Optionally update the window title
        self.root.title("Literal Meaning Analysis")

        # Build the main analysis interface
        self.setup_main_analysis_interface()

    def user_input(self, word, pankti):
        print(f"Opening input window for {word}")
        self.input_submitted = False
        # normalize for repeat-note consistency
        verse_key = unicodedata.normalize(
            "NFC", re.sub(r"\s+", " ", pankti.replace('ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¥Ãƒâ€šÃ‚Â¥', '').strip())
        )
        raw_tokens = pankti.split()
        word_norm = unicodedata.normalize("NFC", word.strip())
        safe_idx = max(0, min(self.current_word_index, len(raw_tokens)))
        occurrence_idx = sum(
            1
            for tok in raw_tokens[:safe_idx]
            if unicodedata.normalize("NFC", tok.strip().replace('ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¥Ãƒâ€šÃ‚Â¥', '')) == word_norm
        )
        if occurrence_idx > 0 and not getattr(self, "_use_inline_literal_banner", True):
            self._maybe_show_repeat_important_note(word_norm, occurrence_idx, verse_key)

        self.input_window = tk.Toplevel(self.root)
        self.input_window.title(f"Input for {word}")
        self.input_window.configure(bg='light gray')
        try:
            self._wm_for(self.input_window).maximize()
        except Exception:
            pass
        self.input_window.resizable(True, True)

        # ---------------------------
        # Display the Pankti on top
        # ---------------------------
        pankti_frame = tk.Frame(self.input_window, bg='light gray')
        pankti_frame.pack(fill=tk.X, padx=20, pady=10)

        pankti_display = tk.Text(
            pankti_frame, wrap=tk.WORD, bg='light gray', font=('Arial', 32),
            height=2, padx=5, pady=5
        )
        pankti_display.pack(fill=tk.X, expand=False)
        pankti_display.insert(tk.END, pankti)
        pankti_display.tag_add("center", "1.0", "end")
        pankti_display.tag_configure("center", justify='center')

        # Highlight the word at self.current_word_index
        words = pankti.split()
        start_idx = 0
        for i, w in enumerate(words):
            # When we reach the word at current_word_index, calculate its start/end
            if i == self.current_word_index:
                end_idx = start_idx + len(w)
                pankti_display.tag_add("highlight", f"1.{start_idx}", f"1.{end_idx}")
                pankti_display.tag_config("highlight", foreground="red", font=('Arial', 32, 'bold'))
                break
            # Move the start index past this word plus one space
            start_idx += len(w) + 1

        pankti_display.config(state=tk.DISABLED)

        # ---------------------------
        # Create a horizontal PanedWindow for split layout
        # ---------------------------
        split_pane = tk.PanedWindow(self.input_window, orient=tk.HORIZONTAL, bg='light gray')
        split_pane.pack(fill=tk.BOTH, expand=True)

        # Left pane: Meanings
        self.left_pane = tk.Frame(split_pane, bg='light gray')
        split_pane.add(self.left_pane, stretch="always")
        tk.Label(self.left_pane, text=f"Meanings for {word}:", bg='light gray',
                font=('Arial', 14, 'bold')).pack(anchor='center', pady=(0, 10))
        self.meanings_scrollbar = tk.Scrollbar(self.left_pane, orient=tk.VERTICAL)
        self.meanings_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.meanings_canvas = tk.Canvas(self.left_pane, bg='light gray', borderwidth=0,
                                        yscrollcommand=self.meanings_scrollbar.set)
        self.meanings_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.meanings_scrollbar.config(command=self.meanings_canvas.yview)
        self.meanings_inner_frame = tk.Frame(self.meanings_canvas, bg='light gray')
        self.meanings_canvas.create_window((0, 0), window=self.meanings_inner_frame, anchor='nw')

        # Right pane: Grammar Options
        right_pane = tk.Frame(split_pane, bg='light gray')
        split_pane.add(right_pane, stretch="always")
        tk.Label(right_pane, text="Select Grammar Options:", bg='light gray',
                font=('Arial', 14, 'bold')).pack(pady=10)
        self.setup_options(
            right_pane,
            "Do you know the Number of the word?",
            [("Singular", "Singular / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢"), ("Plural", "Plural / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â"), ("Not Applicable", "NA")],
            self.number_var
        )
        self.setup_options(
            right_pane,
            "Do you know the Gender of the word?",
            [("Masculine", "Masculine / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â"), ("Feminine", "Feminine / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬"), ("Neutral", "Trans / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢")],
            self.gender_var
        )
        self.setup_options(
            right_pane,
            "Do you know the Part of Speech for the word?",
            [("Noun", "Noun / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Âµ"), ("Adjective", "Adjectives / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¶ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¶ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£"),
            ("Adverb", "Adverb / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â  ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¶ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£"), ("Verb", "Verb / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â "),
            ("Pronoun", "Pronoun / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Âµ"), ("Postposition", "Postposition / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â§ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢"),
            ("Conjunction", "Conjunction / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¯ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢"), ("Interjection", "Interjection / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢")],
            self.pos_var
        )

        # ---------------------------
        # Bottom Button Frame
        # ---------------------------
        button_frame = tk.Frame(self.input_window, bg='light gray')
        button_frame.pack(side=tk.BOTTOM, fill=tk.X, padx=20, pady=(6, 0))
        # Center the action buttons: use expanding spacers on both sides
        tk.Frame(button_frame, bg='light gray').pack(side=tk.LEFT, expand=True)
        tk.Button(button_frame, text="Submit", command=self.submit_input,
                font=('Arial', 12, 'bold'), bg='navy', fg='white',
                padx=30, pady=15).pack(side=tk.LEFT, padx=15)
        tk.Button(button_frame, text="Skip", command=self.skip_input,
                font=('Arial', 12, 'bold'), bg='navy', fg='white',
                padx=30, pady=15).pack(side=tk.LEFT, padx=15)
        tk.Frame(button_frame, bg='light gray').pack(side=tk.LEFT, expand=True)

        # ---------------------------
        # Launch the dictionary lookup in a separate thread
        # ---------------------------
        self.start_progress()
        threading.Thread(target=self.lookup_meanings_thread, args=(word,), daemon=True).start()

        self.input_window.transient(self.root)
        self.input_window.grab_set()
        self.root.wait_window(self.input_window)
        print(f"Input window for {word} closed")

    def lookup_meanings_thread(self, word):
        """
        Performs the dictionary lookup in a separate thread and then
        schedules the update of the meanings UI on the main thread.
        """
        meanings = self.lookup_word_in_dictionary(word)
        # Schedule the UI update on the main thread
        self.root.after(0, lambda: self.update_meanings_ui(meanings))

    def update_meanings_ui(self, meanings):
        """
        Updates the meanings UI with the lookup results.
        Stops the progress bar and populates the meanings section.
        """
        self.stop_progress()  # Stop the progress window now that lookup is complete
        self.accumulate_meanings_data(meanings)
        split_meanings = self.split_meanings_for_display(meanings)
        # Clear any existing widgets in the meanings inner frame
        for widget in self.meanings_inner_frame.winfo_children():
            widget.destroy()
        # Repopulate the meanings UI
        for i, column in enumerate(split_meanings.values()):
            column_frame = tk.Frame(self.meanings_inner_frame, bg='light gray')
            column_frame.grid(row=0, column=i, padx=10, pady=10, sticky='nw')
            for meaning in column:
                tk.Label(column_frame, text=f"- {meaning}", bg='light gray',
                        font=('Arial', 12), wraplength=400, justify=tk.LEFT).pack(anchor='w', padx=15, pady=5)
        self.meanings_inner_frame.update_idletasks()
        self.meanings_canvas.config(scrollregion=self.meanings_inner_frame.bbox("all"))

    def split_meanings_for_display(self, meanings): # Helper function to split meanings into two columns
        # Determine if 'meanings' is a dict or list.
        if isinstance(meanings, dict):
            mlist = meanings.get("meanings", [])
        else:
            mlist = meanings  # Assume it's already a list.
            
        # Now split the list into two halves.
        mid = len(mlist) // 2
        left = mlist[:mid]
        right = mlist[mid:]
        return {"left": left, "right": right}

    def _rule_key_from_entry(self, d):
        return " | ".join([
            d.get("Word",""),
            d.get("\ufeffVowel Ending",""),
            d.get("Number / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨",""),
            d.get("Grammar / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¯ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£",""),
            d.get("Gender / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â",""),
            d.get("Word Root",""),
            d.get("Type",""),
        ]).strip()

    def submit_matches(self):
        any_selection = False
        current_entries = []  # Local list for entries of the current word

        # Process matching rule checkboxes as before
        for var, match in self.match_vars:
            if var.get():
                match_string = match[0]
                self.results_text.insert(tk.END, f"{match_string}\n")
                self.results_text.insert(tk.END, "-" * 50 + "\n")
                data = match_string.split(" | ")
                new_entry = {
                    "Word": data[0],
                    "\ufeffVowel Ending": data[1],
                    "Number / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨": data[2],
                    "Grammar / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¯ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£": data[3],
                    "Gender / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â": data[4],
                    "Word Root": data[5],
                    "Type": data[6],
                }
                current_entries.append(new_entry)
                any_selection = True

        # Also update the accumulated meanings for the current word
        selected_meanings = [meaning for var, meaning in self.meaning_vars if var.get()]
        self.accumulate_meanings_data(selected_meanings)

        # --- NEW BLOCK: Check for repeated word and prompt update for previous occurrences ---
        current_word = self.pankti_words[self.current_word_index]
        # Find the first occurrence index for the current word
        first_index = next((i for i, w in enumerate(self.pankti_words) if w == current_word), self.current_word_index)

        # Get the current selection from the UI for this occurrence.
        current_selected = selected_meanings  # already computed

        # If this is a repeated occurrence, merge prior selections for this word only.
        if self.current_word_index != first_index:
            merged_meanings = []
            for idx in range(first_index, self.current_word_index):
                # Only merge if the word at that index matches the current word.
                if idx < len(self.accumulated_meanings) and self.pankti_words[idx] == current_word:
                    entry = self.accumulated_meanings[idx]
                    if isinstance(entry, dict):
                        merged_meanings.extend(entry.get("meanings", []))
                    else:
                        merged_meanings.extend(entry)
            # Remove duplicates while preserving order.
            prior_meanings = list(dict.fromkeys(merged_meanings))
            
            # Compare the current selection to prior selections.
            if set(current_selected) != set(prior_meanings):
                update_prev = messagebox.askyesno(
                    "Update Previous Meanings",
                    f"You have selected different meanings for the word '{current_word}'.\n"
                    "Do you want to update the meanings for all previous occurrences of this word?"
                )
                if update_prev:
                    for idx in range(first_index, self.current_word_index):
                        if idx < len(self.accumulated_meanings) and self.pankti_words[idx] == current_word:
                            self.accumulated_meanings[idx] = {"word": current_word, "meanings": current_selected}

        # Finally, update (or add) the current occurrence's accumulated meanings with only the current selection.
        if self.current_word_index < len(self.accumulated_meanings):
            self.accumulated_meanings[self.current_word_index] = {"word": current_word, "meanings": current_selected}
        else:
            self.accumulated_meanings.append({"word": current_word, "meanings": current_selected})

        # --- End NEW BLOCK ---

        # --- Assign verse using word index and verse boundaries ---
        verse_boundaries = []
        pointer = 0
        for verse in self.selected_verses:
            verse_words = verse.split()
            start = pointer
            end = pointer + len(verse_words)
            verse_boundaries.append((start, end))
            pointer = end

        current_verse = None
        for i, (start, end) in enumerate(verse_boundaries):
            if start <= self.current_word_index < end:
                current_verse = self.selected_verses[i]
                break

        # Attach current verse to each grammar entry
        for entry in current_entries:
            entry["Verse"] = current_verse

        # Process finalized matches (avoiding duplicates)
        finalized_matches = []
        for var, match in self.match_vars:
            if var.get():
                match_word = match[0].split(" | ")[0]
                for entry in current_entries:
                    if entry["Word"] == match_word and entry not in finalized_matches:
                        finalized_matches.append(entry)
        self.accumulate_finalized_matches(finalized_matches)
            
        if not any_selection:
            messagebox.showwarning("No Selection", "No matches were selected. Please select at least one match.")
        else:
            self.match_window.destroy()
            # Add the current word's entries to the global accumulator
            self.all_new_entries.extend(current_entries)
            self.current_word_index += 1
            self.process_next_word()

    def show_matches(self, matches, pankti, meanings, max_display=30):
        # Destroy any existing match window
        if hasattr(self, 'match_window') and self.match_window.winfo_exists():
            self.match_window.destroy()

        # Create the match window
        self.match_window = tk.Toplevel(self.root)
        self.match_window.title("Select Matches and Meanings")
        self.match_window.configure(bg='light gray')
        try:
            self._wm_for(self.match_window).maximize()
        except Exception:
            pass
        # New window ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ allow a fresh one-time resize binding
        self._inline_resize_bound = False
        try:
            # If the window is destroyed via an atypical path, ensure the next window can rebind
            # and clear any lingering inline banner references.
            def _on_destroy(_e=None):
                try:
                    self._inline_resize_bound = False
                    self.literal_note_frame = None
                    self.literal_note_title = None
                    self.literal_note_body  = None
                except Exception:
                    pass
            self.match_window.bind("<Destroy>", _on_destroy, add="+")
        except Exception:
            pass

        # Reset check-variable lists for matches and meanings
        self.match_vars = []      # For matching rule checkboxes
        self.meaning_vars = []    # For meaning checkboxes
        unique_matches = self.filter_unique_matches(matches)
        self.all_matches.append(unique_matches)

        # ---------------------------
        # Display the complete Pankti at the top
        # ---------------------------
        pankti_frame = tk.Frame(self.match_window, bg='light gray')
        pankti_frame.pack(fill=tk.BOTH, padx=30, pady=20)

        pankti_display = tk.Text(pankti_frame, wrap=tk.WORD, bg='light gray',
                                font=('Arial', 32), height=2, padx=5, pady=5)
        pankti_display.pack(fill=tk.BOTH, expand=False)
        pankti_display.insert(tk.END, f"{pankti}")
        pankti_display.tag_add("center", "1.0", "end")
        pankti_display.tag_configure("center", justify='center')
        pankti_display.config(state=tk.DISABLED)

        # Compute the character offset for the word at self.current_word_index
        words = pankti.split()
        # Align the navigation token stream with the displayed verse tokens so
        # highlighting, indexing, and repeat detection use the same sequence.
        if getattr(self, "pankti_words", None) != words:
            self._norm_words_cache = [self._norm_tok(w) for w in words]
        self.pankti_words = words
        max_idx = len(words) - 1
        self.current_word_index = max(0, min(self.current_word_index, max_idx))
        idx = self.current_word_index
        start_idx = 0
        for i, w in enumerate(words):
            if i == idx:
                break
            # +1 accounts for the space between words
            start_idx += len(w) + 1
        end_idx = start_idx + len(words[idx])

        pankti_display.tag_add("highlight", f"1.{start_idx}", f"1.{end_idx}")
        pankti_display.tag_config("highlight", foreground="red", font=('Arial', 32, 'bold'))
        pankti_display.config(state=tk.DISABLED)

        # ----- Inline Important Note ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â Literal Analysis (conditional replica of reanalysis) -----
        # Anchor just under the Pankti for deterministic placement (mirrors reanalysis)
        banner_anchor = tk.Frame(self.match_window, bg='light gray')
        banner_anchor.pack(fill=tk.X, padx=20, pady=(0, 0))

        # Compute repeat condition for the selected token
        display_word = self.pankti_words[idx] if idx < len(self.pankti_words) else words[idx]
        word_norm    = self._norm_tok(display_word)
        seen_before  = sum(self._norm_tok(w) == word_norm for w in words[:idx]) if word_norm else 0
        total_hits   = sum(self._norm_tok(w) == word_norm for w in words)        if word_norm else 0
        trigger_now  = bool(word_norm) and total_hits >= 2 and seen_before >= 1
        inline_allowed = bool(getattr(self, "_use_inline_literal_banner", True)) \
                         and not bool(getattr(self, "_suppress_repeat_notes_for_verse", False))

        if inline_allowed and trigger_now:
            # Create/reuse the same framed banner style used by reanalysis
            if not (getattr(self, "literal_note_frame", None)
                    and self.literal_note_frame.winfo_exists()
                    and self.literal_note_frame.master is banner_anchor):
                try:
                    if getattr(self, "literal_note_frame", None) and self.literal_note_frame.winfo_exists():
                        self.literal_note_frame.destroy()
                except Exception:
                    pass
                self.literal_note_frame = tk.Frame(banner_anchor, bg='AntiqueWhite', relief='groove', bd=2)
                self.literal_note_title = tk.Label(self.literal_note_frame,
                                                   text="Important Note ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â Literal Analysis",
                                                   font=("Arial", 14, 'bold'),
                                                   bg='AntiqueWhite')
                self.literal_note_title.pack(pady=(5, 0))
                self.literal_note_body  = tk.Label(self.literal_note_frame,
                                                   bg='AntiqueWhite', fg='black',
                                                   font=('Arial', 12), justify=tk.LEFT)
                self.literal_note_body.pack(pady=(0, 10), padx=10)

            explanation_text = (
                "ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¢ Highlighted selections (displayed in Yellow) indicate the meanings or grammar rules that "
                "were previously confirmed in your assessment.\n"
                "ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¢ This helps you quickly recognize which items reflect your earlier choices."
            )
            self.literal_note_body.config(
                text=explanation_text,
                wraplength=self._banner_wraplength(self.match_window)
            )
            if not self.literal_note_frame.winfo_ismapped():
                self.literal_note_frame.pack(fill=tk.X, padx=20, pady=(5, 10))
            try:
                if not getattr(self, "_inline_resize_bound", False):
                    self.match_window.bind("<Configure>", self._on_match_window_resize, add="+")
                    self._inline_resize_bound = True
            except Exception:
                pass
        else:
            # Clean up any stale banner if not needed
            if getattr(self, "literal_note_frame", None):
                try:
                    if self.literal_note_frame.winfo_exists():
                        self.literal_note_frame.destroy()
                except Exception:
                    pass
                self.literal_note_frame = None
                self.literal_note_title = None
                self.literal_note_body  = None
        # ----- end inline note -----

        # ---------------------------
        # Create a main frame to hold both the Meanings and the Matching Rules sections
        # ---------------------------
        main_frame = tk.Frame(self.match_window, bg='light gray')
        main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)

        # ---------------------------
        # Left Pane: Display Meanings as Checkboxes
        # ---------------------------
        meanings_frame = tk.Frame(main_frame, bg='light gray')
        meanings_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=10)

        tk.Label(meanings_frame, text=f"Select Meanings for {self.pankti_words[self.current_word_index]}:",
                bg='light gray', font=('Arial', 14, 'bold')).pack(pady=10)
        # NEW: Add a toggle checkbutton to select/deselect all meanings
        self.select_all_meanings_var = tk.BooleanVar(value=True)
        tk.Checkbutton(meanings_frame, text="Select/Deselect All Meanings",
                    variable=self.select_all_meanings_var, bg='light gray',
                    font=('Arial', 12), command=self.toggle_all_meanings).pack(pady=5)

        meanings_canvas = tk.Canvas(meanings_frame, bg='light gray', borderwidth=0)
        meanings_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        meanings_scrollbar = tk.Scrollbar(meanings_frame, orient=tk.VERTICAL, command=meanings_canvas.yview)
        meanings_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        meanings_canvas.config(yscrollcommand=meanings_scrollbar.set)

        meanings_content = tk.Frame(meanings_canvas, bg='light gray')
        meanings_canvas.create_window((0, 0), window=meanings_content, anchor='nw')

        # Split the meanings into two columns and create a checkbox for each meaning
        split_meanings = self.split_meanings_for_display(meanings)
        # --- Determine if the current word is repeated and merge prior selections ---
        current_word = self.pankti_words[self.current_word_index]
        first_index = next((i for i, w in enumerate(self.pankti_words) if w == current_word), self.current_word_index)

        # Merge meanings from all occurrences from the first occurrence up to (but not including) the current occurrence.
        merged_meanings = []
        for idx in range(first_index, self.current_word_index):
            if idx < len(self.accumulated_meanings):
                entry = self.accumulated_meanings[idx]
                if isinstance(entry, dict):
                    merged_meanings.extend(entry.get("meanings", []))
                else:
                    merged_meanings.extend(entry)
        # Remove duplicates while preserving order
        prior_meanings = list(dict.fromkeys(merged_meanings))

        # If this is a repeated occurrence (current index is not the first occurrence), reorder the meanings.
        if self.current_word_index != first_index:
            # Obtain the current meanings list
            if isinstance(meanings, dict):
                current_meanings = meanings.get("meanings", [])
            else:
                current_meanings = meanings
            # Reorder: meanings from the first occurrence first, then the rest.
            reordered = [m for m in current_meanings if m in prior_meanings]
            reordered += [m for m in current_meanings if m not in prior_meanings]
            if isinstance(meanings, dict):
                meanings["meanings"] = reordered
            else:
                meanings = reordered

        # Now split the meanings for display using the (possibly) reordered meanings.
        split_meanings = self.split_meanings_for_display(meanings)

        # --- Create the checkboxes as before ---
        for i, column in enumerate(split_meanings.values()):
            column_frame = tk.Frame(meanings_content, bg='light gray')
            column_frame.grid(row=0, column=i, padx=10, pady=10, sticky='nw')
            for meaning in column:
                # For repeated occurrences, preselect if the meaning was chosen in the first occurrence,
                # and highlight those checkbuttons with yellow.
                if self.current_word_index != first_index:
                    preselect = meaning in prior_meanings
                    bg_color = "yellow" if preselect else "light gray"
                else:
                    preselect = True
                    bg_color = "light gray"
                var = tk.BooleanVar(value=preselect)
                chk = tk.Checkbutton(
                    column_frame,
                    text=f"- {meaning}",
                    variable=var,
                    bg=bg_color,
                    font=('Arial', 12),
                    wraplength=325,
                    anchor='w',
                    justify=tk.LEFT,
                    selectcolor='light blue'
                )
                chk.pack(anchor='w', padx=15, pady=5)
                self.meaning_vars.append((var, meaning))

        meanings_content.update_idletasks()
        meanings_canvas.config(scrollregion=meanings_canvas.bbox("all"))

        # ---------------------------
        # Right Pane: Display Matching Rules as Checkboxes
        # ---------------------------
        matches_frame = tk.Frame(main_frame, bg='light gray')
        matches_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=10)

        tk.Label(matches_frame, text="Select the matching rules:",
                bg='light gray', font=('Arial', 14, 'bold')).pack(pady=10)

        matches_canvas = tk.Canvas(matches_frame, bg='light gray', borderwidth=0)
        matches_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        matches_scrollbar = tk.Scrollbar(matches_frame, orient=tk.VERTICAL, command=matches_canvas.yview)
        matches_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        matches_canvas.config(yscrollcommand=matches_scrollbar.set)

        matches_content = tk.Frame(matches_canvas, bg='light gray')
        matches_canvas.create_window((0, 0), window=matches_content, anchor='nw')

        # Determine if the current word is repeated and gather prior grammar rules
        prior_rules = set()
        if self.current_word_index != first_index:
            for idx in range(first_index, self.current_word_index):
                if idx < len(self.accumulated_finalized_matches):
                    prior_rules.update({
                        self._rule_key_from_entry(entry)
                        for entry in self.accumulated_finalized_matches[idx]
                    })

        # Display each match with a checkbox
        for match in unique_matches[:max_display]:
            display_str = match[0]
            core = re.sub(r'\s*\(Matching Characters:\s*\d+\)\s*$', '', display_str).strip()
            if self.current_word_index != first_index:
                preselect = core in prior_rules
                bg_color = "yellow" if preselect else "light gray"
            else:
                preselect = False
                bg_color = "light gray"
            var = tk.BooleanVar(value=preselect)
            text_str = display_str if " (Matching Characters:" in display_str else f"{display_str} (Matching Characters: {match[1]})"
            tk.Checkbutton(
                matches_content,
                text=text_str,
                variable=var,
                bg=bg_color,
                selectcolor='light blue',
                anchor='w'
            ).pack(fill=tk.X, padx=10, pady=5)
            self.match_vars.append((var, match))

        matches_content.update_idletasks()
        matches_canvas.config(scrollregion=matches_canvas.bbox("all"))

        # ---------------------------
        # Bottom Button Frame: Submit and Back
        # ---------------------------
        button_frame = tk.Frame(self.match_window, bg='light gray')
        button_frame.pack(pady=10)
        tk.Button(button_frame, text="Submit", command=self.submit_matches,
                font=('Arial', 12, 'bold'), bg='navy', fg='white', padx=20, pady=10
                ).pack(side=tk.LEFT, padx=5)
        tk.Button(button_frame, text="Back", command=lambda: self.back_to_user_input_with_pankti(pankti),
                font=('Arial', 12, 'bold'), bg='navy', fg='white', padx=20, pady=10
                ).pack(side=tk.LEFT, padx=5)

        print("Match window created and populated.")

    def toggle_all_meanings(self):
        """Toggle all meaning checkboxes based on the select-all checkbutton."""
        new_value = self.select_all_meanings_var.get()
        for var, meaning in self.meaning_vars:
            var.set(new_value)

    def load_grammar_data(self, file_path):
        """
        Loads grammar data from a CSV file.

        Args:
        file_path (str): The path to the CSV file containing grammar data.

        Returns:
        list: A list of dictionaries containing the grammar data.

        Raises:
        FileNotFoundError: If the specified file does not exist.
        IOError: If an error occurs while reading the file.
        """
        try:
            with open(file_path, "r", encoding='utf-8') as data_base:
                return list(csv.DictReader(data_base))
        except FileNotFoundError:
            print(f"Error: The file '{file_path}' does not exist.")
            return []
        except IOError as e:
            print(f"Error reading file '{file_path}': {e}")
            return []

    def break_into_characters(self, word):
        """Breaks a word into individual characters."""
        return [char for char in unicodedata.normalize('NFC', word)]

    def match_inflections(self, word, inflection, pos):
        """
        Determines the number of Matching Characters between the word and the inflection pattern,
        considering the part of speech. Matches are weighted by position and continuity.

        Args:
        word (str): The word to check.
        inflection (str): The inflection pattern to match against.
        pos (str): The part of speech to consider.

        Returns:
        int: The number of Matching Characters, weighted by position and continuity.
        """
        # Break the word and inflection into individual characters
        word_chars = self.break_into_characters(word)
        inflection_chars = self.break_into_characters(inflection)

        # Only consider specified endings for nouns
        if (pos == "Noun / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Âµ" or pos == "Adjectives / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¶ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¶ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£") and inflection == 'ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾':
            return 0  # No suffix match needed for this case

        # Initialize the match count
        match_count = 0

        # Determine the minimum length between word and inflection for suffix comparison
        min_length = min(len(word_chars), len(inflection_chars))

        # Iterate through the characters in reverse order (from the end of the word and inflection)
        for i in range(1, min_length + 1):
            if word_chars[-i] == inflection_chars[-i]:
                # Increment the match count with a weight factor based on position
                match_count += i  # Giving more weight to matches that occur further back in the word
            else:
                break  # Stop the loop if a mismatch is found (ensuring continuity)

        return match_count  # Return the weighted match count

    def lookup_word_in_dictionary(self, word):
        """
        Looks up the meanings of a word in the dictionary data.

        Args:
        word (str): The word to look up in the dictionary.

        Returns:
        list: A list of meanings for the word.
        """
        meanings = []
        exact_meanings = []

        # Attempt to find the word directly in the dictionary
        result = self.dictionary_data[self.dictionary_data['Word'] == word]
        if not result.empty:
            exact_meanings = ast.literal_eval(result.iloc[0]['Meanings'])
            print(f"Found exact match for: {word}")

        # Search for the word within combined entries regardless of exact match
        combined_entries = []
        for _, row in self.dictionary_data.iterrows():
            combined_word = row['Word']
            combined_word_list = re.split(r'\s+', combined_word)  # Split by whitespace to handle combined words

            if word in combined_word_list:  # Check if the exact word is in the list of combined words
                print(f"Found exact match in combined entry: {row['Word']}")
                combined_entries.append(row)

        # If combined entries are found, proceed with adjacent word search
        adjacent_word_matches = []
        if combined_entries:
            words_in_pankti = self.accumulated_pankti.split()  # Split pankti into words
            word_index = words_in_pankti.index(word) if word in words_in_pankti else -1

            # Identify adjacent words in the pankti
            adjacent_combinations = []
            if word_index != -1:
                # Two-word combinations (main word + adjacent)
                if word_index > 0:  # Previous word
                    adjacent_combinations.append([words_in_pankti[word_index - 1], word])
                if word_index < len(words_in_pankti) - 1:  # Next word
                    adjacent_combinations.append([word, words_in_pankti[word_index + 1]])

                # Three-word combinations (main word + two adjacents)
                if word_index < len(words_in_pankti) - 2:
                    adjacent_combinations.append([words_in_pankti[word_index], words_in_pankti[word_index + 1], words_in_pankti[word_index + 2]])
                if word_index > 0 and word_index < len(words_in_pankti) - 1:
                    adjacent_combinations.append([words_in_pankti[word_index - 1], words_in_pankti[word_index], words_in_pankti[word_index + 1]])
                if word_index >= 2:
                    adjacent_combinations.append([words_in_pankti[word_index - 2], words_in_pankti[word_index - 1], words_in_pankti[word_index]])

            # Now search within combined entries using adjacent word combinations
            for entry in combined_entries:
                combined_word_list = re.split(r'\s+', entry['Word'])  # Split by whitespace to handle combined words

                for combination in adjacent_combinations:
                    # Convert the combination to possible strings to match
                    combination_strings = [' '.join(combination), ' '.join(combination[::-1])]
                    
                    if any(comb_str in entry['Word'] for comb_str in combination_strings):
                        print(f"Found adjacent match in combined entry: {entry['Word']}")
                        combined_meanings = ast.literal_eval(entry['Meanings'])
                        combined_entry = f"{entry['Word']}: {', '.join(combined_meanings)}"
                        adjacent_word_matches.append(combined_entry)

            # Sort the adjacent matches to prioritize three-word matches
            adjacent_word_matches = sorted(adjacent_word_matches, key=lambda x: len(x.split(' ')), reverse=True)

        # If both exact match and adjacent word matches are found, return adjacent first
        if adjacent_word_matches and exact_meanings:
            return adjacent_word_matches + exact_meanings

        # If only exact match is found, return exact match
        if exact_meanings:
            return exact_meanings

        # If only adjacent word matches are found, return them
        if adjacent_word_matches:
            return adjacent_word_matches

        # If no adjacent word matches are found, return original combined entries meanings
        for entry in combined_entries:
            combined_meanings = ast.literal_eval(entry['Meanings'])
            combined_entry = f"{entry['Word']}: {', '.join(combined_meanings)}"
            meanings.append(combined_entry)

        # If meanings are found, return them
        if meanings:
            return meanings

        # If no meaning is found, return a list with a single string message
        return [f"No meanings found for {word}"]

    def back_to_user_input_with_pankti(self, pankti):
        """
        Allows the user to return to the input stage for the current word, with pankti passed as an argument.
        """
        try:
            if hasattr(self, 'match_window') and self.match_window:
                self.match_window.destroy()  # Close the match window if it exists

            # Ensure the current_word_index is within valid range
            if 0 <= self.current_word_index < len(self.pankti_words):
                word = self.pankti_words[self.current_word_index]  # Get the current word
                self.reset_input_variables()  # Reset selections for the new word
                self.user_input(word, pankti)  # Reopen the input window for the current word and pass pankti
            else:
                messagebox.showerror("Error", "No valid word to return to.")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {e}")

    def fetch_data(self, word, pankti):
        try:
            self.reset_input_variables()  # Reset selections for the new word
            print(f"Opening input window for {word}")

            self.user_input(word, pankti)  # Opens the input window with the Pankti

            # Check if input_window exists before waiting
            if hasattr(self, 'input_window') and self.input_window.winfo_exists():
                self.root.wait_window(self.input_window)  # Wait for the input window to close
            else:
                print(f"Input window for {word} did not open properly.")
                return  # Exit if the input window was not opened

            print(f"Input window for {word} closed")

            if not self.input_submitted:
                print(f"No input submitted for {word}. Skipping to next word.")
                self.current_word_index += 1
                self.process_next_word()
            else:
                # Process the submitted input
                self.handle_submitted_input(word)

        except Exception as e:
            print(f"An error occurred while fetching data for {word}: {str(e)}")
            # Ensure the input window is closed in case of an error
            if hasattr(self, 'input_window') and self.input_window.winfo_exists():
                try:
                    self.input_window.destroy()
                except Exception as close_error:
                    print(f"Failed to close input window: {str(close_error)}")
            # Optionally, log the error to a file for future debugging
            with open("error_log.txt", "a", encoding="utf-8") as log_file:
                log_file.write(f"Error with word '{word}': {str(e)}\n")

    def process_next_word(self):
        """Process the next valid word or prompt for saving if finished."""
        pankti = " ".join(self.pankti_words)
        self.current_pankti = pankti

        if self.current_word_index < len(self.pankti_words):
            word = self.pankti_words[self.current_word_index]
            if self.is_non_word_character(word):
                self.current_word_index += 1  # Skip non-word characters
                self.process_next_word()
            else:
                self.fetch_data(word, pankti)  # Process current word
        else:
            # All words processedÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â€šÂ¬Ã‚Âprompt to save using the global accumulator
            self.save_results_btn.config(state=tk.NORMAL)
            self.prompt_save_results(self.all_new_entries)

    def skip_input(self):
        """
        Handles the action when the user decides to skip the current word.
        """
        try:
            # Ask for user confirmation before skipping
            confirm_skip = messagebox.askyesno("Confirm Skip", "Are you sure you want to skip this word?")
            if not confirm_skip:
                return  # Do nothing if the user cancels the skip

            # Mark input as not submitted and close the input window
            self.input_submitted = False
            self.input_window.destroy()

            # Update the index to move to the next word and process it
            self.current_word_index += 1
            self.process_next_word()
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred while skipping: {e}")

    def submit_input(self):
        self.input_submitted = True
        if hasattr(self, 'input_window') and self.input_window.winfo_exists():
            self.input_window.destroy()

        # Start the progress bar
        self.start_progress()

        # Run the search in a separate thread
        search_thread = threading.Thread(target=self.perform_search_and_finish)
        search_thread.start()

    def perform_search_and_finish(self):
        current_word = self.pankti_words[self.current_word_index]
        number = self.number_var.get()
        gender = self.gender_var.get()
        pos = self.pos_var.get()

        print(f"Processing word: {current_word}, Number: {number}, Gender: {gender}, POS: {pos}")
        
        matches = []
        
        if number == "NA" and gender == "NA" and pos == "NA":
            matches = self.search_by_inflections(current_word)
        else:
            matches = self.search_by_criteria(current_word, number, gender, pos)
            if not matches:
                messagebox.showinfo("No Matches Found", "No matches were found as per the criteria given by you. Now conducting a general search.")
                matches = self.search_by_inflections(current_word)

        # Check if meanings are already accumulated for the current word index
        if len(self.accumulated_meanings) > self.current_word_index:
            entry = self.accumulated_meanings[self.current_word_index]
            # If the entry is a dictionary, extract the meanings list; otherwise assume it's already a list.
            if isinstance(entry, dict):
                meanings = entry.get("meanings", [])
            else:
                meanings = entry
            self.handle_lookup_results(matches, meanings)
        else:
            # Launch the dictionary lookup in a separate thread if no meanings are accumulated.
            self.perform_dictionary_lookup(current_word, lambda meanings: self.handle_lookup_results(matches, meanings))

        # Stop the progress bar once done (use Tkinter's `after` to ensure it runs in the main thread)
        self.root.after(0, self.stop_progress)

        if matches:
            print(f"Found matches for {current_word}: {matches}")
            # Ensure current_pankti and meanings are passed to show_matches
            self.root.after(0, lambda: self.show_matches(matches, self.current_pankti, meanings))
        else:
            self.root.after(0, lambda: messagebox.showinfo("No Matches", f"No matches found for the word: {current_word}"))
            self.current_word_index += 1
            self.root.after(0, self.process_next_word)

    def is_non_word_character(self, word):
        """
        Determines if the given word consists solely of non-word characters.

        Args:
        word (str): The word to be checked.

        Returns:
        bool: True if the word consists only of non-word characters; False otherwise.
        """
        # Regular expression pattern to match non-word characters and digits
        pattern = r"^[^\w\s]*[\dÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¥Ãƒâ€šÃ‚Â¥]+[^\w\s]*$"

        # Check if the word matches the pattern
        return re.match(pattern, word) is not None

    def search_by_criteria(self, word, number, gender, pos):
        matches = []
        seen = set()  # To store unique combinations

        # Part of Speech: Noun, Verb
        if pos in ["Noun / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Âµ", "Verb / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â "]:
            specified_endings = [
                "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¦Ã¢â‚¬â„¢", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â ", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã¢â‚¬Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã¢â‚¬Â ", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â",
                "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â ", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã¢â‚¬Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡",
                "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿",
                "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â ", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã¢â‚¬Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã¢â‚¬Â ", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨"
            ]

            # Determine if the word is truly inflectionless
            is_inflectionless = all(not word.endswith(ending) for ending in specified_endings)

            # Iterate through each rule in the grammar data
            for rule in self.grammar_data:
                current_number = number if number != "NA" else rule['Number / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨']
                current_gender = gender if gender != "NA" else rule['Gender / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â']
                current_pos = pos if pos != "NA" else rule['Type']

                # Handle the 'ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾' case
                include_mukta = is_inflectionless and current_pos == "Noun / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Âµ"

                if include_mukta and rule['\ufeffVowel Ending'] == "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾" and rule['Number / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨'] == current_number and rule['Gender / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â'] == current_gender and rule['Type'] == current_pos:
                    result = " | ".join([
                        word,
                        rule.get('\ufeffVowel Ending', ""),
                        rule.get('Number / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨', ""),
                        rule.get('Grammar / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¯ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£', ""),
                        rule.get('Gender / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â', ""),
                        rule.get('Word Root', ""),
                        rule.get('Type', "")
                    ])
                    match_count, match_percentage = (1, 100.0)
                    matches.append((result, match_count, match_percentage))
                elif not include_mukta and rule['Number / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨'] == current_number and rule['Gender / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â'] == current_gender and rule['Type'] == current_pos:
                    # Regular inflection matching
                    inflections = rule['\ufeffVowel Ending'].split()
                    for inflection in inflections:
                        match_count, match_percentage = self.calculate_match_metrics(word, inflection)
                        if match_count > 0:
                            result = " | ".join([
                                word,
                                inflection,
                                rule.get('Number / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨', ""),
                                rule.get('Grammar / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¯ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£', ""),
                                rule.get('Gender / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â', ""),
                                rule.get('Word Root', ""),
                                rule.get('Type', "")
                            ])
                            matches.append((result, match_count, match_percentage))

        # Part of Speech: Adjective (Always perform both searches)
        elif pos == "Adjectives / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¶ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¶ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£":
            specified_endings = [
                "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¦Ã¢â‚¬â„¢", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â ", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã¢â‚¬Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã¢â‚¬Â ", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â",
                "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â ", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã¢â‚¬Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡",
                "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿",
                "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â ", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã¢â‚¬Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã¢â‚¬Â ", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨"
            ]

            # Determine if the word is truly inflectionless
            is_inflectionless = all(not word.endswith(ending) for ending in specified_endings)

            for rule in self.grammar_data:
                current_number = number if number != "NA" else rule['Number / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨']
                current_gender = gender if gender != "NA" else rule['Gender / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â']
                current_pos = pos if pos != "NA" else rule['Type']

                # Handle the 'ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾' case
                include_mukta = is_inflectionless and current_pos == "Adjectives / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¶ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¶ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£"

                # Handle inflections (like Nouns)
                if include_mukta and rule['\ufeffVowel Ending'] == "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾" and rule['Number / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨'] == current_number and rule['Gender / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â'] == current_gender and rule['Type'] == current_pos:
                    result = " | ".join([
                        word,
                        rule.get('\ufeffVowel Ending', ""),
                        rule.get('Number / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨', ""),
                        rule.get('Grammar / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¯ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£', ""),
                        rule.get('Gender / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â', ""),
                        rule.get('Word Root', ""),
                        rule.get('Type', "")
                    ])
                    match_count, match_percentage = (1, 100.0)
                    matches.append((result, match_count, match_percentage))
                elif not include_mukta and rule['Number / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨'] == current_number and rule['Gender / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â'] == current_gender and rule['Type'] == current_pos:
                    inflections = rule['\ufeffVowel Ending'].split()
                    for inflection in inflections:
                        match_count, match_percentage = self.calculate_match_metrics(word, inflection)
                        if match_count > 0:
                            result = " | ".join([
                                word,
                                inflection,
                                rule.get('Number / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨', ""),
                                rule.get('Grammar / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¯ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£', ""),
                                rule.get('Gender / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â', ""),
                                rule.get('Word Root', ""),
                                rule.get('Type', "")
                            ])
                            matches.append((result, match_count, match_percentage))

                # Also check for exact matches (like Pronouns)
                if word in rule['\ufeffVowel Ending'] and rule['Number / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨'] == current_number and rule['Gender / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â'] == current_gender and rule['Type'] == current_pos:
                    result = " | ".join([
                        word,
                        rule.get('\ufeffVowel Ending', ""),
                        rule.get('Number / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨', ""),
                        rule.get('Grammar / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¯ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£', ""),
                        rule.get('Gender / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â', ""),
                        rule.get('Word Root', ""),
                        rule.get('Type', "")
                    ])
                    match_count, match_percentage = self.calculate_match_metrics(word, rule['\ufeffVowel Ending'])
                    matches.append((result, match_count, match_percentage))

        # Part of Speech: Pronoun
        elif pos == "Pronoun / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Âµ":
            for rule in self.grammar_data:
                current_number = number if number != "NA" else rule['Number / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨']
                current_gender = gender if gender != "NA" else rule['Gender / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â']

                if word in rule['\ufeffVowel Ending'] and rule['Number / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨'] == current_number and rule['Gender / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â'] == current_gender and rule['Type'] == pos:
                    result = " | ".join([
                        word,
                        rule.get('\ufeffVowel Ending', ""),
                        rule.get('Number / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨', ""),
                        rule.get('Grammar / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¯ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£', ""),
                        rule.get('Gender / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â', ""),
                        rule.get('Word Root', ""),
                        rule.get('Type', "")
                    ])
                    match_count, match_percentage = self.calculate_match_metrics(word, rule['\ufeffVowel Ending'])
                    matches.append((result, match_count, match_percentage))

        # Part of Speech: Adverb, Postposition, Conjunction
        elif pos in ["Adverb / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â  ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¶ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£", "Postposition / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â§ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢", "Conjunction / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¯ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢", "Interjection / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢"]:
            for rule in self.grammar_data:
                if word in rule['\ufeffVowel Ending'] and rule['Type'] == pos:
                    result = " | ".join([
                        word,
                        rule.get('\ufeffVowel Ending', ""),
                        rule.get('Number / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨', ""),
                        rule.get('Grammar / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¯ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£', ""),
                        rule.get('Gender / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â', ""),
                        rule.get('Word Root', ""),
                        rule.get('Type', "")
                    ])
                    match_count, match_percentage = self.calculate_match_metrics(word, rule['\ufeffVowel Ending'])
                    matches.append((result, match_count, match_percentage))

        # Use filter_unique_matches to remove duplicates and sort the results
        unique_sorted_matches = self.filter_unique_matches(matches)

        return unique_sorted_matches

    def search_by_inflections(self, word):
        """
        Searches for inflection matches for the given word within the grammar data.

        Args:
        word (str): The word to search for inflections.

        Returns:
        list: A list of tuples representing matched grammatical rules for the word,
            along with match count and match percentage.
        """
        matches = []
        seen = set()  # To store unique combinations

        # Define the specified endings for inflectionless check
        specified_endings = [
            "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¦Ã¢â‚¬â„¢", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â ", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã¢â‚¬Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã¢â‚¬Â ", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â",
            "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â ", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã¢â‚¬Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡",
            "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿",
            "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¹Ã¢â‚¬Â ", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã¢â‚¬Â ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¦ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¹Ã¢â‚¬Â ", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â°", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â ", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡", "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨"
        ]

        # Determine if the word is truly inflectionless
        try:
            is_inflectionless = all(not word.endswith(ending) for ending in specified_endings)
        except Exception as e:
            print(f"Error determining if the word '{word}' is inflectionless: {str(e)}")
            is_inflectionless = False  # Default to False if there's an error

        for rule in self.grammar_data:
            rule_pos = rule['Type']

            # Noun, Adjective, and Verb processing
            if rule_pos in ["Noun / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Âµ", "Adjectives / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¶ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¶ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£", "Verb / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â "]:
                include_mukta = is_inflectionless and (rule_pos == "Noun / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Âµ" or rule_pos == "Adjectives / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¶ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¶ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£")

                if include_mukta and rule['\ufeffVowel Ending'] == "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾":
                    # Handle the 'ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚ÂÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¤ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾' case
                    result = " | ".join([
                        word,
                        rule.get('\ufeffVowel Ending', ""),
                        rule.get('Number / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨', ""),
                        rule.get('Grammar / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¯ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£', ""),
                        rule.get('Gender / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â', ""),
                        rule.get('Word Root', ""),
                        rule.get('Type', "")
                    ])
                    match_count, match_percentage = (1, 100.0)
                    matches.append((result, match_count, match_percentage))
                else:
                    # Regular inflection matching
                    inflections = rule['\ufeffVowel Ending'].split()
                    for inflection in inflections:
                        match_count, match_percentage = self.calculate_match_metrics(word, inflection)
                        if match_count > 0:
                            result = " | ".join([
                                word,
                                inflection,
                                rule.get('Number / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨', ""),
                                rule.get('Grammar / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¯ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£', ""),
                                rule.get('Gender / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â', ""),
                                rule.get('Word Root', ""),
                                rule.get('Type', "")
                            ])
                            matches.append((result, match_count, match_percentage))
                    # Hybrid handling for Adjectives
                    if rule_pos == "Adjectives / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¶ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¶ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£" and word in rule['\ufeffVowel Ending']:
                        result = " | ".join([
                            word,
                            rule.get('\ufeffVowel Ending', ""),
                            rule.get('Number / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨', ""),
                            rule.get('Grammar / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¯ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£', ""),
                            rule.get('Gender / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â', ""),
                            rule.get('Word Root', ""),
                            rule.get('Type', "")
                        ])
                        match_count, match_percentage = self.calculate_match_metrics(word, rule['\ufeffVowel Ending'])
                        matches.append((result, match_count, match_percentage))

            # Pronoun processing
            elif rule_pos == "Pronoun / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂªÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Âµ":
                if word in rule['\ufeffVowel Ending']:
                    result = " | ".join([
                        word,
                        rule.get('\ufeffVowel Ending', ""),
                        rule.get('Number / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨', ""),
                        rule.get('Grammar / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¯ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£', ""),
                        rule.get('Gender / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â', ""),
                        rule.get('Word Root', ""),
                        rule.get('Type', "")
                    ])
                    match_count, match_percentage = self.calculate_match_metrics(word, rule['\ufeffVowel Ending'])
                    matches.append((result, match_count, match_percentage))

            # Adverb, Postposition, and Conjunction processing
            elif rule_pos in ["Adverb / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â  ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¶ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£", "Postposition / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â§ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢", "Conjunction / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¯ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã¢â‚¬Å“ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢", "Interjection / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¸ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â®ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢"]:
                if word in rule['\ufeffVowel Ending']:
                    result = " | ".join([
                        word,
                        rule.get('\ufeffVowel Ending', ""),
                        rule.get('Number / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨', ""),
                        rule.get('Grammar / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¯ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£', ""),
                        rule.get('Gender / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â', ""),
                        rule.get('Word Root', ""),
                        rule.get('Type', "")
                    ])
                    match_count, match_percentage = self.calculate_match_metrics(word, rule['\ufeffVowel Ending'])
                    matches.append((result, match_count, match_percentage))

        # Use filter_unique_matches to remove duplicates and sort the results
        unique_sorted_matches = self.filter_unique_matches(matches)

        return unique_sorted_matches

    def analyze_pankti(self):
        """
        1) Get userÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬ÃƒÂ¢Ã¢â‚¬Å¾Ã‚Â¢s typed verse/pankti.
        2) Fuzzy-match it against SGGS data.
        3) Show radio buttons for each match so user can select exactly one.
        """
        user_input = self.pankti_entry.get().strip()
        if not user_input:
            messagebox.showerror("Error", "Please enter some text to analyze.")
            return

        # 1) Fuzzy-match the verse in SGGS data (example function)
        candidate_matches = self.match_sggs_verse(user_input)
        if not candidate_matches:
            messagebox.showinfo("No SGGS Match", "No matching verses found in SGGS. Continuing with grammar analysis.")
            # If you still want to do grammar analysis with the typed input:
            self.finish_grammar_analysis(user_input)
            return
        else:
            # Show a new window with radio buttons
            self.show_sggs_matches_option(candidate_matches, user_input)

    def load_sggs_data(self):
        """
        Loads the SGGS data from an Excel file while displaying a modal progress bar.
        The main window is disabled (to prevent user interaction) while the heavy work runs in a background thread.
        The main thread enters a loop to update the UI (so the progress bar animates) until the data is loaded.
        """
        # Disable the main window to prevent interaction
        self.root.attributes("-disabled", True)

        # Show the progress bar modally on the main thread
        self.start_progress()
        self.root.update()  # Ensure the progress window appears immediately

        # This flag will be set when loading is complete.
        self.loading_done = False

        import threading

        def heavy_work():
            # Perform the heavy work (reading and processing the Excel file)
            data = pd.read_excel("1.1.3 sggs_extracted_with_page_numbers.xlsx")
            headers = list(data.columns)
            data['NormalizedVerse'] = (
                data['Verse']
                .astype(str)
                .str.lower()
                .str.strip()
            )
            # Schedule the finalization on the main thread.
            def finish():
                self.sggs_data = data
                self.sggs_headers = headers
                self.loading_done = True
            self.root.after(0, finish)

        # Run the heavy work in a background thread
        threading.Thread(target=heavy_work, daemon=True).start()

        # Process the event loop until loading is done (this allows the progress bar to animate)
        while not self.loading_done:
            self.root.update_idletasks()
            self.root.update()

        # Re-enable the main window now that the heavy work is complete
        self.root.attributes("-disabled", False)
        self.stop_progress()

    def match_sggs_verse(self, user_input, max_results=10, min_score=25):
        """
        Fuzzy-match the user's input (pankti) against the SGGS 'Verse' column.
        Return a tuple (headers, candidate_matches) where headers is the list
        of all column names from the Excel file, and candidate_matches is a list
        (up to max_results) of best matches above the min_score similarity.
        """
        # Ensure we have loaded the data and headers
        if not hasattr(self, 'sggs_data'):
            self.load_sggs_data()
        
        headers = self.sggs_headers
        normalized_input = user_input.lower().strip()
        candidate_matches = []

        # Disable the main window to block user interaction during matching
        self.root.attributes("-disabled", True)
        # Start the modal progress bar
        self.start_progress()
        self.root.update()  # Ensure progress window appears

        total_rows = len(self.sggs_data)
        # Iterate over the rows in SGGS data
        for i, (_, row) in enumerate(self.sggs_data.iterrows()):
            verse_text = row['NormalizedVerse']
            # Remove extra spaces around numbers within "ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¥Ãƒâ€šÃ‚Â¥" markers
            verse_text = re.sub(r'ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¥Ãƒâ€šÃ‚Â¥\s*(\d+)\s*ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¥Ãƒâ€šÃ‚Â¥', r'ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¥Ãƒâ€šÃ‚Â¥\1ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¥Ãƒâ€šÃ‚Â¥', verse_text)
            score = fuzz.token_sort_ratio(normalized_input, verse_text)
            if score >= min_score:
                candidate_matches.append({
                    'S. No.': row['S. No.'],
                    'Verse': row['Verse'],
                    'Verse No.': row.get('Verse No.'),
                    'Stanza No.': row['Stanza No.'],
                    'Text Set No.': row.get('Text Set No.'),
                    'Raag (Fixed)': row['Raag (Fixed)'],
                    'Sub-Raag': row.get('Sub-Raag'),
                    'Writer (Fixed)': row['Writer (Fixed)'],
                    'Verse Configuration (Optional)': row.get('Verse Configuration (Optional)'),
                    'Stanza Configuration (Optional)': row.get('Stanza Configuration (Optional)'),
                    'Bani Name': row['Bani Name'],
                    'Musical Note Configuration': row.get('Musical Note Configuration'),
                    'Special Type Demonstrator': row.get('Special Type Demonstrator'),
                    'Type': row.get('Type'),
                    'Page Number': row['Page Number'],
                    'Score': score
                })
            if i % 50 == 0:
                self.root.update_idletasks()
                self.root.update()

        # Sort the candidate matches by descending score and select the top results.
        candidate_matches.sort(key=lambda x: x['Score'], reverse=True)

        # Stop the progress bar and re-enable the main window.
        self.stop_progress()
        self.root.attributes("-disabled", False)

        return headers, candidate_matches[:max_results]

    def show_sggs_matches_option(self, candidate_matches, user_input):
        """
        Display fuzzy SGGS matches as radio buttons so the user can choose one.
        The header will also include the user input verse.
        """

        # 1) If the first item is a header list, strip it off:
        if candidate_matches and isinstance(candidate_matches[0], list):
            # This is presumably the header row
            self.header_row = candidate_matches[0]
            # If there's a second element and it's a list of dicts, use that
            if len(candidate_matches) > 1 and isinstance(candidate_matches[1], list):
                candidate_matches = candidate_matches[1]
            else:
                # No real data after the header
                candidate_matches = []

        # 2) Store matches for later reference
        self.candidate_matches = candidate_matches

        # 3) Destroy if there's an existing option window
        if hasattr(self, 'sggs_option_window') and self.sggs_option_window.winfo_exists():
            self.sggs_option_window.destroy()

        # 4) Create a new Toplevel window
        self.sggs_option_window = tk.Toplevel(self.root)
        self.sggs_option_window.title("Select One Matching Verse")
        self.sggs_option_window.configure(bg='light gray')
        try:
            self._wm_for(self.sggs_option_window).maximize()
        except Exception:
            pass

        # 5) A Tk variable that holds the index of the selected match
        self.sggs_option_var = tk.IntVar(value=-1)  # -1 = nothing selected

        # 6) Header label
        header_label = tk.Label(
            self.sggs_option_window,
            text=f"Fuzzy Matches Found for '{user_input}'. Please select one:",
            bg='dark slate gray',
            fg='white',
            font=('Arial', 16, 'bold'),
            pady=10
        )
        header_label.pack(fill=tk.X)

        # 7) Scrollable 2-column cards layout (visual parity with _populate_cards)
        middle = tk.Frame(self.sggs_option_window, bg='light gray')
        middle.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)

        canvas = tk.Canvas(middle, bg='light gray', highlightthickness=0)
        vsb    = tk.Scrollbar(middle, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        cards_frame = tk.Frame(canvas, bg='light gray')
        cards_window = canvas.create_window((0, 0), window=cards_frame, anchor="n")

        # two equal-width columns
        # Enforce equal column widths using a uniform group
        try:
            cards_frame.grid_columnconfigure(0, weight=1, minsize=450, uniform='cards')
            cards_frame.grid_columnconfigure(1, weight=1, minsize=450, uniform='cards')
        except Exception:
            cards_frame.grid_columnconfigure(0, weight=1, minsize=450)
            cards_frame.grid_columnconfigure(1, weight=1, minsize=450)

        def _on_cards_configure(event):
            canvas.configure(scrollregion=canvas.bbox("all"))
        cards_frame.bind("<Configure>", _on_cards_configure)

        def _on_canvas_resize(event):
            # Keep the cards frame horizontally centered within the visible canvas
            canvas.coords(cards_window, event.width // 2, 0)
            # Optionally make the window at least as wide as the canvas to avoid left-hugging
            try:
                canvas.itemconfigure(cards_window, width=event.width)
            except Exception:
                pass
        canvas.bind("<Configure>", _on_canvas_resize)

        # Ensure initial centering after first layout
        canvas.update_idletasks()
        try:
            canvas.coords(cards_window, canvas.winfo_width() // 2, 0)
            canvas.itemconfigure(cards_window, width=canvas.winfo_width())
        except Exception:
            pass

        # 8) Populate cards mirroring _populate_cards, but keep selection via radio buttons
        total_cards = len(candidate_matches)
        for idx, match in enumerate(candidate_matches):
            # Normalize match into a dict-like mapping
            if isinstance(match, dict):
                m = match
            else:
                # Fallback mapping for list/tuple structures (preserve prior behavior)
                score_val = match[7] if len(match) > 7 else '?'
                verse_val = match[1] if len(match) > 1 else ''
                raag_val  = match[2] if len(match) > 2 else ''
                writer_val= match[3] if len(match) > 3 else ''
                bani_val  = match[4] if len(match) > 4 else ''
                page_val  = match[5] if len(match) > 5 else ''
                m = {
                    'Score': score_val,
                    'Verse': verse_val,
                    'Raag (Fixed)': raag_val,
                    'Writer (Fixed)': writer_val,
                    'Bani Name': bani_val,
                    'Page Number': page_val,
                }

            # Build each card
            row, col = divmod(idx, 2)
            card = tk.Frame(
                cards_frame,
                bd=1,
                relief="solid",
                bg="white",
                padx=8,
                pady=8
            )
            # If odd number of cards and this is the last one, span both columns for visual centering
            if (total_cards % 2 == 1) and (idx == total_cards - 1):
                card.grid(row=row, column=0, columnspan=2, padx=10, pady=10, sticky="nsew")
            else:
                card.grid(row=row, column=col, padx=10, pady=10, sticky="nsew")

            # Verse label
            tk.Label(
                card,
                text=str(m.get("Verse", "")).strip(),
                font=("Arial", 14, "bold"),
                wraplength=500,
                justify="center",
                bg="white"
            ).pack(pady=(14,4), padx=(28,8))

            # Radio button at top-left for selection (placed after verse to ensure on top)
            rb = tk.Radiobutton(
                card,
                variable=self.sggs_option_var,
                value=idx,
                bg="white",
                activebackground="white",
                selectcolor='light blue'
            )
            rb.place(x=6, y=6)
            try:
                rb.lift()
            except Exception:
                pass

            # Metadata line (Raag, Writer, Bani, Page) + Match%
            fields = [
                ("Raag",   "Raag (Fixed)"),
                ("Writer", "Writer (Fixed)"),
                ("Bani",   "Bani Name"),
                ("Page",   "Page Number"),
            ]
            meta_parts = []
            for label, key in fields:
                v = m.get(key)
                if v is None or (isinstance(v, float) and math.isnan(v)):
                    continue
                if str(v).lower() == 'nan':
                    continue
                meta_parts.append(f"{label}: {v}")

            # score formatting
            try:
                score_val = float(m.get('Score', 0))
            except Exception:
                score_val = 0.0
            meta_parts.append(f"Match: {score_val:.1f}%")

            tk.Label(
                card,
                text="   |   ".join(meta_parts),
                font=("Arial", 12),
                bg="white"
            ).pack()

        # Ensure scrollregion correct
        cards_frame.update_idletasks()
        canvas.configure(scrollregion=canvas.bbox("all"))

        # 9) Buttons
        btn_frame = tk.Frame(self.sggs_option_window, bg='light gray')
        btn_frame.pack(pady=10)

        tk.Button(
            btn_frame, text="Submit",
            command=self.handle_sggs_option_submit,
            font=('Arial', 12, 'bold'), bg='navy', fg='white',
            padx=20, pady=10
        ).pack(side=tk.LEFT, padx=10)

        tk.Button(
            btn_frame, text="Cancel",
            command=self.sggs_option_window.destroy,
            font=('Arial', 12, 'bold'), bg='red', fg='white',
            padx=20, pady=10
        ).pack(side=tk.LEFT, padx=10)

        print("Match window created and populated.")

    def handle_sggs_option_submit(self):
        idx = self.sggs_option_var.get()
        if idx < 0:
            messagebox.showwarning("No Selection", "Please select one verse match.")
            return

        chosen_match = self.candidate_matches[idx]
        final_input = chosen_match.get('Verse', '')

        if not final_input.strip():
            messagebox.showerror("Error", "The selected verse text is empty. Cannot proceed.")
            return

        # Optionally close the first match window
        self.sggs_option_window.destroy()

        # Now, show the consecutive lines selection
        self.show_consecutive_verses_option(chosen_match, final_input)

    def show_consecutive_verses_option(self, chosen_match, main_verse_text):
        """
        After the user picks the main verse from fuzzy matches,
        let them select additional consecutive lines from the same stanza or text set.
        Ensure the main verse is highlighted and only consecutive verses can be selected.
        """
        # 1) Query your data to find consecutive lines
        stanza_lines = self.fetch_stanza_lines(chosen_match)
        self.chosen_match = stanza_lines
        if not stanza_lines:
            messagebox.showinfo("No Consecutive Lines", "No additional consecutive verses found. Proceeding.")
            self.finish_grammar_analysis(main_verse_text)
            return

        # 2) Store the main user input so that we can return to SGGS matches later.
        self.last_user_input = main_verse_text

        # 3) Create a window to show these lines as checkboxes
        self.consecutive_window = tk.Toplevel(self.root)
        self.consecutive_window.title("Select Consecutive Verses")
        self.consecutive_window.configure(bg='light gray')
        try:
            self._wm_for(self.consecutive_window).maximize()
        except Exception:
            pass

        # Adjust header text based on the stored user's choice.
        if hasattr(self, 'verses_choice') and self.verses_choice:
            header_text = f"Select Additional Lines from the '{chosen_match['Special Type Demonstrator']}'"
        else:
            header_text = "Select Additional Lines from the Stanza"

        header_label = tk.Label(
            self.consecutive_window,
            text=header_text,
            bg='dark slate gray',
            fg='white',
            font=('Arial', 16, 'bold'),
            pady=10
        )
        header_label.pack(fill=tk.X)

        frame = tk.Frame(self.consecutive_window, bg='light gray')
        frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)

        canvas = tk.Canvas(frame, bg='light gray', borderwidth=0)
        vsb = tk.Scrollbar(frame, orient="vertical", command=canvas.yview)
        scroll_frame = tk.Frame(canvas, bg='light gray')

        canvas.configure(yscrollcommand=vsb.set)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)

        canvas.create_window((0, 0), window=scroll_frame, anchor="nw")

        # Populate with checkboxes for each consecutive line.
        # Highlight the main verse (pre-selected and disabled) and add normal checkboxes for others.
        self.stanza_checkvars = []
        for idx, line_info in enumerate(stanza_lines):
            line_text = line_info.get('Verse', '')
            if line_text.strip() == main_verse_text.strip():
                # Highlight the main verse
                var = tk.BooleanVar(value=True)
                chk = tk.Checkbutton(
                    scroll_frame,
                    text=line_text,
                    variable=var,
                    bg='yellow',           # highlight color
                    font=('Arial', 12, 'bold'),
                    anchor='w',
                    justify=tk.LEFT,
                    wraplength=800,
                    state='disabled',      # force it to remain selected
                    selectcolor='light blue'
                )
            else:
                var = tk.BooleanVar(value=False)
                chk = tk.Checkbutton(
                    scroll_frame,
                    text=line_text,
                    variable=var,
                    bg='light gray',
                    font=('Arial', 12),
                    anchor='w',
                    justify=tk.LEFT,
                    wraplength=800,
                    selectcolor='light blue'
                )
            chk.pack(fill=tk.X, padx=10, pady=5)
            self.stanza_checkvars.append((var, line_text))

        scroll_frame.update_idletasks()
        canvas.config(scrollregion=canvas.bbox("all"))

        # ---------------------------
        # Button Frame: Submit, Cancel, and Back
        # ---------------------------
        btn_frame = tk.Frame(self.consecutive_window, bg='light gray')
        btn_frame.pack(pady=10)

        tk.Button(
            btn_frame,
            text="Submit",
            command=lambda: self.validate_and_submit_consecutive(main_verse_text),
            font=('Arial', 12, 'bold'),
            bg='navy',
            fg='white',
            padx=20,
            pady=10
        ).pack(side=tk.LEFT, padx=10)

        tk.Button(
            btn_frame,
            text="Cancel",
            command=self.consecutive_window.destroy,
            font=('Arial', 12, 'bold'),
            bg='red',
            fg='white',
            padx=20,
            pady=10
        ).pack(side=tk.LEFT, padx=10)

        tk.Button(
            btn_frame,
            text="Back",
            command=self.back_to_sggs_matches_option,
            font=('Arial', 12, 'bold'),
            bg='gray',
            fg='white',
            padx=20,
            pady=10
        ).pack(side=tk.LEFT, padx=10)

        print("Consecutive verses window created and populated.")

    def validate_and_submit_consecutive(self, main_verse_text):
        """
        Validate that the selected verses form a consecutive block that includes the main verse.
        If valid, proceed with handling the submission.
        """
        # Get indices of all selected verses
        selected_indices = [i for i, (var, _) in enumerate(self.stanza_checkvars) if var.get()]
        if not selected_indices:
            messagebox.showwarning("No Selection", "Please ensure at least the main verse is selected.")
            return

        # Find the index of the main verse (which should be pre-selected)
        main_indices = [i for i, (_, line_text) in enumerate(self.stanza_checkvars)
                        if line_text.strip() == main_verse_text.strip()]
        if not main_indices:
            messagebox.showerror("Error", "Main verse not found in the list.")
            return
        main_index = main_indices[0]

        # Check if the selected indices form a consecutive block.
        if max(selected_indices) - min(selected_indices) != len(selected_indices) - 1:
            messagebox.showerror("Non-Consecutive Selection", "Please select only consecutive verses.")
            return

        # Ensure the consecutive block includes the main verse.
        if main_index < min(selected_indices) or main_index > max(selected_indices):
            messagebox.showerror("Selection Error", "The selected consecutive block must include the main verse.")
            return

        # If validation passes, proceed with the submission.
        self.handle_consecutive_submit(main_verse_text)

    def back_to_sggs_matches_option(self):
        """
        Close the consecutive verses window and re-display the SGGS matches option window.
        """
        self.consecutive_window.destroy()
        # Re-open the SGGS matches option window using the stored candidate matches and user input.
        self.show_sggs_matches_option(self.candidate_matches, self.last_user_input)

    def handle_consecutive_submit(self, main_verse_text):
        """
        Combine the main verse text with the lines selected by the user,
        then proceed with grammar analysis. Ensures that the main verse is
        not duplicated if it was also selected by the user.
        Also, store the individual selected verses in self.selected_verses 
        (as a list of tuples: (start_index, end_index, verse_text)) for later use.
        """
        # Strip the main verse text
        main_line = main_verse_text.strip()

        # Gather the lines that were checked; strip extra whitespace
        selected_lines = [line.strip() for var, line in self.stanza_checkvars if var.get()]

        # Store each selected line as an individual verse in self.selected_verses.
        self.selected_verses = list(selected_lines)

        # Keep only matches where the verse is in self.selected_verses.
        self.chosen_match = [match for match in self.chosen_match if match['Verse'] in self.selected_verses]

        # Filter out any line that is exactly the same as the main verse
        extra_lines = [line for line in selected_lines if line != main_line]

        # Combine all selected lines (including the main verse, already in correct order) into a single string,
        # ensuring the original sequence of the verse is maintained.
        if extra_lines:
            combined_text = " ".join(selected_lines)
        else:
            combined_text = main_line
        
        # Close the consecutive window
        self.consecutive_window.destroy()

        # Proceed with grammar analysis using the combined text
        self.finish_grammar_analysis(combined_text)

    def fetch_stanza_lines(self, chosen_match):
        """
        Return a list of dictionaries representing consecutive lines from either the current stanza or the entire text set,
        based on the user's choice.
        This function uses the SGGS data loaded by load_sggs_data() (stored in self.sggs_data).
        """
        # Ensure the SGGS data is loaded
        if not hasattr(self, 'sggs_data'):
            self.load_sggs_data()

        # Retrieve the Stanza No. and Text Set No. from the chosen match.
        stanza_no = chosen_match.get('Stanza No.')
        text_set_no = chosen_match.get('Text Set No.')

        if stanza_no is None or text_set_no is None:
            print("Chosen match does not contain required 'Stanza No.' or 'Text Set No.' information.")
            return []

        # Get the Special Type Demonstrator value.
        special_type = chosen_match.get('Special Type Demonstrator', '')

        # If it's 'ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¶ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢', then we don't ask the user because a ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¶ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ is always a stanza.
        if special_type == 'ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¶ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¹ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢':
            choice = False
        else:
            choice = messagebox.askyesno(
                "Select Verses",
                f"Do you want to fetch verses from the entire '{special_type}'?\n\n"
                f"(Yes = Entire '{special_type}', No = Only the current Stanza)"
            )
        # Store the user's choice for later use
        self.verses_choice = choice

        if choice:
            # User selected the entire text set.
            subset = self.sggs_data[self.sggs_data['Text Set No.'] == text_set_no]
        else:
            # User selected only the current stanza.
            subset = self.sggs_data[
                (self.sggs_data['Stanza No.'] == stanza_no) &
                (self.sggs_data['Text Set No.'] == text_set_no)
            ]

        # Optionally sort by 'Verse No.' if available
        if 'Verse No.' in subset.columns:
            subset = subset.sort_values(by='Verse No.')

        lines = subset.to_dict(orient='records')

        return lines

    def finish_grammar_analysis(self, user_input):
        """
        Runs grammar analysis after the user has selected or typed a final verse.
        """
        self._repeat_note_shown = set()
        self._suppress_repeat_notes_for_verse = False
        self.pankti_words = user_input.split()
        self.accumulate_pankti_data(user_input)
        self.current_word_index = 0
        self.all_new_entries = []  # Reset global accumulator

        self.update_navigation_buttons()
        self.process_next_word()

    def prompt_for_assessment(self, metadata_entry):
        """
        Opens a modal window that lets the user paste the analysis result (translation)
        and choose options for 'Framework?' and 'Explicit?'.
        The metadata_entry is a dictionary containing the existing metadata (e.g., Word, Vowel Ending, etc.).
        """
        assessment_win = tk.Toplevel(self.root)
        assessment_win.title("Enter Translation Assessment")
        assessment_win.configure(bg='light gray')
        try:
            self._wm_for(assessment_win)
        except Exception:
            pass

        instruction_label = tk.Label(assessment_win, 
                                    text="Paste the analysis result below:",
                                    font=("Helvetica", 14), bg="light gray")
        instruction_label.pack(pady=10)

        analysis_text = scrolledtext.ScrolledText(assessment_win, width=80, height=10,
                                                font=("Helvetica", 12), wrap=tk.WORD)
        analysis_text.pack(padx=20, pady=10)

        cb_frame = tk.Frame(assessment_win, bg="light gray")
        cb_frame.pack(pady=10)

        framework_var = tk.BooleanVar()
        explicit_var = tk.BooleanVar()

        framework_cb = tk.Checkbutton(cb_frame, text="Framework?", variable=framework_var,
                                    font=("Helvetica", 12), bg="light gray")
        framework_cb.pack(side=tk.LEFT, padx=10)

        explicit_cb = tk.Checkbutton(cb_frame, text="Explicit?", variable=explicit_var,
                                    font=("Helvetica", 12), bg="light gray")
        explicit_cb.pack(side=tk.LEFT, padx=10)

        def on_save():
            translation = analysis_text.get("1.0", tk.END).strip()
            if not translation:
                messagebox.showerror("Error", "Please paste the analysis result.")
                return
            # Merge the metadata with the new assessment data
            new_entry = metadata_entry.copy()
            new_entry["\ufeffVowel Ending"] = self._norm_get(new_entry, "\ufeffVowel Ending")
            new_entry.pop("Vowel Ending", None)
            new_entry["Type"] = self._norm_get(new_entry, "Type")
            new_entry.pop("Word Type", None)
            new_entry["Translation"] = translation
            new_entry["Framework?"] = framework_var.get()
            new_entry["Explicit?"] = explicit_var.get()
            # Revision is computed in save_assessment_data
            self.save_assessment_data(new_entry)
            assessment_win.destroy()

        save_btn = tk.Button(assessment_win, text="Save Assessment",
                            command=on_save, font=("Helvetica", 14, "bold"),
                            bg="#007acc", fg="white", padx=20, pady=10)
        save_btn.pack(pady=20)

        assessment_win.transient(self.root)
        assessment_win.grab_set()
        self.root.wait_window(assessment_win)

    def load_existing_assessment_data(self, file_path):
        expected_columns = [
            "Verse", "Translation", "Translation Revision",
            "Word", "Selected Darpan Meaning", "\ufeffVowel Ending", "Number / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨", "Grammar / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¯ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£", "Gender / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â", "Word Root", "Type", "Grammar Revision", "Word Index",
            "S. No.", "Verse No.", "Stanza No.", "Text Set No.", "Raag (Fixed)", "Sub-Raag", "Writer (Fixed)",
            "Verse Configuration (Optional)", "Stanza Configuration (Optional)", "Bani Name", "Musical Note Configuration",
            "Special Type Demonstrator", "Verse Type", "Page Number",
            "Framework?", "Explicit?"
        ]
        if os.path.exists(file_path):
            try:
                df = pd.read_excel(file_path)
                df.rename(columns={"Vowel Ending": "\ufeffVowel Ending", "Word Type": "Type"}, inplace=True)
                if df.empty or len(df.columns) == 0:
                    df = pd.DataFrame(columns=expected_columns)
                else:
                    # Ensure the DataFrame has exactly the expected columns.
                    df = df.reindex(columns=expected_columns)
                return df
            except Exception as e:
                print(f"Error reading {file_path}: {e}")
        return pd.DataFrame(columns=expected_columns)

    def save_assessment_data(self, new_entry):
        """
        Saves a new assessment entry to an Excel file with the following behavior:
        - For a given verse, update the "Translation" field for all entries.
        - For the specific word being assessed in that verse, group all matching rows.
        - If any grammar field differs from the new entry, increment the Grammar Revision once for the group
            and update all matching rows with the new grammar details.
        - Update "Selected Darpan Meaning" for the specific word occurrence.
        - If any word's grammar is revised in the verse, increment the "Translation Revision" for all words of that verse.
        - If no matching entry exists for the word in that verse, append the new entry with a Grammar Revision of 1,
            and initialize the Translation Revision.
        """
        file_path = "1.2.1 assessment_data.xlsx"
        df_existing = self.load_existing_assessment_data(file_path)
        
        # Define the grammar keys to compare (excluding Translation, which is updated separately).
        grammar_keys = [
            '\ufeffVowel Ending', 'Number / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨', 'Grammar / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¯ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£',
            'Gender / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â', 'Word Root', 'Type'
        ]
        
        # Update the Translation for all rows in the same verse.
        df_existing.loc[df_existing["Verse"] == new_entry["Verse"], "Translation"] = new_entry["Translation"]
        
        # Filter for rows with the same word, same verse, and same word index.
        matching_rows = df_existing[
            (df_existing["Word"] == new_entry["Word"]) &
            (df_existing["Verse"] == new_entry["Verse"]) &
            (df_existing["Word Index"] == new_entry["Word Index"])
        ]
        
        if not matching_rows.empty:
            # For repeated occurrences, pick the highest Grammar Revision as a representative.
            latest_idx = matching_rows["Grammar Revision"].idxmax()
            latest_row = df_existing.loc[latest_idx]
            
            # Check if any grammar field is different.
            differences = any(new_entry.get(key) != self._norm_get(latest_row, key) for key in grammar_keys)
            
            if differences:
                # Compute new Grammar Revision number for the group.
                new_grammar_revision = matching_rows["Grammar Revision"].max() + 1
                new_entry["Grammar Revision"] = new_grammar_revision
                
                # Update all matching rows with new grammar values and new Grammar Revision.
                for idx in matching_rows.index:
                    for key in grammar_keys:
                        df_existing.at[idx, key] = new_entry.get(key)
                    df_existing.at[idx, "Grammar Revision"] = new_grammar_revision
                    # Update additional fields from new_entry if needed.
                    for key, value in new_entry.items():
                        if key not in grammar_keys and key not in ["Translation"]:
                            if key in ("Framework?", "Explicit?"):
                                df_existing.at[idx, key] = int(value)  # Cast Boolean to int (False->0, True->1)
                            else:
                                df_existing.at[idx, key] = value
                # Update Selected Darpan Meaning for these rows.
                for idx in matching_rows.index:
                    # --- Compute correct global index for the current word using the current verse ---
                    current_verse_words = self.accumulated_pankti.split()

                    def find_sublist_index(haystack, needle):
                        for i in range(len(haystack) - len(needle) + 1):
                            if haystack[i:i+len(needle)] == needle:
                                return i
                        return -1

                    start_index = find_sublist_index(self.pankti_words, current_verse_words)
                    if start_index == -1:
                        start_index = 0

                    # Assume new_entry["Word Index"] is the local index within the verse.
                    global_index = start_index + new_entry.get("Word Index", 0)

                    # Retrieve the selected Darpan meaning using the global index.
                    if len(self.accumulated_meanings) > global_index:
                        acc_entry = self.accumulated_meanings[global_index]
                        if isinstance(acc_entry, dict):
                            selected_meaning = "| ".join(acc_entry.get("meanings", []))
                        else:
                            selected_meaning = "| ".join(acc_entry)
                    else:
                        selected_meaning = ""

                    df_existing.at[idx, "Selected Darpan Meaning"] = selected_meaning

                # Now update the Translation Revision for all rows in the verse 
                # to reflect the latest Grammar Revision (without extra +1).
                verse_mask = df_existing["Verse"] == new_entry["Verse"]
                latest_grammar_revision = df_existing.loc[verse_mask, "Grammar Revision"].max()
                df_existing.loc[verse_mask, "Translation Revision"] = latest_grammar_revision
            else:
                # No differences detected; no update necessary.
                return
        else:
            # No existing entry for this word in the verse; add it with Grammar Revision 1 as well as Translation Revision 1.
            new_entry["Grammar Revision"] = 1
            new_entry["Translation Revision"] = 1

            # --- Compute correct global index for the current word using the current verse ---
            current_verse_words = self.accumulated_pankti.split()
            
            def find_sublist_index(haystack, needle):
                for i in range(len(haystack) - len(needle) + 1):
                    if haystack[i:i+len(needle)] == needle:
                        return i
                return -1

            start_index = find_sublist_index(self.pankti_words, current_verse_words)
            if start_index == -1:
                start_index = 0

            # Assume new_entry["Word Index"] holds the local index within the current verse.
            global_index = start_index + new_entry.get("Word Index", 0)

            # Retrieve the selected Darpan meaning using the global index.
            if len(self.accumulated_meanings) > global_index:
                acc_entry = self.accumulated_meanings[global_index]
                if isinstance(acc_entry, dict):
                    selected_meaning = "| ".join(acc_entry.get("meanings", []))
                else:
                    selected_meaning = "| ".join(acc_entry)
            else:
                selected_meaning = ""

            new_entry["Selected Darpan Meaning"] = selected_meaning

            # Determine Translation Revision for the verse.
            current_translation_revision = df_existing[df_existing["Verse"] == new_entry["Verse"]]["Translation Revision"].max()
            new_entry["Translation Revision"] = (current_translation_revision + 1) if current_translation_revision is not None else 1

            # Append new_entry to the DataFrame.
            df_existing = pd.concat([df_existing, pd.DataFrame([new_entry])], ignore_index=True)


        try:
            df_existing.to_excel(file_path, index=False)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save assessment data: {e}")

    def setup_options(self, parent_frame, label_text, options, variable):
        """
        Sets up radio button options in the specified parent frame.

        Args:
        parent_frame (tk.Frame): The frame in which to pack the radio buttons.
        label_text (str): The label text to display above the radio buttons.
        options (list of tuple): A list of tuples where each tuple contains the option text and the value to set in the variable.
        variable (tk.StringVar): The control variable for the group of radio buttons.
        """
        # Create a label for the option group with styling
        tk.Label(parent_frame, text=label_text, bg='light gray', font=('Arial', 12)).pack(pady=(10, 5))

        # Create radio buttons for each option
        for opt_text, opt_value in options:
            tk.Radiobutton(
                parent_frame,
                text=opt_text,
                variable=variable,
                value=opt_value,
                bg='light gray',
                selectcolor='light blue',
                anchor='w',
                font=('Arial', 11)
            ).pack(anchor='w', padx=20, pady=2)

    def navigation_controls(self, parent_frame):
        """
        Add navigation controls to the parent frame to navigate through words.
        """
        nav_frame = tk.Frame(parent_frame, bg='light gray')
        nav_frame.pack(pady=10)

        tk.Button(nav_frame, text="Previous", command=self.previous_word, bg='dark gray', fg='white').pack(side='left', padx=5)
        tk.Button(nav_frame, text="Next", command=self.next_word, bg='dark gray', fg='white').pack(side='left', padx=5)

    def update_navigation_buttons(self):
        """Enable or disable navigation buttons based on the current word index."""
        print(f"Updating buttons, current index: {self.current_word_index}, total words: {len(self.pankti_words)}")
        
        # Disable 'Previous' button if at the start
        if self.current_word_index <= 0:
            self.prev_button.config(state=tk.DISABLED)
        else:
            self.prev_button.config(state=tk.NORMAL)
        
        # Disable 'Next' button if at the end
        if self.current_word_index >= len(self.pankti_words) - 1:
            self.next_button.config(state=tk.DISABLED)
        else:
            self.next_button.config(state=tk.NORMAL)

    def update_current_word_label(self):
        """Update the word label to display the current word."""
        if hasattr(self, 'pankti_words') and self.pankti_words:
            # Clamp the index to valid range:
            if self.current_word_index < 0:
                self.current_word_index = 0
            if self.current_word_index >= len(self.pankti_words):
                self.current_word_index = len(self.pankti_words) - 1
            current_word = self.pankti_words[self.current_word_index]
            self.word_label.config(text=current_word)
        else:
            self.word_label.config(text="No Word Available")

    def prev_word(self):
        if self.current_word_index > 0:
            self.current_word_index -= 1
            self.update_current_word_label()
        self.update_navigation_buttons()

    def next_word(self):
        if self.current_word_index < len(self.pankti_words) - 1:
            self.current_word_index += 1
            self.update_current_word_label()
        self.update_navigation_buttons()

    def select_current_word(self):
        """Trigger analysis for the currently displayed word."""
        if hasattr(self, 'pankti_words') and self.pankti_words:
            # Ensure current_word_index is within range:
            if self.current_word_index >= len(self.pankti_words):
                self.current_word_index = len(self.pankti_words) - 1
            word = self.pankti_words[self.current_word_index]
            self.fetch_data(word, " ".join(self.pankti_words))
        else:
            print("No word available for selection.")

    def close_window(self, window):
        """Closes the given Tkinter window."""
        if window and window.winfo_exists():
            window.destroy()

    def reset_input_variables(self):
        """Reset input variables for number, gender, and part of speech."""
        self.number_var.set("NA")
        self.gender_var.set("NA")
        self.pos_var.set("NA")

    def compose_clipboard_text_for_chatgpt(self):
        clipboard_text = "### Detailed Analysis & Literal Translation\n\n"
        clipboard_text += (
            f"The verse **'{self.accumulated_pankti}'** holds deep meaning. Below is a breakdown of each word with "
            "user-selected meanings and grammar details, which together form the basis for a literal translation prompt.\n\n"
        )
        
        # --- Preceding Verses & Translations ---
        # Load existing assessment data.
        existing_data = self.load_existing_assessment_data("1.2.1 assessment_data.xlsx")
        
        # Identify the candidate matching the current verse.
        current_candidate = next((cand for cand in self.chosen_match 
                                if cand.get("Verse", "").strip() == self.accumulated_pankti.strip()), None)
        
        preceding_verses_text = ""
        if current_candidate:
            text_set_no = current_candidate.get("Text Set No.")
            try:
                current_verse_no = int(current_candidate.get("Verse No."))
            except (ValueError, TypeError):
                current_verse_no = None
            
            if current_verse_no is not None:
                # Filter for the same Text Set No.
                filtered_data = existing_data[existing_data["Text Set No."] == text_set_no]
                consecutive_verses = []
                target_verse_no = current_verse_no - 1
                # Collect consecutive preceding verses.
                while True:
                    row = filtered_data[filtered_data["Verse No."] == target_verse_no]
                    if row.empty:
                        break
                    row_data = row.iloc[0]
                    consecutive_verses.insert(0, row_data)  # earlier verses first
                    target_verse_no -= 1
                
                if consecutive_verses:
                    preceding_verses_text += "\n### Preceding Verses & Translations\n\n"
                    for row_data in consecutive_verses:
                        verse_no = row_data.get("Verse No.", "")
                        verse_text = row_data.get("Verse", "")
                        translation = row_data.get("Translation", "")
                        preceding_verses_text += f"**Verse {verse_no}:** {verse_text}\n"
                        preceding_verses_text += f"**Translation:** {translation}\n\n"
        
        clipboard_text += preceding_verses_text
        
        # --- Current Verse Analysis ---
        current_verse_words = self.accumulated_pankti.split()
        
        def find_sublist_index(haystack, needle):
            # Find a consecutive occurrence of 'needle' in 'haystack'
            for i in range(len(haystack) - len(needle) + 1):
                if haystack[i:i+len(needle)] == needle:
                    return i
            return -1

        start_index = find_sublist_index(self.pankti_words, current_verse_words)
        if start_index == -1:
            start_index = 0  # Fallback if no match is found

        for i, word in enumerate(current_verse_words):
            actual_index = start_index + i
            clipboard_text += f"**Word {i+1}: {word}**\n"
            
            # Retrieve the entry for the current word (if available)
            acc_entry = self.accumulated_meanings[actual_index] if actual_index < len(self.accumulated_meanings) else {}
            # If the entry is a dictionary, extract the 'meanings' list; otherwise, assume it's already a list
            if isinstance(acc_entry, dict):
                meanings_list = acc_entry.get("meanings", [])
            else:
                meanings_list = acc_entry
            # Create a string of meanings, or a default message if none are available
            meanings_str = ", ".join(meanings_list) if meanings_list else "No user-selected meanings available"
            clipboard_text += f"- **User-Selected Meanings:** {meanings_str}\n"
            
            clipboard_text += "- **Grammar Options:**\n"
            finalized_matches_list = self.accumulated_finalized_matches[actual_index] if actual_index < len(self.accumulated_finalized_matches) else []
            
            if finalized_matches_list:
                for option_idx, match in enumerate(finalized_matches_list, start=1):
                    clipboard_text += (
                        f"  - **Option {option_idx}:**\n"
                        f"      - **Word:** {self._norm_get(match, 'Word') or 'N/A'}\n"
                        f"      - **Vowel Ending:** {self._norm_get(match, '\\ufeffVowel Ending') or 'N/A'}\n"
                        f"      - **Number / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨:** {match.get('Number / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨', 'N/A')}\n"
                        f"      - **Grammar / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¯ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£:** {match.get('Grammar / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¯ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£', 'N/A')}\n"
                        f"      - **Gender / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â:** {match.get('Gender / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â', 'N/A')}\n"
                        f"      - **Word Root:** {match.get('Word Root', 'N/A')}\n"
                        f"      - **Type:** {self._norm_get(match, 'Type') or 'N/A'}\n"
                    )
                    clipboard_text += (
                        f"      - **Literal Translation (Option {option_idx}):** The word '{word}' functions as a "
                        f"'{self._norm_get(match, 'Type') or 'N/A'}' with '{match.get('Grammar / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¯ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£', 'N/A')}' usage, in the "
                        f"'{match.get('Number / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨', 'N/A')}' form and '{match.get('Gender / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â', 'N/A')}' gender. Translation: ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â¦\n"
                    )
            else:
                clipboard_text += "  - No finalized grammar options available\n"
            
            clipboard_text += "\n"
        
        if 'ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¥Ãƒâ€šÃ‚Â¥' in current_verse_words:
            clipboard_text += (
                "**Symbol:** ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¥Ãƒâ€šÃ‚Â¥\n"
                "- **Meaning:** End of verse or sentence\n"
                "- **Context:** Denotes the conclusion of the verse.\n\n"
            )
        
        clipboard_text += "\n### Literal Translation Prompt\n"
        clipboard_text += (
            f"Using the above user-selected meanings and grammar details for the verse '{self.accumulated_pankti}', "
            "please generate a literal translation that adheres strictly to the grammatical structure, "
            "capturing the tense, number, gender, and function accurately."
        )
        
        return clipboard_text

    def prompt_copy_to_clipboard(self):
        print("Prompting to copy text to clipboard...")
        copy_prompt = messagebox.askyesno(
            "Copy to Clipboard", 
            f"Would you like to copy the detailed analysis for the verse '{self.accumulated_pankti}' to your clipboard?"
        )
        if copy_prompt:
            if not self.accumulated_pankti or not self.accumulated_meanings or not self.accumulated_grammar_matches or not self.accumulated_finalized_matches:
                print("Error: Accumulated data is not populated.")
                messagebox.showerror("Error", "Failed to copy data to clipboard: Data not populated.")
            else:
                clipboard_text = self.compose_clipboard_text_for_chatgpt()
                pyperclip.copy(clipboard_text)
                messagebox.showinfo("Copied", "The analysis has been copied to the clipboard!")
                print("Clipboard text copied successfully!")

    def prompt_for_final_grammar(self, word_entries):
        """
        Opens a modal window showing all grammar options for a given word.
        Generates a prompt text for ChatGPT to help finalize the grammar choice,
        copies it to the clipboard (with a button to re-copy if needed), and then
        allows the user to select the final applicable grammar.
        
        word_entries: a list of dictionaries (each corresponding to one grammar option for the word)
        """
        final_choice = {}

        final_win = tk.Toplevel(self.root)
        final_win.title(f"Finalize Grammar for '{word_entries[0]['Word']}'")
        final_win.configure(bg='light gray')
        try:
            self._wm_for(final_win)
        except Exception:
            pass

        # --- Build the ChatGPT prompt text ---
        prompt_lines = []
        prompt_lines.append(f"Finalise the applicable grammar for the word: {word_entries[0]['Word']}")
        prompt_lines.append("The following grammar options are available:")
        for idx, entry in enumerate(word_entries, start=1):
            summary = " | ".join([
                self._norm_get(entry, "\ufeffVowel Ending") or "",
                self._norm_get(entry, "Number / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨") or "",
                self._norm_get(entry, "Grammar / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¯ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£") or "",
                self._norm_get(entry, "Gender / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â") or "",
                self._norm_get(entry, "Word Root") or "",
                self._norm_get(entry, "Type") or ""
            ])
            prompt_lines.append(f"Option {idx}: {summary}")
        prompt_text = "\n".join(prompt_lines)
        # --- End building prompt text ---

        # --- Display the prompt text and add a copy button ---
        prompt_frame = tk.Frame(final_win, bg="light gray")
        prompt_frame.pack(padx=20, pady=10, fill=tk.BOTH, expand=True)

        prompt_label = tk.Label(prompt_frame, 
                                text="ChatGPT Prompt for Grammar Finalisation:",
                                font=("Helvetica", 14, "bold"),
                                bg="light gray")
        prompt_label.pack(anchor="w", pady=(0,5))

        prompt_text_widget = scrolledtext.ScrolledText(prompt_frame, width=80, height=6,
                                                        font=("Helvetica", 12), wrap=tk.WORD)
        prompt_text_widget.pack(fill=tk.BOTH, expand=True)
        prompt_text_widget.insert(tk.END, prompt_text)
        prompt_text_widget.config(state=tk.DISABLED)

        def copy_prompt():
            self.root.clipboard_clear()
            self.root.clipboard_append(prompt_text)
            messagebox.showinfo("Copied", "Prompt text copied to clipboard!")

        copy_btn = tk.Button(prompt_frame, text="Copy Prompt", command=copy_prompt,
                            font=("Helvetica", 12, "bold"),
                            bg="#007acc", fg="white", padx=10, pady=5)
        copy_btn.pack(anchor="e", pady=5)
        # --- End prompt display ---

        # Instruction for selecting final grammar
        instruction = tk.Label(final_win,
                            text="Multiple grammar options found for this word.\nPlease select the final applicable grammar:",
                            font=("Helvetica", 14),
                            bg="light gray")
        instruction.pack(pady=10)

        # Tk variable to hold the selected option index
        choice_var = tk.IntVar(value=0)

        # --- Create a scrollable area for the radio buttons ---
        options_container = tk.Frame(final_win, bg="light gray")
        options_container.pack(padx=20, pady=10, fill=tk.BOTH, expand=True)

        # Create a canvas in the container
        canvas = tk.Canvas(options_container, bg="light gray", highlightthickness=0)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Add a vertical scrollbar to the container
        vsb = tk.Scrollbar(options_container, orient="vertical", command=canvas.yview)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        canvas.configure(yscrollcommand=vsb.set)

        # Create a frame inside the canvas to hold the radio buttons
        options_frame = tk.Frame(canvas, bg="light gray")
        canvas.create_window((0,0), window=options_frame, anchor="nw")

        # Update the scroll region when the frame's size changes
        def on_frame_configure(event):
            canvas.configure(scrollregion=canvas.bbox("all"))
        options_frame.bind("<Configure>", on_frame_configure)
        # --- End scrollable area creation ---

        # Create radio buttons with option text as "Option {number}: {summary}"
        for idx, entry in enumerate(word_entries):
            summary = " | ".join([
                self._norm_get(entry, "\ufeffVowel Ending") or "",
                self._norm_get(entry, "Number / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨") or "",
                self._norm_get(entry, "Grammar / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¯ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£") or "",
                self._norm_get(entry, "Gender / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â") or "",
                self._norm_get(entry, "Word Root") or "",
                self._norm_get(entry, "Type") or ""
            ])
            rb_text = f"Option {idx+1}: {summary}"
            rb = tk.Radiobutton(options_frame,
                                text=rb_text,
                                variable=choice_var,
                                value=idx,
                                bg="light gray",
                                font=("Helvetica", 12),
                                anchor='w',
                                justify=tk.LEFT,
                                selectcolor='light blue')
            rb.pack(anchor="w", padx=10, pady=5)

        def on_save():
            selected_index = choice_var.get()
            nonlocal final_choice
            final_choice = word_entries[selected_index]
            final_win.destroy()

        save_btn = tk.Button(final_win, text="Save Choice",
                            command=on_save,
                            font=("Helvetica", 14, "bold"),
                            bg="#007acc", fg="white", padx=20, pady=10)
        save_btn.pack(pady=20)

        final_win.transient(self.root)
        final_win.grab_set()
        self.root.wait_window(final_win)
        return final_choice

    def prompt_save_results(self, new_entries, skip_copy=False):
        """
        For each verse in self.selected_verses, prompts the user to save new entries (accumulated from all words),
        checking for duplicates first. Then opens a modal prompt for assessment and saves the finalized data
        (including verse-level metadata) to an Excel file.
        """
        def _s(val):
            if val is None:
                return ""
            return unicodedata.normalize("NFC", str(val).strip())

        file_path = "1.2.1 assessment_data.xlsx"
        existing_data = self.load_existing_assessment_data(file_path)
        
        # Save the original accumulated_pankti so it can be restored later.
        original_accumulated_pankti = self.accumulated_pankti

        # Process each verse in the selected verses
        for verse in self.selected_verses:
            # Update the current verse for processing.
            self.accumulated_pankti = verse
            verse_norm = _s(verse)

            # reset repeat-note tracking for this verse
            self._repeat_note_shown = set()
            self._suppress_repeat_notes_for_verse = False

            # normalize for repeat-note consistency
            verse_key = unicodedata.normalize(
                "NFC", re.sub(r"\s+", " ", verse_norm.replace('ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¥Ãƒâ€šÃ‚Â¥', '').strip())
            )
            current_verse_words = verse_key.split()

            # Normalize tokens once and use everywhere to avoid Unicode/spacing mismatches
            normalized_tokens = [_s(w) for w in current_verse_words]
            normalized_words = set(normalized_tokens)

            from collections import Counter
            word_counts = Counter(normalized_tokens)

            # Filter new_entries to only those whose "Word" is present in the current verse.
            filtered_new_entries = [
                entry for entry in new_entries
                if _s(entry.get("Word")) in normalized_words and _s(entry.get("Verse")) == verse_norm
            ]

            duplicate_entries = []
            unique_entries = []

            # Duplicate check: for each filtered entry, compare against existing data.
            for new_entry in filtered_new_entries:
                new_word = self._norm_get(new_entry, "Word")
                new_ve = self._norm_get(new_entry, "\ufeffVowel Ending")
                new_num = self._norm_get(new_entry, "Number / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨")
                new_grammar = self._norm_get(new_entry, "Grammar / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¯ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£")
                new_gender = self._norm_get(new_entry, "Gender / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â")
                new_root = self._norm_get(new_entry, "Word Root")
                new_type = self._norm_get(new_entry, "Type")
                new_verse = self._norm_get(new_entry, "Verse")

                if any(
                    new_word == self._norm_get(existing_entry, "Word") and
                    new_ve == self._norm_get(existing_entry, "\ufeffVowel Ending") and
                    new_num == self._norm_get(existing_entry, "Number / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¨") and
                    new_grammar == self._norm_get(existing_entry, "Grammar / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚ÂµÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¯ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¾ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¢ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â£") and
                    new_gender == self._norm_get(existing_entry, "Gender / ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â²ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨Ãƒâ€šÃ‚Â¿ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â©Ãƒâ€šÃ‚Â°ÃƒÆ’Ã‚Â Ãƒâ€šÃ‚Â¨ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â") and
                    new_root == self._norm_get(existing_entry, "Word Root") and
                    new_type == self._norm_get(existing_entry, "Type") and
                    new_verse == self._norm_get(existing_entry, "Verse")
                    for existing_entry in existing_data.to_dict('records')
                ):
                    duplicate_entries.append(new_entry)
                else:
                    unique_entries.append(new_entry)

            if duplicate_entries:
                duplicate_message = "Some entries are already present:\n" + "\n".join(map(str, duplicate_entries))
                messagebox.showinfo("Duplicates Found", duplicate_message)

            if not skip_copy:
                self.prompt_copy_to_clipboard()

            if unique_entries:
                save = messagebox.askyesno(
                    "Save Results",
                    f"Would you like to save the new entries for the following verse?\n\n{verse_norm}"
                )
                if save:
                    # Open one assessment prompt for the current verse.
                    assessment_data = self.prompt_for_assessment_once()

                    # --- Extract verse metadata from candidate matches ---
                    verse_to_match = _s(self.accumulated_pankti)
                    candidate = None
                    if hasattr(self, 'candidate_matches') and hasattr(self, 'chosen_match') and self.chosen_match:
                        for cand in self.chosen_match:
                            if _s(cand.get("Verse")) == verse_to_match:
                                candidate = cand
                                break
                        if candidate is None:
                            candidate = self.chosen_match[0]
                        verse_metadata = {
                            "Verse": verse_to_match,
                            "S. No.": candidate.get("S. No.", ""),
                            "Verse No.": candidate.get("Verse No.", ""),
                            "Stanza No.": candidate.get("Stanza No.", ""),
                            "Text Set No.": candidate.get("Text Set No.", ""),
                            "Raag (Fixed)": candidate.get("Raag (Fixed)", ""),
                            "Sub-Raag": candidate.get("Sub-Raag", ""),
                            "Writer (Fixed)": candidate.get("Writer (Fixed)", ""),
                            "Verse Configuration (Optional)": candidate.get("Verse Configuration (Optional)", ""),
                            "Stanza Configuration (Optional)": candidate.get("Stanza Configuration (Optional)", ""),
                            "Bani Name": candidate.get("Bani Name", ""),
                            "Musical Note Configuration": candidate.get("Musical Note Configuration", ""),
                            "Special Type Demonstrator": candidate.get("Special Type Demonstrator", ""),
                            "Verse Type": candidate.get("Type", ""),
                            "Page Number": candidate.get("Page Number", "")
                        }
                    else:
                        verse_metadata = {}
                    # -------------------------------------------------------

                    # --- Finalize grammar options per word (handling repeated words by occurrence order via clustering) ---
                    # --- Group unique_entries by word ---
                    word_groups = {}
                    for entry in unique_entries:
                        word = _s(entry["Word"])
                        word_groups.setdefault(word, []).append(entry)
                        
                    final_entries = []
                    occurrence_mapping = {}  # Mapping from (word, occurrence_position) to list of entries (options)

                    # For each unique word in the current verse, partition its entries into clusters based on occurrence count.
                    for word in set(normalized_tokens):
                        count = word_counts[word]
                        if word not in word_groups:
                            continue
                        entries_list = word_groups[word]  # all unique entries for that word (in sequence)
                        n = len(entries_list)
                        k = count  # expected number of clusters
                        groups = []
                        start = 0
                        # Partition entries_list into k groups using a chunking method.
                        # If n isn't exactly divisible by k, distribute the remainder to the first few groups.
                        group_size = n // k
                        remainder = n % k
                        for i in range(k):
                            size = group_size + (1 if i < remainder else 0)
                            group = entries_list[start:start+size]
                            groups.append(group)
                            start += size
                        # Find the indices (positions) in normalized_tokens where this word occurs, in order.
                        occurrence_positions = [i for i, w in enumerate(normalized_tokens) if w == word]
                        for occ, pos in zip(range(k), occurrence_positions):
                            occurrence_mapping[(word, pos)] = groups[occ]

                    occurrence_counters = {}
                    # Now, iterate over normalized_tokens (which are in order) and process each occurrence.
                    for idx, word in enumerate(normalized_tokens):
                        occ_idx = occurrence_counters.get(word, 0)
                        occurrence_counters[word] = occ_idx + 1
                        key = (word, idx)  # Unique key for the occurrence at position idx.
                        entries = occurrence_mapping.get(key, [])
                        if not entries:
                            continue  # No entries for this occurrence.

                        if word_counts.get(word, 0) > 1 and occ_idx > 0:
                            if not getattr(self, "_use_inline_literal_banner", True):
                                self._maybe_show_repeat_important_note(word, occ_idx, verse_key)
                            else:
                                self._repeat_note_shown.add((verse_key, word, "second"))

                        # Remove duplicate entries within this occurrence group (local deduplication).
                        dedup_entries = []
                        seen = set()
                        for entry in entries:
                            entry_tuple = tuple(sorted(entry.items()))
                            if entry_tuple not in seen:
                                seen.add(entry_tuple)
                                dedup_entries.append(entry)
                        entries = dedup_entries

                        # If more than one unique option exists for this occurrence, prompt the user to choose.
                        if len(entries) > 1:
                            chosen_entry = self.prompt_for_final_grammar(entries)
                        else:
                            chosen_entry = entries[0]

                        # Capture the occurrence index.
                        chosen_entry['Word Index'] = idx

                        # Ensure self.accumulated_finalized_matches is long enough.
                        if len(self.accumulated_finalized_matches) <= idx:
                            self.accumulated_finalized_matches.extend([[]] * (idx - len(self.accumulated_finalized_matches) + 1))
                        self.accumulated_finalized_matches[idx] = [chosen_entry]
                        final_entries.append(chosen_entry)
                    # -------------------------------------------------------

                    # Now update each finalized entry with the assessment and verse metadata, then save.
                    for entry in final_entries:
                        entry.update(assessment_data)
                        entry.update(verse_metadata)
                        self.save_assessment_data(entry)
                    messagebox.showinfo("Saved", "Assessment data saved successfully for verse:\n" + verse_norm)

        # Restore the original accumulated_pankti after processing all verses.
        self.accumulated_pankti = original_accumulated_pankti

        if hasattr(self, 'copy_button') and self.copy_button.winfo_exists():
            self.copy_button.config(state=tk.NORMAL)

    def prompt_for_assessment_once(self):
        """Opens a modal prompt for the entire verse assessment and returns the collected data."""
        assessment_win = tk.Toplevel(self.root)
        assessment_win.title(f"Enter Translation Assessment for: '{self.accumulated_pankti}'")
        assessment_win.configure(bg='light gray')
        try:
            self._wm_for(assessment_win)
        except Exception:
            pass

        instruction_label = tk.Label(
            assessment_win, 
            text="Paste the analysis result for the entire verse below:",
            font=("Helvetica", 14), bg="light gray"
        )
        instruction_label.pack(pady=10)

        analysis_text = scrolledtext.ScrolledText(assessment_win, width=80, height=10,
                                                font=("Helvetica", 12), wrap=tk.WORD)
        analysis_text.pack(padx=20, pady=10)

        cb_frame = tk.Frame(assessment_win, bg="light gray")
        cb_frame.pack(pady=10)

        framework_var = tk.BooleanVar()
        explicit_var = tk.BooleanVar()

        framework_cb = tk.Checkbutton(cb_frame, text="Framework?", variable=framework_var,
                                    font=("Helvetica", 12), bg="light gray")
        framework_cb.pack(side=tk.LEFT, padx=10)

        explicit_cb = tk.Checkbutton(cb_frame, text="Explicit?", variable=explicit_var,
                                    font=("Helvetica", 12), bg="light gray")
        explicit_cb.pack(side=tk.LEFT, padx=10)

        assessment_data = {}

        def on_save():
            translation = analysis_text.get("1.0", tk.END).strip()
            if not translation:
                messagebox.showerror("Error", "Please paste the analysis result.")
                return
            assessment_data["Translation"] = translation
            assessment_data["Framework?"] = framework_var.get()
            assessment_data["Explicit?"] = explicit_var.get()
            assessment_win.destroy()

        save_btn = tk.Button(assessment_win, text="Save Assessment",
                            command=on_save, font=("Helvetica", 14, "bold"),
                            bg="#007acc", fg="white", padx=20, pady=10)
        save_btn.pack(pady=20)

        assessment_win.transient(self.root)
        assessment_win.grab_set()
        self.root.wait_window(assessment_win)

        return assessment_data

    def accumulate_pankti_data(self, pankti):
        self.accumulated_pankti = pankti

    def accumulate_meanings_data(self, meanings):
        """
        Accumulate the meanings for the current word, along with the word itself and its index.
        This creates a mapping for each word occurrence.
        """
        # Ensure the list is long enough for the current word index.
        while len(self.accumulated_meanings) <= self.current_word_index:
            self.accumulated_meanings.append({"word": None, "meanings": []})
        
        # Store the word if it hasn't been stored yet.
        if self.accumulated_meanings[self.current_word_index]["word"] is None:
            # Assuming self.pankti_words is defined and holds the current verse's words.
            self.accumulated_meanings[self.current_word_index]["word"] = self.pankti_words[self.current_word_index]
        
        # Update the meanings for the current word occurrence.
        self.accumulated_meanings[self.current_word_index]["meanings"] = meanings

    def accumulate_grammar_matches(self, matches):
        self.accumulated_grammar_matches.append(matches)

    def accumulate_finalized_matches(self, finalized_matches):
        # Ensure the list is long enough to hold the current index
        if len(self.accumulated_finalized_matches) <= self.current_word_index:
            # Expand the list to the required length with empty lists
            self.accumulated_finalized_matches.extend([[]] * (self.current_word_index - len(self.accumulated_finalized_matches) + 1))
        
        # Store the finalized matches at the correct index
        self.accumulated_finalized_matches[self.current_word_index] = finalized_matches

    def start_progress(self):
        self.progress_window = tk.Toplevel(self.root)
        self.progress_window.title("Please Wait...")
        self.progress_window.geometry("350x120")
        # Allow at least one resizable dimension so users can expand if needed
        try:
            self.progress_window.resizable(True, False)
        except Exception:
            pass
        self.progress_window.attributes("-topmost", True)
        self.progress_window.configure(bg="#f0f0f0")
        self.progress_window.attributes('-alpha', 0.0)  # Start fully transparent
        try:
            # Bind F11 toggle; do not auto-maximize progress window
            self._wm_for(self.progress_window)
        except Exception:
            pass

        # Center the progress window
        self.progress_window.update_idletasks()
        x = (self.progress_window.winfo_screenwidth() - self.progress_window.winfo_width()) // 2
        y = (self.progress_window.winfo_screenheight() - self.progress_window.winfo_height()) // 3
        self.progress_window.geometry(f"+{x}+{y}")

        # Fade-in effect
        def fade_in(window, alpha=0.0):
            alpha = round(alpha + 0.05, 2)
            if alpha <= 1.0:
                window.attributes('-alpha', alpha)
                window.after(30, lambda: fade_in(window, alpha))

        fade_in(self.progress_window)

        # Custom style
        style = ttk.Style(self.progress_window)
        style.theme_use('default')

        style.configure("Custom.Horizontal.TProgressbar",
                        troughcolor="#e0e0e0",
                        bordercolor="#e0e0e0",
                        background="#4a90e2",
                        lightcolor="#4a90e2",
                        darkcolor="#4a90e2",
                        thickness=20)

        label = ttk.Label(self.progress_window, text="Processing, please wait...", background="#f0f0f0", font=("Segoe UI", 10))
        label.pack(pady=(20, 5))

        self.progress_bar = ttk.Progressbar(self.progress_window,
                                            mode='indeterminate',
                                            style="Custom.Horizontal.TProgressbar")
        self.progress_bar.pack(padx=30, fill=tk.X)
        self.progress_bar.start(7)

        self.root.update_idletasks()

    def stop_progress(self):
        self.progress_bar.stop()
        self.progress_window.destroy()

    def handle_submitted_input(self, word):
        # Implement your logic for handling the submitted input here
        print(f"Handling submitted input for: {word}")
        # You can add your logic here, such as saving or processing the input further

    def filter_unique_matches(self, matches):
        """
        Eliminate duplicates based on the grammatical parts and return unique matches.

        Args:
        matches (list): The list of matches to filter.

        Returns:
        list: The list of unique matches.
        """
        # Sort matches by match_count first and then by match_percentage, both in descending order
        matches_sorted = sorted(matches, key=lambda x: (x[1], x[2]), reverse=True)

        unique_matches = []
        seen_grammar = set()
        for match in matches_sorted:
            grammar_info = match[0].split(" | ")[2:]  # Extract grammar info (e.g., Number, Gender, POS, etc.)
            grammar_tuple = tuple(grammar_info)
            if grammar_tuple not in seen_grammar:
                unique_matches.append(match)
                seen_grammar.add(grammar_tuple)

        # Accumulate Grammar Matches data
        self.accumulate_grammar_matches(unique_matches)
        return unique_matches

    def perform_dictionary_lookup(self, word, callback):
        """
        Runs the dictionary lookup (with its progress bar) in a separate thread.
        Once the lookup completes, the callback is called on the main thread with the lookup result.
        """
        def lookup_worker():
            # Call your lookup function (which already shows the progress bar)
            meanings = self.lookup_word_in_dictionary(word)
            # Schedule the callback on the main thread with the obtained meanings
            self.root.after(0, lambda: callback(meanings))
        threading.Thread(target=lookup_worker, daemon=True).start()

    def handle_lookup_results(self, matches, meanings):
        """
        Called when the dictionary lookup is finished.
        'matches' are those computed earlier and 'meanings' is the result from the lookup.
        This method updates the UI by calling show_matches.
        """
        print(f"Lookup completed for {self.pankti_words[self.current_word_index]}. Meanings: {meanings}")
        self.show_matches(matches, self.current_pankti, meanings)

    def calculate_match_metrics(self, word, vowel_ending):
        """
        Calculates the number of matching characters from the end of vowel_ending and word, 
        and the percentage of matching characters with respect to the matched part of vowel_ending.
        
        Parameters:
        word (str): The word to be compared.
        vowel_ending (str): The vowel ending to be compared, possibly containing multiple parts.
        
        Returns:
        tuple: A tuple containing:
            - match_count (int): Number of matching characters from the end.
            - match_percentage (float): Percentage of matching characters based on the matched part of vowel_ending.
        """
        word_chars = list(word)  # Convert word to a list of characters
        vowel_parts = vowel_ending.split()  # Split vowel ending into parts

        total_match_count = 0
        max_match_percentage = 0.0

        # Iterate through each part of the vowel ending
        for part in vowel_parts:
            part_chars = list(part)  # Convert each part to a list of characters

            match_count = 0

            # Reverse iterate through both word and part to compare characters from the end
            for i in range(1, min(len(word_chars), len(part_chars)) + 1):
                if word_chars[-i] == part_chars[-i]:
                    match_count += 1
                else:
                    break  # Stop when characters no longer match

            if match_count > 0:
                # Calculate match percentage based on the length of the matched part
                match_percentage = (match_count / len(part_chars)) * 100
                total_match_count += match_count
                
                # Track the highest match percentage found
                if match_percentage > max_match_percentage:
                    max_match_percentage = match_percentage

            # Stop if there was any match in the current part
            if match_count > 0:
                break

        return total_match_count, max_match_percentage


if __name__ == "__main__":
    root = tk.Tk()
    app = GrammarApp(root)
    root.mainloop()
    def _do_prompt_whats_new_fixed(self, state: dict):
        """Prompt the user about recent UI changes with ASCII-safe quotes."""
        try:
            if messagebox.askyesno(
                "What's New",
                (
                    "We've improved verse selection cards: centered layout, equal column widths, ",
                    "and radios no longer overlap text. View details now?"
                ),
            ):
                self.show_whats_new()
        finally:
            try:
                state["last_whats_new"] = WHATS_NEW_ID
                self._save_state(state)
            except Exception:
                pass
        return


